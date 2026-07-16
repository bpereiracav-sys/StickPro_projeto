from fastapi import FastAPI, APIRouter, HTTPException, Depends, UploadFile, File, Form, Query
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from fastapi.staticfiles import StaticFiles
from fastapi.responses import StreamingResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from enum import Enum
from core.database import db, client
import os
import io
import logging
import asyncio
import resend
import secrets
import hashlib
from pathlib import Path
from pydantic import BaseModel, Field, EmailStr, ConfigDict, field_validator
from typing import List, Optional, Literal, Dict, Any
import uuid
from datetime import datetime, timezone, timedelta
import jwt
import bcrypt
import base64
import shutil
import httpx
from bs4 import BeautifulSoup
import re
import pandas as pd
from services.communication_service import CommunicationService
from services.recipient_service import RecipientService

# Import RBAC permissions module
from permissions import PermissionChecker, get_permission_checker, Role, ROLE_PERMISSIONS

# Phase E2: new modular activation email helper (services.emails-backed).
# Imported lazily-friendly so app boot fails loudly if services/ is missing.
from services.activation_emails import send_activation_email
from services.family_invitation_emails import send_family_invitation_email

# Phase E3: password reset email helper.
from services.password_reset_emails import send_password_reset_email

# P0 startup hardening: validate email-related env vars at module import.
# In production, this raises EmailConfigError and aborts boot when any of
# RESEND_API_KEY / SENDER_EMAIL / FRONTEND_URL is missing. In development
# and test environments, validate_email_config() only logs a warning and
# returns — outbound mail falls back to dry-run mode via services.emails.
from services.emails import validate_email_config as _validate_email_config
_validate_email_config()

# Phase S1: Stripe configuration validation. Mirrors the email validator —
# in production, missing required Stripe env vars abort startup; in dev
# they only emit a warning so pods that don't need Stripe keep booting.
# Live keys outside production are always refused.
from services.stripe_config import validate_stripe_config as _validate_stripe_config
_validate_stripe_config()

ROOT_DIR = Path(__file__).parent
UPLOADS_DIR = ROOT_DIR / "uploads"
UPLOADS_DIR.mkdir(exist_ok=True)
load_dotenv(ROOT_DIR / '.env')

# JWT Configuration
# Security: JWT_SECRET MUST be set in production. In development/testing only,
# an insecure fallback is used so the app can boot — never deploy this way.
ENVIRONMENT = os.environ.get('ENVIRONMENT', 'development').lower()
JWT_SECRET = os.environ.get('JWT_SECRET')
if not JWT_SECRET:
    if ENVIRONMENT == 'production':
        raise RuntimeError(
            "JWT_SECRET environment variable is required in production. "
            "Set JWT_SECRET to a strong random value (min 32 chars) before starting the app."
        )
    JWT_SECRET = 'dev-only-insecure-jwt-secret-change-me'
    logging.getLogger(__name__).warning(
        "JWT_SECRET not set - using insecure development fallback. "
        "DO NOT use this in production. Set JWT_SECRET in backend/.env."
    )
JWT_ALGORITHM = "HS256"
JWT_EXPIRATION_HOURS = 24

# Create the main app
app = FastAPI(title="Roller Hockey Hub API")

# Mount uploads folder for static files - use /api/uploads so it's accessible via proxy
app.mount("/api/uploads", StaticFiles(directory=str(UPLOADS_DIR)), name="uploads")

# Create a router with the /api prefix
api_router = APIRouter(prefix="/api")

# Security
security = HTTPBearer()

# Communication Service
communication_service = CommunicationService(db)
recipient_service = RecipientService(db)

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# ==================== MODELS ====================

UserRole = Literal["admin", "gestor_desportivo", "treinador", "treinador_adjunto", "delegado", "jogador", "responsavel"]
EventType = Literal["treino", "jogo_campeonato", "jogo_amigavel", "torneio", "evento_administrativo","birthday","outro"]
AttendanceStatus = Literal["confirmado", "ausente", "pendente", "faltou_sem_aviso"]
MatchLocation = Literal["casa", "fora", "neutro"]
PlayerPosition = Literal["GR", "JC"]
ChampionshipFormat = Literal["5x5", "3x3"]
ConvocationType = Literal["automatica", "manual"]
EquipmentSize = str  # Free text: S/M/L/XL or 8/10/12 etc

# Admin-level roles (have full permissions)
ADMIN_ROLES = ["admin", "gestor_desportivo"]

def is_admin_role(role: str) -> bool:
    """Check if a role has admin-level permissions"""
    return role in ADMIN_ROLES

# ==================== COMPETITION PERMISSION SETTINGS ====================

DEFAULT_COMPETITION_PERMISSIONS = {
    "coach_can_create_competitions": False,
    "coach_can_archive_competitions": False,
    "coach_can_edit_standings": False,

    "delegate_can_create_competitions": False,
    "delegate_can_edit_competitions": False,
    "delegate_can_archive_competitions": False,
    "delegate_can_create_games": False,
    "delegate_can_edit_games": False,
    "delegate_can_edit_results": False,
    "delegate_can_import_gamesheet": False,
    "delegate_can_edit_statistics": False,
    "delegate_can_edit_standings": False,
}


class CompetitionPermissionSettingsUpdate(BaseModel):
    coach_can_create_competitions: Optional[bool] = None
    coach_can_archive_competitions: Optional[bool] = None
    coach_can_edit_standings: Optional[bool] = None

    delegate_can_create_competitions: Optional[bool] = None
    delegate_can_edit_competitions: Optional[bool] = None
    delegate_can_archive_competitions: Optional[bool] = None
    delegate_can_create_games: Optional[bool] = None
    delegate_can_edit_games: Optional[bool] = None
    delegate_can_edit_results: Optional[bool] = None
    delegate_can_import_gamesheet: Optional[bool] = None
    delegate_can_edit_statistics: Optional[bool] = None
    delegate_can_edit_standings: Optional[bool] = None


async def get_current_club_for_user(current_user: dict) -> Optional[dict]:
    club_id = current_user.get("club_id")

    if club_id:
        club = await db.clubs.find_one({"id": club_id}, {"_id": 0})
        if club:
            return club

    return await db.clubs.find_one({}, {"_id": 0})


async def get_competition_permission_settings(current_user: dict) -> dict:
    club = await get_current_club_for_user(current_user)

    settings = DEFAULT_COMPETITION_PERMISSIONS.copy()

    if club:
        stored = club.get("competition_permissions") or {}
        if isinstance(stored, dict):
            for key in DEFAULT_COMPETITION_PERMISSIONS.keys():
                if key in stored:
                    settings[key] = bool(stored[key])

    return settings


def get_user_competition_role(current_user: dict) -> str:
    return current_user.get("role") or "jogador"


def user_has_team_access(current_user: dict, team_id: Optional[str]) -> bool:
    if not team_id:
        return False

    if is_admin_role(current_user.get("role")):
        return True

    return team_id in (current_user.get("team_ids") or [])


async def can_view_competition(current_user: dict, championship: dict) -> bool:
    if is_admin_role(current_user.get("role")):
        return True

    return user_has_team_access(current_user, championship.get("team_id"))


async def can_create_competition(current_user: dict, team_id: Optional[str]) -> bool:
    role = get_user_competition_role(current_user)

    if is_admin_role(role):
        return True

    if role in ["jogador", "responsavel"]:
        return False

    if not user_has_team_access(current_user, team_id):
        return False

    settings = await get_competition_permission_settings(current_user)

    if role in ["treinador", "treinador_adjunto"]:
        return settings["coach_can_create_competitions"]

    if role == "delegado":
        return settings["delegate_can_create_competitions"]

    return False


async def can_edit_competition(current_user: dict, championship: dict) -> bool:
    role = get_user_competition_role(current_user)

    if is_admin_role(role):
        return True

    if role in ["jogador", "responsavel"]:
        return False

    if not user_has_team_access(current_user, championship.get("team_id")):
        return False

    settings = await get_competition_permission_settings(current_user)

    if role in ["treinador", "treinador_adjunto"]:
        return True

    if role == "delegado":
        return settings["delegate_can_edit_competitions"]

    return False


async def can_archive_competition(current_user: dict, championship: dict) -> bool:
    role = get_user_competition_role(current_user)

    if is_admin_role(role):
        return True

    if role in ["jogador", "responsavel"]:
        return False

    if not user_has_team_access(current_user, championship.get("team_id")):
        return False

    settings = await get_competition_permission_settings(current_user)

    if role in ["treinador", "treinador_adjunto"]:
        return settings["coach_can_archive_competitions"]

    if role == "delegado":
        return settings["delegate_can_archive_competitions"]

    return False


async def can_create_competition_game(current_user: dict, championship: dict) -> bool:
    role = get_user_competition_role(current_user)

    if is_admin_role(role):
        return True

    if role in ["jogador", "responsavel"]:
        return False

    if not user_has_team_access(current_user, championship.get("team_id")):
        return False

    settings = await get_competition_permission_settings(current_user)

    if role in ["treinador", "treinador_adjunto"]:
        return True

    if role == "delegado":
        return settings["delegate_can_create_games"]

    return False


async def can_edit_competition_game(current_user: dict, championship: dict) -> bool:
    role = get_user_competition_role(current_user)

    if is_admin_role(role):
        return True

    if role in ["jogador", "responsavel"]:
        return False

    if not user_has_team_access(current_user, championship.get("team_id")):
        return False

    settings = await get_competition_permission_settings(current_user)

    if role in ["treinador", "treinador_adjunto"]:
        return True

    if role == "delegado":
        return settings["delegate_can_edit_games"]

    return False


async def can_edit_competition_result(current_user: dict, championship: dict) -> bool:
    role = get_user_competition_role(current_user)

    if is_admin_role(role):
        return True

    if role in ["jogador", "responsavel"]:
        return False

    if not user_has_team_access(current_user, championship.get("team_id")):
        return False

    settings = await get_competition_permission_settings(current_user)

    if role in ["treinador", "treinador_adjunto"]:
        return True

    if role == "delegado":
        return settings["delegate_can_edit_results"]

    return False


async def can_import_competition_gamesheet(current_user: dict, championship: dict) -> bool:
    role = get_user_competition_role(current_user)

    if is_admin_role(role):
        return True

    if role in ["jogador", "responsavel"]:
        return False

    if not user_has_team_access(current_user, championship.get("team_id")):
        return False

    settings = await get_competition_permission_settings(current_user)

    if role in ["treinador", "treinador_adjunto"]:
        return True

    if role == "delegado":
        return settings["delegate_can_import_gamesheet"]

    return False


async def can_edit_competition_statistics(current_user: dict, championship: dict) -> bool:
    role = get_user_competition_role(current_user)

    if is_admin_role(role):
        return True

    if role in ["jogador", "responsavel"]:
        return False

    if not user_has_team_access(current_user, championship.get("team_id")):
        return False

    settings = await get_competition_permission_settings(current_user)

    if role in ["treinador", "treinador_adjunto"]:
        return True

    if role == "delegado":
        return settings["delegate_can_edit_statistics"]

    return False


async def can_edit_competition_standings(current_user: dict, championship: dict) -> bool:
    role = get_user_competition_role(current_user)

    if is_admin_role(role):
        return True

    if role in ["jogador", "responsavel"]:
        return False

    if not user_has_team_access(current_user, championship.get("team_id")):
        return False

    settings = await get_competition_permission_settings(current_user)

    if role in ["treinador", "treinador_adjunto"]:
        return settings["coach_can_edit_standings"]

    if role == "delegado":
        return settings["delegate_can_edit_standings"]

    return False

async def enrich_championship_for_response(
    championship: dict,
    current_user: dict
) -> dict:
    enriched = dict(championship)

    team = None
    team_id = championship.get("team_id")

    if team_id:
        team = await db.teams.find_one(
            {"id": team_id},
            {"_id": 0, "id": 1, "name": 1, "category": 1, "season": 1}
        )

    enriched["team"] = team
    enriched["team_name"] = team.get("name") if team else None
    enriched["team_category"] = team.get("category") if team else None
    enriched["team_season"] = team.get("season") if team else None

    enriched["permissions"] = {
        "can_view": await can_view_competition(current_user, championship),
        "can_edit": await can_edit_competition(current_user, championship),
        "can_archive": await can_archive_competition(current_user, championship),
        "can_create_games": await can_create_competition_game(current_user, championship),
        "can_edit_games": await can_edit_competition_game(current_user, championship),
        "can_edit_results": await can_edit_competition_result(current_user, championship),
        "can_import_gamesheet": await can_import_competition_gamesheet(current_user, championship),
        "can_edit_statistics": await can_edit_competition_statistics(current_user, championship),
        "can_edit_standings": await can_edit_competition_standings(current_user, championship),
    }

    return enriched

# ==================== PERMISSION SYSTEM ====================

# Default permissions by role
DEFAULT_PERMISSIONS = {
    "admin": {
        "can_view_all": True,
        "can_edit_all": True,
        "can_manage_permissions": True,
        "can_view_family_data": True,
        "can_edit_family_data": True,
        "can_manage_teams": True,
        "can_manage_championships": True,
        "can_manage_events": True,
        "can_manage_members": True,
    },
    "gestor_desportivo": {
        "can_view_all": True,
        "can_edit_all": True,
        "can_manage_permissions": True,
        "can_view_family_data": True,
        "can_edit_family_data": True,
        "can_manage_teams": True,
        "can_manage_championships": True,
        "can_manage_events": True,
        "can_manage_members": True,
    },
    "treinador": {
        "can_view_all": False,
        "can_edit_all": False,
        "can_manage_permissions": False,
        "can_view_family_data": False,
        "can_edit_family_data": False,
        "can_manage_teams": True,
        "can_manage_championships": True,
        "can_manage_events": True,
        "can_manage_members": True,
    },
    "treinador_adjunto": {
        "can_view_all": False,
        "can_edit_all": False,
        "can_manage_permissions": False,
        "can_view_family_data": False,
        "can_edit_family_data": False,
        "can_manage_teams": True,
        "can_manage_championships": True,
        "can_manage_events": True,
        "can_manage_members": True,
    },
    "delegado": {
        "can_view_all": False,
        "can_edit_all": False,
        "can_manage_permissions": False,
        "can_view_family_data": False,
        "can_edit_family_data": False,
        "can_manage_teams": True,
        "can_manage_championships": True,
        "can_manage_events": True,
        "can_manage_members": True,
    },
    "jogador": {
        "can_view_all": False,
        "can_edit_all": False,
        "can_manage_permissions": False,
        "can_view_family_data": False,
        "can_edit_family_data": False,
        "can_manage_teams": False,
        "can_manage_championships": False,
        "can_manage_events": False,
        "can_manage_members": False,
        "can_edit_own_profile": True,
    },
    "responsavel": {
        "can_view_all": False,
        "can_edit_all": False,
        "can_manage_permissions": False,
        "can_view_family_data": True,
        "can_edit_family_data": True,
        "can_manage_teams": False,
        "can_manage_championships": False,
        "can_manage_events": False,
        "can_manage_members": False,
    }
}

class Permissions(BaseModel):
    can_view_all: bool = False
    can_edit_all: bool = False
    can_manage_permissions: bool = False
    can_view_family_data: bool = False
    can_edit_family_data: bool = False
    can_manage_teams: bool = False
    can_manage_championships: bool = False
    can_manage_events: bool = False
    can_manage_members: bool = False
    can_edit_own_profile: bool = True

class UserCreate(BaseModel):
    email: EmailStr
    password: str
    name: str
    surname: Optional[str] = None
    role: UserRole = "jogador"
    phone: Optional[str] = None
    additional_roles: List[UserRole] = []

class UserLogin(BaseModel):
    email: EmailStr
    password: str

# Family member model
class FamilyMember(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    first_name: str
    surname: str
    email: Optional[EmailStr] = None
    phone: Optional[str] = None
    relationship: str = "pai"  # pai, mae, outro

class LinkPlayerRequest(BaseModel):
    player_id: str
    
# Extended User Profile
class UserProfile(BaseModel):
    # Identity
    photo_url: Optional[str] = None
    first_name: Optional[str] = None
    surname: Optional[str] = None
    nickname: Optional[str] = None
    birth_date: Optional[str] = None  # ISO date string
    gender: Optional[str] = None  # masculino, feminino
    nationality: Optional[str] = None  # NEW: Nationality
    fpp_license: Optional[str] = None  # Federação Portuguesa de Patinagem
    
    # Family members
    family_members: List[FamilyMember] = []
    
    # Biometric data
    weight: Optional[float] = None  # kg
    height: Optional[float] = None  # cm
    shoe_size: Optional[str] = None  # Free text
    
    # Sports info
    year_joined_club: Optional[int] = None
    fpp_number: Optional[str] = None
    function: Optional[UserRole] = None  # jogador, treinador, etc
    position: Optional[str] = None  # GR/JC - changed to str for flexibility
    jersey_number: Optional[str] = None  # Changed to str for flexibility
    
    # Equipment sizes (free text)
    training_kit_size: Optional[str] = None
    tracksuit_size: Optional[str] = None
    polo_size: Optional[str] = None
    training_sock_size: Optional[str] = None

    model_config = ConfigDict(extra="ignore")

    # Validator to handle empty strings and type coercion
    @field_validator('weight', 'height', mode='before')
    @classmethod
    def empty_str_to_none_float(cls, v):
        if v == '' or v is None:
            return None
        try:
            return float(v)
        except (ValueError, TypeError):
            return None

    @field_validator('year_joined_club', mode='before')
    @classmethod
    def empty_str_to_none_int(cls, v):
        if v == '' or v is None:
            return None
        try:
            return int(v)
        except (ValueError, TypeError):
            return None

    @field_validator('jersey_number', 'position', mode='before')
    @classmethod
    def coerce_to_string(cls, v):
        if v is None or v == '':
            return None
        return str(v)

class User(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    email: EmailStr
    name: str
    surname: Optional[str] = None
    role: UserRole  # Primary/global role (admin stays here)
    additional_roles: List[UserRole] = []
    phone: Optional[str] = None
    avatar_url: Optional[str] = None
    team_ids: List[str] = []
    team_roles: Dict[str, UserRole] = {}  # NEW: team_id -> role mapping
    club_id: Optional[str] = None  # Club association
    associated_accounts: List[str] = []
    parent_account_id: Optional[str] = None
    linked_player_id: Optional[str] = None  # For family_members: linked player's ID
    linked_player_ids: List[str] = []  # NEW: Multiple linked players for family accounts
    
    # Extended profile data
    profile: Optional[UserProfile] = None
    
    # Custom permissions (if admin has modified defaults)
    custom_permissions: Optional[Dict[str, bool]] = None

    # Phase O1 — Admin onboarding wizard. Null until the admin (or
    # gestor_desportivo) finishes the onboarding flow at least once.
    onboarding_completed_at: Optional[datetime] = None

    # Phase O2 — Per-step onboarding state so the admin can resume the
    # wizard between sessions. None until the wizard is touched the first
    # time. Keys:
    #   current_step:     int (default 0)
    #   completed_steps:  List[str] (e.g. ["club", "season"])
    #   club_id:          str | None (set after Club step)
    #   season_id:        str | None (set after Season step)
    onboarding_state: Optional[Dict[str, Any]] = None

    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class UserResponse(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    email: str
    name: str
    surname: Optional[str] = None
    role: UserRole
    additional_roles: List[UserRole] = []
    phone: Optional[str] = None
    avatar_url: Optional[str] = None
    team_ids: List[str] = []
    team_roles: Dict[str, UserRole] = {}  # NEW
    club_id: Optional[str] = None
    associated_accounts: List[str] = []
    linked_player_id: Optional[str] = None
    linked_player_ids: List[str] = []  # NEW
    profile: Optional[UserProfile] = None
    permissions: Optional[Dict[str, bool]] = None
    # Phase O1 — surfaced so the frontend can route admins to /onboarding
    # on first login without an extra request.
    onboarding_completed_at: Optional[datetime] = None

class AcceptFamilyInviteRequest(BaseModel):
    token: str
    name: str
    password: str

class AssociateAccountRequest(BaseModel):
    child_user_id: str
    relationship: str = "filho/a"  # filho/a, atleta, etc.

class ActiveProfileRequest(BaseModel):
    profile_type: str  # "self" or "associated"
    associated_user_id: Optional[str] = None  # If viewing as associated account
    active_role: Optional[UserRole] = None  # Which role to use
    team_id: Optional[str] = None  # Which team context

class AuthResponse(BaseModel):
    token: str
    user: UserResponse
    available_profiles: List[dict] = []  # List of profiles user can access

class TeamCreate(BaseModel):
    name: str
    category: str
    season: str
    photo_url: Optional[str] = None
    club_id: Optional[str] = None

class TeamUpdate(BaseModel):
    name: Optional[str] = None
    category: Optional[str] = None
    season: Optional[str] = None
    photo_url: Optional[str] = None

class Team(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    category: str
    season: str
    photo_url: Optional[str] = None
    coach_ids: List[str] = []
    assistant_coach_ids: List[str] = []  # Treinador adjunto
    delegate_ids: List[str] = []
    player_ids: List[str] = []
    club_id: Optional[str] = None  # Reference to club
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

# Club Model
class ClubCreate(BaseModel):
    name: str
    acronym: Optional[str] = None
    logo_url: Optional[str] = None
    address: Optional[str] = None
    city: Optional[str] = None
    country: str = "Portugal"
    founded_year: Optional[int] = None
    website: Optional[str] = None
    email: Optional[str] = None
    phone: Optional[str] = None

class Club(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    acronym: Optional[str] = None
    logo_url: Optional[str] = None
    address: Optional[str] = None
    city: Optional[str] = None
    country: str = "Portugal"
    founded_year: Optional[int] = None
    website: Optional[str] = None
    email: Optional[str] = None
    phone: Optional[str] = None
    # Pavilhão/Arena do clube
    venue_name: Optional[str] = None  # Nome do pavilhão
    venue_location: Optional[str] = None  # Localização/morada do pavilhão
    admin_ids: List[str] = []  # Users with admin access to this club
    # Theme colors
    primary_color: Optional[str] = "#006D5B"  # Default teal
    secondary_color: Optional[str] = "#FFD700"  # Default gold
    accent_color: Optional[str] = "#1a1a2e"  # Default dark
    theme_mode: Optional[str] = "light"  # light or dark
    # Timezone
    timezone: Optional[str] = "Europe/Lisbon"  # Default timezone
    # Sidebar accent color for active item text
    sidebar_accent_color: Optional[str] = "#22d3ee"  # Default cyan
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class ClubUpdate(BaseModel):
    name: Optional[str] = None
    logo_url: Optional[str] = None
    address: Optional[str] = None
    city: Optional[str] = None
    country: Optional[str] = None
    founded_year: Optional[int] = None
    website: Optional[str] = None
    email: Optional[str] = None
    phone: Optional[str] = None
    venue_name: Optional[str] = None
    venue_location: Optional[str] = None
    primary_color: Optional[str] = None
    secondary_color: Optional[str] = None
    accent_color: Optional[str] = None
    theme_mode: Optional[str] = None
    timezone: Optional[str] = None
    sidebar_accent_color: Optional[str] = None

# Season Models
class SeasonCreate(BaseModel):
    name: str  # e.g., "2024/2025"
    start_date: str  # ISO date string
    end_date: str  # ISO date string
    is_active: bool = False

class Season(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    club_id: str
    name: str
    start_date: str
    end_date: str
    is_active: bool = False
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class SeasonUpdate(BaseModel):
    name: Optional[str] = None
    start_date: Optional[str] = None
    end_date: Optional[str] = None
    is_active: Optional[bool] = None

# Subscription Models
class SubscriptionPlan(str, Enum):
    standard = "standard"
    plus = "plus"

class SubscriptionStatus(str, Enum):
    active = "active"
    expired = "expired"
    cancelled = "cancelled"
    pending = "pending"

class PaymentMethod(str, Enum):
    credit_card = "credit_card"
    bank_transfer = "bank_transfer"

class InvoiceStatus(str, Enum):
    pending = "pending"
    paid = "paid"
    overdue = "overdue"
    cancelled = "cancelled"

class SubscriptionCreate(BaseModel):
    plan_type: SubscriptionPlan = SubscriptionPlan.standard
    payment_method: PaymentMethod = PaymentMethod.bank_transfer

class Subscription(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    club_id: str
    plan_type: str = "standard"  # "standard" or "plus"
    start_date: str  # ISO date string
    end_date: str  # ISO date string (1 year after start)
    status: str = "active"  # "active", "expired", "cancelled", "pending"
    payment_method: str = "bank_transfer"  # "credit_card" or "bank_transfer"
    member_count: int = 0  # Number of subscribed members
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class SubscriptionUpdate(BaseModel):
    plan_type: Optional[str] = None
    payment_method: Optional[str] = None
    status: Optional[str] = None
    member_count: Optional[int] = None

class SubscriptionInvoice(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    subscription_id: str
    club_id: str
    invoice_number: str  # e.g., "INV-2024-001"
    start_date: str  # Billing period start
    end_date: str  # Billing period end
    paying_members: int
    price_per_member: float  # e.g., 2.50
    total_due: float
    total_paid: float = 0.0
    status: str = "pending"  # "pending", "paid", "overdue", "cancelled"
    file_url: Optional[str] = None  # PDF invoice link
    paid_at: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class SubscriptionInvoiceCreate(BaseModel):
    start_date: str
    end_date: str
    paying_members: int
    price_per_member: float
    total_due: float

# Library Models
class LibraryItemType(str, Enum):
    pdf = "pdf"
    link = "link"
    video = "video"

class LibraryItemCreate(BaseModel):
    title: str
    description: Optional[str] = None
    item_type: LibraryItemType
    url: str  # For links/videos or file path for PDFs
    category: Optional[str] = None  # e.g., "Regras", "Táticas", "Treino"
    tags: List[str] = []

class LibraryItem(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    title: str
    description: Optional[str] = None
    item_type: LibraryItemType
    url: str
    category: Optional[str] = None
    tags: List[str] = []
    thumbnail_url: Optional[str] = None
    created_by: str
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

# Match Lineup Models (for coaches to manage game periods)
class LineupPosition(BaseModel):
    position: str  # "guarda_redes", "defesa_esquerda", "defesa_direita", "avancado_esquerda", "avancado_direita"
    player_id: Optional[str] = None
    player_name: Optional[str] = None

class MatchPeriod(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str  # "1ª Parte", "2ª Parte", "Prolongamento", etc.
    order: int
    positions: List[LineupPosition] = []
    notes: Optional[str] = None

# Lineup visibility options
LineupVisibility = Literal["coach_only", "assistant", "delegate", "assistant_and_delegate"]

class MatchLineupCreate(BaseModel):
    match_id: str
    periods: List[MatchPeriod] = []
    visibility: LineupVisibility = "coach_only"

class MatchLineup(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    match_id: str
    team_id: str
    periods: List[dict] = []
    visibility: LineupVisibility = "coach_only"
    created_by: str
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

# Competition Team Models (equipas participantes nas competições)
class TeamKitColors(BaseModel):
    primary_shirt: Optional[str] = None
    secondary_shirt: Optional[str] = None
    primary_shorts: Optional[str] = None
    secondary_shorts: Optional[str] = None
    primary_socks: Optional[str] = None
    secondary_socks: Optional[str] = None

class CompetitionTeamCreate(BaseModel):
    name: str
    pavilion_name: Optional[str] = None
    pavilion_address: Optional[str] = None
    field_player_kit: Optional[TeamKitColors] = None
    goalkeeper_kit: Optional[TeamKitColors] = None

class CompetitionTeam(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    championship_id: str
    name: str
    pavilion_name: Optional[str] = None
    pavilion_address: Optional[str] = None
    field_player_kit: Optional[dict] = None
    goalkeeper_kit: Optional[dict] = None
    created_by: str
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

# AI Chat Models
class AIChatMessage(BaseModel):
    role: str  # "user" or "assistant"
    content: str
    timestamp: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class AIChatRequest(BaseModel):
    message: str
    session_id: Optional[str] = None
    language: Optional[str] = "pt"

# Championship Models
class CompetitionRules(BaseModel):
    model_config = ConfigDict(extra="ignore")

    # normal | rtp | apl_cup | custom
    game_format: str = "normal"

    # número de segmentos/semi-partes
    segments_count: int = 2

    # jogadores por segmento
    players_per_segment: int = 5

    # existe obrigatoriedade de participação?
    mandatory_participation: bool = False

    # segmentos onde todos têm de participar
    mandatory_segments: List[int] = Field(default_factory=list)

    # segmentos livres
    free_segments: List[int] = Field(default_factory=list)

    # validação automática pelo assistente
    automatic_validation: bool = True


class ChampionshipCreate(BaseModel):
    name: str
    season: str
    team_id: str
    description: Optional[str] = None
    format: ChampionshipFormat = "5x5"
    location: Optional[str] = None
    convocation_type: ConvocationType = "manual"
    age_group: Optional[str] = None
    competition_type: Optional[str] = "campeonato_distrital"

class Championship(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    season: str
    team_id: str
    description: Optional[str] = None
    format: ChampionshipFormat = "5x5"
    location: Optional[str] = None
    convocation_type: ConvocationType = "manual"
    age_group: Optional[str] = None
    competition_type: Optional[str] = "campeonato_distrital"
    competition_rules: CompetitionRules = Field(
        default_factory=CompetitionRules
    )
    participating_teams: List[str] = []
    created_by: str
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class ChampionshipMatchCreate(BaseModel):
    championship_id: str
    home_team: Optional[str] = None  # Nome da equipa da casa (pode ser qualquer equipa)
    away_team: Optional[str] = None
    club_side: Optional[str] = None  # home | away | neutral
    official_match_url: Optional[str] = None    
    opponent_team: str  # Nome da equipa visitante
    match_date: datetime
    match_time: Optional[str] = None  # Hora do jogo (HH:MM)
    location: MatchLocation
    venue: Optional[str] = None
    is_club_match: bool = True  # Se é jogo da equipa do clube ou jogo entre outras equipas
    bonus_points: int = 0
    penalty_points: int = 0
    matchday: Optional[int] = None  # Número da jornada

class ChampionshipMatch(BaseModel):
    model_config = ConfigDict(extra="ignore")

    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    championship_id: str
    team_id: str

    home_team: Optional[str] = None
    away_team: Optional[str] = None
    club_side: Optional[str] = None  # home | away | neutral
    opponent_team: str

    match_date: datetime
    match_time: Optional[str] = None
    location: MatchLocation
    venue: Optional[str] = None
    matchday: Optional[int] = None

    home_score: Optional[int] = None
    away_score: Optional[int] = None
    is_completed: bool = False
    is_club_match: bool = True

    # Sprint 2.7A — estado operacional do jogo
    match_status: str = "scheduled"
    match_status_updated_at: Optional[datetime] = None
    match_status_updated_by: Optional[str] = None

    bonus_points: int = 0
    penalty_points: int = 0

    official_match_url: Optional[str] = None
    gamesheet_url: Optional[str] = None

    # Sprint 2.0C — origem, sincronização e validação
    source: str = "manual"  # manual | apl | fpp | official
    source_url: Optional[str] = None
    external_match_id: Optional[str] = None

    is_verified: bool = False
    verified_by: Optional[str] = None
    verified_at: Optional[datetime] = None

    last_synced_at: Optional[datetime] = None
    last_sync_error: Optional[str] = None

    sync_status: str = "manual"
    # manual | pending | syncing | synced | error
    # podem continuar a existir valores antigos:
    # imported | conflict

    created_at: datetime = Field(
        default_factory=lambda: datetime.now(timezone.utc)
    )

class MatchResultUpdate(BaseModel):
    home_score: int
    away_score: int
    bonus_points: int = 0
    penalty_points: int = 0

# Player Match Stats - comprehensive stats per match
class PlayerMatchStatsCreate(BaseModel):
    match_id: str
    player_id: str
    position: Optional[PlayerPosition] = None  # Made optional for stats import
    started_match: bool = False
    minutes_played: int = 0
    goals: int = 0
    own_goals: int = 0
    assists: int = 0
    penalties_scored: int = 0
    penalties_missed: int = 0
    penalties_saved: int = 0
    penalties_conceded: int = 0
    free_kicks_scored: int = 0
    free_kicks_missed: int = 0
    free_kicks_saved: int = 0
    free_kicks_conceded: int = 0
    direct_free_kicks: int = 0
    saves: int = 0
    blue_cards: int = 0
    yellow_cards: int = 0
    white_cards: int = 0
    red_cards: int = 0

class PlayerMatchStats(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    match_id: str
    player_id: str
    team_id: str
    championship_id: str
    position: Optional[PlayerPosition] = None
    started_match: bool = False
    minutes_played: int = 0
    goals: int = 0
    own_goals: int = 0
    assists: int = 0
    penalties_scored: int = 0
    penalties_missed: int = 0
    penalties_saved: int = 0
    penalties_conceded: int = 0
    free_kicks_scored: int = 0
    free_kicks_missed: int = 0
    free_kicks_saved: int = 0
    free_kicks_conceded: int = 0
    direct_free_kicks: int = 0
    saves: int = 0
    blue_cards: int = 0
    yellow_cards: int = 0
    white_cards: int = 0
    red_cards: int = 0
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class MatchLineup(BaseModel):
    model_config = ConfigDict(extra="ignore")

    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    match_id: str
    championship_id: str
    team_id: str

    starting_five: List[str] = Field(default_factory=list)
    bench: List[str] = Field(default_factory=list)

    captain_id: Optional[str] = None
    vice_captain_id: Optional[str] = None

    goalkeeper_starting_id: Optional[str] = None
    goalkeeper_bench_id: Optional[str] = None

    penalty_order: List[str] = Field(default_factory=list)
    free_kick_order: List[str] = Field(default_factory=list)

    ball_center_id: Optional[str] = None
    last_free_kick_id: Optional[str] = None
    timeout_leader_id: Optional[str] = None
    penalty_main_id: Optional[str] = None
    free_kick_main_id: Optional[str] = None

    rotation_plan: List[dict] = Field(default_factory=list)

    tactical_plan: Optional[str] = None
    coach_notes: Optional[str] = None
    assistant_notes: Optional[str] = None

    status: str = "draft"
    version: int = 1

    created_by: str
    updated_by: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    
# Event Models
class EventCreate(BaseModel):
    team_id: str
    event_type: EventType
    title: str
    description: Optional[str] = None
    location: str
    start_time: datetime
    end_time: Optional[datetime] = None
    opponent: Optional[str] = None
    championship_id: Optional[str] = None
    status: Optional[str] = "scheduled"  # scheduled, postponed, cancelled
    postponed_to_start_time: Optional[datetime] = None
    postponed_to_end_time: Optional[datetime] = None
    postponement_reason: Optional[str] = None
    original_event_id: Optional[str] = None
        
class Event(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    team_id: str
    event_type: EventType
    title: str
    description: Optional[str] = None
    location: str
    start_time: datetime
    end_time: Optional[datetime] = None
    opponent: Optional[str] = None
    championship_id: Optional[str] = None
    championship_match_id: Optional[str] = None
    status: str = "scheduled"  # scheduled, postponed, cancelled
    postponed_to_start_time: Optional[datetime] = None
    postponed_to_end_time: Optional[datetime] = None
    postponement_reason: Optional[str] = None
    original_event_id: Optional[str] = None
    created_by: str
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

# Attendance Models
class Attendance(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    event_id: str
    convocation_id: Optional[str] = None
    player_id: str
    team_id: str
    event_type: str
    championship_id: Optional[str] = None
    status: AttendanceStatus = "pendente"
    reason: Optional[str] = None
    event_date: datetime
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class AttendanceUpdate(BaseModel):
    status: AttendanceStatus
    reason: Optional[str] = None

# Convocation Models
class ConvocationVisibility(str, Enum):
    players = "players"
    delegates = "delegates"
    all = "all"
    private = "private"  # Sprint 3.3.1: lista visível apenas para convocados/equipa técnica

class ConvocationStatus(str, Enum):
    draft = "draft"
    published = "published"
    closed = "closed"
    cancelled = "cancelled"

class ConvocationCreate(BaseModel):
    event_id: str
    player_ids: List[str]
    message: Optional[str] = None
    visibility: ConvocationVisibility = ConvocationVisibility.all
    # Sprint 3.3.2 — compatibilidade com frontend atual e futuro.
    # O frontend pode enviar visibility=private, is_private=True ou privacy=private.
    is_private: Optional[bool] = False
    privacy: Optional[str] = None
    # O endpoint existente /convocations representa "lançar convocatória".
    # Mantém comportamento histórico: cria attendance e publica imediatamente.
    publish_immediately: Optional[bool] = True

class Convocation(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    event_id: str
    player_ids: List[str]
    message: Optional[str] = None
    visibility: ConvocationVisibility = ConvocationVisibility.all
    status: ConvocationStatus = ConvocationStatus.draft
    published_at: Optional[datetime] = None
    created_by: str
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class TrainingFeedbackCreate(BaseModel):
    event_id: str
    rating: str
    comment: Optional[str] = None


class TrainingFeedback(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    event_id: str
    player_id: str
    team_id: str
    rating: str
    comment: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))


# Evaluation Models — Sprint 4.2.1
EvaluationCriterionCategory = Literal[
    "technical",
    "tactical",
    "physical",
    "psychological",
    "attitude",
    "other"
]

EvaluationVisibility = Literal[
    "coach_only",
    "technical_staff",
    "player",
    "guardian",
    "all"
]


class EvaluationCriterionCreate(BaseModel):
    name: str
    description: Optional[str] = None
    category: EvaluationCriterionCategory = "technical"
    scale_min: int = 1
    scale_max: int = 5
    weight: float = 1.0
    team_id: Optional[str] = None
    is_active: bool = True


class EvaluationCriterion(BaseModel):
    model_config = ConfigDict(extra="ignore")

    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    description: Optional[str] = None
    category: EvaluationCriterionCategory = "technical"
    scale_min: int = 1
    scale_max: int = 5
    weight: float = 1.0
    team_id: Optional[str] = None
    club_id: Optional[str] = None
    is_active: bool = True
    created_by: str
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))


class EvaluationCriterionUpdate(BaseModel):
    name: Optional[str] = None
    description: Optional[str] = None
    category: Optional[EvaluationCriterionCategory] = None
    scale_min: Optional[int] = None
    scale_max: Optional[int] = None
    weight: Optional[float] = None
    team_id: Optional[str] = None
    is_active: Optional[bool] = None


class PlayerEvaluationScore(BaseModel):
    criterion_id: str
    score: float
    comment: Optional[str] = None


class PlayerEvaluationCreate(BaseModel):
    player_id: str
    team_id: str
    event_id: Optional[str] = None
    period_label: Optional[str] = None
    visibility: EvaluationVisibility = "coach_only"
    scores: List[PlayerEvaluationScore] = []
    general_comment: Optional[str] = None

    # Regra StickPro:
    # atleta e responsável veem exatamente a mesma versão motivacional.
    share_with_player: bool = False
    share_with_guardian: bool = False

    # Camada motivacional, visível apenas quando partilhada.
    public_summary: Optional[str] = None
    strengths: List[str] = []
    improvement_goals: List[str] = []
    motivational_message: Optional[str] = None


class PlayerEvaluation(BaseModel):
    model_config = ConfigDict(extra="ignore")

    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    player_id: str
    team_id: str
    event_id: Optional[str] = None
    period_label: Optional[str] = None
    visibility: EvaluationVisibility = "coach_only"
    scores: List[dict] = []
    general_comment: Optional[str] = None

    # Mantemos os dois campos por compatibilidade, mas serão sincronizados.
    share_with_player: bool = False
    share_with_guardian: bool = False

    # Informação pedagógica/motivacional para atleta e responsáveis.
    public_summary: Optional[str] = None
    strengths: List[str] = []
    improvement_goals: List[str] = []
    motivational_message: Optional[str] = None

    overall_score: Optional[float] = None
    created_by: str
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))


class PlayerEvaluationUpdate(BaseModel):
    period_label: Optional[str] = None
    visibility: Optional[EvaluationVisibility] = None
    scores: Optional[List[PlayerEvaluationScore]] = None
    general_comment: Optional[str] = None
    share_with_player: Optional[bool] = None
    share_with_guardian: Optional[bool] = None
    public_summary: Optional[str] = None
    strengths: Optional[List[str]] = None
    improvement_goals: Optional[List[str]] = None
    motivational_message: Optional[str] = None


# Evaluation Plan Models — Sprint 4.2.3.1
EvaluationPlanCategory = Literal[
    "training",
    "match",
    "goalkeeper",
    "technical",
    "tactical",
    "physical",
    "custom"
]


class EvaluationPlanCriterion(BaseModel):
    criterion_id: str
    weight: float = 1.0
    required: bool = True
    order: int = 0


class EvaluationPlanCreate(BaseModel):
    name: str
    description: Optional[str] = None
    category: EvaluationPlanCategory = "training"
    team_id: Optional[str] = None
    criteria: List[EvaluationPlanCriterion] = []
    estimated_minutes: Optional[int] = None
    is_active: bool = True


class EvaluationPlan(BaseModel):
    model_config = ConfigDict(extra="ignore")

    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    description: Optional[str] = None
    category: EvaluationPlanCategory = "training"
    team_id: Optional[str] = None
    club_id: Optional[str] = None
    criteria: List[dict] = []
    estimated_minutes: Optional[int] = None
    is_active: bool = True
    created_by: str
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))


class EvaluationPlanUpdate(BaseModel):
    name: Optional[str] = None
    description: Optional[str] = None
    category: Optional[EvaluationPlanCategory] = None
    team_id: Optional[str] = None
    criteria: Optional[List[EvaluationPlanCriterion]] = None
    estimated_minutes: Optional[int] = None
    is_active: Optional[bool] = None


# Evaluation Execution Models — Sprint 4.2.4.1
class EvaluationFromPlanScore(BaseModel):
    criterion_id: str
    score: float
    comment: Optional[str] = None


class PlayerEvaluationFromPlanItem(BaseModel):
    player_id: str
    scores: List[EvaluationFromPlanScore] = []
    general_comment: Optional[str] = None
    share_with_player: bool = False
    share_with_guardian: bool = False
    public_summary: Optional[str] = None
    strengths: List[str] = []
    improvement_goals: List[str] = []
    motivational_message: Optional[str] = None


class BulkEvaluationFromPlanCreate(BaseModel):
    plan_id: str
    team_id: str
    event_id: Optional[str] = None
    period_label: Optional[str] = None
    visibility: EvaluationVisibility = "coach_only"
    evaluations: List[PlayerEvaluationFromPlanItem] = []


# Unavailability Models
UnavailabilityReason = Literal["ferias", "doenca", "escola", "outro"]

class UnavailabilityCreate(BaseModel):
    start_date: datetime
    end_date: datetime
    reason: str  # ferias, doenca, escola, outro
    notes: Optional[str] = None  # Free text for additional details

class Unavailability(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    user_id: str
    team_ids: List[str] = []  # Teams affected
    start_date: datetime
    end_date: datetime
    reason: str
    notes: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

# Event Reminder Models
class EventReminder(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    event_id: str
    team_id: str
    reminder_type: str  # "no_convocation_4h"
    sent_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    notified_user_ids: List[str] = []

# Message Models with attachments
class MessageCreate(BaseModel):
    team_id: str
    content: str
    recipient_ids: List[str] = []  # Empty = all team members
    attachment_name: Optional[str] = None
    attachment_data: Optional[str] = None  # Base64 encoded

class Message(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    team_id: str
    sender_id: str
    sender_name: str
    content: str
    recipient_ids: List[str] = []
    attachment_name: Optional[str] = None
    attachment_url: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

# ==================== PAYMENT MODELS ====================

PaymentStatus = Literal["pending", "paid", "overdue"]
PaymentType = Literal["monthly_fee", "custom"]

class MonthlyFeeCreate(BaseModel):
    user_id: str
    amount: float
    month: int  # 1-12
    year: int
    due_date: datetime
    notes: Optional[str] = None

class MonthlyFee(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    user_id: str
    amount: float
    month: int
    year: int
    due_date: datetime
    status: PaymentStatus = "pending"
    paid_at: Optional[datetime] = None
    proof_url: Optional[str] = None
    proof_filename: Optional[str] = None
    notes: Optional[str] = None
    created_by: str
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class CustomPaymentCreate(BaseModel):
    user_id: str
    title: str
    description: Optional[str] = None
    amount: float
    due_date: datetime

class CustomPayment(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    user_id: str
    title: str
    description: Optional[str] = None
    amount: float
    due_date: datetime
    status: PaymentStatus = "pending"
    paid_at: Optional[datetime] = None
    proof_url: Optional[str] = None
    proof_filename: Optional[str] = None
    created_by: str
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class PaymentSettingsUpdate(BaseModel):
    payments_disabled: Optional[bool] = None
    default_monthly_fee: Optional[float] = None

# ==================== AUTH HELPERS ====================

def hash_password(password: str) -> str:
    return bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')

def verify_password(password: str, hashed: str) -> bool:
    return bcrypt.checkpw(password.encode('utf-8'), hashed.encode('utf-8'))

def create_token(user_id: str, email: str, role: str) -> str:
    payload = {
        "user_id": user_id,
        "email": email,
        "role": role,
        "exp": datetime.now(timezone.utc) + timedelta(hours=JWT_EXPIRATION_HOURS)
    }
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGORITHM)

async def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)) -> dict:
    try:
        payload = jwt.decode(credentials.credentials, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        user = await db.users.find_one({"id": payload["user_id"]}, {"_id": 0, "password": 0})
        if not user:
            raise HTTPException(status_code=401, detail="Utilizador não encontrado")
        return user
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expirado")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Token inválido")

# ==================== EMAIL MOCK ====================

async def send_email_notification(to_email: str, subject: str, html_content: str, attachment_name: str = None, attachment_data: bytes = None):
    """Send email using Resend API - falls back gracefully if not configured"""
    resend_api_key = os.environ.get('RESEND_API_KEY')
    sender_email = os.environ.get('SENDER_EMAIL', 'onboarding@resend.dev')
    
    if not resend_api_key:
        logger.warning(f"[EMAIL SKIPPED] RESEND_API_KEY not configured. Would send to: {to_email}, Subject: {subject}")
        return False
    
    try:
        # Configure Resend with API key
        resend.api_key = resend_api_key
        
        # Build email params
        params = {
            "from": sender_email,
            "to": [to_email],
            "subject": subject,
            "html": html_content
        }
        
        # Add attachment if provided
        if attachment_name and attachment_data:
            params["attachments"] = [{
                "filename": attachment_name,
                "content": base64.b64encode(attachment_data).decode('utf-8')
            }]
        
        # Send email using thread to keep async non-blocking
        email_response = await asyncio.to_thread(resend.Emails.send, params)
        
        logger.info(f"[EMAIL SENT] To: {to_email}, Subject: {subject}, ID: {email_response.get('id', 'unknown')}")
        return True
        
    except Exception as e:
        # Log error but don't break the app flow
        logger.error(f"[EMAIL ERROR] Failed to send to {to_email}: {str(e)}")
        return False


def build_email_template(title: str, content: str, footer_text: str = None) -> str:
    """Build a clean, professional email template"""
    footer = footer_text or "Esta é uma mensagem automática do StickPro."
    
    return f"""
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
    </head>
    <body style="margin: 0; padding: 0; font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, 'Helvetica Neue', Arial, sans-serif; background-color: #f4f4f5;">
        <table width="100%" cellpadding="0" cellspacing="0" style="background-color: #f4f4f5; padding: 20px 0;">
            <tr>
                <td align="center">
                    <table width="600" cellpadding="0" cellspacing="0" style="background-color: #ffffff; border-radius: 8px; overflow: hidden; box-shadow: 0 2px 4px rgba(0,0,0,0.1);">
                        <!-- Header -->
                        <tr>
                            <td style="background: linear-gradient(135deg, #0f172a 0%, #1e293b 100%); padding: 24px 32px; text-align: center;">
                                <h1 style="margin: 0; color: #22d3ee; font-size: 24px; font-weight: 700; letter-spacing: 2px;">STICK PRO</h1>
                            </td>
                        </tr>
                        <!-- Title -->
                        <tr>
                            <td style="padding: 32px 32px 16px 32px;">
                                <h2 style="margin: 0; color: #0f172a; font-size: 20px; font-weight: 600;">{title}</h2>
                            </td>
                        </tr>
                        <!-- Content -->
                        <tr>
                            <td style="padding: 0 32px 32px 32px; color: #374151; font-size: 15px; line-height: 1.6;">
                                {content}
                            </td>
                        </tr>
                        <!-- Footer -->
                        <tr>
                            <td style="background-color: #f8fafc; padding: 20px 32px; border-top: 1px solid #e5e7eb;">
                                <p style="margin: 0; color: #6b7280; font-size: 13px; text-align: center;">
                                    {footer}
                                </p>
                            </td>
                        </tr>
                    </table>
                </td>
            </tr>
        </table>
    </body>
    </html>
    """

# ==================== PUSH NOTIFICATIONS HELPER ====================

async def send_push_to_users(user_ids: List[str], title: str, body: str, url: str = "/"):
    """Send push notifications to specific users"""
    from pywebpush import webpush, WebPushException
    import json
    
    vapid_private_key = os.environ.get('VAPID_PRIVATE_KEY')
    vapid_claims_email = os.environ.get('VAPID_CLAIMS_EMAIL', 'noreply@stickpro.com')
    
    if not vapid_private_key:
        logger.warning("VAPID_PRIVATE_KEY not configured, skipping push notifications")
        return
    
    # Get subscriptions for these users
    subscriptions = await db.push_subscriptions.find(
        {"user_id": {"$in": user_ids}}, 
        {"_id": 0}
    ).to_list(1000)
    
    if not subscriptions:
        logger.info(f"No push subscriptions found for users: {user_ids}")
        return
    
    payload = json.dumps({
        "title": title,
        "body": body,
        "url": url,
        "icon": "/icons/icon-192x192.png"
    })
    
    for sub in subscriptions:
        try:
            webpush(
                subscription_info={
                    "endpoint": sub['endpoint'],
                    "keys": sub['keys']
                },
                data=payload,
                vapid_private_key=vapid_private_key,
                vapid_claims={"sub": f"mailto:{vapid_claims_email}"}
            )
            logger.info(f"Push sent to user {sub['user_id']}")
        except WebPushException as e:
            logger.error(f"Push failed for user {sub['user_id']}: {e}")
            # Remove invalid subscriptions
            if e.response and e.response.status_code in [404, 410]:
                await db.push_subscriptions.delete_one({"endpoint": sub['endpoint']})
        except Exception as e:
            logger.error(f"Push error: {e}")


async def notify_guardians_of_team_event(team_id: str, event_title: str, event_type: str, event_time: str):
    """
    Notify all guardians (parents) whose children are members of a team when an event is created.
    
    Args:
        team_id: ID of the team the event is for
        event_type: Type of event (treino, jogo_campeonato, etc.)
        event_title: Title of the event
        event_time: When the event will happen
    """
    try:
        # Get the team
        team = await db.teams.find_one({"id": team_id}, {"_id": 0, "name": 1, "player_ids": 1, "member_ids": 1})
        if not team:
            logger.warning(f"Team {team_id} not found for notification")
            return
        
        team_name = team.get('name', 'Equipa')
        # Use player_ids or member_ids (whichever has data)
        member_ids = team.get('player_ids', []) or team.get('member_ids', [])
        
        # If no members in team, find users who have this team in their team_ids
        if not member_ids:
            member_ids = []
            async for user in db.users.find({"team_ids": team_id}, {"_id": 0, "id": 1}):
                member_ids.append(user['id'])
        
        if not member_ids:
            logger.info(f"No members found in team {team_name} for notification")
            return
        
        logger.info(f"Found {len(member_ids)} members in team {team_name}")
        
        # Find all guardians who have children in this team
        guardian_ids = []
        async for user in db.users.find(
            {
                "role": "responsavel",
                "$or": [
                    {"linked_player_ids": {"$in": member_ids}},
                    {"linked_player_id": {"$in": member_ids}}
                ]
            },
            {"_id": 0, "id": 1, "email": 1, "name": 1, "linked_player_ids": 1, "linked_player_id": 1}
        ):
            guardian_ids.append(user['id'])
            
            # Get children names for the email
            all_linked = user.get('linked_player_ids', [])
            if user.get('linked_player_id'):
                all_linked = list(set(all_linked + [user['linked_player_id']]))
            
            # Filter to only children in this team
            children_in_team = [pid for pid in all_linked if pid in member_ids]
            
            if children_in_team and user.get('email'):
                # Get child names
                child_names = []
                for child_id in children_in_team:
                    child = await db.users.find_one({"id": child_id}, {"_id": 0, "name": 1})
                    if child:
                        child_names.append(child.get('name', 'Atleta').split(' ')[0])
                
                children_str = ', '.join(child_names) if child_names else 'o seu filho'
                
                # Send email notification
                event_type_names = {
                    'treino': 'Treino',
                    'jogo_campeonato': 'Jogo de Campeonato',
                    'jogo_amigavel': 'Jogo Amigável',
                    'torneio': 'Torneio',
                    'outro': 'Evento'
                }
                event_type_name = event_type_names.get(event_type, 'Evento')
                
                html_content = f"""
                <html>
                <body style="font-family: Arial, sans-serif; line-height: 1.6; color: #333;">
                    <div style="max-width: 600px; margin: 0 auto; padding: 20px;">
                        <h2 style="color: #006D5B;">Novo Evento para {children_str}</h2>
                        <p>Olá {user.get('name', 'Responsável').split(' ')[0]},</p>
                        <p>Foi criado um novo evento para a equipa <strong>{team_name}</strong>:</p>
                        <div style="background-color: #f5f5f5; padding: 15px; border-radius: 8px; margin: 20px 0;">
                            <p style="margin: 5px 0;"><strong>Tipo:</strong> {event_type_name}</p>
                            <p style="margin: 5px 0;"><strong>Título:</strong> {event_title}</p>
                            <p style="margin: 5px 0;"><strong>Data/Hora:</strong> {event_time}</p>
                        </div>
                        <p>Aceda à aplicação para confirmar a presença de {children_str}.</p>
                        <p style="color: #666; font-size: 12px; margin-top: 30px;">
                            Este email foi enviado automaticamente pelo StickPro.
                        </p>
                    </div>
                </body>
                </html>
                """
                
                try:
                    await send_email_notification(
                        to_email=user['email'],
                        subject=f"Novo {event_type_name} - {team_name}",
                        html_content=html_content
                    )
                    logger.info(f"Email notification sent to guardian {user['email']} about event {event_title}")
                except Exception as e:
                    logger.error(f"Failed to send email to guardian {user['email']}: {e}")
        
        # Send push notifications to all guardians
        if guardian_ids:
            event_type_names = {
                'treino': 'Treino',
                'jogo_campeonato': 'Jogo',
                'jogo_amigavel': 'Jogo',
                'torneio': 'Torneio',
                'outro': 'Evento'
            }
            event_type_name = event_type_names.get(event_type, 'Evento')
            
            await send_push_to_users(
                user_ids=guardian_ids,
                title=f"Novo {event_type_name} - {team_name}",
                body=f"{event_title}",
                url="/dashboard"
            )
            logger.info(f"Push notifications sent to {len(guardian_ids)} guardians")
        else:
            logger.info(f"No guardians found to notify for team {team_name}")
    
    except Exception as e:
        logger.error(f"Error notifying guardians: {e}")

# ==================== AUTH ROUTES ====================

async def build_available_profiles(user: dict) -> List[dict]:
    """Build list of all profiles a user can access."""
    profiles = []

    async def get_teams(team_ids: list) -> list:
        teams = []
        for team_id in team_ids or []:
            team = await db.teams.find_one({"id": team_id}, {"_id": 0})
            if team:
                teams.append(team)
        return teams

    user_teams = await get_teams(user.get("team_ids", []))

    # Own profiles: main role + additional roles
    all_roles = []
    if user.get("role"):
        all_roles.append(user["role"])

    for role in user.get("additional_roles", []):
        if role not in all_roles:
            all_roles.append(role)

    for role in all_roles:
        role_name = getRoleNamePt(role)

        profiles.append({
            "profile_id": f"self:{user['id']}:{role}",
            "type": "self",
            "user_id": user["id"],
            "user_name": user.get("name", ""),
            "role": role,
            "role_name": role_name,
            "label": f"{role_name}",
            "description": user.get("name", ""),
            "teams": user_teams,
            "team_ids": [team.get("id") for team in user_teams if team.get("id")]
        })

    # Associated profiles: children / athletes linked to this account
    for assoc_id in user.get("associated_accounts", []):
        assoc_user = await db.users.find_one(
            {"id": assoc_id},
            {"_id": 0, "hashed_password": 0}
        )

        if not assoc_user:
            continue

        assoc_teams = await get_teams(assoc_user.get("team_ids", []))

        relationship = assoc_user.get("relationship") or "responsavel"

        profiles.append({
            "profile_id": f"associated:{assoc_user['id']}:responsavel",
            "type": "associated",
            "user_id": assoc_user["id"],
            "user_name": assoc_user.get("name", ""),
            "role": "responsavel",
            "role_name": "Responsável",
            "relationship": relationship,
            "label": assoc_user.get("name", ""),
            "description": f"Responsável de {assoc_user.get('name', '')}",
            "teams": assoc_teams,
            "team_ids": [team.get("id") for team in assoc_teams if team.get("id")]
        })

    return profiles

def getRoleNamePt(role: str) -> str:
    roles = {
        "admin": "Administrador",
        "gestor_desportivo": "Gestor Desportivo",
        "diretor_tecnico": "Diretor Técnico",
        "treinador": "Treinador",
        "delegado": "Delegado",
        "jogador": "Atleta",
        "atleta": "Atleta",
        "responsavel": "Responsável"
    }
    return roles.get(role, role)


class ActivateAccountRequest(BaseModel):
    token: str
    password: str


@api_router.post("/auth/activate")
async def activate_account(data: ActivateAccountRequest):
    user = await db.users.find_one({"invite_token": data.token}, {"_id": 0})

    if not user:
        raise HTTPException(status_code=400, detail="Convite inválido")

    if user.get("is_activated"):
        raise HTTPException(status_code=400, detail="Conta já ativada")

    expires_at = user.get("invite_expires_at")
    if not expires_at:
        raise HTTPException(status_code=400, detail="Convite inválido")

    if datetime.fromisoformat(expires_at) < datetime.now(timezone.utc):
        raise HTTPException(status_code=400, detail="Convite expirado")

    hashed_password = bcrypt.hashpw(
        data.password.encode("utf-8"),
        bcrypt.gensalt()
    ).decode("utf-8")

    await db.users.update_one(
        {"id": user["id"]},
        {
            "$set": {
                "hashed_password": hashed_password,
                "is_activated": True,
                "activated_at": datetime.now(timezone.utc).isoformat()
            },
            "$unset": {
                "invite_token": "",
                "invite_expires_at": ""
            }
        }
    )

    return {"message": "Conta ativada com sucesso"}


# Phase E2 — public, unauthenticated endpoint to request a fresh activation link.
# Security properties:
#   * Always returns the same generic message regardless of whether the email
#     exists in the database — prevents account enumeration.
#   * Only sends a link for accounts that exist AND are NOT yet activated.
#   * Refreshes the invite_token if it's missing or expired; otherwise reuses
#     the still-valid one (token rotation is bounded by the 7-day expiry).
#   * Throttles repeat sends per account to once every 60 seconds via the
#     last_activation_email_sent_at field — the API still returns success so
#     attackers can't probe the throttle either.
class RequestActivationLinkRequest(BaseModel):
    email: EmailStr


_GENERIC_ACTIVATION_RESPONSE = {
    "message": (
        "Se existir uma conta inativa associada a este email, "
        "enviámos um novo link de ativação."
    )
}


@api_router.post("/auth/request-new-activation-link")
async def request_new_activation_link(data: RequestActivationLinkRequest):
    email = data.email.strip().lower()
    user = await db.users.find_one({"email": email}, {"_id": 0})

    if not user or user.get("is_activated"):
        return _GENERIC_ACTIVATION_RESPONSE

    # Throttle: at most one activation email every 60 seconds per account.
    last_sent_iso = user.get("last_activation_email_sent_at")
    if last_sent_iso:
        try:
            last_sent = datetime.fromisoformat(last_sent_iso)
            if (datetime.now(timezone.utc) - last_sent).total_seconds() < 60:
                return _GENERIC_ACTIVATION_RESPONSE
        except Exception:
            pass  # bad data — fall through and treat as no throttle

    # Refresh token only if missing or expired; otherwise reuse the valid one.
    invite_token = user.get("invite_token")
    invite_expires_at = user.get("invite_expires_at")
    token_valid = False
    if invite_token and invite_expires_at:
        try:
            token_valid = datetime.fromisoformat(invite_expires_at) > datetime.now(timezone.utc)
        except Exception:
            token_valid = False
    if not token_valid:
        invite_token = secrets.token_urlsafe(32)
        invite_expires_at = (datetime.now(timezone.utc) + timedelta(days=7)).isoformat()
        await db.users.update_one(
            {"id": user["id"]},
            {"$set": {"invite_token": invite_token, "invite_expires_at": invite_expires_at}}
        )

    sent_at_iso = datetime.now(timezone.utc).isoformat()
    try:
        await send_activation_email(
            to_email=email,
            name=user.get("name", "Atleta"),
            token=invite_token,
            idempotency_key=f"resend-{user['id']}-{invite_token[:8]}",
        )
    except Exception as e:
        logger.warning(f"[ACTIVATION EMAIL] request-new-link: {e}")
    else:
        await db.users.update_one(
            {"id": user["id"]},
            {"$set": {"last_activation_email_sent_at": sent_at_iso}}
        )

    return _GENERIC_ACTIVATION_RESPONSE


# ============================================================================
# Phase E3 — Password Reset
# ============================================================================
#
# Two endpoints, both public/unauthenticated:
#
#   POST /api/auth/forgot-password { email }    → always returns generic 200
#   POST /api/auth/reset-password  { token, password } → 204 on success,
#                                                        400 on invalid/expired/used
#
# Security properties:
#   * forgot-password never reveals whether an account exists.
#   * The raw token is sent in the email; only its SHA-256 hash is stored.
#     A leaked DB snapshot therefore does not let an attacker reset passwords.
#   * Tokens are single-use: marked `used_at` atomically on successful reset.
#   * Tokens expire after 1 hour.
#   * forgot-password is rate-limited to one send / 60 s per account via the
#     `last_password_reset_email_sent_at` field on the user document.
#   * Every reset attempt (success or failure) is logged to the
#     `password_reset_audit` MongoDB collection with timestamp, masked email,
#     IP-agnostic outcome code and reason.
#
# Token lifecycle storage on the user document:
#   {
#     "password_reset_token_hash": "<hex sha256>",  # cleared on success
#     "password_reset_expires_at": "<iso8601 UTC>",
#     "last_password_reset_email_sent_at": "<iso8601 UTC>"
#   }
#
PASSWORD_RESET_TOKEN_TTL = timedelta(hours=1)
PASSWORD_RESET_THROTTLE_SECONDS = 60
PASSWORD_RESET_MIN_LEN = 8

_GENERIC_FORGOT_RESPONSE = {
    "message": (
        "Se existir uma conta associada a este email, "
        "enviámos um link para redefinir a palavra-passe."
    )
}


def _hash_reset_token(token: str) -> str:
    """SHA-256 hex digest of a reset token. Stored in DB; raw token in email."""
    return hashlib.sha256(token.encode("utf-8")).hexdigest()


def _mask_email_for_audit(email: str) -> str:
    """Return a redacted form like 'j***@example.com' for audit logs."""
    if not email or "@" not in email:
        return "***"
    local, _, domain = email.partition("@")
    if len(local) <= 1:
        return f"*@{domain}"
    return f"{local[0]}***@{domain}"


async def _audit_password_reset(
    *, email: str, outcome: str, reason: str = "", user_id: Optional[str] = None
) -> None:
    """Append a row to the password_reset_audit collection. Never raises."""
    try:
        await db.password_reset_audit.insert_one(
            {
                "id": str(uuid.uuid4()),
                "email_masked": _mask_email_for_audit(email or ""),
                "user_id": user_id,
                "outcome": outcome,
                "reason": reason,
                "timestamp": datetime.now(timezone.utc).isoformat(),
            }
        )
    except Exception as e:  # pragma: no cover — audit must never block flow
        logger.error(f"[AUDIT] failed to write password_reset_audit row: {e}")


class ForgotPasswordRequest(BaseModel):
    email: EmailStr


class ResetPasswordRequest(BaseModel):
    token: str = Field(..., min_length=10)
    password: str = Field(..., min_length=PASSWORD_RESET_MIN_LEN)


@api_router.post("/auth/forgot-password")
async def forgot_password(data: ForgotPasswordRequest):
    """Issue a password reset token + email. Always returns generic 200."""
    email = data.email.strip().lower()
    user = await db.users.find_one({"email": email}, {"_id": 0})

    if not user:
        await _audit_password_reset(email=email, outcome="ignored", reason="unknown_email")
        return _GENERIC_FORGOT_RESPONSE

    if not user.get("is_activated", False):
        await _audit_password_reset(
            email=email,
            outcome="ignored",
            reason="account_not_activated",
            user_id=user.get("id"),
        )
        return _GENERIC_FORGOT_RESPONSE

    last_sent_iso = user.get("last_password_reset_email_sent_at")
    if last_sent_iso:
        try:
            last_sent = datetime.fromisoformat(last_sent_iso)
            if (datetime.now(timezone.utc) - last_sent).total_seconds() < PASSWORD_RESET_THROTTLE_SECONDS:
                await _audit_password_reset(
                    email=email,
                    outcome="ignored",
                    reason="throttled",
                    user_id=user.get("id"),
                )
                return _GENERIC_FORGOT_RESPONSE
        except Exception:
            pass

    raw_token = secrets.token_urlsafe(32)
    token_hash = _hash_reset_token(raw_token)
    expires_at_iso = (datetime.now(timezone.utc) + PASSWORD_RESET_TOKEN_TTL).isoformat()
    sent_at_iso = datetime.now(timezone.utc).isoformat()

    await db.users.update_one(
        {"id": user["id"]},
        {
            "$set": {
                "password_reset_token_hash": token_hash,
                "password_reset_expires_at": expires_at_iso,
            }
        },
    )

    delivered = False
    try:
        delivered = await send_password_reset_email(
            to_email=email,
            name=user.get("name", "Atleta"),
            token=raw_token,
            idempotency_key=f"forgot-{user['id']}-{token_hash[:8]}",
        )
    except Exception as e:
        logger.warning(f"[PASSWORD RESET] send failed: {e}")

    if delivered:
        await db.users.update_one(
            {"id": user["id"]},
            {"$set": {"last_password_reset_email_sent_at": sent_at_iso}},
        )
        await _audit_password_reset(
            email=email,
            outcome="email_sent",
            reason="",
            user_id=user["id"],
        )
    else:
        await _audit_password_reset(
            email=email,
            outcome="email_failed",
            reason="delivery_failed",
            user_id=user["id"],
        )

    return _GENERIC_FORGOT_RESPONSE
    
@api_router.post("/auth/reset-password", status_code=204)
async def reset_password(data: ResetPasswordRequest):
    """Consume a reset token and set the new password. Single-use."""
    token_hash = _hash_reset_token(data.token)
    user = await db.users.find_one(
        {"password_reset_token_hash": token_hash}, {"_id": 0}
    )

    if not user:
        await _audit_password_reset(
            email="?", outcome="reset_rejected", reason="invalid_token",
        )
        raise HTTPException(status_code=400, detail="Link inválido ou já utilizado")

    # Expiry check
    expires_at_iso = user.get("password_reset_expires_at")
    if not expires_at_iso:
        await _audit_password_reset(
            email=user.get("email", ""), outcome="reset_rejected",
            reason="missing_expiry", user_id=user.get("id"),
        )
        raise HTTPException(status_code=400, detail="Link inválido ou já utilizado")
    try:
        expires_at = datetime.fromisoformat(expires_at_iso)
    except Exception:
        await _audit_password_reset(
            email=user.get("email", ""), outcome="reset_rejected",
            reason="bad_expiry", user_id=user.get("id"),
        )
        raise HTTPException(status_code=400, detail="Link inválido ou já utilizado")
    if expires_at < datetime.now(timezone.utc):
        # Clean up the expired token to make the state explicit and reduce
        # future enumeration surface area.
        await db.users.update_one(
            {"id": user["id"]},
            {"$unset": {
                "password_reset_token_hash": "",
                "password_reset_expires_at": "",
            }},
        )
        await _audit_password_reset(
            email=user.get("email", ""), outcome="reset_rejected",
            reason="expired", user_id=user.get("id"),
        )
        raise HTTPException(status_code=400, detail="Link expirado")

    # Atomic single-use: only consume if the hash field is still present.
    new_hashed = hash_password(data.password)
    update_result = await db.users.update_one(
        {
            "id": user["id"],
            "password_reset_token_hash": token_hash,
        },
        {
            "$set": {
                "hashed_password": new_hashed,
                "is_activated": True,  # ensures user can log in after reset
                "password_reset_used_at": datetime.now(timezone.utc).isoformat(),
            },
            "$unset": {
                "password_reset_token_hash": "",
                "password_reset_expires_at": "",
            },
        },
    )
    if update_result.modified_count != 1:
        await _audit_password_reset(
            email=user.get("email", ""), outcome="reset_rejected",
            reason="already_used", user_id=user.get("id"),
        )
        raise HTTPException(status_code=400, detail="Link inválido ou já utilizado")

    await _audit_password_reset(
        email=user.get("email", ""), outcome="reset_succeeded",
        user_id=user.get("id"),
    )
    # No body: 204 status. Clients should redirect to /login.
    return None

@api_router.post("/auth/register")
async def register(user_data: UserCreate):
    email = user_data.email.strip().lower()

    existing_user = await db.users.find_one({"email": email}, {"_id": 0})
    if existing_user:
        raise HTTPException(status_code=400, detail="Já existe uma conta com este email")

    user = User(
        email=email,
        name=user_data.name,
        surname=user_data.surname,
        role=user_data.role,
        phone=user_data.phone,
        additional_roles=user_data.additional_roles
    )

    user_dict = user.model_dump()
    user_dict["hashed_password"] = hash_password(user_data.password)
    user_dict["is_activated"] = True
    user_dict["created_at"] = user_dict["created_at"].isoformat()

    await db.users.insert_one(user_dict)

    token = create_token(user_dict["id"], user_dict["email"], user_dict["role"])
    profiles = await build_available_profiles(user_dict)

    return {
        "token": token,
        "user": {
            "id": user_dict["id"],
            "email": user_dict["email"],
            "name": user_dict["name"],
            "role": user_dict["role"],
            "additional_roles": user_dict.get("additional_roles", []),
            "phone": user_dict.get("phone"),
            "avatar_url": user_dict.get("avatar_url"),
            "team_ids": user_dict.get("team_ids", []),
            "associated_accounts": user_dict.get("associated_accounts", [])
        },
        "available_profiles": profiles
    }

@api_router.get("/family-invitations/{token}")
async def get_family_invitation(token: str):
    """Get public family invitation details for the accept page."""
    if not token:
        raise HTTPException(status_code=400, detail="Token de convite inválido")

    guardian_link = await db.guardian_links.find_one(
        {
            "invite_token": token,
            "status": "pending"
        },
        {"_id": 0}
    )

    if not guardian_link:
        raise HTTPException(status_code=404, detail="Convite não encontrado ou já utilizado")

    expires_at = guardian_link.get("invite_expires_at")
    if expires_at:
        try:
            expires_dt = datetime.fromisoformat(expires_at)
            if expires_dt < datetime.now(timezone.utc):
                raise HTTPException(status_code=400, detail="Convite expirado")
        except ValueError:
            raise HTTPException(status_code=400, detail="Convite inválido")

    return {
        "player_name": guardian_link.get("player_name"),
        "guardian_name": guardian_link.get("guardian_name"),
        "guardian_email": guardian_link.get("guardian_email"),
        "relationship": guardian_link.get("relationship"),
        "language": guardian_link.get("language", "pt"),
        "status": guardian_link.get("status"),
        "expires_at": guardian_link.get("invite_expires_at")
    }
    
@api_router.post("/family-invitations/accept")
async def accept_family_invitation(data: AcceptFamilyInviteRequest):
    """Accept a family invitation, create or reuse guardian user account and link to athlete."""
    if not data.token:
        raise HTTPException(status_code=400, detail="Token de convite inválido")

    guardian_link = await db.guardian_links.find_one(
        {
            "invite_token": data.token,
            "status": "pending"
        },
        {"_id": 0}
    )

    if not guardian_link:
        raise HTTPException(status_code=404, detail="Convite não encontrado ou já utilizado")

    expires_at = guardian_link.get("invite_expires_at")
    if expires_at:
        try:
            expires_dt = datetime.fromisoformat(expires_at)
            if expires_dt < datetime.now(timezone.utc):
                raise HTTPException(status_code=400, detail="Convite expirado")
        except ValueError:
            raise HTTPException(status_code=400, detail="Convite inválido")

    guardian_email = guardian_link.get("guardian_email")
    if not guardian_email:
        raise HTTPException(status_code=400, detail="Convite sem email associado")

    guardian_email = guardian_email.strip().lower()
    player_id = guardian_link.get("player_id")

    if not player_id:
        raise HTTPException(status_code=400, detail="Convite sem atleta associado")

    existing_user = await db.users.find_one(
        {"email": guardian_email},
        {"_id": 0}
    )

    if existing_user:
        guardian_user_id = existing_user["id"]

        await db.users.update_one(
            {"id": guardian_user_id},
            {
                "$addToSet": {
                    "linked_player_ids": player_id,
                    "associated_accounts": player_id
                },
                "$set": {
                    "updated_at": datetime.now(timezone.utc).isoformat()
                }
            }
        )

        user_response = await db.users.find_one(
            {"id": guardian_user_id},
            {"_id": 0, "hashed_password": 0, "password": 0}
        )

    else:
        guardian_user_id = str(uuid.uuid4())

        user = {
            "id": guardian_user_id,
            "name": data.name,
            "email": guardian_email,
            "hashed_password": hash_password(data.password),
            "role": "responsavel",
            "additional_roles": [],
            "club_id": guardian_link.get("club_id"),
            "team_ids": [],
            "linked_player_ids": [player_id],
            "associated_accounts": [player_id],
            "is_activated": True,
            "login_enabled": True,
            "has_real_email": True,
            "created_at": datetime.now(timezone.utc).isoformat(),
            "updated_at": datetime.now(timezone.utc).isoformat()
        }

        await db.users.insert_one(user)

        user_response = {
            k: v for k, v in user.items()
            if k not in ("hashed_password", "password", "_id")
        }

    await db.guardian_links.update_one(
        {"id": guardian_link["id"]},
        {
            "$set": {
                "guardian_user_id": guardian_user_id,
                "guardian_email": guardian_email,
                "status": "accepted",
                "accepted_at": datetime.now(timezone.utc).isoformat()
            }
        }
    )

    token = create_token(
        guardian_user_id,
        guardian_email,
        user_response.get("role", "responsavel")
    )

    profiles = await build_available_profiles(user_response)

    safe_user = {
        k: v for k, v in user_response.items()
        if k not in ("hashed_password", "password", "_id")
    }

    return {
        "token": token,
        "user": safe_user,
        "available_profiles": profiles,
        "message": "Convite aceite com sucesso"
    }
@api_router.post("/auth/login")
async def login(credentials: UserLogin):
    email = credentials.email.strip().lower()

    users = await db.users.find(
        {
            "$expr": {
                "$eq": [
                    {"$toLower": "$email"},
                    email
                ]
            }
        },
        {"_id": 0}
    ).sort("created_at", -1).to_list(50)

    matched_user = None
    matched_hash = None

    for candidate in users:
        stored_hash = candidate.get("hashed_password") or candidate.get("password")

        if not stored_hash:
            continue

        if verify_password(credentials.password, stored_hash):
            matched_user = candidate
            matched_hash = stored_hash
            break

    if not matched_user or not matched_hash:
        raise HTTPException(status_code=401, detail="Credenciais inválidas")

    if not matched_user.get("is_activated", True):
        raise HTTPException(status_code=401, detail="Conta ainda não ativada")

    if matched_user.get("password") and not matched_user.get("hashed_password"):
        await db.users.update_one(
            {"id": matched_user["id"]},
            {
                "$set": {"hashed_password": matched_hash},
                "$unset": {"password": ""}
            }
        )

    token = create_token(
        matched_user["id"],
        matched_user["email"],
        matched_user["role"]
    )

    profiles = await build_available_profiles(matched_user)

    return {
        "token": token,
        "user": {
            "id": matched_user["id"],
            "email": matched_user["email"],
            "name": matched_user["name"],
            "role": matched_user["role"],
            "additional_roles": matched_user.get("additional_roles", []),
            "phone": matched_user.get("phone"),
            "avatar_url": matched_user.get("avatar_url"),
            "team_ids": matched_user.get("team_ids", []),
            "associated_accounts": matched_user.get("associated_accounts", []),
            "linked_player_ids": matched_user.get("linked_player_ids", [])
        },
        "available_profiles": profiles
    }

@api_router.post("/debug/link-valid-player")
async def debug_link_valid_player():
    parent_user_id = "895fc2c9-6fbf-4ee6-a11e-01e4d76e2602"
    player_id = "0b9f6435-2065-43c6-a571-88d0682ca1ff"

    result = await db.users.update_one(
        {"id": parent_user_id},
        {
            "$set": {
                "linked_player_ids": [player_id]
            }
        }
    )

    return {
        "message": "Ligação criada com sucesso.",
        "parent_user_id": parent_user_id,
        "player_id": player_id,
        "matched_count": result.matched_count,
        "modified_count": result.modified_count
    }

@api_router.post("/debug/create-feedback-test")
async def debug_create_feedback_test():
    player_id = "0b9f6435-2065-43c6-a571-88d0682ca1ff"
    team_id = "debug-team"
    
    now = datetime.now(timezone.utc)
    start_time = now - timedelta(hours=2)
    end_time = now - timedelta(hours=1)

    event_id = str(uuid.uuid4())

    event = {
        "id": event_id,
        "title": "Treino de Teste Feedback",
        "team_id": team_id,
        "event_type": "training",
        "start_time": start_time.isoformat(),
        "end_time": end_time.isoformat(),
        "created_at": now.isoformat()
    }

    attendance = {
        "id": str(uuid.uuid4()),
        "event_id": event_id,
        "convocation_id": "debug-convocation",
        "player_id": player_id,
        "team_id": team_id,
        "event_type": "training",
        "event_date": start_time.isoformat(),
        "status": "confirmado",
        "updated_at": now.isoformat()
    }

    await db.events.insert_one(event)
    await db.attendance.insert_one(attendance)

    return {
        "message": "Treino de teste criado com presença confirmada.",
        "event": event,
        "attendance": attendance
    }

@api_router.get("/debug/team-feedback/{team_id}")
async def debug_team_feedback(team_id: str):
    feedbacks = await db.training_feedback.find(
        {"team_id": team_id},
        {"_id": 0}
    ).sort("created_at", -1).to_list(500)

    return feedbacks

@api_router.get("/debug/teams")
async def debug_list_teams():
    teams = await db.teams.find(
        {},
        {
            "_id": 0,
            "id": 1,
            "name": 1,
            "category": 1,
            "season": 1,
            "club_id": 1
        }
    ).to_list(100)

    return teams

@api_router.post("/debug/create-team-feedback-test")
async def debug_create_team_feedback_test():
    team_id = "de795245-cf52-4c13-8359-60fe7606de5e"
    player_id = "0b9f6435-2065-43c6-a571-88d0682ca1ff"

    now = datetime.now(timezone.utc)
    start_time = now - timedelta(hours=2)
    end_time = now - timedelta(hours=1)

    event_id = str(uuid.uuid4())

    event = {
        "id": event_id,
        "title": "Treino Técnico Escolares A",
        "team_id": team_id,
        "event_type": "training",
        "start_time": start_time.isoformat(),
        "end_time": end_time.isoformat(),
        "created_at": now.isoformat()
    }

    feedback = {
        "id": str(uuid.uuid4()),
        "event_id": event_id,
        "player_id": player_id,
        "team_id": team_id,
        "rating": "positive",
        "comment": "Gostei muito dos exercícios e do ritmo do treino.",
        "created_at": now.isoformat()
    }

    await db.events.insert_one(event.copy())
    await db.training_feedback.insert_one(feedback.copy())

    return {
        "message": "Feedback de teste criado para Escolares A.",
        "event": event,
        "feedback": feedback
    }
    
@api_router.get("/debug/latest-family-invite")
async def debug_latest_family_invite():

    invite = await db.guardian_links.find_one(
        {"status": "pending", "invite_token": {"$exists": True}},
        {"_id": 0},
        sort=[("created_at", -1)]
    )

    if not invite:
        raise HTTPException(status_code=404, detail="Nenhum convite pendente encontrado")

    frontend_url = os.environ.get("FRONTEND_URL", "").rstrip("/")
    accept_link = f"{frontend_url}/accept-family-invite?token={invite['invite_token']}"

    return {
        "player_name": invite.get("player_name"),
        "guardian_email": invite.get("guardian_email"),
        "relationship": invite.get("relationship"),
        "token": invite.get("invite_token"),
        "accept_link": accept_link,
        "expires_at": invite.get("invite_expires_at"),
        "status": invite.get("status")
    }

@api_router.get("/debug/associated-user-lite/{user_id}")
async def debug_associated_user_lite(
    user_id: str,
    current_user: dict = Depends(get_current_user)
):
    allowed_accounts = current_user.get("associated_accounts", [])

    if user_id not in allowed_accounts and current_user.get("role") not in ["admin", "gestor_desportivo"]:
        raise HTTPException(status_code=403, detail="Sem acesso a este utilizador")

    user = await db.users.find_one(
        {"id": user_id},
        {
            "_id": 0,
            "id": 1,
            "name": 1,
            "role": 1,
            "team_id": 1,
            "team_ids": 1,
            "teams": 1,
            "profile": 1,
            "associated_accounts": 1,
            "linked_player_ids": 1,
        }
    )

    if not user:
        raise HTTPException(status_code=404, detail="Utilizador não encontrado")

    return user

@api_router.get("/auth/me")
async def get_me(current_user: dict = Depends(get_current_user)):
    profiles = await build_available_profiles(current_user)
    
    # Build user permissions
    checker = get_permission_checker(current_user)
    permissions = {
        "is_admin": checker.is_admin,
        "is_coach": checker.is_coach,
        "is_assistant_coach": checker.is_assistant_coach,
        "is_delegate": checker.is_delegate,
        "is_player": checker.is_player,
        "is_family_member": checker.is_family_member,
        "is_staff": checker.is_staff,
        "can_manage_team": checker.can_manage_team,
        "can_manage_events": checker.can_manage_events,
        "can_manage_stats": checker.can_manage_stats,
        "can_manage_attendance": checker.can_manage_attendance,
        "can_create_convocations": checker.can_create_convocations,
        "can_manage_lineups": checker.can_manage_lineups,
        "can_import_data": checker.can_import_data,
        "can_manage_club": checker.can_manage_club,
    }
    
    return {
        **UserResponse(**current_user).model_dump(),
        "available_profiles": profiles,
        "permissions": permissions,
        "accessible_team_ids": list(checker.team_ids) if not checker.is_admin else None
    }

@api_router.post("/profile/link-player")
async def link_player_to_current_user(
    data: LinkPlayerRequest,
    current_user: dict = Depends(get_current_user)
):
    player = await db.users.find_one({"id": data.player_id}, {"_id": 0})

    if not player:
        raise HTTPException(status_code=404, detail="Atleta não encontrado.")

    if player.get("role") not in ["atleta", "player", "jogador"]:
        raise HTTPException(
            status_code=400,
            detail="Só é possível associar contas de atleta."
        )

    linked_player_ids = current_user.get("linked_player_ids") or []

    if data.player_id not in linked_player_ids:
        linked_player_ids.append(data.player_id)

    await db.users.update_one(
        {"id": current_user["id"]},
        {
            "$set": {
                "linked_player_ids": linked_player_ids
            }
        }
    )

    return {
        "message": "Atleta associado com sucesso.",
        "linked_player_ids": linked_player_ids
    }

@api_router.get("/debug/attendance/{player_id}")
async def debug_attendance(player_id: str):
    attendances = await db.attendance.find(
        {"player_id": player_id},
        {"_id": 0}
    ).to_list(100)

    return attendances
    

@api_router.get("/debug/players")
async def debug_list_players():
    players = await db.users.find(
        {},
        {
            "_id": 0,
            "id": 1,
            "name": 1,
            "surname": 1,
            "email": 1,
            "role": 1,
            "team_ids": 1,
            "club_id": 1,
            "linked_player_ids": 1
        }
    ).to_list(100)

    return players
    players = await db.users.find(
        {
            "role": {"$in": ["atleta", "player", "jogador"]}
        },
        {
            "_id": 0,
            "id": 1,
            "name": 1,
            "surname": 1,
            "email": 1,
            "role": 1,
            "team_ids": 1,
            "club_id": 1,
            "linked_player_ids": 1
        }
    ).to_list(100)

    return players

@api_router.post("/debug/link-player")
async def debug_link_player():
    parent_user_id = "895fc2c9-6fbf-4ee6-a11e-01e4d76e2602"
    player_id = "0b9f6435-2065-43c6-a571-88d0682ca1ff"

    await db.users.update_one(
        {"id": parent_user_id},
        {
            "$set": {
                "linked_player_ids": [player_id]
            }
        }
    )

    return {
        "message": "Ligação criada com sucesso.",
        "parent_user_id": parent_user_id,
        "linked_player_ids": [player_id]
    }

@api_router.get("/debug/current-user")
async def debug_current_user():
    users = await db.users.find(
        {"email": "bpereiracav@gmail.com"},
        {
            "_id": 0,
            "id": 1,
            "email": 1,
            "name": 1,
            "role": 1,
            "linked_player_ids": 1,
            "hashed_password": 1,
            "password": 1,
            "created_at": 1
        }
    ).to_list(100)

    return users

@api_router.get("/debug/communication-logs")
async def debug_communication_logs(
    current_user: dict = Depends(get_current_user)
):
    checker = get_permission_checker(current_user)

    if not checker.is_admin:
        raise HTTPException(
            status_code=403,
            detail="Sem permissão"
        )

    logs = await db.communication_logs.find(
        {},
        {"_id": 0}
    ).sort("created_at", -1).to_list(100)

    return logs

@api_router.get("/debug/event-recipients/{event_id}")
async def debug_event_recipients(event_id: str, current_user: dict = Depends(get_current_user)):
    checker = get_permission_checker(current_user)

    if not checker.is_admin:
        raise HTTPException(status_code=403, detail="Sem permissão")

    event = await db.events.find_one({"id": event_id}, {"_id": 0})
    if not event:
        raise HTTPException(status_code=404, detail="Evento não encontrado")

    recipients = await recipient_service.get_event_recipients(event)

    return {
        "event_id": event_id,
        "team_id": event.get("team_id"),
        "team_ids": event.get("team_ids"),
        "recipients_count": len(recipients),
        "recipients": recipients,
    }

@api_router.get("/auth/permissions")
async def get_my_permissions(current_user: dict = Depends(get_current_user)):
    """Get current user's permissions"""
    checker = get_permission_checker(current_user)
    
    return {
        "role": current_user.get("role"),
        "additional_roles": current_user.get("additional_roles", []),
        "team_ids": list(checker.team_ids),
        "is_admin": checker.is_admin,
        "is_coach": checker.is_coach,
        "is_assistant_coach": checker.is_assistant_coach,
        "is_delegate": checker.is_delegate,
        "is_player": checker.is_player,
        "is_family_member": checker.is_family_member,
        "is_staff": checker.is_staff,
        "can_manage_team": checker.can_manage_team,
        "can_manage_events": checker.can_manage_events,
        "can_manage_stats": checker.can_manage_stats,
        "can_manage_attendance": checker.can_manage_attendance,
        "can_create_convocations": checker.can_create_convocations,
        "can_manage_lineups": checker.can_manage_lineups,
        "can_import_data": checker.can_import_data,
        "can_manage_club": checker.can_manage_club,
        "linked_player_id": current_user.get("linked_player_id"),
    }


@api_router.get("/auth/profiles")
async def get_my_profiles(current_user: dict = Depends(get_current_user)):
    """Get all available profiles for the current user"""
    return await build_available_profiles(current_user)


# ==================== ONBOARDING ROUTES (Phase O1 + O2) ====================
# Admin onboarding wizard — shell + routing (O1) plus per-step state so the
# wizard can resume across sessions (O2). Phases O3..O4 layer real Teams,
# Members and Invitations on top of the same /state surface.

ONBOARDING_ALLOWED_ROLES = {"admin", "gestor_desportivo"}
# Whitelisted step keys the client can mark as completed via PATCH /state.
ONBOARDING_STEP_KEYS = {
    "welcome", "club", "season", "teams", "members", "summary",
}


def _ensure_onboarding_role(current_user: dict) -> None:
    """Reject non-admin roles. The wizard is admin/gestor_desportivo only."""
    role = current_user.get("role")
    if role not in ONBOARDING_ALLOWED_ROLES:
        raise HTTPException(
            status_code=403,
            detail="Apenas administradores podem aceder ao onboarding",
        )


def _normalize_onboarding_state(raw: Optional[Dict[str, Any]]) -> Dict[str, Any]:
    """Return a safe, fully-shaped onboarding_state dict.

    Defensive against legacy users (None / partial dicts) so the frontend
    never has to guard against missing keys.
    """
    raw = raw or {}
    current_step = raw.get("current_step", 0)
    try:
        current_step = int(current_step)
    except (TypeError, ValueError):
        current_step = 0
    completed_steps = raw.get("completed_steps") or []
    if not isinstance(completed_steps, list):
        completed_steps = []
    return {
        "current_step": current_step,
        "completed_steps": [s for s in completed_steps if isinstance(s, str)],
        "club_id": raw.get("club_id") or None,
        "season_id": raw.get("season_id") or None,
    }


class OnboardingStatePatch(BaseModel):
    """Body for PATCH /api/onboarding/state. All fields optional / additive."""
    current_step: Optional[int] = None
    completed_step: Optional[str] = None
    club_id: Optional[str] = None
    season_id: Optional[str] = None


@api_router.get("/onboarding/status")
async def get_onboarding_status(current_user: dict = Depends(get_current_user)):
    """Return whether the current admin has finished the onboarding wizard
    plus the per-step resume state (current_step, completed_steps, club_id,
    season_id)."""
    _ensure_onboarding_role(current_user)
    completed_at = current_user.get("onboarding_completed_at")
    completed_at_iso: Optional[str] = None
    if isinstance(completed_at, datetime):
        completed_at_iso = completed_at.isoformat()
    elif isinstance(completed_at, str) and completed_at:
        completed_at_iso = completed_at

    state = _normalize_onboarding_state(current_user.get("onboarding_state"))
    return {
        "completed": completed_at_iso is not None,
        "completed_at": completed_at_iso,
        **state,
    }


@api_router.patch("/onboarding/state")
async def patch_onboarding_state(
    patch: OnboardingStatePatch,
    current_user: dict = Depends(get_current_user),
):
    """Merge per-step onboarding progress into the user document.

    Idempotent: marking the same step completed twice keeps it in the list
    exactly once. Returns the resulting normalized state.
    """
    _ensure_onboarding_role(current_user)

    state = _normalize_onboarding_state(current_user.get("onboarding_state"))

    if patch.current_step is not None:
        if patch.current_step < 0:
            raise HTTPException(status_code=400, detail="current_step inválido")
        state["current_step"] = patch.current_step

    if patch.completed_step is not None:
        if patch.completed_step not in ONBOARDING_STEP_KEYS:
            raise HTTPException(
                status_code=400,
                detail=f"completed_step inválido: {patch.completed_step}",
            )
        if patch.completed_step not in state["completed_steps"]:
            state["completed_steps"].append(patch.completed_step)

    if patch.club_id is not None:
        state["club_id"] = patch.club_id or None
    if patch.season_id is not None:
        state["season_id"] = patch.season_id or None

    await db.users.update_one(
        {"id": current_user["id"]},
        {"$set": {"onboarding_state": state}},
    )
    return state

@api_router.post("/onboarding/complete")
async def complete_onboarding(current_user: dict = Depends(get_current_user)):
    _ensure_onboarding_role(current_user)

    existing = current_user.get("onboarding_completed_at")
    if existing:
        completed_at_iso = (
            existing.isoformat()
            if isinstance(existing, datetime)
            else existing
        )
        return {
            "completed": True,
            "completed_at": completed_at_iso
        }

    now = datetime.now(timezone.utc)

    await db.users.update_one(
        {"id": current_user["id"]},
        {
            "$set": {
                "onboarding_completed_at": now.isoformat(),
                "onboarding_state.completed": True
            }
        }
    )

    return {
        "completed": True,
        "completed_at": now.isoformat()
    }

# ---- Phase O4 — Invitations preview + batch dispatch ---------------------

class OnboardingSendInvitesRequest(BaseModel):
    """Body for POST /api/onboarding/send-invites.

    ``member_ids`` is optional: when null/empty the endpoint sends to every
    pending (non-activated) member of the admin's onboarding club.
    """
    member_ids: Optional[List[str]] = None


def _is_email_dry_run() -> bool:
    """Mirror the dry-run heuristic used by services.emails: in non-prod
    environments without a Resend key, mail goes nowhere. The endpoint
    surfaces this so the wizard can warn the admin."""
    env = os.environ.get("ENVIRONMENT", "development").lower()
    if env == "production":
        return False
    return not os.environ.get("RESEND_API_KEY")


def _parse_invite_expiry(raw) -> Optional[datetime]:
    if not raw:
        return None
    if isinstance(raw, datetime):
        return raw if raw.tzinfo else raw.replace(tzinfo=timezone.utc)
    if isinstance(raw, str):
        try:
            parsed = datetime.fromisoformat(raw)
            return parsed if parsed.tzinfo else parsed.replace(tzinfo=timezone.utc)
        except ValueError:
            return None
    return None


@api_router.get("/onboarding/invite-preview")
async def get_invite_preview(current_user: dict = Depends(get_current_user)):
    """Return the list of members the admin's onboarding wizard created so
    the Summary step can render a preview table before dispatching invites.

    Only returns non-activated members of the admin's onboarding club —
    activated users no longer need an invite and admins themselves are
    excluded so the operator never accidentally invites their own account.
    """
    _ensure_onboarding_role(current_user)

    state = _normalize_onboarding_state(current_user.get("onboarding_state"))
    club_id = state.get("club_id")
    if not club_id:
        return {"club_id": None, "members": [], "dry_run": _is_email_dry_run()}

    members = await db.users.find(
        {
            "club_id": club_id,
            "role": {"$nin": ["admin", "gestor_desportivo"]},
            "is_activated": {"$ne": True},
        },
        {"_id": 0, "hashed_password": 0, "password_reset_token_hash": 0},
    ).to_list(500)

    # Cross-reference with teams to attach a team name (best-effort).
    team_names: Dict[str, str] = {}
    for tid_list in (m.get("team_ids", []) for m in members):
        for tid in tid_list or []:
            team_names.setdefault(tid, "")
    if team_names:
        team_docs = await db.teams.find(
            {"id": {"$in": list(team_names.keys())}}, {"_id": 0, "id": 1, "name": 1}
        ).to_list(500)
        for td in team_docs:
            team_names[td["id"]] = td.get("name", "")

    out = []
    for m in members:
        team_id = (m.get("team_ids") or [None])[0]
        out.append({
            "id": m["id"],
            "name": m.get("name"),
            "email": m.get("email"),
            "role": m.get("role"),
            "team_id": team_id,
            "team_name": team_names.get(team_id, "") if team_id else "",
            "has_token": bool(m.get("invite_token")),
        })

    return {
        "club_id": club_id,
        "members": out,
        "dry_run": _is_email_dry_run(),
    }


@api_router.post("/onboarding/send-invites")
async def send_onboarding_invites(
    request: OnboardingSendInvitesRequest,
    current_user: dict = Depends(get_current_user),
):
    """Batch-send activation invites to pending members of the admin's
    onboarding club.

    * Admin / gestor_desportivo only.
    * If ``member_ids`` is provided, only those IDs are processed; otherwise
      every pending member of the club is invited.
    * Already-activated members are skipped (returned in ``skipped``).
    * Members not belonging to the admin's onboarding club_id are skipped
      with reason ``foreign_club`` — never silently invite someone outside
      the wizard's scope.
    * Existing invite tokens are reused if still valid; new tokens are
      generated otherwise so links never come back as expired.
    * Returns a per-member result list plus aggregate counters.
    """
    _ensure_onboarding_role(current_user)

    state = _normalize_onboarding_state(current_user.get("onboarding_state"))
    club_id = state.get("club_id")
    if not club_id:
        raise HTTPException(
            status_code=400,
            detail="Onboarding sem clube configurado",
        )

    query: Dict[str, Any] = {
        "club_id": club_id,
        "role": {"$nin": ["admin", "gestor_desportivo"]},
    }
    if request.member_ids:
        query["id"] = {"$in": request.member_ids}

    candidates = await db.users.find(
        query, {"_id": 0, "hashed_password": 0}
    ).to_list(500)

    # If the caller specified ids that don't exist for this club, surface
    # them as failed rows so the wizard can show "foreign_club" reasons.
    found_ids = {c["id"] for c in candidates}
    missing_ids: List[str] = []
    if request.member_ids:
        missing_ids = [mid for mid in request.member_ids if mid not in found_ids]

    sent: List[Dict[str, Any]] = []
    skipped: List[Dict[str, Any]] = []
    failed: List[Dict[str, Any]] = []
    dry_run = _is_email_dry_run()
    now = datetime.now(timezone.utc)

    for member in candidates:
        if member.get("is_activated") is True:
            skipped.append({
                "id": member["id"],
                "email": member.get("email"),
                "reason": "already_activated",
            })
            continue

        # Reuse the existing token if still in the future, otherwise mint
        # a fresh one. Tokens always renew the 7-day window so the link
        # the recipient receives is valid for the next week.
        existing_token = member.get("invite_token")
        existing_expiry = _parse_invite_expiry(member.get("invite_expires_at"))
        token_reused = (
            bool(existing_token)
            and existing_expiry is not None
            and existing_expiry > now
        )
        token = existing_token if token_reused else secrets.token_urlsafe(32)
        expires_at = (now + timedelta(days=7)).isoformat()

        await db.users.update_one(
            {"id": member["id"]},
            {"$set": {"invite_token": token, "invite_expires_at": expires_at}},
        )

        try:
            ok = await send_activation_email(
                to_email=member["email"],
                name=member.get("name") or "Atleta",
                token=token,
                idempotency_key=f"onboarding-invite-{member['id']}-{token[:8]}",
            )
        except Exception as exc:  # noqa: BLE001
            failed.append({
                "id": member["id"],
                "email": member.get("email"),
                "reason": "send_exception",
                "error": f"{type(exc).__name__}: {exc}",
            })
            continue

        if ok:
            sent.append({
                "id": member["id"],
                "email": member.get("email"),
                "token_reused": token_reused,
            })
        else:
            failed.append({
                "id": member["id"],
                "email": member.get("email"),
                "reason": "send_failed",
            })

    for mid in missing_ids:
        failed.append({"id": mid, "email": None, "reason": "foreign_club"})

    return {
        "dry_run": dry_run,
        "sent_count": len(sent),
        "skipped_count": len(skipped),
        "failed_count": len(failed),
        "sent": sent,
        "skipped": skipped,
        "failed": failed,
    }


# ==================== USER ROUTES ====================

@api_router.get("/users", response_model=List[UserResponse])
async def get_users(role: Optional[str] = None, current_user: dict = Depends(get_current_user)):
    query = {}
    if role:
        query["role"] = role
    users = await db.users.find(query, {"_id": 0, "hashed_password": 0}).to_list(1000)
    return [UserResponse(**u) for u in users]


# ==================== ASSOCIATED ACCOUNTS ROUTES ====================
# NOTE: These routes MUST be defined BEFORE /users/{user_id} to avoid route conflicts

@api_router.get("/users/associated")
async def get_associated_accounts(current_user: dict = Depends(get_current_user)):
    """Get all accounts associated with the current user (children/athletes)"""
    associated_ids = current_user.get("associated_accounts", [])
    
    if not associated_ids:
        return []
    
    associated_users = await db.users.find(
        {"id": {"$in": associated_ids}},
        {"_id": 0, "hashed_password": 0}
    ).to_list(100)
    
    # Add relationship info
    for user in associated_users:
        user["relationship"] = "filho/a"
    
    return associated_users


@api_router.post("/users/associate")
async def associate_account(request: AssociateAccountRequest, current_user: dict = Depends(get_current_user)):
    """Associate a child/athlete account with the current user (parent/guardian)"""
    
    child = await db.users.find_one({"id": request.child_user_id}, {"_id": 0})
    if not child:
        raise HTTPException(status_code=404, detail="Conta não encontrada")
    
    if request.child_user_id in current_user.get("associated_accounts", []):
        raise HTTPException(status_code=400, detail="Conta já está associada")
    
    if child.get("parent_account_id"):
        raise HTTPException(status_code=400, detail="Esta conta já tem um responsável associado")
    
    await db.users.update_one(
        {"id": current_user["id"]},
        {"$addToSet": {"associated_accounts": request.child_user_id}}
    )
    
    await db.users.update_one(
        {"id": request.child_user_id},
        {"$set": {"parent_account_id": current_user["id"]}}
    )
    
    return {"message": f'Conta de {child["name"]} associada com sucesso', "child": child}


@api_router.post("/users/associate/search")
async def search_user_to_associate(email: str, current_user: dict = Depends(get_current_user)):
    """Search for a user by email to associate"""
    user = await db.users.find_one({"email": email}, {"_id": 0, "hashed_password": 0})
    
    if not user:
        raise HTTPException(status_code=404, detail="Utilizador não encontrado com este email")
    
    if user["id"] == current_user["id"]:
        raise HTTPException(status_code=400, detail="Não pode associar a sua própria conta")
    
    if user["id"] in current_user.get("associated_accounts", []):
        raise HTTPException(status_code=400, detail="Esta conta já está associada")
    
    if user.get("parent_account_id"):
        raise HTTPException(status_code=400, detail="Esta conta já tem um responsável")
    
    return {
        "id": user["id"],
        "name": user["name"],
        "email": user["email"],
        "role": user["role"],
        "team_ids": user.get("team_ids", [])
    }


@api_router.delete("/users/associate/{child_id}")
async def remove_association(child_id: str, current_user: dict = Depends(get_current_user)):
    """Remove association with a child account"""
    
    if child_id not in current_user.get("associated_accounts", []):
        raise HTTPException(status_code=404, detail="Associação não encontrada")
    
    await db.users.update_one(
        {"id": current_user["id"]},
        {"$pull": {"associated_accounts": child_id}}
    )
    
    await db.users.update_one(
        {"id": child_id},
        {"$set": {"parent_account_id": None}}
    )
    
    return {"message": "Associação removida com sucesso"}


@api_router.put("/users/{user_id}/admin-role")
async def toggle_admin_role(user_id: str, role_data: dict, current_user: dict = Depends(get_current_user)):
    """Grant or revoke admin role - only admins can do this"""
    checker = get_permission_checker(current_user)
    
    if not checker.is_admin:
        raise HTTPException(status_code=403, detail="Apenas administradores podem alterar roles de admin")
    
    # Can't remove own admin role
    if user_id == current_user["id"] and not role_data.get("is_admin", True):
        raise HTTPException(status_code=400, detail="Não pode remover o seu próprio role de admin")
    
    user = await db.users.find_one({"id": user_id}, {"_id": 0})
    if not user:
        raise HTTPException(status_code=404, detail="Utilizador não encontrado")
    
    is_admin = role_data.get("is_admin", False)
    
    if is_admin:
        # Grant admin role
        await db.users.update_one(
            {"id": user_id},
            {"$set": {"role": "admin"}}
        )
        return {"message": f'Role de admin concedido a {user["name"]}', "role": "admin"}
    else:
        # Remove admin role - set to most common role in their teams or jogador
        new_role = "jogador"
        team_roles = user.get("team_roles", {})
        if team_roles:
            # Use the most common role from their team roles
            roles_count = {}
            for role in team_roles.values():
                roles_count[role] = roles_count.get(role, 0) + 1
            if roles_count:
                new_role = max(roles_count, key=roles_count.get)
        
        await db.users.update_one(
            {"id": user_id},
            {"$set": {"role": new_role}}
        )
        return {"message": f'Role de admin removido de {user["name"]}', "role": new_role}


class LinkPlayerRequest(BaseModel):
    """Request to link a family member to a player"""
    player_id: str


@api_router.post("/users/link-player")
async def link_family_member_to_player(request: LinkPlayerRequest, current_user: dict = Depends(get_current_user)):
    """Link a family member (responsavel) to a player they are responsible for"""
    
    # Only family members (responsavel) can be linked to players
    if current_user.get("role") != "responsavel":
        raise HTTPException(status_code=400, detail="Apenas responsáveis/familiares podem ser ligados a jogadores")
    
    # Check if player exists and is a player
    player = await db.users.find_one({"id": request.player_id}, {"_id": 0})
    if not player:
        raise HTTPException(status_code=404, detail="Jogador não encontrado")
    
    if player.get("role") != "jogador":
        raise HTTPException(status_code=400, detail="O utilizador selecionado não é um jogador")
    
    # Update the family member with linked_player_id
    await db.users.update_one(
        {"id": current_user["id"]},
        {"$set": {
            "linked_player_id": request.player_id,
            "team_ids": player.get("team_ids", [])  # Give family member access to player's teams
        }}
    )
    
    return {
        "message": f'Ligado com sucesso ao jogador {player["name"]}',
        "linked_player": {
            "id": player["id"],
            "name": player["name"],
            "team_ids": player.get("team_ids", [])
        }
    }


@api_router.post("/users/link-players")
async def link_multiple_players(request: dict, current_user: dict = Depends(get_current_user)):
    """Link a family member to multiple players - for family accounts"""
    
    # Only family members (responsavel/familiar) can be linked to players
    if current_user.get("role") not in ["responsavel", "familiar"]:
        raise HTTPException(status_code=400, detail="Apenas responsáveis/familiares podem ser ligados a jogadores")
    
    player_ids = request.get("player_ids", [])
    if not player_ids:
        raise HTTPException(status_code=400, detail="Deve fornecer pelo menos um jogador")
    
    # Verify all players exist
    players = await db.users.find({"id": {"$in": player_ids}, "role": "jogador"}, {"_id": 0}).to_list(100)
    
    if len(players) != len(player_ids):
        raise HTTPException(status_code=404, detail="Um ou mais jogadores não encontrados")
    
    # Collect all team_ids from linked players
    all_team_ids = set()
    for player in players:
        all_team_ids.update(player.get("team_ids", []))
    
    # Update the family member with linked_player_ids
    await db.users.update_one(
        {"id": current_user["id"]},
        {"$set": {
            "linked_player_ids": player_ids,
            "linked_player_id": player_ids[0] if player_ids else None,  # Keep backwards compatibility
            "team_ids": list(all_team_ids)  # Give family member access to all linked players' teams
        }}
    )
    
    return {
        "message": f"Ligado com sucesso a {len(players)} jogador(es)",
        "linked_players": [{"id": p["id"], "name": p["name"]} for p in players]
    }


@api_router.delete("/users/link-player")
async def unlink_family_member_from_player(current_user: dict = Depends(get_current_user)):
    """Remove the link between a family member and a player"""
    
    if not current_user.get("linked_player_id"):
        raise HTTPException(status_code=400, detail="Não está ligado a nenhum jogador")
    
    await db.users.update_one(
        {"id": current_user["id"]},
        {"$set": {"linked_player_id": None}}
    )
    
    return {"message": "Ligação removida com sucesso"}


@api_router.post("/auth/switch-profile")
async def switch_profile(request: ActiveProfileRequest, current_user: dict = Depends(get_current_user)):
    """Switch to a different profile (self or associated account)"""
    
    if request.profile_type == "self":
        target_user = current_user
        active_role = request.active_role or current_user["role"]
    elif request.profile_type == "associated":
        if not request.associated_user_id:
            raise HTTPException(status_code=400, detail="ID da conta associada é obrigatório")
        
        if request.associated_user_id not in current_user.get("associated_accounts", []):
            raise HTTPException(status_code=403, detail="Conta não está associada a si")
        
        target_user = await db.users.find_one({"id": request.associated_user_id}, {"_id": 0, "hashed_password": 0})
        if not target_user:
            raise HTTPException(status_code=404, detail="Conta associada não encontrada")
        
        active_role = "responsavel"
    else:
        raise HTTPException(status_code=400, detail="Tipo de perfil inválido")
    
    teams = []
    for team_id in target_user.get("team_ids", []):
        team = await db.teams.find_one({"id": team_id}, {"_id": 0})
        if team:
            teams.append(team)
    
    return {
        "profile_type": request.profile_type,
        "viewing_as": {
            "id": target_user["id"],
            "name": target_user["name"],
            "role": active_role,
            "teams": teams
        },
        "original_user": {
            "id": current_user["id"],
            "name": current_user["name"]
        }
    }

# ==================== USER BY ID ROUTES ====================

@api_router.get("/users/{user_id}", response_model=UserResponse)
async def get_user(user_id: str, current_user: dict = Depends(get_current_user)):
    user = await db.users.find_one({"id": user_id}, {"_id": 0, "hashed_password": 0})
    if not user:
        raise HTTPException(status_code=404, detail="Utilizador não encontrado")
    return UserResponse(**user)

@api_router.put("/users/{user_id}")
async def update_user(user_id: str, updates: dict, current_user: dict = Depends(get_current_user)):
    if current_user['id'] != user_id and not is_admin_role(current_user['role']):
        raise HTTPException(status_code=403, detail="Sem permissão")
    
    # Check if user has permission to edit this data
    user_permissions = get_user_permissions(current_user)
    
    allowed_fields = ['name', 'surname', 'phone', 'avatar_url']
    
    # Handle profile updates
    if 'profile' in updates:
        profile_data = updates.pop('profile')
        
        # Filter family data based on permissions
        if 'family_members' in profile_data:
            if not user_permissions.get('can_edit_family_data', False) and current_user['id'] != user_id:
                del profile_data['family_members']
        
        # Update profile in database
        if profile_data:
            # First ensure profile object exists (to avoid "Cannot create field in null" error)
            await db.users.update_one(
                {"id": user_id, "profile": None},
                {"$set": {"profile": {}}}
            )
            # Then update the profile fields
            await db.users.update_one(
                {"id": user_id},
                {"$set": {f"profile.{k}": v for k, v in profile_data.items()}}
            )
    
    # Filter basic fields
    filtered_updates = {k: v for k, v in updates.items() if k in allowed_fields}
    
    if filtered_updates:
        await db.users.update_one({"id": user_id}, {"$set": filtered_updates})
    
    return {"message": "Utilizador atualizado"}

@api_router.put("/users/{user_id}/role")
async def update_user_role(user_id: str, role_data: dict, current_user: dict = Depends(get_current_user)):
    """Update user role - Admin only"""
    if not is_admin_role(current_user['role']):
        raise HTTPException(status_code=403, detail="Apenas administradores podem alterar permissões")
    
    new_role = role_data.get('role')
    if new_role not in ['admin', 'gestor_desportivo', 'treinador', 'treinador_adjunto', 'delegado', 'jogador', 'responsavel']:
        raise HTTPException(status_code=400, detail="Role inválido")
    
    # Cannot demote yourself
    if current_user['id'] == user_id and not is_admin_role(new_role):
        raise HTTPException(status_code=400, detail="Não podes remover o teu próprio privilégio de admin")
    
    user = await db.users.find_one({"id": user_id})
    if not user:
        raise HTTPException(status_code=404, detail="Utilizador não encontrado")
    
    await db.users.update_one({"id": user_id}, {"$set": {"role": new_role}})
    
    return {"message": f"Permissão alterada para {new_role}"}

def get_user_permissions(user: dict) -> dict:
    """Get effective permissions for a user"""
    role = user.get('role', 'jogador')
    base_permissions = DEFAULT_PERMISSIONS.get(role, DEFAULT_PERMISSIONS['jogador']).copy()
    
    # Apply custom permissions if set
    custom = user.get('custom_permissions', {})
    if custom:
        base_permissions.update(custom)
    
    return base_permissions

# ==================== CLUB ROUTES ====================

@api_router.post("/clubs")
async def create_club(club_data: ClubCreate, current_user: dict = Depends(get_current_user)):
    if not is_admin_role(current_user['role']):
        raise HTTPException(status_code=403, detail="Apenas administradores podem criar clubes")
    
    club = Club(**club_data.model_dump())
    club.admin_ids.append(current_user['id'])
    
    club_dict = club.model_dump()
    club_dict['created_at'] = club_dict['created_at'].isoformat()
    await db.clubs.insert_one(club_dict)
    club_dict.pop('_id', None)
    
    return club_dict

@api_router.get("/clubs")
async def get_clubs(current_user: dict = Depends(get_current_user)):
    clubs = await db.clubs.find({}, {"_id": 0}).to_list(100)
    return clubs

@api_router.get("/clubs/{club_id}")
async def get_club(club_id: str, current_user: dict = Depends(get_current_user)):
    club = await db.clubs.find_one({"id": club_id}, {"_id": 0})
    if not club:
        raise HTTPException(status_code=404, detail="Clube não encontrado")
    return club

@api_router.put("/clubs/{club_id}")
async def update_club(club_id: str, updates: dict, current_user: dict = Depends(get_current_user)):
    club = await db.clubs.find_one({"id": club_id})
    if not club:
        raise HTTPException(status_code=404, detail="Clube não encontrado")
    
    if not is_admin_role(current_user['role']) and current_user['id'] not in club.get('admin_ids', []):
        raise HTTPException(status_code=403, detail="Sem permissão")
    
    allowed_fields = ['name', 'acronym', 'logo_url', 'address', 'city', 'country', 'founded_year', 'website', 'email', 'phone', 'venue_name', 'venue_location', 'primary_color', 'secondary_color', 'accent_color', 'theme_mode', 'timezone', 'sidebar_accent_color']
    filtered_updates = {k: v for k, v in updates.items() if k in allowed_fields}
    
    if filtered_updates:
        await db.clubs.update_one({"id": club_id}, {"$set": filtered_updates})
    
    return {"message": "Clube atualizado"}

# ==================== SEASONS ROUTES ====================

@api_router.post("/clubs/{club_id}/seasons")
async def create_season(club_id: str, season_data: SeasonCreate, current_user: dict = Depends(get_current_user)):
    """Create a new season for a club"""
    if current_user['role'] not in ['admin', 'gestor_desportivo']:
        raise HTTPException(status_code=403, detail="Sem permissão para criar temporadas")
    
    club = await db.clubs.find_one({"id": club_id})
    if not club:
        raise HTTPException(status_code=404, detail="Clube não encontrado")
    
    # If this season is active, deactivate all other seasons
    if season_data.is_active:
        await db.seasons.update_many(
            {"club_id": club_id},
            {"$set": {"is_active": False}}
        )
    
    season = Season(
        club_id=club_id,
        name=season_data.name,
        start_date=season_data.start_date,
        end_date=season_data.end_date,
        is_active=season_data.is_active
    )
    
    await db.seasons.insert_one(season.model_dump())
    
    return {"message": "Temporada criada", "season": season.model_dump()}

@api_router.get("/clubs/{club_id}/seasons")
async def get_seasons(club_id: str, current_user: dict = Depends(get_current_user)):
    """Get all seasons for a club"""
    seasons = await db.seasons.find({"club_id": club_id}, {"_id": 0}).sort("start_date", -1).to_list(100)
    return seasons

@api_router.get("/clubs/{club_id}/seasons/active")
async def get_active_season(club_id: str, current_user: dict = Depends(get_current_user)):
    """Get the active season for a club"""
    season = await db.seasons.find_one({"club_id": club_id, "is_active": True}, {"_id": 0})
    return season

@api_router.put("/clubs/{club_id}/seasons/{season_id}")
async def update_season(club_id: str, season_id: str, updates: SeasonUpdate, current_user: dict = Depends(get_current_user)):
    """Update a season"""
    if current_user['role'] not in ['admin', 'gestor_desportivo']:
        raise HTTPException(status_code=403, detail="Sem permissão para editar temporadas")
    
    season = await db.seasons.find_one({"id": season_id, "club_id": club_id})
    if not season:
        raise HTTPException(status_code=404, detail="Temporada não encontrada")
    
    update_data = updates.model_dump(exclude_unset=True)
    
    # If setting this season as active, deactivate others
    if update_data.get('is_active'):
        await db.seasons.update_many(
            {"club_id": club_id, "id": {"$ne": season_id}},
            {"$set": {"is_active": False}}
        )
    
    if update_data:
        await db.seasons.update_one({"id": season_id}, {"$set": update_data})
    
    return {"message": "Temporada atualizada"}

@api_router.delete("/clubs/{club_id}/seasons/{season_id}")
async def delete_season(club_id: str, season_id: str, current_user: dict = Depends(get_current_user)):
    """Delete a season"""
    if current_user['role'] not in ['admin', 'gestor_desportivo']:
        raise HTTPException(status_code=403, detail="Sem permissão para eliminar temporadas")
    
    result = await db.seasons.delete_one({"id": season_id, "club_id": club_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Temporada não encontrada")
    
    return {"message": "Temporada eliminada"}

@api_router.put("/clubs/{club_id}/seasons/{season_id}/activate")
async def activate_season(club_id: str, season_id: str, current_user: dict = Depends(get_current_user)):
    """Set a season as active (deactivates all others)"""
    if current_user['role'] not in ['admin', 'gestor_desportivo']:
        raise HTTPException(status_code=403, detail="Sem permissão para ativar temporadas")
    
    season = await db.seasons.find_one({"id": season_id, "club_id": club_id})
    if not season:
        raise HTTPException(status_code=404, detail="Temporada não encontrada")
    
    # Deactivate all seasons
    await db.seasons.update_many({"club_id": club_id}, {"$set": {"is_active": False}})
    
    # Activate this season
    await db.seasons.update_one({"id": season_id}, {"$set": {"is_active": True}})
    
    return {"message": "Temporada ativada"}

# ==================== SUBSCRIPTION ROUTES ====================

@api_router.get("/subscription")
async def get_subscription(current_user: dict = Depends(get_current_user)):
    """Get subscription for current user's club"""
    if current_user['role'] not in ['admin', 'gestor_desportivo']:
        raise HTTPException(status_code=403, detail="Sem permissão para ver subscrição")
    
    # Get club
    club = await db.clubs.find_one({}, {"_id": 0})
    if not club:
        raise HTTPException(status_code=404, detail="Clube não encontrado")
    
    # Get or create subscription
    subscription = await db.subscriptions.find_one({"club_id": club['id']}, {"_id": 0})
    
    if not subscription:
        # Create default subscription
        from datetime import timedelta
        today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
        one_year_later = (datetime.now(timezone.utc) + timedelta(days=365)).strftime("%Y-%m-%d")
        
        # Count active members
        member_count = await db.users.count_documents({"is_archived": {"$ne": True}})
        
        new_subscription = Subscription(
            club_id=club['id'],
            plan_type="standard",
            start_date=today,
            end_date=one_year_later,
            status="active",
            payment_method="bank_transfer",
            member_count=member_count
        )
        await db.subscriptions.insert_one(new_subscription.model_dump())
        subscription = new_subscription.model_dump()
    
    # Update member count
    member_count = await db.users.count_documents({"is_archived": {"$ne": True}})
    if subscription.get('member_count') != member_count:
        await db.subscriptions.update_one(
            {"id": subscription['id']},
            {"$set": {"member_count": member_count}}
        )
        subscription['member_count'] = member_count
    
    return subscription

@api_router.patch("/subscription")
async def update_subscription(updates: SubscriptionUpdate, current_user: dict = Depends(get_current_user)):
    """Update subscription settings"""
    if current_user['role'] not in ['admin', 'gestor_desportivo']:
        raise HTTPException(status_code=403, detail="Sem permissão para editar subscrição")
    
    club = await db.clubs.find_one({}, {"_id": 0})
    if not club:
        raise HTTPException(status_code=404, detail="Clube não encontrado")
    
    subscription = await db.subscriptions.find_one({"club_id": club['id']})
    if not subscription:
        raise HTTPException(status_code=404, detail="Subscrição não encontrada")
    
    update_data = updates.model_dump(exclude_unset=True)
    if update_data:
        await db.subscriptions.update_one(
            {"id": subscription['id']},
            {"$set": update_data}
        )
    
    return {"message": "Subscrição atualizada"}

@api_router.post("/subscription/cancel")
async def cancel_subscription(current_user: dict = Depends(get_current_user)):
    """Cancel the subscription"""
    if not is_admin_role(current_user['role']):
        raise HTTPException(status_code=403, detail="Apenas administradores podem cancelar subscrições")
    
    club = await db.clubs.find_one({}, {"_id": 0})
    if not club:
        raise HTTPException(status_code=404, detail="Clube não encontrado")
    
    subscription = await db.subscriptions.find_one({"club_id": club['id']})
    if not subscription:
        raise HTTPException(status_code=404, detail="Subscrição não encontrada")
    
    await db.subscriptions.update_one(
        {"id": subscription['id']},
        {"$set": {"status": "cancelled"}}
    )
    
    return {"message": "Subscrição cancelada"}

@api_router.get("/subscription/invoices")
async def get_invoices(current_user: dict = Depends(get_current_user)):
    """Get all invoices for the subscription"""
    if current_user['role'] not in ['admin', 'gestor_desportivo']:
        raise HTTPException(status_code=403, detail="Sem permissão para ver faturas")
    
    club = await db.clubs.find_one({}, {"_id": 0})
    if not club:
        raise HTTPException(status_code=404, detail="Clube não encontrado")
    
    subscription = await db.subscriptions.find_one({"club_id": club['id']}, {"_id": 0})
    if not subscription:
        return []
    
    invoices = await db.subscription_invoices.find(
        {"subscription_id": subscription['id']},
        {"_id": 0}
    ).sort("created_at", -1).to_list(100)
    
    return invoices

@api_router.post("/subscription/invoices")
async def create_invoice(invoice_data: SubscriptionInvoiceCreate, current_user: dict = Depends(get_current_user)):
    """Create a new invoice (admin only)"""
    if not is_admin_role(current_user['role']):
        raise HTTPException(status_code=403, detail="Apenas administradores podem criar faturas")
    
    club = await db.clubs.find_one({}, {"_id": 0})
    if not club:
        raise HTTPException(status_code=404, detail="Clube não encontrado")
    
    subscription = await db.subscriptions.find_one({"club_id": club['id']})
    if not subscription:
        raise HTTPException(status_code=404, detail="Subscrição não encontrada")
    
    # Generate invoice number
    count = await db.subscription_invoices.count_documents({"club_id": club['id']})
    invoice_number = f"INV-{datetime.now().year}-{str(count + 1).zfill(3)}"
    
    invoice = SubscriptionInvoice(
        subscription_id=subscription['id'],
        club_id=club['id'],
        invoice_number=invoice_number,
        start_date=invoice_data.start_date,
        end_date=invoice_data.end_date,
        paying_members=invoice_data.paying_members,
        price_per_member=invoice_data.price_per_member,
        total_due=invoice_data.total_due,
        status="pending"
    )
    
    await db.subscription_invoices.insert_one(invoice.model_dump())
    
    return {"message": "Fatura criada", "invoice": invoice.model_dump()}

@api_router.get("/subscription/invoices/{invoice_id}")
async def get_invoice(invoice_id: str, current_user: dict = Depends(get_current_user)):
    """Get a specific invoice"""
    if current_user['role'] not in ['admin', 'gestor_desportivo']:
        raise HTTPException(status_code=403, detail="Sem permissão para ver faturas")
    
    invoice = await db.subscription_invoices.find_one({"id": invoice_id}, {"_id": 0})
    if not invoice:
        raise HTTPException(status_code=404, detail="Fatura não encontrada")
    
    return invoice

@api_router.get("/subscription/invoices/{invoice_id}/download")
async def download_invoice(invoice_id: str, current_user: dict = Depends(get_current_user)):
    """Get download link for invoice"""
    if current_user['role'] not in ['admin', 'gestor_desportivo']:
        raise HTTPException(status_code=403, detail="Sem permissão para descarregar faturas")
    
    invoice = await db.subscription_invoices.find_one({"id": invoice_id}, {"_id": 0})
    if not invoice:
        raise HTTPException(status_code=404, detail="Fatura não encontrada")
    
    if not invoice.get('file_url'):
        raise HTTPException(status_code=404, detail="Ficheiro da fatura não disponível")
    
    return {"download_url": invoice['file_url']}

@api_router.patch("/subscription/invoices/{invoice_id}")
async def update_invoice(invoice_id: str, updates: dict, current_user: dict = Depends(get_current_user)):
    """Update an invoice (mark as paid, etc.)"""
    if not is_admin_role(current_user['role']):
        raise HTTPException(status_code=403, detail="Apenas administradores podem editar faturas")
    
    invoice = await db.subscription_invoices.find_one({"id": invoice_id})
    if not invoice:
        raise HTTPException(status_code=404, detail="Fatura não encontrada")
    
    allowed_fields = ['status', 'total_paid', 'paid_at', 'file_url']
    filtered_updates = {k: v for k, v in updates.items() if k in allowed_fields}
    
    if filtered_updates:
        await db.subscription_invoices.update_one(
            {"id": invoice_id},
            {"$set": filtered_updates}
        )
    
    return {"message": "Fatura atualizada"}

# ==================== PERMISSIONS ROUTES ====================

@api_router.get("/permissions/defaults")
async def get_default_permissions(current_user: dict = Depends(get_current_user)):
    """Get default permissions for all roles"""
    if not is_admin_role(current_user['role']):
        raise HTTPException(status_code=403, detail="Apenas administradores podem ver permissões")
    return DEFAULT_PERMISSIONS

@api_router.get("/permissions/{user_id}")
async def get_user_permissions_endpoint(user_id: str, current_user: dict = Depends(get_current_user)):
    """Get effective permissions for a specific user"""
    if not is_admin_role(current_user['role']) and current_user['id'] != user_id:
        raise HTTPException(status_code=403, detail="Sem permissão")
    
    user = await db.users.find_one({"id": user_id}, {"_id": 0, "password": 0})
    if not user:
        raise HTTPException(status_code=404, detail="Utilizador não encontrado")
    
    return get_user_permissions(user)

@api_router.put("/permissions/{user_id}")
async def update_user_permissions(user_id: str, permissions: dict, current_user: dict = Depends(get_current_user)):
    """Update custom permissions for a user (admin only)"""
    if not is_admin_role(current_user['role']):
        raise HTTPException(status_code=403, detail="Apenas administradores podem modificar permissões")
    
    user = await db.users.find_one({"id": user_id})
    if not user:
        raise HTTPException(status_code=404, detail="Utilizador não encontrado")
    
    # Validate permission keys
    valid_keys = set(DEFAULT_PERMISSIONS['admin'].keys())
    filtered_permissions = {k: v for k, v in permissions.items() if k in valid_keys}
    
    await db.users.update_one(
        {"id": user_id},
        {"$set": {"custom_permissions": filtered_permissions}}
    )
    
    return {"message": "Permissões atualizadas"}

# ==================== GUARDIAN (PARENT) ROUTES ====================

@api_router.get("/guardian/children")
async def get_guardian_children(current_user: dict = Depends(get_current_user)):
    """
    Get list of children (linked players) for a guardian/parent user.
    Returns children with their teams count.
    """
    user_role = current_user.get('role')
    
    # Only responsavel (parent/guardian) can access this
    if user_role != 'responsavel':
        raise HTTPException(status_code=403, detail="Apenas responsáveis podem aceder a esta funcionalidade")
    
    # Get linked player IDs from user
    linked_player_ids = current_user.get('linked_player_ids', [])
    linked_player_id = current_user.get('linked_player_id')
    
    # Combine both fields (backwards compatibility)
    all_linked = list(set(linked_player_ids + ([linked_player_id] if linked_player_id else [])))
    
    if not all_linked:
        return []
    
    # Fetch children data
    children = []
    for child_id in all_linked:
        child = await db.users.find_one({"id": child_id}, {"_id": 0, "password": 0})
        if child:
            # Count teams
            team_ids = child.get('team_ids', [])
            children.append({
                "id": child['id'],
                "name": child.get('name', 'Atleta'),
                "avatar_url": child.get('avatar_url'),
                "email": child.get('email'),
                "role": child.get('role'),
                "teams_count": len(team_ids),
                "team_ids": team_ids
            })
    
    return children

@api_router.get("/guardian/children/{child_id}/teams")
async def get_guardian_child_teams(child_id: str, current_user: dict = Depends(get_current_user)):
    """
    Get teams and clubs for a specific child.
    Returns same structure as "As minhas equipas".
    """
    user_role = current_user.get('role')
    
    # Only responsavel (parent/guardian) can access this
    if user_role != 'responsavel':
        raise HTTPException(status_code=403, detail="Apenas responsáveis podem aceder a esta funcionalidade")
    
    # Verify that child is linked to this parent
    linked_player_ids = current_user.get('linked_player_ids', [])
    linked_player_id = current_user.get('linked_player_id')
    all_linked = list(set(linked_player_ids + ([linked_player_id] if linked_player_id else [])))
    
    if child_id not in all_linked:
        raise HTTPException(status_code=403, detail="Este atleta não está associado à sua conta")
    
    # Fetch child data
    child = await db.users.find_one({"id": child_id}, {"_id": 0})
    if not child:
        raise HTTPException(status_code=404, detail="Atleta não encontrado")
    
    # Get child's teams
    child_team_ids = child.get('team_ids', [])
    
    teams = []
    if child_team_ids:
        teams = await db.teams.find({"id": {"$in": child_team_ids}}, {"_id": 0}).to_list(100)
        
        # Add child's role in each team
        for team in teams:
            team_role = 'jogador'  # Default
            if child_id in team.get('coach_ids', []):
                team_role = 'treinador'
            elif child_id in team.get('assistant_coach_ids', []):
                team_role = 'treinador_adjunto'
            elif child_id in team.get('delegate_ids', []):
                team_role = 'delegado'
            team['child_role'] = team_role
            team['child_name'] = child.get('name', 'Atleta').split(' ')[0]
    
    # Get club data
    clubs = await db.clubs.find({}, {"_id": 0}).to_list(10)
    club = clubs[0] if clubs else None
    
    return {
        "child": {
            "id": child['id'],
            "name": child.get('name'),
            "avatar_url": child.get('avatar_url')
        },
        "teams": teams,
        "club": club
    }

# ==================== TEAM ROUTES ====================

@api_router.post("/teams", response_model=Team)
async def create_team(team_data: TeamCreate, current_user: dict = Depends(get_current_user)):
    checker = get_permission_checker(current_user)
    
    if not checker.can_manage_team:
        raise HTTPException(status_code=403, detail="Sem permissão para criar equipas")
    
    team = Team(**team_data.model_dump())
    if checker.is_coach and not checker.is_admin:
        team.coach_ids.append(current_user['id'])
    
    team_dict = team.model_dump()
    team_dict['created_at'] = team_dict['created_at'].isoformat()
    await db.teams.insert_one(team_dict)
    return team

@api_router.get("/teams", response_model=List[Team])
async def get_teams(current_user: dict = Depends(get_current_user)):
    checker = get_permission_checker(current_user)
    
    if checker.is_admin:
        teams = await db.teams.find({}, {"_id": 0}).to_list(100)
    else:
        # Filter teams based on user's team_ids
        user_team_ids = list(checker.team_ids)
        if user_team_ids:
            teams = await db.teams.find({"id": {"$in": user_team_ids}}, {"_id": 0}).to_list(100)
        else:
            # Legacy fallback: check if user is in team's coach/delegate/player lists
            user_id = current_user['id']
            teams = await db.teams.find({
                "$or": [{"coach_ids": user_id}, {"delegate_ids": user_id}, {"player_ids": user_id}]
            }, {"_id": 0}).to_list(100)
    
    for team in teams:
        if isinstance(team.get('created_at'), str):
            team['created_at'] = datetime.fromisoformat(team['created_at'])
    return teams

@api_router.get("/teams/{team_id}")
async def get_team(team_id: str, current_user: dict = Depends(get_current_user)):
    checker = get_permission_checker(current_user)
    
    # Check team access (admin can access all, others need team assignment)
    if not checker.is_admin and not checker.can_access_team(team_id):
        raise HTTPException(status_code=403, detail="Sem acesso a esta equipa")
    
    team = await db.teams.find_one({"id": team_id}, {"_id": 0})
    if not team:
        raise HTTPException(status_code=404, detail="Equipa não encontrada")
    if isinstance(team.get('created_at'), str):
        team['created_at'] = datetime.fromisoformat(team['created_at'])
    return team

@api_router.put("/teams/{team_id}")
async def update_team(team_id: str, team_data: TeamUpdate, current_user: dict = Depends(get_current_user)):
    checker = get_permission_checker(current_user)
    
    if not checker.can_manage_team:
        raise HTTPException(status_code=403, detail="Sem permissão para editar equipas")
    
    if not checker.is_admin and not checker.can_access_team(team_id):
        raise HTTPException(status_code=403, detail="Sem acesso a esta equipa")
    
    team = await db.teams.find_one({"id": team_id})
    if not team:
        raise HTTPException(status_code=404, detail="Equipa não encontrada")
    
    # Build update dict with non-None values
    update_data = {k: v for k, v in team_data.model_dump().items() if v is not None}
    
    if update_data:
        await db.teams.update_one({"id": team_id}, {"$set": update_data})
    
    updated_team = await db.teams.find_one({"id": team_id}, {"_id": 0})
    return updated_team

@api_router.delete("/teams/{team_id}")
async def delete_team(team_id: str, current_user: dict = Depends(get_current_user)):
    checker = get_permission_checker(current_user)
    
    if not checker.can_manage_club:
        raise HTTPException(status_code=403, detail="Apenas administradores podem eliminar equipas")
    
    team = await db.teams.find_one({"id": team_id})
    if not team:
        raise HTTPException(status_code=404, detail="Equipa não encontrada")
    
    # Remove team from all users
    await db.users.update_many(
        {"team_ids": team_id},
        {"$pull": {"team_ids": team_id}}
    )
    
    # Delete related events
    await db.events.delete_many({"team_id": team_id})
    
    # Delete related championships
    await db.championships.delete_many({"team_id": team_id})
    
    # Delete the team
    await db.teams.delete_one({"id": team_id})
    
    return {"message": "Equipa eliminada com sucesso"}


@api_router.post("/teams/{team_id}/members")
async def add_team_member(team_id: str, member_data: dict, current_user: dict = Depends(get_current_user)):
    checker = get_permission_checker(current_user)
    
    if not checker.can_manage_team:
        raise HTTPException(status_code=403, detail="Sem permissão para gerir membros")
    
    if not checker.is_admin and not checker.can_access_team(team_id):
        raise HTTPException(status_code=403, detail="Sem acesso a esta equipa")
    
    team = await db.teams.find_one({"id": team_id})
    if not team:
        raise HTTPException(status_code=404, detail="Equipa não encontrada")
    
    user_id = member_data.get('user_id')
    role = member_data.get('role', 'jogador')
    
    # Map role to team field
    field_map = {'treinador': 'coach_ids', 'treinador_adjunto': 'coach_ids', 'delegado': 'delegate_ids', 'jogador': 'player_ids', 'responsavel': 'player_ids', 'familiar': 'player_ids'}
    field = field_map.get(role, 'player_ids')
    
    # Add to team
    await db.teams.update_one({"id": team_id}, {"$addToSet": {field: user_id}})
    
    # Update user: add team_id and set team_roles mapping
    await db.users.update_one(
        {"id": user_id}, 
        {
            "$addToSet": {"team_ids": team_id},
            "$set": {f"team_roles.{team_id}": role}
        }
    )
    
    return {"message": "Membro adicionado à equipa", "role": role}

@api_router.put("/teams/{team_id}/members/{user_id}/role")
async def update_team_member_role(team_id: str, user_id: str, role_data: dict, current_user: dict = Depends(get_current_user)):
    """Update a member's role within a specific team"""
    checker = get_permission_checker(current_user)
    
    if not checker.can_manage_team:
        raise HTTPException(status_code=403, detail="Sem permissão para gerir membros")
    
    if not checker.is_admin and not checker.can_access_team(team_id):
        raise HTTPException(status_code=403, detail="Sem acesso a esta equipa")
    
    new_role = role_data.get('role', 'jogador')
    
    # Remove from old field and add to new field in team
    await db.teams.update_one(
        {"id": team_id}, 
        {"$pull": {"coach_ids": user_id, "delegate_ids": user_id, "player_ids": user_id}}
    )
    
    field_map = {'treinador': 'coach_ids', 'treinador_adjunto': 'coach_ids', 'delegado': 'delegate_ids', 'jogador': 'player_ids', 'responsavel': 'player_ids', 'familiar': 'player_ids'}
    field = field_map.get(new_role, 'player_ids')
    
    await db.teams.update_one({"id": team_id}, {"$addToSet": {field: user_id}})
    
    # Update team_roles mapping
    await db.users.update_one(
        {"id": user_id},
        {"$set": {f"team_roles.{team_id}": new_role}}
    )
    
    return {"message": "Role atualizado", "role": new_role}

@api_router.delete("/teams/{team_id}/members/{user_id}")
async def remove_team_member(team_id: str, user_id: str, current_user: dict = Depends(get_current_user)):
    checker = get_permission_checker(current_user)
    
    if not checker.can_manage_team:
        raise HTTPException(status_code=403, detail="Sem permissão para gerir membros")
    
    if not checker.is_admin and not checker.can_access_team(team_id):
        raise HTTPException(status_code=403, detail="Sem acesso a esta equipa")
    
    await db.teams.update_one({"id": team_id}, {"$pull": {"coach_ids": user_id, "delegate_ids": user_id, "player_ids": user_id}})
    await db.users.update_one({"id": user_id}, {"$pull": {"team_ids": team_id}})
    
    return {"message": "Membro removido da equipa"}

@api_router.get("/teams/{team_id}/members")
async def get_team_members(team_id: str, current_user: dict = Depends(get_current_user)):
    checker = get_permission_checker(current_user)
    
    # Check team access
    if not checker.is_admin and not checker.can_access_team(team_id):
        raise HTTPException(status_code=403, detail="Sem acesso a esta equipa")
    
    team = await db.teams.find_one({"id": team_id}, {"_id": 0})
    if not team:
        raise HTTPException(status_code=404, detail="Equipa não encontrada")
    
    all_ids = team.get('coach_ids', []) + team.get('delegate_ids', []) + team.get('player_ids', [])
    
    # For family members, only show linked player
    if checker.is_family_member and checker.linked_player_id:
        if checker.linked_player_id in all_ids:
            all_ids = [checker.linked_player_id]
        else:
            all_ids = []
    
    # For players, show all team members (read-only view)
    members = await db.users.find({"id": {"$in": all_ids}}, {"_id": 0, "password": 0}).to_list(100)
    
    result = []
    for member in members:
        member_role = "jogador"
        if member['id'] in team.get('coach_ids', []):
            member_role = "treinador"
        elif member['id'] in team.get('delegate_ids', []):
            member_role = "delegado"
        result.append({**member, "team_role": member_role})
    
    return result

class MemberCreate(BaseModel):
    name: str
    email: Optional[EmailStr] = None
    role: UserRole = "jogador"
    team_id: Optional[str] = None
    club_id: Optional[str] = None
    jersey_number: Optional[str] = None
    position: Optional[str] = None
    phone: Optional[str] = None
    nationalities: Optional[List[str]] = None

    # StickPro v2.0 — responsável / tutor associado ao atleta
    guardian_name: Optional[str] = None
    guardian_email: Optional[EmailStr] = None
    guardian_relationship: Optional[str] = None
    language: Optional[str] = "pt"

    # Phase O3 — onboarding wizard sets this to True so the activation
    # email is deferred to the Invitations step. Default keeps historical
    # behaviour for every other caller.
    suppress_invite: bool = False

class FamilyMemberCreate(BaseModel):
    first_name: str
    surname: Optional[str] = ""
    email: Optional[EmailStr] = None
    phone: Optional[str] = ""
    relationship: str = "pai"

class MemberUpdate(BaseModel):
    name: Optional[str] = None
    email: Optional[EmailStr] = None
    role: Optional[UserRole] = None
    jersey_number: Optional[str] = None
    position: Optional[str] = None
    phone: Optional[str] = None
    nationalities: Optional[List[str]] = None
    is_archived: Optional[bool] = None
    is_activated: Optional[bool] = None
    invite_token: Optional[str] = None
    invite_expires_at: Optional[str] = None


@api_router.post("/members")
async def create_member(data: MemberCreate, current_user: dict = Depends(get_current_user)):
    """Create a new member (user) associated with the club with pending activation"""
    checker = get_permission_checker(current_user)

    if not checker.can_manage_team:
        raise HTTPException(status_code=403, detail="Sem permissão para criar membros")

    if data.team_id and not checker.is_admin and not checker.can_access_team(data.team_id):
        raise HTTPException(status_code=403, detail="Sem acesso a esta equipa")

    club_id = data.club_id
    if not club_id:
        club = await db.clubs.find_one({}, {"_id": 0, "id": 1})
        if club:
            club_id = club["id"]

    member_email = data.email.strip().lower() if data.email else None

    if data.role != "jogador" and not member_email:
        raise HTTPException(
            status_code=400,
            detail="O email é obrigatório para funções técnicas e administrativas"
        )
    
    if data.role == "jogador" and not member_email and not data.guardian_email:
        raise HTTPException(
            status_code=400,
            detail="Para criar um atleta sem email próprio, indique o email de um responsável"
        )
    
    if member_email:
        existing_user = await db.users.find_one({"email": member_email}, {"_id": 0})
        if existing_user:
            raise HTTPException(
                status_code=400,
                detail="Já existe um utilizador com este email. Use 'Adicionar membros do clube'."
            )
    
    user_id = str(uuid.uuid4())
    
    # Email técnico interno para atletas menores sem email próprio.
    # Não deve ser usado para login nem enviado ao utilizador.
    if not member_email:
        member_email = f"player-{user_id}@stickpro.local"
    
    invite_token = secrets.token_urlsafe(32)
    invite_expires_at = (datetime.now(timezone.utc) + timedelta(days=7)).isoformat()
    
    guardian_email = data.guardian_email.strip().lower() if data.guardian_email else None
    
    user = {
        "id": user_id,
        "name": data.name,
        "email": member_email,
        "has_real_email": bool(data.email),
        "login_enabled": bool(data.email),
        "hashed_password": None,
        "role": data.role,
        "club_id": club_id,
        "team_ids": [data.team_id] if data.team_id else [],
        "is_activated": False,
        "invite_token": invite_token if data.email else None,
        "invite_expires_at": invite_expires_at if data.email else None,
        "guardian_name": data.guardian_name,
        "guardian_email": guardian_email,
        "guardian_relationship": data.guardian_relationship,
        "guardian_emails": [guardian_email] if guardian_email else [],
        "profile": {
            "photo_url": "",
            "first_name": data.name,
            "surname": "",
            "nickname": "",
            "birth_date": "",
            "gender": "",
            "nationality": data.nationalities[0] if data.nationalities else "",
            "nationalities": data.nationalities or [],
            "fpp_license": "",
            "family_members": [
                {
                    "id": str(uuid.uuid4()),
                    "first_name": data.guardian_name or "",
                    "surname": "",
                    "email": guardian_email,
                    "phone": "",
                    "relationship": data.guardian_relationship or "pai"
                }
            ] if guardian_email else [],
            "weight": "",
            "height": "",
            "shoe_size": "",
            "year_joined_club": "",
            "fpp_number": "",
            "function": data.role,
            "position": data.position or "",
            "jersey_number": data.jersey_number or "",
            "training_kit_size": "",
            "tracksuit_size": "",
            "polo_size": "",
            "training_sock_size": ""
        },
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.users.insert_one(user)

    if data.role == "jogador" and guardian_email:
    
        guardian_invite_token = secrets.token_urlsafe(32)
    
        guardian_invite_expires_at = (
            datetime.now(timezone.utc) + timedelta(days=7)
        ).isoformat()
    
        guardian_link = {
            "id": str(uuid.uuid4()),
            "player_id": user_id,
            "player_name": data.name,
            "guardian_user_id": None,
            "guardian_name": data.guardian_name or "",
            "guardian_email": guardian_email,
            "relationship": data.guardian_relationship or "pai",
            "language": data.language or "pt",
            "status": "pending",
            "is_primary": True,
            "club_id": club_id,
            "team_id": data.team_id,
            "invite_token": guardian_invite_token,
            "invite_expires_at": guardian_invite_expires_at,
            "permissions": {
                "receive_notifications": True,
                "respond_convocations": True,
                "justify_absences": True,
                "view_calendar": True,
                "view_feedback": True,
                "view_evaluations": True,
                "edit_player_profile": False
            },
            "created_at": datetime.now(timezone.utc).isoformat()
        }
    
        await db.guardian_links.insert_one(guardian_link)
    
        try:
            await send_family_invitation_email(
                to_email=guardian_email,
                guardian_name=data.guardian_name or "",
                player_name=data.name,
                relationship=data.guardian_relationship or "pai",
                token=guardian_invite_token,
                idempotency_key=f"guardian-invite-{guardian_link['id']}",
            )
        except Exception as e:
            logger.warning(
                f"[FAMILY INVITE EMAIL] failed to send to {guardian_email}: {e}"
            )
            
    if data.team_id:
        field_map = {
            "treinador": "coach_ids",
            "treinador_adjunto": "assistant_coach_ids",
            "delegado": "delegate_ids"
        }
        field = field_map.get(data.role, "player_ids")
        await db.teams.update_one({"id": data.team_id}, {"$addToSet": {field: user_id}})

    frontend_url = os.environ.get("FRONTEND_URL", "").rstrip("/")
    activation_link = f"{frontend_url}/activate-account?token={invite_token}" if frontend_url else ""

    # Phase E2: dispatch activation email via the new service. Failure to
    # send must NOT abort the user creation flow; activation_link is still
    # returned so the operator can hand it over manually if needed.
    # Phase O3: the onboarding wizard sets ``suppress_invite=True`` so the
    # email is not fired here — it is dispatched later by the Invitations
    # step (O4). The activation token is still persisted on the user.
    if not data.suppress_invite and data.email:
        try:
            await send_activation_email(
                to_email=user["email"],
                name=user["name"],
                token=invite_token,
                idempotency_key=f"member-create-{user_id}",
            )
        except Exception as e:
            logger.warning(f"[ACTIVATION EMAIL] failed to send to {user['email']}: {e}")

    safe_user = {k: v for k, v in user.items() if k not in ("hashed_password", "_id")}

    return {
        "user": safe_user,
        "activation_link": activation_link,
        "invite_token": invite_token
    }


@api_router.get("/clubs/{club_id}/members")
async def get_club_members(club_id: str, current_user: dict = Depends(get_current_user)):
    """Get all members that belong to the club"""
    members = await db.users.find(
        {"club_id": club_id, "role": {"$ne": "admin"}},
        {"_id": 0, "hashed_password": 0}
    ).to_list(500)
    return members


@api_router.post("/members/{member_id}/send-invite")
async def send_member_invite(member_id: str, current_user: dict = Depends(get_current_user)):
    checker = get_permission_checker(current_user)

    if not checker.can_manage_team:
        raise HTTPException(status_code=403, detail="Sem permissão para enviar convites")

    user = await db.users.find_one({"id": member_id}, {"_id": 0, "hashed_password": 0})
    if not user:
        raise HTTPException(status_code=404, detail="Membro não encontrado")

    if user.get("is_activated"):
        raise HTTPException(status_code=400, detail="A conta já está ativada")

    invite_token = secrets.token_urlsafe(32)
    invite_expires_at = (datetime.now(timezone.utc) + timedelta(days=7)).isoformat()

    await db.users.update_one(
        {"id": member_id},
        {
            "$set": {
                "invite_token": invite_token,
                "invite_expires_at": invite_expires_at
            }
        }
    )

    frontend_url = os.environ.get("FRONTEND_URL", "").rstrip("/")
    activation_link = f"{frontend_url}/activate-account?token={invite_token}" if frontend_url else ""

    # Phase E2: dispatch the activation email. Email failure does not abort
    # the flow — operator still gets the activation_link in the response.
    try:
        await send_activation_email(
            to_email=user["email"],
            name=user.get("name", "Atleta"),
            token=invite_token,
            idempotency_key=f"send-invite-{member_id}-{invite_token[:8]}",
        )
    except Exception as e:
        logger.warning(f"[ACTIVATION EMAIL] send-invite to {user.get('email')}: {e}")

    return {
        "message": "Convite gerado com sucesso",
        "activation_link": activation_link
    }


@api_router.get("/clubs/{club_id}/members/search")
async def search_club_members(
    club_id: str,
    query: str,
    current_user: dict = Depends(get_current_user)
):
    checker = get_permission_checker(current_user)

    if not checker.can_manage_team:
        raise HTTPException(status_code=403, detail="Sem permissão")

    members = await db.users.find(
        {
            "club_id": club_id,
            "role": {"$ne": "admin"},
            "$or": [
                {"name": {"$regex": query, "$options": "i"}},
                {"email": {"$regex": query, "$options": "i"}}
            ]
        },
        {"_id": 0, "hashed_password": 0}
    ).to_list(50)

    return members


@api_router.post("/teams/{team_id}/add-existing-member/{member_id}")
async def add_existing_member_to_team(
    team_id: str,
    member_id: str,
    current_user: dict = Depends(get_current_user)
):
    checker = get_permission_checker(current_user)

    if not checker.can_manage_team:
        raise HTTPException(status_code=403, detail="Sem permissão")

    if not checker.is_admin and not checker.can_access_team(team_id):
        raise HTTPException(status_code=403, detail="Sem acesso a esta equipa")

    team = await db.teams.find_one({"id": team_id}, {"_id": 0})
    if not team:
        raise HTTPException(status_code=404, detail="Equipa não encontrada")

    user = await db.users.find_one({"id": member_id}, {"_id": 0, "hashed_password": 0})
    if not user:
        raise HTTPException(status_code=404, detail="Membro não encontrado")

    role = user.get("role", "jogador")
    field_map = {
        "treinador": "coach_ids",
        "treinador_adjunto": "assistant_coach_ids",
        "delegado": "delegate_ids"
    }
    field = field_map.get(role, "player_ids")

    await db.teams.update_one({"id": team_id}, {"$addToSet": {field: member_id}})
    await db.users.update_one({"id": member_id}, {"$addToSet": {"team_ids": team_id}})

    return {"message": "Membro adicionado à equipa com sucesso"}


@api_router.post("/members/import")
async def import_members(file: UploadFile = File(...), team_id: str = None, club_id: str = None, current_user: dict = Depends(get_current_user)):
    """Import members from Excel/CSV file - members are associated with the club with pending activation"""
    checker = get_permission_checker(current_user)
    
    if not checker.can_import_data:
        raise HTTPException(status_code=403, detail="Sem permissão para importar dados")
    
    if team_id and not checker.is_admin and not checker.can_access_team(team_id):
        raise HTTPException(status_code=403, detail="Sem acesso a esta equipa")
    
    if not club_id:
        club = await db.clubs.find_one({}, {"_id": 0, "id": 1})
        if club:
            club_id = club["id"]
    
    content = await file.read()
    results = {"success": 0, "errors": [], "created": [], "warnings": []}
    
    try:
        if file.filename.endswith(".csv"):
            import csv
            reader = csv.DictReader(io.StringIO(content.decode("utf-8")))
            rows = list(reader)
        else:
            try:
                import openpyxl
                wb = openpyxl.load_workbook(io.BytesIO(content))
                ws = wb.active
                headers = [cell.value for cell in ws[1]]
                rows = []
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if any(row):
                        rows.append(dict(zip(headers, row)))
            except ImportError:
                import csv
                reader = csv.DictReader(io.StringIO(content.decode("utf-8")))
                rows = list(reader)
        
        for row in rows:
            try:
                nome = (row.get("Nome") or row.get("nome") or row.get("Nombre") or row.get("nombre") or 
                       row.get("Prénom") or row.get("prenom") or row.get("prénom") or 
                       row.get("First Name") or row.get("first_name") or row.get("name") or "")
                
                apelido = (row.get("Apelido") or row.get("apelido") or row.get("Sobrenome") or row.get("sobrenome") or
                          row.get("Apellido") or row.get("apellido") or 
                          row.get("Nom") or row.get("nom") or 
                          row.get("Cognome") or row.get("cognome") or
                          row.get("Surname") or row.get("surname") or row.get("last_name") or "")
                
                data_nascimento = (row.get("Data de Nascimento") or row.get("data_nascimento") or row.get("nascimento") or
                                  row.get("Fecha de Nacimiento") or row.get("fecha_nacimiento") or
                                  row.get("Date de Naissance") or row.get("date_naissance") or
                                  row.get("Data di Nascita") or row.get("data_nascita") or
                                  row.get("Date of Birth") or row.get("birth_date") or row.get("dob") or "")
                
                email = (row.get("Email") or row.get("email") or row.get("E-mail") or row.get("e-mail") or
                        row.get("Correo") or row.get("correo") or "")
                
                funcao = (row.get("Função") or row.get("funcao") or row.get("função") or
                         row.get("Rol") or row.get("rol") or
                         row.get("Fonction") or row.get("fonction") or
                         row.get("Ruolo") or row.get("ruolo") or
                         row.get("Role") or row.get("role") or "jogador")
                
                numero = (row.get("Número") or row.get("numero") or row.get("Nº") or row.get("nº") or row.get("N") or row.get("n") or
                         row.get("Numero") or
                         row.get("Numéro") or row.get("numéro") or
                         row.get("Number") or row.get("number") or row.get("Jersey") or row.get("jersey") or "")
                
                posicao = (row.get("Posição") or row.get("posicao") or row.get("posição") or
                          row.get("Posición") or row.get("posicion") or
                          row.get("Poste") or row.get("poste") or
                          row.get("Posizione") or row.get("posizione") or
                          row.get("Position") or row.get("position") or "")
                
                telefone = (row.get("Telefone") or row.get("telefone") or row.get("contacto") or
                           row.get("Teléfono") or row.get("telefono") or
                           row.get("Téléphone") or row.get("téléphone") or row.get("telephone") or
                           row.get("Telefono") or
                           row.get("Phone") or row.get("phone") or "")
                
                nacionalidade = (row.get("Nacionalidade") or row.get("nacionalidade") or
                                row.get("Nacionalidad") or row.get("nacionalidad") or
                                row.get("Nationalité") or row.get("nationalité") or row.get("nationalite") or
                                row.get("Nazionalità") or row.get("nazionalità") or row.get("nazionalita") or
                                row.get("Nationality") or row.get("nationality") or "")
                
                sexo = (row.get("Sexo") or row.get("sexo") or
                       row.get("Sexe") or row.get("sexe") or
                       row.get("Sesso") or row.get("sesso") or
                       row.get("Gender") or row.get("gender") or row.get("Sex") or row.get("sex") or "")
                
                full_name = f"{nome} {apelido}".strip() if apelido else nome.strip()
                
                if not full_name or not email:
                    results["errors"].append(f"Linha sem nome ou email: {row}")
                    continue

                normalized_email = email.strip().lower()
                existing_user = await db.users.find_one({"email": normalized_email}, {"_id": 0})
                if existing_user:
                    results["warnings"].append(f"Email já existente no clube/sistema: {normalized_email}")
                    continue
                
                user_id = str(uuid.uuid4())
                invite_token = secrets.token_urlsafe(32)
                invite_expires_at = (datetime.now(timezone.utc) + timedelta(days=7)).isoformat()
                
                funcao_map = {
                    "administrador": "admin",
                    "admin": "admin",
                    "gestor desportivo": "gestor_desportivo",
                    "gestor_desportivo": "gestor_desportivo",
                    "diretor desportivo": "gestor_desportivo",
                    "treinador": "treinador",
                    "treinador principal": "treinador",
                    "treinador adjunto": "treinador_adjunto",
                    "adjunto": "treinador_adjunto",
                    "delegado": "delegado",
                    "jogador": "jogador",
                    "atleta": "jogador",
                    "responsavel": "responsavel",
                    "responsável": "responsavel",
                    "pai": "responsavel",
                    "mãe": "responsavel",
                    "encarregado": "responsavel",
                    "administrator": "admin",
                    "sports manager": "gestor_desportivo",
                    "sports director": "gestor_desportivo",
                    "coach": "treinador",
                    "head coach": "treinador",
                    "assistant coach": "treinador_adjunto",
                    "assistant": "treinador_adjunto",
                    "delegate": "delegado",
                    "team manager": "delegado",
                    "player": "jogador",
                    "guardian": "responsavel",
                    "parent": "responsavel",
                    "family": "responsavel",
                    "gestor deportivo": "gestor_desportivo",
                    "director deportivo": "gestor_desportivo",
                    "entrenador": "treinador",
                    "entrenador principal": "treinador",
                    "entrenador asistente": "treinador_adjunto",
                    "entrenador adjunto": "treinador_adjunto",
                    "jugador": "jogador",
                    "responsable": "responsavel",
                    "familiar": "responsavel",
                    "administrateur": "admin",
                    "responsable sportif": "gestor_desportivo",
                    "directeur sportif": "gestor_desportivo",
                    "entraîneur": "treinador",
                    "entraineur": "treinador",
                    "entraîneur adjoint": "treinador_adjunto",
                    "entraineur adjoint": "treinador_adjunto",
                    "délégué": "delegado",
                    "delegue": "delegado",
                    "joueur": "jogador",
                    "amministratore": "admin",
                    "responsabile sportivo": "gestor_desportivo",
                    "direttore sportivo": "gestor_desportivo",
                    "allenatore": "treinador",
                    "allenatore in seconda": "treinador_adjunto",
                    "vice allenatore": "treinador_adjunto",
                    "delegato": "delegado",
                    "giocatore": "jogador",
                }
                role = funcao_map.get(funcao.lower().strip(), "jogador") if funcao else "jogador"
                
                posicao_map = {
                    "guarda-redes": "GR",
                    "gr": "GR",
                    "goalkeeper": "GR",
                    "portero": "GR",
                    "jogador de campo": "JC",
                    "jc": "JC",
                    "field player": "JC",
                    "jugador de campo": "JC",
                    "avançado": "JC",
                    "defesa": "JC",
                }
                normalized_position = posicao_map.get(str(posicao).lower().strip(), str(posicao).strip().upper() if posicao else "")
                
                sexo_map = {
                    "masculino": "Masculino",
                    "m": "Masculino",
                    "male": "Masculino",
                    "hombre": "Masculino",
                    "feminino": "Feminino",
                    "f": "Feminino",
                    "female": "Feminino",
                    "mujer": "Feminino",
                }
                normalized_gender = sexo_map.get(str(sexo).lower().strip(), "") if sexo else ""
                
                nationalities = []
                if nacionalidade:
                    nationality_map = {
                        "portuguesa": "PT", "portugal": "PT", "pt": "PT",
                        "espanhola": "ES", "espanha": "ES", "spain": "ES", "es": "ES", "española": "ES",
                        "francesa": "FR", "frança": "FR", "france": "FR", "fr": "FR",
                        "brasileira": "BR", "brasil": "BR", "brazil": "BR", "br": "BR",
                        "italiana": "IT", "itália": "IT", "italy": "IT", "it": "IT",
                        "alemã": "DE", "alemanha": "DE", "germany": "DE", "de": "DE",
                        "inglesa": "GB", "inglaterra": "GB", "uk": "GB", "gb": "GB", "england": "GB",
                        "americana": "US", "eua": "US", "usa": "US", "us": "US",
                        "angolana": "AO", "angola": "AO", "ao": "AO",
                        "moçambicana": "MZ", "moçambique": "MZ", "mozambique": "MZ", "mz": "MZ",
                        "cabo-verdiana": "CV", "cabo verde": "CV", "cv": "CV",
                        "guineense": "GW", "guiné-bissau": "GW", "gw": "GW",
                        "são-tomense": "ST", "são tomé": "ST", "st": "ST",
                        "holandesa": "NL", "holanda": "NL", "netherlands": "NL", "nl": "NL",
                        "belga": "BE", "bélgica": "BE", "belgium": "BE", "be": "BE",
                        "suíça": "CH", "switzerland": "CH", "ch": "CH",
                        "argentina": "AR", "ar": "AR",
                        "marroquina": "MA", "marrocos": "MA", "morocco": "MA", "ma": "MA",
                        "romena": "RO", "roménia": "RO", "romania": "RO", "ro": "RO",
                    }
                    nat_parts = [n.strip() for n in str(nacionalidade).replace(";", ",").split(",")]
                    for nat in nat_parts:
                        if nat:
                            code = nationality_map.get(nat.lower(), nat.upper()[:2] if len(nat) == 2 else None)
                            if code and code not in nationalities:
                                nationalities.append(code)
                
                user = {
                    "id": user_id,
                    "name": full_name,
                    "email": normalized_email,
                    "hashed_password": None,
                    "role": role,
                    "club_id": club_id,
                    "team_ids": [team_id] if team_id else [],
                    "nationalities": nationalities[:2],
                    "is_activated": False,
                    "invite_token": invite_token,
                    "invite_expires_at": invite_expires_at,
                    "profile": {
                        "sports_info": {
                            "jersey_number": str(numero).strip() if numero else "",
                            "position": normalized_position
                        },
                        "identity": {
                            "phone": str(telefone).strip() if telefone else "",
                            "birth_date": str(data_nascimento).strip() if data_nascimento else "",
                            "gender": normalized_gender
                        }
                    },
                    "created_at": datetime.now(timezone.utc).isoformat()
                }
                
                await db.users.insert_one(user)
                
                if team_id:
                    field_map = {
                        "treinador": "coach_ids",
                        "treinador_adjunto": "assistant_coach_ids",
                        "delegado": "delegate_ids"
                    }
                    field = field_map.get(role, "player_ids")
                    await db.teams.update_one({"id": team_id}, {"$addToSet": {field: user_id}})
                
                frontend_url = os.environ.get("FRONTEND_URL", "").rstrip("/")
                activation_link = f"{frontend_url}/activate-account?token={invite_token}" if frontend_url else ""

                results["success"] += 1
                results["created"].append({
                    "name": full_name,
                    "email": normalized_email,
                    "activation_link": activation_link
                })
                
            except Exception as e:
                results["errors"].append(f"Erro na linha: {str(e)}")
        
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Erro ao processar ficheiro: {str(e)}")
    
    return results

@api_router.post("/members/{member_id}/teams/{team_id}")
async def add_member_to_team(member_id: str, team_id: str, current_user: dict = Depends(get_current_user)):
    """Add a club member to a team"""
    checker = get_permission_checker(current_user)
    
    if not checker.can_manage_team:
        raise HTTPException(status_code=403, detail="Sem permissão para gerir membros")
    
    # Check team access
    if not checker.is_admin and not checker.can_access_team(team_id):
        raise HTTPException(status_code=403, detail="Sem acesso a esta equipa")
    
    # Get the member
    member = await db.users.find_one({"id": member_id}, {"_id": 0})
    if not member:
        raise HTTPException(status_code=404, detail="Membro não encontrado")
    
    # Get the team
    team = await db.teams.find_one({"id": team_id}, {"_id": 0})
    if not team:
        raise HTTPException(status_code=404, detail="Equipa não encontrada")
    
    # Add team to member's team_ids
    await db.users.update_one({"id": member_id}, {"$addToSet": {"team_ids": team_id}})
    
    # Add member to team's appropriate list
    role = member.get('role', 'jogador')
    field_map = {'treinador': 'coach_ids', 'treinador_adjunto': 'coach_ids', 'delegado': 'delegate_ids'}
    field = field_map.get(role, 'player_ids')
    await db.teams.update_one({"id": team_id}, {"$addToSet": {field: member_id}})
    
    return {"message": "Membro adicionado à equipa"}

@api_router.delete("/members/{member_id}/teams/{team_id}")
async def remove_member_from_team(member_id: str, team_id: str, current_user: dict = Depends(get_current_user)):
    """Remove a member from a team"""
    checker = get_permission_checker(current_user)
    
    if not checker.can_manage_team:
        raise HTTPException(status_code=403, detail="Sem permissão para gerir membros")
    
    # Check team access
    if not checker.is_admin and not checker.can_access_team(team_id):
        raise HTTPException(status_code=403, detail="Sem acesso a esta equipa")
    
    # Remove team from member's team_ids
    await db.users.update_one({"id": member_id}, {"$pull": {"team_ids": team_id}})
    
    # Remove member from all team lists
    await db.teams.update_one(
        {"id": team_id}, 
        {"$pull": {"player_ids": member_id, "coach_ids": member_id, "delegate_ids": member_id}}
    )
    
    return {"message": "Membro removido da equipa"}

@api_router.delete("/members/{member_id}")
async def delete_member_permanently(member_id: str, current_user: dict = Depends(get_current_user)):
    """Permanently delete a member and all associated data (Admin only)"""
    checker = get_permission_checker(current_user)
    
    # Only admin can permanently delete members
    if not checker.is_admin:
        raise HTTPException(status_code=403, detail="Apenas administradores podem eliminar membros permanentemente")
    
    # Check if member exists
    member = await db.users.find_one({"id": member_id})
    if not member:
        raise HTTPException(status_code=404, detail="Membro não encontrado")
    
    # Cannot delete yourself
    if member_id == current_user.get('id'):
        raise HTTPException(status_code=400, detail="Não pode eliminar a sua própria conta")
    
    # Cannot delete admin accounts
    if member.get('role') in ['admin']:
        raise HTTPException(status_code=400, detail="Não pode eliminar contas de administrador")
    
    # Remove member from all teams
    team_ids = member.get('team_ids', [])
    for team_id in team_ids:
        await db.teams.update_one(
            {"id": team_id},
            {"$pull": {"player_ids": member_id, "coach_ids": member_id, "delegate_ids": member_id}}
        )
    
    # Remove from linked_player_ids (guardians)
    await db.users.update_many(
        {"linked_player_ids": member_id},
        {"$pull": {"linked_player_ids": member_id}}
    )
    
    # Remove attendance records
    await db.attendance.delete_many({"user_id": member_id})
    
    # Remove player stats
    await db.player_stats.delete_many({"player_id": member_id})
    
    # Remove unavailabilities
    await db.unavailabilities.delete_many({"user_id": member_id})
    
    # Remove payments
    await db.payments.delete_many({"user_id": member_id})
    await db.monthly_fees.delete_many({"user_id": member_id})
    
    # Remove push subscriptions
    await db.push_subscriptions.delete_many({"user_id": member_id})
    
    # Remove user messages (private ones)
    await db.messages.delete_many({"sender_id": member_id})
    
    # Finally, delete the user
    result = await db.users.delete_one({"id": member_id})
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=500, detail="Erro ao eliminar membro")
    
    return {"message": "Membro eliminado permanentemente", "deleted_id": member_id}

@api_router.get("/members")
async def get_members_paginated(
    page: int = 1,
    per_page: int = 20,
    search: Optional[str] = None,
    team_id: Optional[str] = None,
    club_id: Optional[str] = None,
    include_archived: bool = False,
    current_user: dict = Depends(get_current_user)
):
    """Get paginated members list with search"""
    checker = get_permission_checker(current_user)
    
    query = {"role": {"$ne": "admin"}}
    
    # Filter by archived status
    if not include_archived:
        query["is_archived"] = {"$ne": True}
    
    # Apply team filter
    if team_id:
        if not checker.is_admin and not checker.can_access_team(team_id):
            raise HTTPException(status_code=403, detail="Sem acesso a esta equipa")
        query["team_ids"] = team_id
    elif club_id:
        query["club_id"] = club_id
    elif not checker.is_admin:
        # Non-admin can only see members from their teams
        user_teams = list(checker.team_ids)
        if user_teams:
            query["team_ids"] = {"$in": user_teams}
        else:
            return {"members": [], "total": 0, "page": page, "per_page": per_page, "total_pages": 0}
    
    # Apply search filter
    if search:
        query["name"] = {"$regex": search, "$options": "i"}
    
    # Get total count
    total = await db.users.count_documents(query)
    total_pages = (total + per_page - 1) // per_page
    
    # Get paginated results sorted alphabetically
    skip = (page - 1) * per_page
    members = await db.users.find(query, {"_id": 0, "password": 0}).sort("name", 1).skip(skip).limit(per_page).to_list(per_page)
    
    return {
        "members": members,
        "total": total,
        "page": page,
        "per_page": per_page,
        "total_pages": total_pages
    }

@api_router.get("/members/archived")
async def get_archived_members(
    page: int = 1,
    per_page: int = 20,
    search: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Get archived members - admin only"""
    checker = get_permission_checker(current_user)
    
    if not checker.is_admin:
        raise HTTPException(status_code=403, detail="Apenas administradores podem ver membros arquivados")
    
    query = {"is_archived": True, "role": {"$ne": "admin"}}
    
    if search:
        query["name"] = {"$regex": search, "$options": "i"}
    
    total = await db.users.count_documents(query)
    total_pages = (total + per_page - 1) // per_page
    
    skip = (page - 1) * per_page
    members = await db.users.find(query, {"_id": 0, "password": 0}).sort("name", 1).skip(skip).limit(per_page).to_list(per_page)
    
    return {
        "members": members,
        "total": total,
        "page": page,
        "per_page": per_page,
        "total_pages": total_pages
    }

@api_router.get("/members/export")
async def export_members_excel(
    team_id: Optional[str] = Query(None, description="Filter by team ID"),
    role: Optional[str] = Query(None, description="Filter by role"),
    search: Optional[str] = Query(None, description="Search by name or email"),
    current_user: dict = Depends(get_current_user)
):
    """Export members to Excel file - admin only"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    
    checker = get_permission_checker(current_user)
    
    if not checker.is_admin:
        raise HTTPException(status_code=403, detail="Apenas administradores podem exportar membros")
    
    # Build query
    query = {"is_archived": {"$ne": True}}
    
    if team_id:
        query["team_ids"] = team_id
    
    if role:
        query["role"] = role
    
    # Get members
    members = await db.users.find(query, {"_id": 0}).to_list(1000)
    
    # Get all teams for team names
    teams = await db.teams.find({}, {"_id": 0, "id": 1, "name": 1}).to_list(100)
    team_map = {t["id"]: t["name"] for t in teams}
    
    # Apply search filter
    if search:
        search_lower = search.lower()
        members = [m for m in members if 
                   search_lower in m.get('name', '').lower() or 
                   search_lower in m.get('email', '').lower()]
    
    # Role translations
    role_names = {
        'admin': 'Administrador',
        'treinador': 'Treinador',
        'treinador_adjunto': 'Treinador Adjunto',
        'jogador': 'Jogador',
        'delegado': 'Delegado',
        'familiar': 'Familiar'
    }
    
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Membros"
    
    # Define headers
    headers = [
        "Nome",
        "Email",
        "Equipa(s)",
        "Função",
        "Nacionalidade",
        "Data de Nascimento",
        "Telefone",
        "Número de Jogador",
        "Posição"
    ]
    
    # Header styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="0D9488", end_color="0D9488", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Write data
    for row_idx, member in enumerate(members, 2):
        profile = member.get('profile') or {}
        identity = profile.get('identity') or {}
        sports = profile.get('sports') or {}
        
        # Get team names
        team_ids = member.get('team_ids', [])
        team_names = [team_map.get(tid, '') for tid in team_ids if team_map.get(tid)]
        
        # Format birth date
        birth_date = identity.get('birth_date', '')
        if birth_date:
            try:
                dt = datetime.fromisoformat(birth_date.replace('Z', '+00:00'))
                birth_date = dt.strftime('%d/%m/%Y')
            except:
                pass
        
        row_data = [
            member.get('name', ''),
            member.get('email', ''),
            ', '.join(team_names),
            role_names.get(member.get('role', ''), member.get('role', '')),
            identity.get('nationality', ''),
            birth_date,
            identity.get('phone', ''),
            sports.get('player_number', ''),
            sports.get('position', '')
        ]
        
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")
    
    # Adjust column widths
    column_widths = [25, 30, 25, 18, 15, 14, 15, 12, 15]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    
    # Freeze header row
    ws.freeze_panes = 'A2'
    
    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Generate filename with timestamp
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f"membros_export_{timestamp}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@api_router.get("/members/{member_id}")
async def get_member_detail(member_id: str, current_user: dict = Depends(get_current_user)):
    """Get member details including statistics"""
    checker = get_permission_checker(current_user)
    
    member = await db.users.find_one({"id": member_id}, {"_id": 0, "password": 0})
    if not member:
        raise HTTPException(status_code=404, detail="Membro não encontrado")
    
    # Check access - admin can see all, others can see if in same team or is self
    if not checker.is_admin:
        if member_id != current_user['id']:
            member_teams = set(member.get('team_ids', []))
            user_teams = set(checker.team_ids)
            if not member_teams.intersection(user_teams):
                raise HTTPException(status_code=403, detail="Sem acesso a este membro")
    
    # Get statistics summary
    stats = {
        "total_events": 0,
        "attendance_rate": 0,
        "goals": 0,
        "assists": 0
    }
    
    # Count attendance
    attendances = await db.attendance.find({"player_id": member_id}, {"_id": 0}).to_list(500)
    if attendances:
        stats["total_events"] = len(attendances)
        confirmed = sum(1 for a in attendances if a.get('status') == 'confirmado')
        stats["attendance_rate"] = round((confirmed / len(attendances)) * 100, 1) if attendances else 0
    
    # Get player stats
    player_stats = await db.player_match_stats.find({"player_id": member_id}, {"_id": 0}).to_list(200)
    for ps in player_stats:
        stats["goals"] += ps.get('goals', 0)
        stats["assists"] += ps.get('assists', 0)
    
    return {
        "member": member,
        "statistics": stats
    }

@api_router.put("/members/{member_id}")
async def update_member(member_id: str, data: MemberUpdate, current_user: dict = Depends(get_current_user)):
    """Update member data"""
    checker = get_permission_checker(current_user)
    
    member = await db.users.find_one({"id": member_id}, {"_id": 0})
    if not member:
        raise HTTPException(status_code=404, detail="Membro não encontrado")
    
    # Check permissions
    is_own_profile = member_id == current_user['id']
    can_edit = checker.is_admin or (is_own_profile and not data.is_archived and not data.role)
    
    if not can_edit:
        # Staff can edit members in their teams (except role and archive)
        if checker.is_staff:
            member_teams = set(member.get('team_ids', []))
            user_teams = set(checker.team_ids)
            if member_teams.intersection(user_teams) and not data.is_archived and not data.role:
                can_edit = True
    
    if not can_edit:
        raise HTTPException(status_code=403, detail="Sem permissão para editar este membro")
    
    # Only admin can archive/unarchive or change role
    if (data.is_archived is not None or data.role is not None) and not checker.is_admin:
        raise HTTPException(status_code=403, detail="Apenas administradores podem arquivar membros ou alterar roles")
    
    # Build update dict
    update_data = {}
    if data.name:
        update_data["name"] = data.name
    if data.email:
        update_data["email"] = data.email
    if data.role:
        update_data["role"] = data.role
    if data.nationalities is not None:
        update_data["nationalities"] = data.nationalities[:2]  # Max 2
    if data.is_archived is not None:
        update_data["is_archived"] = data.is_archived
        if data.is_archived:
            update_data["archived_at"] = datetime.now(timezone.utc).isoformat()
        else:
            update_data["archived_at"] = None
    if data.is_activated is not None:
        update_data["is_activated"] = data.is_activated
    
    # Update profile fields
    if data.jersey_number is not None:
        update_data["profile.jersey_number"] = data.jersey_number
    if data.position is not None:
        update_data["profile.position"] = data.position
    if data.phone is not None:
        update_data["phone"] = data.phone
    
    if update_data:
        await db.users.update_one({"id": member_id}, {"$set": update_data})
    
    return {"message": "Membro atualizado"}

@api_router.post("/members/{member_id}/family")
async def add_family_member(member_id: str, data: FamilyMemberCreate, current_user: dict = Depends(get_current_user)):
    """Add a family member/responsible contact to an athlete profile and create guardian link if email exists."""
    checker = get_permission_checker(current_user)

    member = await db.users.find_one({"id": member_id}, {"_id": 0})
    if not member:
        raise HTTPException(status_code=404, detail="Membro não encontrado")

    if member.get("role") != "jogador":
        raise HTTPException(status_code=400, detail="Familiares só podem ser associados a atletas")

    can_edit = checker.is_admin

    if not can_edit and checker.is_staff:
        member_teams = set(member.get("team_ids", []))
        user_teams = set(checker.team_ids)
        can_edit = bool(member_teams.intersection(user_teams))

    if not can_edit and member_id == current_user.get("id"):
        can_edit = True

    if not can_edit:
        raise HTTPException(status_code=403, detail="Sem permissão para editar familiares deste atleta")

    family_member = {
        "id": str(uuid.uuid4()),
        "first_name": data.first_name,
        "surname": data.surname or "",
        "email": data.email.strip().lower() if data.email else "",
        "phone": data.phone or "",
        "relationship": data.relationship or "pai"
    }

    await db.users.update_one(
        {"id": member_id},
        {"$addToSet": {"profile.family_members": family_member}}
    )

    guardian_link = None

    if data.email:
        guardian_email = data.email.strip().lower()

        existing_link = await db.guardian_links.find_one({
            "player_id": member_id,
            "guardian_email": guardian_email
        })

        if not existing_link:
            guardian_link = {
                "id": str(uuid.uuid4()),
                "player_id": member_id,
                "player_name": member.get("name"),
                "guardian_user_id": None,
                "guardian_name": f"{data.first_name} {data.surname or ''}".strip(),
                "guardian_email": guardian_email,
                "relationship": data.relationship or "pai",
                "status": "pending",
                "is_primary": False,
                "club_id": member.get("club_id"),
                "team_id": member.get("team_ids", [None])[0] if member.get("team_ids") else None,
                "permissions": {
                    "receive_notifications": True,
                    "respond_convocations": True,
                    "justify_absences": True,
                    "view_calendar": True,
                    "view_feedback": True,
                    "view_evaluations": True,
                    "edit_player_profile": False
                },
                "created_at": datetime.now(timezone.utc).isoformat()
            }

            await db.guardian_links.insert_one(guardian_link)

    return {
        "message": "Familiar adicionado",
        "family_member": family_member,
        "guardian_link": guardian_link
    }

@api_router.put("/members/{member_id}/family/{family_member_id}")
async def update_family_member(
    member_id: str,
    family_member_id: str,
    data: FamilyMemberCreate,
    current_user: dict = Depends(get_current_user)
):
    """Update a family member/responsible contact."""
    checker = get_permission_checker(current_user)

    member = await db.users.find_one({"id": member_id}, {"_id": 0})
    if not member:
        raise HTTPException(status_code=404, detail="Membro não encontrado")

    can_edit = checker.is_admin

    if not can_edit and checker.is_staff:
        member_teams = set(member.get("team_ids", []))
        user_teams = set(checker.team_ids)
        can_edit = bool(member_teams.intersection(user_teams))

    if not can_edit:
        raise HTTPException(status_code=403, detail="Sem permissão para editar familiares deste atleta")

    updated_family_member = {
        "id": family_member_id,
        "first_name": data.first_name,
        "surname": data.surname or "",
        "email": data.email.strip().lower() if data.email else "",
        "phone": data.phone or "",
        "relationship": data.relationship or "pai"
    }

    result = await db.users.update_one(
        {"id": member_id, "profile.family_members.id": family_member_id},
        {"$set": {"profile.family_members.$": updated_family_member}}
    )

    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Familiar não encontrado")

    if data.email:
        guardian_email = data.email.strip().lower()
        await db.guardian_links.update_one(
            {"player_id": member_id, "guardian_email": guardian_email},
            {
                "$set": {
                    "guardian_name": f"{data.first_name} {data.surname or ''}".strip(),
                    "relationship": data.relationship or "pai",
                    "updated_at": datetime.now(timezone.utc).isoformat()
                }
            }
        )

    return {
        "message": "Familiar atualizado",
        "family_member": updated_family_member
    }
    
@api_router.post("/members/{member_id}/archive")
async def archive_member(member_id: str, current_user: dict = Depends(get_current_user)):
    """Archive a member without deleting statistics - admin only"""
    checker = get_permission_checker(current_user)
    
    if not checker.is_admin:
        raise HTTPException(status_code=403, detail="Apenas administradores podem arquivar membros")
    
    member = await db.users.find_one({"id": member_id}, {"_id": 0})
    if not member:
        raise HTTPException(status_code=404, detail="Membro não encontrado")
    
    if member.get('role') == 'admin':
        raise HTTPException(status_code=400, detail="Não é possível arquivar administradores")
    
    # Archive member - remove from teams but keep statistics
    team_ids = member.get('team_ids', [])
    
    # Store previous teams for potential restore
    await db.users.update_one({"id": member_id}, {
        "$set": {
            "is_archived": True,
            "archived_at": datetime.now(timezone.utc).isoformat(),
            "archived_team_ids": team_ids,
            "team_ids": []
        }
    })
    
    # Remove from all teams
    for team_id in team_ids:
        await db.teams.update_one(
            {"id": team_id},
            {"$pull": {"player_ids": member_id, "coach_ids": member_id, "delegate_ids": member_id}}
        )
    
    return {"message": "Membro arquivado com sucesso. Estatísticas mantidas."}

@api_router.post("/members/{member_id}/restore")
async def restore_member(member_id: str, team_id: Optional[str] = None, current_user: dict = Depends(get_current_user)):
    """Restore an archived member - admin only"""
    checker = get_permission_checker(current_user)
    
    if not checker.is_admin:
        raise HTTPException(status_code=403, detail="Apenas administradores podem restaurar membros")
    
    member = await db.users.find_one({"id": member_id}, {"_id": 0})
    if not member:
        raise HTTPException(status_code=404, detail="Membro não encontrado")
    
    if not member.get('is_archived'):
        raise HTTPException(status_code=400, detail="Membro não está arquivado")
    
    # Determine which team to restore to
    restore_team_id = team_id or (member.get('archived_team_ids', []) or [None])[0]
    
    update_data = {
        "is_archived": False,
        "archived_at": None
    }
    
    if restore_team_id:
        update_data["team_ids"] = [restore_team_id]
        
        # Add to team
        role = member.get('role', 'jogador')
        field_map = {'treinador': 'coach_ids', 'treinador_adjunto': 'coach_ids', 'delegado': 'delegate_ids'}
        field = field_map.get(role, 'player_ids')
        await db.teams.update_one({"id": restore_team_id}, {"$addToSet": {field: member_id}})
    
    await db.users.update_one({"id": member_id}, {"$set": update_data})
    
    return {"message": "Membro restaurado com sucesso", "team_id": restore_team_id}

@api_router.post("/members/{member_id}/send-activation-reminder")
async def send_activation_reminder(member_id: str, current_user: dict = Depends(get_current_user)):
    """Send push notification reminder to activate account - admin only"""
    checker = get_permission_checker(current_user)
    
    if not checker.is_admin:
        raise HTTPException(status_code=403, detail="Apenas administradores podem enviar lembretes")
    
    member = await db.users.find_one({"id": member_id}, {"_id": 0})
    if not member:
        raise HTTPException(status_code=404, detail="Membro não encontrado")
    
    if member.get('is_activated'):
        raise HTTPException(status_code=400, detail="Conta já está ativada")
    
    # Send push notification
    try:
        await send_push_to_users(
            user_ids=[member_id],
            title="Ativa a tua conta!",
            body="Por favor, atualiza a tua palavra-passe para ativar a tua conta no StickPro.",
            url="/settings"
        )
    except Exception as e:
        logging.error(f"Failed to send activation reminder: {e}")
    
    # Send email
    # Phase E2: ensure a valid (non-expired) invite_token exists, then send
    # the activation email through services.activation_emails. The legacy
    # inline HTML/send_email_notification path is kept disabled to avoid
    # duplicate sends.
    try:
        invite_token = member.get("invite_token")
        invite_expires_at = member.get("invite_expires_at")
        token_valid = False
        if invite_token and invite_expires_at:
            try:
                token_valid = datetime.fromisoformat(invite_expires_at) > datetime.now(timezone.utc)
            except Exception:
                token_valid = False
        if not token_valid:
            invite_token = secrets.token_urlsafe(32)
            invite_expires_at = (datetime.now(timezone.utc) + timedelta(days=7)).isoformat()
            await db.users.update_one(
                {"id": member_id},
                {"$set": {"invite_token": invite_token, "invite_expires_at": invite_expires_at}}
            )

        await send_activation_email(
            to_email=member.get("email"),
            name=member.get("name", "Atleta"),
            token=invite_token,
            idempotency_key=f"reminder-{member_id}-{invite_token[:8]}",
        )
    except Exception as e:
        logging.error(f"Failed to send activation email: {e}")
    
    return {"message": "Lembrete enviado"}

# ==================== COMPETITION PERMISSION ROUTES ====================

@api_router.get("/championships/permissions/settings")
async def get_championship_permission_settings(
    current_user: dict = Depends(get_current_user)
):
    if not is_admin_role(current_user.get("role")):
        raise HTTPException(
            status_code=403,
            detail="Apenas administradores e coordenadores técnicos podem ver estas permissões"
        )

    settings = await get_competition_permission_settings(current_user)

    return {
        "settings": settings,
        "defaults": DEFAULT_COMPETITION_PERMISSIONS,
    }


@api_router.put("/championships/permissions/settings")
async def update_championship_permission_settings(
    updates: CompetitionPermissionSettingsUpdate,
    current_user: dict = Depends(get_current_user)
):
    if not is_admin_role(current_user.get("role")):
        raise HTTPException(
            status_code=403,
            detail="Apenas administradores e coordenadores técnicos podem alterar estas permissões"
        )

    club = await get_current_club_for_user(current_user)
    if not club:
        raise HTTPException(status_code=404, detail="Clube não encontrado")

    update_data = updates.model_dump(exclude_unset=True)

    filtered = {
        key: bool(value)
        for key, value in update_data.items()
        if key in DEFAULT_COMPETITION_PERMISSIONS
    }

    if filtered:
        await db.clubs.update_one(
            {"id": club["id"]},
            {
                "$set": {
                    **{
                        f"competition_permissions.{key}": value
                        for key, value in filtered.items()
                    },
                    "competition_permissions_updated_at": datetime.now(timezone.utc).isoformat(),
                    "competition_permissions_updated_by": current_user["id"],
                }
            }
        )

    settings = await get_competition_permission_settings(current_user)

    return {
        "message": "Permissões das competições atualizadas",
        "settings": settings,
    }

# ==================== CHAMPIONSHIP ROUTES ====================

@api_router.post("/championships")
async def create_championship(
    data: ChampionshipCreate,
    current_user: dict = Depends(get_current_user)
):
    if not await can_create_competition(current_user, data.team_id):
        raise HTTPException(
            status_code=403,
            detail="Sem permissão para criar competições nesta equipa"
        )

    championship = Championship(**data.model_dump(), created_by=current_user["id"])
    champ_dict = championship.model_dump()
    champ_dict["created_at"] = champ_dict["created_at"].isoformat()
    champ_dict["is_archived"] = False
    champ_dict["updated_at"] = datetime.now(timezone.utc).isoformat()
    champ_dict["updated_by"] = current_user["id"]

    await db.championships.insert_one(champ_dict)
    champ_dict.pop("_id", None)

    return champ_dict

def parse_match_datetime(value):
    """
    Converte datas guardadas como datetime ou string ISO.
    Devolve None quando o valor não é válido.
    """
    if isinstance(value, datetime):
        if value.tzinfo is None:
            return value.replace(tzinfo=timezone.utc)
        return value

    if isinstance(value, str):
        try:
            parsed = datetime.fromisoformat(value.replace("Z", "+00:00"))

            if parsed.tzinfo is None:
                parsed = parsed.replace(tzinfo=timezone.utc)

            return parsed
        except ValueError:
            return None

    return None


def build_match_team_names(match: dict, club_team_name: str):
    """
    Normaliza os nomes oficiais da equipa da casa e visitante,
    mantendo compatibilidade com jogos antigos.
    """
    is_club_match = match.get("is_club_match", True)

    if not is_club_match:
        home_team = (
            match.get("home_team")
            or "Equipa da casa"
        )

        away_team = (
            match.get("away_team")
            or match.get("opponent_team")
            or "Equipa visitante"
        )

        return {
            "home_team": home_team,
            "away_team": away_team,
            "club_side": None,
        }

    club_side = (
        match.get("club_side")
        or (
            "away"
            if match.get("location") == "fora"
            else "neutral"
            if match.get("location") == "neutro"
            else "home"
        )
    )

    opponent_team = (
        match.get("opponent_team")
        or "Adversário"
    )

    if club_side == "away":
        home_team = (
            match.get("home_team")
            or opponent_team
        )

        away_team = (
            match.get("away_team")
            or club_team_name
        )

    elif club_side == "neutral":
        home_team = (
            match.get("home_team")
            or club_team_name
        )

        away_team = (
            match.get("away_team")
            or opponent_team
        )

    else:
        home_team = (
            match.get("home_team")
            or club_team_name
        )

        away_team = (
            match.get("away_team")
            or opponent_team
        )

    return {
        "home_team": home_team,
        "away_team": away_team,
        "club_side": club_side,
    }


def get_club_match_result(
    match: dict,
    club_side: Optional[str]
):
    """
    Calcula vitória, empate ou derrota na perspetiva da equipa do clube.
    """
    if (
        match.get("is_club_match", True) is False
        or not match.get("is_completed")
    ):
        return None

    home_score = match.get("home_score")
    away_score = match.get("away_score")

    if home_score is None or away_score is None:
        return None

    if club_side == "away":
        club_score = away_score
        opponent_score = home_score
    else:
        club_score = home_score
        opponent_score = away_score

    if club_score > opponent_score:
        return "win"

    if club_score < opponent_score:
        return "loss"

    return "draw"


async def enrich_championship_dashboard(
    championship: dict
):
    """
    Acrescenta ao campeonato os indicadores necessários
    para o dashboard vivo da página de Competições.
    """
    championship_id = championship.get("id")
    team_id = championship.get("team_id")

    team = await db.teams.find_one(
        {"id": team_id},
        {
            "_id": 0,
            "name": 1,
        }
    )

    club_team_name = (
        team.get("name")
        if team
        else championship.get("team_name")
        or "Nossa Equipa"
    )

    matches = await db.championship_matches.find(
        {
            "championship_id": championship_id,
            "archived": {"$ne": True},
        },
        {"_id": 0}
    ).to_list(500)

    now = datetime.now(timezone.utc)

    club_matches = [
        match
        for match in matches
        if match.get("is_club_match", True) is not False
    ]

    completed_matches = [
        match
        for match in club_matches
        if match.get("is_completed") is True
    ]

    pending_matches = [
        match
        for match in club_matches
        if match.get("is_completed") is not True
    ]

    wins = 0
    draws = 0
    losses = 0

    for match in completed_matches:
        names = build_match_team_names(
            match,
            club_team_name
        )

        result = get_club_match_result(
            match,
            names.get("club_side")
        )

        if result == "win":
            wins += 1
        elif result == "draw":
            draws += 1
        elif result == "loss":
            losses += 1

    pending_gamesheets = sum(
        1
        for match in completed_matches
        if not match.get("gamesheet_url")
    )

    dated_pending_matches = []

    for match in pending_matches:
        match_datetime = parse_match_datetime(
            match.get("match_date")
        )

        if match_datetime:
            dated_pending_matches.append(
                (match_datetime, match)
            )

    future_matches = [
        item
        for item in dated_pending_matches
        if item[0] >= now
    ]

    if future_matches:
        future_matches.sort(
            key=lambda item: item[0]
        )
        next_match = future_matches[0][1]
    elif dated_pending_matches:
        dated_pending_matches.sort(
            key=lambda item: item[0]
        )
        next_match = dated_pending_matches[0][1]
    else:
        next_match = None

    dated_completed_matches = []

    for match in completed_matches:
        match_datetime = parse_match_datetime(
            match.get("match_date")
        )

        if match_datetime:
            dated_completed_matches.append(
                (match_datetime, match)
            )

    dated_completed_matches.sort(
        key=lambda item: item[0],
        reverse=True
    )

    last_match = (
        dated_completed_matches[0][1]
        if dated_completed_matches
        else None
    )

    next_match_data = None

    if next_match:
        next_names = build_match_team_names(
            next_match,
            club_team_name
        )

        next_match_data = {
            "id": next_match.get("id"),
            "match_date": next_match.get("match_date"),
            "match_time": next_match.get("match_time"),
            "matchday": next_match.get("matchday"),
            "venue": next_match.get("venue"),
            "location": next_match.get("location"),
            "home_team": next_names["home_team"],
            "away_team": next_names["away_team"],
            "club_side": next_names["club_side"],
        }

    last_result_data = None

    if last_match:
        last_names = build_match_team_names(
            last_match,
            club_team_name
        )

        last_result_data = {
            "id": last_match.get("id"),
            "match_date": last_match.get("match_date"),
            "matchday": last_match.get("matchday"),
            "home_team": last_names["home_team"],
            "away_team": last_names["away_team"],
            "home_score": last_match.get("home_score"),
            "away_score": last_match.get("away_score"),
            "club_side": last_names["club_side"],
            "club_result": get_club_match_result(
                last_match,
                last_names["club_side"]
            ),
        }

    dashboard = {
        "summary": {
            "matches_total": len(club_matches),
            "matches_completed": len(completed_matches),
            "matches_pending": len(pending_matches),
            "wins": wins,
            "draws": draws,
            "losses": losses,
            "pending_gamesheets": pending_gamesheets,
        },
        "next_match": next_match_data,
        "last_result": last_result_data,
        "standing_position": None,
    }

    return {
        **championship,
        "dashboard": dashboard,
    }

@api_router.get("/championships")
async def get_championships(
    team_id: Optional[str] = None,
    season: Optional[str] = None,
    include_archived: bool = False,
    current_user: dict = Depends(get_current_user)
):
    checker = get_permission_checker(current_user)

    query = {}

    if not include_archived:
        query["is_archived"] = {"$ne": True}

    if team_id:
        if not checker.is_admin and not checker.can_access_team(team_id):
            raise HTTPException(status_code=403, detail="Sem acesso a esta equipa")
        query["team_id"] = team_id

    elif not checker.is_admin:
        user_teams = list(checker.team_ids)
        if user_teams:
            query["team_id"] = {"$in": user_teams}
        else:
            return []

    if season:
        query["season"] = season

    championships = await db.championships.find(
        query,
        {"_id": 0}
    ).sort("created_at", -1).to_list(500)

    enriched = []

    for championship in championships:
        if not await can_view_competition(
            current_user,
            championship
        ):
            continue
    
        championship_response = (
            await enrich_championship_for_response(
                championship,
                current_user
            )
        )
    
        championship_response = (
            await enrich_championship_dashboard(
                championship_response
            )
        )
    
        enriched.append(championship_response)
    
    return enriched

@api_router.get("/championships/{championship_id}")
async def get_championship(
    championship_id: str,
    current_user: dict = Depends(get_current_user)
):
    championship = await db.championships.find_one(
        {"id": championship_id},
        {"_id": 0}
    )

    if not championship:
        raise HTTPException(status_code=404, detail="Campeonato não encontrado")

    if not await can_view_competition(current_user, championship):
        raise HTTPException(status_code=403, detail="Sem acesso a esta competição")

    return await enrich_championship_for_response(championship, current_user)

@api_router.put("/championships/{championship_id}")
async def update_championship(
    championship_id: str,
    updates: dict,
    current_user: dict = Depends(get_current_user)
):
    championship = await db.championships.find_one(
        {"id": championship_id},
        {"_id": 0}
    )

    if not championship:
        raise HTTPException(status_code=404, detail="Campeonato não encontrado")

    if not await can_edit_competition(current_user, championship):
        raise HTTPException(
            status_code=403,
            detail="Sem permissão para editar esta competição"
        )

    allowed = [
        "name",
        "description",
        "season",
        "format",
        "location",
        "convocation_type",
        "age_group",
        "competition_type",
        "participating_teams",
        "team_id",
    ]

    filtered = {k: v for k, v in updates.items() if k in allowed}

    if "team_id" in filtered:
        if not user_has_team_access(current_user, filtered["team_id"]):
            raise HTTPException(status_code=403, detail="Sem acesso à nova equipa")

    if filtered:
        filtered["updated_at"] = datetime.now(timezone.utc).isoformat()
        filtered["updated_by"] = current_user["id"]

        await db.championships.update_one(
            {"id": championship_id},
            {"$set": filtered}
        )

    updated = await db.championships.find_one(
        {"id": championship_id},
        {"_id": 0}
    )

    return updated

@api_router.delete("/championships/{championship_id}")
async def delete_championship(
    championship_id: str,
    current_user: dict = Depends(get_current_user)
):
    championship = await db.championships.find_one(
        {"id": championship_id},
        {"_id": 0}
    )

    if not championship:
        raise HTTPException(status_code=404, detail="Campeonato não encontrado")

    if not await can_archive_competition(current_user, championship):
        raise HTTPException(
            status_code=403,
            detail="Sem permissão para arquivar esta competição"
        )

    await db.championships.update_one(
        {"id": championship_id},
        {
            "$set": {
                "is_archived": True,
                "archived_at": datetime.now(timezone.utc).isoformat(),
                "archived_by": current_user["id"],
                "updated_at": datetime.now(timezone.utc).isoformat(),
                "updated_by": current_user["id"],
            }
        }
    )

    return {"message": "Competição arquivada com sucesso"}

# ==================== CHAMPIONSHIP MATCH ROUTES ====================

@api_router.post(
    "/championships/{championship_id}/matches"
)
async def create_championship_match(
    championship_id: str,
    data: ChampionshipMatchCreate,
    current_user: dict = Depends(get_current_user)
):
    championship = await db.championships.find_one(
        {"id": championship_id},
        {"_id": 0}
    )

    if not championship:
        raise HTTPException(
            status_code=404,
            detail="Campeonato não encontrado"
        )

    if not await can_create_competition_game(
        current_user,
        championship
    ):
        raise HTTPException(
            status_code=403,
            detail=(
                "Sem permissão para criar jogos "
                "nesta competição"
            )
        )

    club_team_name = (
        await get_championship_club_team_name(
            championship
        )
    )

    is_club_match = data.is_club_match

    if is_club_match is None:
        is_club_match = True

    if is_club_match:
        club_side = data.club_side

        if not club_side:
            club_side = {
                "casa": "home",
                "fora": "away",
                "neutro": "neutral",
            }.get(
                data.location,
                "home"
            )

        opponent_team = (
            data.opponent_team or ""
        ).strip()

        if not opponent_team:
            if club_side == "home":
                opponent_team = (
                    data.away_team or ""
                ).strip()
            elif club_side == "away":
                opponent_team = (
                    data.home_team or ""
                ).strip()
            else:
                candidate_home = (
                    data.home_team or ""
                ).strip()
                candidate_away = (
                    data.away_team or ""
                ).strip()

                opponent_team = (
                    candidate_away
                    if normalize_team_name(
                        candidate_home
                    )
                    == normalize_team_name(
                        club_team_name
                    )
                    else candidate_home
                )

        if not opponent_team:
            raise HTTPException(
                status_code=400,
                detail="Indique a equipa adversária"
            )

        if (
            normalize_team_name(opponent_team)
            == normalize_team_name(
                club_team_name
            )
        ):
            raise HTTPException(
                status_code=400,
                detail=(
                    "A equipa adversária não pode "
                    "ser igual à equipa do clube"
                )
            )

        if club_side == "home":
            home_team = club_team_name
            away_team = opponent_team
            location = "casa"

        elif club_side == "away":
            home_team = opponent_team
            away_team = club_team_name
            location = "fora"

        else:
            requested_home = (
                data.home_team or ""
            ).strip()

            requested_away = (
                data.away_team or ""
            ).strip()

            if (
                requested_home
                and requested_away
            ):
                home_team = requested_home
                away_team = requested_away
            else:
                home_team = club_team_name
                away_team = opponent_team

            location = "neutro"

    else:
        # Jogo entre duas equipas externas.
        home_team = (
            data.home_team or ""
        ).strip()

        away_team = (
            data.away_team or ""
        ).strip()

        if not home_team or not away_team:
            raise HTTPException(
                status_code=400,
                detail=(
                    "Indique a equipa da casa e "
                    "a equipa visitante"
                )
            )

        if (
            normalize_team_name(home_team)
            == normalize_team_name(away_team)
        ):
            raise HTTPException(
                status_code=400,
                detail=(
                    "As equipas não podem ser iguais"
                )
            )

        club_side = "neutral"
        location = "neutro"
        
        # O modelo ChampionshipMatch exige opponent_team como string.
        # Num jogo entre equipas externas, usa-se a equipa visitante
        # como referência técnica.
        opponent_team = away_team

    match = ChampionshipMatch(
        championship_id=championship_id,
        team_id=championship["team_id"],
        home_team=home_team,
        away_team=away_team,
        opponent_team=opponent_team,
        club_side=club_side,
        official_match_url=(
            data.official_match_url
        ),
        match_date=data.match_date,
        match_time=data.match_time,
        location=location,
        venue=data.venue,
        is_club_match=is_club_match,
        bonus_points=data.bonus_points,
        penalty_points=data.penalty_points,
        matchday=data.matchday,
    )

    match_dict = match.model_dump()

    match_dict["match_date"] = (
        match_dict["match_date"].isoformat()
    )
    match_dict["created_at"] = (
        match_dict["created_at"].isoformat()
    )
    match_dict["archived"] = False

    official_url = (
        data.official_match_url or ""
    ).strip()
    
    match_dict["official_match_url"] = (
        official_url or None
    )
    
    match_dict["gamesheet_url"] = (
        official_url or None
    )
    
    match_dict["sync_status"] = (
        "pending"
        if official_url
        else "manual"
    )
    
    match_dict.update(
        build_match_status_payload(
            "scheduled",
            updated_by=current_user["id"],
        )
    )

    official_url = (
        data.official_match_url or ""
    ).strip()
    
    match_source = infer_match_source(
        official_url
    )
    
    match_dict.update({
        "official_match_url": (
            official_url or None
        ),
    
        # Compatibilidade com o importador atual.
        "gamesheet_url": (
            official_url or None
        ),
    
        "source": match_source,
    
        "sync_status": (
            "pending"
            if official_url
            else "manual"
        ),
    
        "last_synced_at": None,
        "last_sync_error": None,
    
        "is_verified": False,
        "verified_by": None,
        "verified_at": None,
    
        "external_match_id": None,
    
        "source_url": (
            official_url or None
        ),
    })

    await db.championship_matches.insert_one(
        match_dict
    )

    match_dict.pop("_id", None)

    await sync_calendar_event_from_match(
        match.id,
        championship=championship,
        match_data=match_dict,
        current_user_id=current_user["id"]
    )

    return match_dict

@api_router.get("/championships/{championship_id}/matches")
async def get_championship_matches(
    championship_id: str,
    current_user: dict = Depends(get_current_user)
):
    checker = get_permission_checker(current_user)

    championship = await db.championships.find_one(
        {"id": championship_id},
        {"_id": 0}
    )

    if not championship:
        raise HTTPException(status_code=404, detail="Campeonato não encontrado")

    champ_team_id = championship.get("team_id")
    if champ_team_id and not checker.is_admin and not checker.can_access_team(champ_team_id):
        raise HTTPException(status_code=403, detail="Sem acesso a esta competição")

    matches = await db.championship_matches.find(
        {
            "championship_id": championship_id,
            "archived": {"$ne": True}
        },
        {"_id": 0}
    ).sort("match_date", 1).to_list(100)

    for match in matches:
        if isinstance(match.get("match_date"), str):
            match["match_date"] = datetime.fromisoformat(match["match_date"])

    return matches

async def sync_calendar_event_from_match(
    match_id: str,
    *,
    championship: dict,
    match_data: dict,
    current_user_id: str
) -> None:
    """
    Mantém o evento do calendário sincronizado com o jogo.

    Compatibilidade:
    - procura primeiro por championship_match_id;
    - usa championship_id + team_id como fallback para dados antigos;
    - cria o evento caso não exista.
    """
    home_team = (
        match_data.get("home_team")
        or championship.get("team_name")
        or "Equipa"
    )

    away_team = (
        match_data.get("away_team")
        or match_data.get("opponent_team")
        or "Adversário"
    )

    match_date = parse_match_datetime(match_data.get("match_date"))

    if not match_date:
        raise HTTPException(
            status_code=400,
            detail="Data do jogo inválida"
        )

    location_value = (
        match_data.get("venue")
        or (
            "Casa"
            if match_data.get("location") == "casa"
            else "Fora"
            if match_data.get("location") == "fora"
            else "Neutro"
        )
    )

    event_update = {
        "team_id": championship["team_id"],
        "event_type": "jogo_campeonato",
        "title": f"{home_team} vs {away_team}",
        "location": location_value,
        "start_time": match_date.isoformat(),
        "opponent": match_data.get("opponent_team"),
        "championship_id": championship["id"],
        "championship_match_id": match_id,
        "updated_at": datetime.now(timezone.utc).isoformat(),
        "updated_by": current_user_id,
        "home_team": home_team,
        "away_team": away_team,
        "home_score": match_data.get("home_score"),
        "away_score": match_data.get("away_score"),
        "is_completed": bool(match_data.get("is_completed")),
        "match_status": infer_match_status(match_data),
        "home_score": match_data.get(
        "home_score"
            ),
            "away_score": match_data.get(
                "away_score"
            ),
            "is_completed": match_data.get(
                "is_completed",
                False
            ),
            "match_status": match_data.get(
                "match_status"
            ),
            "official_match_url": (
                match_data.get("official_match_url")
            ),
            "gamesheet_url": (
                match_data.get("gamesheet_url")
            ),
            "source_url": match_data.get(
                "source_url"
            ),
            "source": match_data.get(
                "source",
                "manual"
            ),
            "sync_status": match_data.get(
                "sync_status",
                "manual"
            ),
            "is_verified": match_data.get(
                "is_verified",
                False
            ),
            "last_synced_at": match_data.get(
                "last_synced_at"
            ),
            "referee": (
                match_data.get("referee")
                or (
                    match_data.get(
                        "gamesheet_raw_data",
                        {}
                    ).get("referee")
                )
            ),
            "gamesheet_raw_data": match_data.get(
                "gamesheet_raw_data"
            ),
    }

    event = await db.events.find_one(
        {"championship_match_id": match_id},
        {"_id": 0}
    )

    if not event:
        event = await db.events.find_one(
            {
                "championship_id": championship["id"],
                "team_id": championship["team_id"],
                "start_time": match_data.get("match_date"),
                "opponent": match_data.get("opponent_team"),
            },
            {"_id": 0}
        )

    if event:
        await db.events.update_one(
            {"id": event["id"]},
            {"$set": event_update}
        )
        return

    new_event = {
        "id": str(uuid.uuid4()),
        **event_update,
        "status": (
            "completed"
            if match_data.get("is_completed")
            else "scheduled"
        ),
        "created_by": current_user_id,
        "created_at": datetime.now(timezone.utc).isoformat(),
    }

    await db.events.insert_one(new_event)

@api_router.put("/championships/matches/{match_id}/result")
async def update_match_result(
    match_id: str,
    result: MatchResultUpdate,
    current_user: dict = Depends(get_current_user)
):
    match = await db.championship_matches.find_one(
        {"id": match_id, "archived": {"$ne": True}},
        {"_id": 0}
    )

    if not match:
        raise HTTPException(
            status_code=404,
            detail="Jogo não encontrado"
        )

    championship = await db.championships.find_one(
        {"id": match["championship_id"]},
        {"_id": 0}
    )

    if not championship:
        raise HTTPException(
            status_code=404,
            detail="Competição não encontrada"
        )

    if not await can_edit_competition_result(
        current_user,
        championship
    ):
        raise HTTPException(
            status_code=403,
            detail="Sem permissão para atualizar resultados desta competição"
        )

    now = datetime.now(timezone.utc)

    update_data = {
        "home_score": result.home_score,
        "away_score": result.away_score,
        "bonus_points": result.bonus_points,
        "penalty_points": result.penalty_points,
        "is_completed": True,
        "match_status": "finished",
        "match_status_label": MATCH_STATUS_LABELS["finished"],
        "match_status_updated_at": now.isoformat(),
        "match_status_updated_by": current_user["id"],
        "result_source": "manual",
        "result_updated_at": now.isoformat(),
        "result_updated_by": current_user["id"],

        # Um resultado manual fica validado pelo utilizador que o inseriu.
        "source": match.get("source") or "manual",
        "sync_status": (
            match.get("sync_status")
            if match.get("source") in ["apl", "fpp"]
            else "manual"
        ),
        "is_verified": True,
        "verified_by": current_user["id"],
        "verified_at": now.isoformat(),

        "updated_at": now.isoformat(),
        "updated_by": current_user["id"],
    }

    await db.championship_matches.update_one(
        {"id": match_id},
        {"$set": update_data}
    )

    updated_match = await db.championship_matches.find_one(
        {"id": match_id},
        {"_id": 0}
    )

    await update_match_workflow(
        match_id=match_id,
        stage="finished",
        updated_by=current_user["id"],
    )

    updated_match = (
        await db.championship_matches.find_one(
            {"id": match_id},
            {"_id": 0}
        )
    )
    
    await sync_calendar_event_from_match(
        match_id,
        championship=championship,
        match_data=updated_match,
        current_user_id=current_user["id"]
    )
    
    return {
        "message": "Resultado atualizado",
        "match": updated_match
    }

def normalize_team_name(value: Optional[str]) -> str:
    return re.sub(
        r"\s+",
        " ",
        (value or "").strip()
    ).casefold()
    
def infer_match_source(
    official_url: Optional[str]
) -> str:
    """
    Identifica a origem provável da ficha oficial.

    Esta identificação não faz ainda qualquer pedido externo.
    Apenas analisa o domínio do URL.
    """
    normalized_url = (
        official_url or ""
    ).strip().lower()

    if not normalized_url:
        return "manual"

    if (
        "fpp.pt" in normalized_url
        or "hoqueipatins.pt" in normalized_url
    ):
        return "fpp"

    if (
        "aplisboa" in normalized_url
        or "aplisboa.pt" in normalized_url
    ):
        return "apl"

    return "official"

async def get_championship_club_team_name(
    championship: dict
) -> str:
    team = await db.teams.find_one(
        {"id": championship.get("team_id")},
        {"_id": 0, "name": 1}
    )

    if team and team.get("name"):
        return team["name"].strip()

    if championship.get("team_name"):
        return championship["team_name"].strip()

    return "Equipa do clube"


def get_existing_opponent_name(
    match: dict,
    club_team_name: str
) -> str:
    explicit_opponent = (
        match.get("opponent_team") or ""
    ).strip()

    if explicit_opponent:
        return explicit_opponent

    current_side = match.get("club_side")

    if current_side == "home":
        return (
            match.get("away_team")
            or "Adversário"
        ).strip()

    if current_side == "away":
        return (
            match.get("home_team")
            or "Adversário"
        ).strip()

    home_team = (
        match.get("home_team") or ""
    ).strip()

    away_team = (
        match.get("away_team") or ""
    ).strip()

    if (
        normalize_team_name(home_team)
        != normalize_team_name(club_team_name)
    ):
        return home_team or "Adversário"

    return away_team or "Adversário"
    
class MatchUpdate(BaseModel):
    home_team: Optional[str] = None
    away_team: Optional[str] = None
    club_side: Optional[str] = None
    official_match_url: Optional[str] = None
    opponent_team: Optional[str] = None

    match_date: Optional[datetime] = None
    match_time: Optional[str] = None
    location: Optional[MatchLocation] = None
    venue: Optional[str] = None
    matchday: Optional[int] = None

    is_club_match: Optional[bool] = None
    bonus_points: Optional[int] = None
    penalty_points: Optional[int] = None


@api_router.put("/championships/matches/{match_id}")
async def update_match(
    match_id: str,
    updates: MatchUpdate,
    current_user: dict = Depends(get_current_user)
):
    match = await db.championship_matches.find_one(
        {
            "id": match_id,
            "archived": {"$ne": True}
        },
        {"_id": 0}
    )

    if not match:
        raise HTTPException(
            status_code=404,
            detail="Jogo não encontrado"
        )

    championship = await db.championships.find_one(
        {"id": match["championship_id"]},
        {"_id": 0}
    )

    if not championship:
        raise HTTPException(
            status_code=404,
            detail="Competição não encontrada"
        )

    if not await can_edit_competition_game(
        current_user,
        championship
    ):
        raise HTTPException(
            status_code=403,
            detail=(
                "Sem permissão para editar jogos "
                "desta competição"
            )
        )

    update_data = updates.model_dump(
        exclude_unset=True
    )
    
    if "official_match_url" in update_data:
        official_url = (
            update_data.get("official_match_url")
            or ""
        ).strip()
    
        update_data["official_match_url"] = (
            official_url or None
        )
    
        # Compatibilidade com o importador atual.
        update_data["gamesheet_url"] = (
            official_url or None
        )
    
        update_data["source_url"] = (
            official_url or None
        )
    
        update_data["source"] = (
            infer_match_source(
                official_url
            )
        )
    
        update_data["sync_status"] = (
            "pending"
            if official_url
            else "manual"
        )
    
        # Um novo URL reinicia o estado técnico.
        update_data["last_sync_error"] = None
    
        if not official_url:
            update_data["last_synced_at"] = None
    
    if "match_date" in update_data:
        update_data["match_date"] = (
            update_data["match_date"].isoformat()
        )
        
    # A localização determina o lado, salvo quando
    # club_side é enviado explicitamente.
    if (
        "location" in update_data
        and "club_side" not in update_data
    ):
        update_data["club_side"] = {
            "casa": "home",
            "fora": "away",
            "neutro": "neutral",
        }.get(
            update_data["location"],
            "neutral"
        )

    merged_match = {
        **match,
        **update_data,
    }

    is_club_match = merged_match.get(
        "is_club_match",
        True
    )

    if is_club_match:
        club_team_name = (
            await get_championship_club_team_name(
                championship
            )
        )

        opponent_team = (
            update_data.get("opponent_team")
            or get_existing_opponent_name(
                match,
                club_team_name
            )
        ).strip()

        if (
            normalize_team_name(opponent_team)
            == normalize_team_name(club_team_name)
        ):
            raise HTTPException(
                status_code=400,
                detail=(
                    "A equipa adversária não pode "
                    "ser igual à equipa do clube"
                )
            )

        club_side = merged_match.get(
            "club_side",
            "home"
        )

        if club_side == "home":
            update_data["home_team"] = (
                club_team_name
            )
            update_data["away_team"] = (
                opponent_team
            )
            update_data["location"] = "casa"

        elif club_side == "away":
            update_data["home_team"] = (
                opponent_team
            )
            update_data["away_team"] = (
                club_team_name
            )
            update_data["location"] = "fora"

        else:
            # Num jogo da nossa equipa em campo neutro,
            # preserva-se a ordem escolhida pelo utilizador.
            requested_home = (
                update_data.get("home_team")
                or merged_match.get("home_team")
            )
            requested_away = (
                update_data.get("away_team")
                or merged_match.get("away_team")
            )

            if (
                normalize_team_name(requested_home)
                == normalize_team_name(
                    club_team_name
                )
            ):
                update_data["home_team"] = (
                    club_team_name
                )
                update_data["away_team"] = (
                    opponent_team
                )

            elif (
                normalize_team_name(requested_away)
                == normalize_team_name(
                    club_team_name
                )
            ):
                update_data["home_team"] = (
                    opponent_team
                )
                update_data["away_team"] = (
                    club_team_name
                )

            else:
                update_data["home_team"] = (
                    club_team_name
                )
                update_data["away_team"] = (
                    opponent_team
                )

            update_data["location"] = "neutro"

        update_data["opponent_team"] = opponent_team

    else:
        # Jogo entre equipas externas.
        home_team = (
            update_data.get("home_team")
            or merged_match.get("home_team")
            or ""
        ).strip()

        away_team = (
            update_data.get("away_team")
            or merged_match.get("away_team")
            or ""
        ).strip()

        if not home_team or not away_team:
            raise HTTPException(
                status_code=400,
                detail=(
                    "Indique a equipa da casa e "
                    "a equipa visitante"
                )
            )

        if (
            normalize_team_name(home_team)
            == normalize_team_name(away_team)
        ):
            raise HTTPException(
                status_code=400,
                detail=(
                    "As equipas não podem ser iguais"
                )
            )

        update_data["home_team"] = home_team
        update_data["away_team"] = away_team
        
        # Compatibilidade com o modelo ChampionshipMatch,
        # que exige opponent_team como string.
        update_data["opponent_team"] = away_team
        
        # Mantém o objeto merged_match coerente durante
        # o restante processamento desta função.
        merged_match["opponent_team"] = away_team
        
        update_data["is_club_match"] = False
        update_data["club_side"] = "neutral"
        update_data["location"] = "neutro"

    now_iso = datetime.now(
        timezone.utc
    ).isoformat()

    update_data["updated_at"] = now_iso
    update_data["updated_by"] = current_user["id"]

    await db.championship_matches.update_one(
        {"id": match_id},
        {"$set": update_data}
    )

    updated_match = (
        await db.championship_matches.find_one(
            {"id": match_id},
            {"_id": 0}
        )
    )

    await sync_calendar_event_from_match(
        match_id,
        championship=championship,
        match_data=updated_match,
        current_user_id=current_user["id"]
    )

    return {
        "message": "Jogo e calendário atualizados",
        "match": updated_match
    }
    
@api_router.post("/championships/{championship_id}/matches/fix-home-away")
async def fix_championship_matches_home_away(
    championship_id: str,
    current_user: dict = Depends(get_current_user)
):
    checker = get_permission_checker(current_user)

    if not checker.is_admin:
        raise HTTPException(
            status_code=403,
            detail="Apenas administradores podem executar esta migração"
        )

    championship = await db.championships.find_one(
        {"id": championship_id},
        {"_id": 0}
    )

    if not championship:
        raise HTTPException(status_code=404, detail="Competição não encontrada")

    team = await db.teams.find_one(
        {"id": championship.get("team_id")},
        {"_id": 0}
    )

    club_team_name = (
        team.get("name")
        if team
        else championship.get("team_name") or championship.get("name") or "Equipa"
    )

    matches = await db.championship_matches.find(
        {
            "championship_id": championship_id,
            "archived": {"$ne": True}
        },
        {"_id": 0}
    ).to_list(500)

    fixed = 0
    unchanged = 0
    inspected = []

    for match in matches:
        location = match.get("location", "casa")
        opponent_team = match.get("opponent_team") or match.get("away_team") or "Adversário"

        if location == "fora":
            club_side = "away"
            home_team = opponent_team
            away_team = club_team_name
        elif location == "neutro":
            club_side = "neutral"
            home_team = match.get("home_team") or club_team_name
            away_team = match.get("away_team") or opponent_team
        else:
            club_side = "home"
            home_team = club_team_name
            away_team = opponent_team

        update_data = {
            "home_team": home_team,
            "away_team": away_team,
            "club_side": club_side,
            "updated_at": datetime.now(timezone.utc).isoformat(),
            "updated_by": current_user["id"],
            "home_away_fixed_at": datetime.now(timezone.utc).isoformat()
        }

        needs_update = (
            match.get("home_team") != home_team or
            match.get("away_team") != away_team or
            match.get("club_side") != club_side
        )

        if needs_update:
            await db.championship_matches.update_one(
                {"id": match["id"]},
                {"$set": update_data}
            )
            fixed += 1
        else:
            unchanged += 1

        inspected.append({
            "match_id": match.get("id"),
            "location": location,
            "before": {
                "home_team": match.get("home_team"),
                "away_team": match.get("away_team"),
                "club_side": match.get("club_side"),
                "opponent_team": match.get("opponent_team"),
            },
            "after": {
                "home_team": home_team,
                "away_team": away_team,
                "club_side": club_side,
            },
            "updated": needs_update
        })

    return {
        "message": "Migração casa/fora concluída",
        "championship_id": championship_id,
        "club_team_name": club_team_name,
        "total": len(matches),
        "fixed": fixed,
        "unchanged": unchanged,
        "matches": inspected
    }

@api_router.put("/championships/matches/{match_id}/archive")
async def archive_match(
    match_id: str,
    current_user: dict = Depends(get_current_user)
):
    """
    Arquiva o jogo sem apagar o histórico e retira o evento do calendário.
    """
    match = await db.championship_matches.find_one(
        {"id": match_id, "archived": {"$ne": True}},
        {"_id": 0}
    )

    if not match:
        raise HTTPException(
            status_code=404,
            detail="Jogo não encontrado"
        )

    championship = await db.championships.find_one(
        {"id": match["championship_id"]},
        {"_id": 0}
    )

    if not championship:
        raise HTTPException(
            status_code=404,
            detail="Competição não encontrada"
        )

    # CORREÇÃO:
    # arquivar jogo é gestão de jogo, não edição de estatísticas.
    if not await can_edit_competition_game(
        current_user,
        championship
    ):
        raise HTTPException(
            status_code=403,
            detail="Sem permissão para arquivar jogos desta competição"
        )

    now = datetime.now(timezone.utc).isoformat()

    status_payload = build_match_status_payload(
        "archived",
        updated_by=current_user["id"],
    )

    now_iso = datetime.now(timezone.utc).isoformat()

    await db.championship_matches.update_one(
        {"id": match_id},
        {
            "$set": {
                **status_payload,
                "archived": True,
                "archived_at": now_iso,
                "archived_by": current_user["id"],
                "updated_at": now_iso,
                "updated_by": current_user["id"],
            }
        },
    )

    await db.events.update_many(
        {"championship_match_id": match_id},
        {
            "$set": {
                "status": "archived",
                "is_archived": True,
                "archived_at": now_iso,
                "archived_by": current_user["id"],
                "updated_at": now_iso,
                "updated_by": current_user["id"],
            }
        },
    )

    return {
        "message": "Jogo arquivado e removido do calendário"
    }

@api_router.get("/championships/matches/import-template")
async def download_matches_import_template(token: str = None, current_user: dict = None):
    """Download Excel template for importing matches"""
    # Allow token via query param for direct download links
    if token and not current_user:
        try:
            payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
            current_user = await db.users.find_one({"id": payload["user_id"]}, {"_id": 0, "password": 0})
        except:
            pass
    
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    import io
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Modelo Importação Jogos"
    
    # Define headers (Portuguese as main, others in comments)
    headers = [
        "Jornada",           # Round / Matchday
        "Data",              # Date
        "Hora",              # Time
        "Equipa Casa",       # Home Team
        "Adversário",        # Opponent / Away Team
        "Local",             # Venue / Pavilion
        "Localização"        # Location (casa/fora/neutro)
    ]
    
    # Style definitions
    header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")  # Dark blue
    header_font = Font(color="FFFFFF", bold=True, size=11)
    example_fill = PatternFill(start_color="E8F4FD", end_color="E8F4FD", fill_type="solid")  # Light blue
    hint_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")  # Light yellow
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')
    
    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border
    
    # Example data rows
    example_data = [
        [1, "2026-01-15", "15:00", "Escolares", "UD Vilafranquense", "Pavilhão Municipal", "casa"],
        [1, "2026-01-15", "16:30", "SC Torres", "Escolares", "Pavilhão SC Torres", "fora"],
        [2, "2026-01-22", "10:00", "Escolares", "Benfica", "Pavilhão Municipal", "casa"],
        [2, "2026-01-22", "17:00", "Sporting", "Escolares", "Pavilhão João Rocha", "fora"],
        [3, "2026-01-29", "14:00", "Escolares", "Porto", "Campo Neutro", "neutro"],
    ]
    
    for row_idx, row_data in enumerate(example_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = example_fill
            cell.border = thin_border
            cell.alignment = center_align if col_idx in [1, 2, 3, 7] else left_align
    
    # Add hints row
    hints = [
        "Número da jornada (1, 2, 3...)",
        "AAAA-MM-DD ou DD/MM/AAAA",
        "HH:MM (24h)",
        "Nome da equipa da casa",
        "Nome do adversário",
        "Nome do pavilhão (opcional)",
        "casa, fora ou neutro"
    ]
    
    hint_row = len(example_data) + 3
    ws.cell(row=hint_row, column=1, value="INSTRUÇÕES:").font = Font(bold=True, size=10)
    
    for col, hint in enumerate(hints, 1):
        cell = ws.cell(row=hint_row + 1, column=col, value=hint)
        cell.fill = hint_fill
        cell.font = Font(size=9, italic=True)
        cell.alignment = left_align
    
    # Multi-language support note
    ws.cell(row=hint_row + 3, column=1, value="Cabeçalhos alternativos suportados:").font = Font(bold=True, size=10)
    alt_headers = [
        "PT: Jornada | ES: Jornada | EN: Round/Matchday | FR: Journée | IT: Giornata",
        "PT: Data | ES: Fecha | EN: Date | FR: Date | IT: Data",
        "PT: Hora | ES: Hora | EN: Time | FR: Heure | IT: Ora",
        "PT: Equipa Casa | ES: Equipo Local | EN: Home Team | FR: Équipe Domicile | IT: Squadra Casa",
        "PT: Adversário | ES: Rival | EN: Opponent/Away Team | FR: Adversaire | IT: Avversario",
        "PT: Local/Pavilhão | ES: Lugar/Pabellón | EN: Venue/Pavilion | FR: Lieu | IT: Luogo",
        "PT: Localização | ES: Ubicación | EN: Location | FR: Localisation | IT: Localizzazione",
    ]
    for idx, text in enumerate(alt_headers):
        ws.cell(row=hint_row + 4 + idx, column=1, value=text).font = Font(size=9, color="666666")
    
    # Adjust column widths
    column_widths = [10, 14, 8, 25, 25, 30, 15]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    
    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    from fastapi.responses import StreamingResponse
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": "attachment; filename=modelo_importacao_jogos.xlsx"
        }
    )

@api_router.post("/championships/{championship_id}/matches/import")
async def import_championship_matches(
    championship_id: str, 
    file: UploadFile = File(...),
    current_user: dict = Depends(get_current_user)
):
    """Import matches from Excel/CSV file with multilingual header support"""
    championship = await db.championships.find_one(
        {"id": championship_id},
        {"_id": 0}
    )
    
    if not championship:
        raise HTTPException(status_code=404, detail="Campeonato não encontrado")
    
    if not await can_create_competition_game(current_user, championship):
        raise HTTPException(
            status_code=403,
            detail="Sem permissão para importar jogos nesta competição"
        )
    
    # Read file
    content = await file.read()
    filename = file.filename.lower()
    
    rows = []
    if filename.endswith('.csv'):
        import csv
        import io
        decoded = content.decode('utf-8-sig')
        reader = csv.DictReader(io.StringIO(decoded))
        rows = list(reader)
    elif filename.endswith('.xlsx'):
        from openpyxl import load_workbook
        import io
        wb = load_workbook(io.BytesIO(content), read_only=True)
        ws = wb.active
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if any(cell is not None for cell in row):
                rows.append(dict(zip(headers, row)))
    else:
        raise HTTPException(status_code=400, detail="Formato não suportado. Use .csv ou .xlsx")
    
    results = {"success": 0, "errors": [], "imported": []}
    
    for row in rows:
        try:
            # Multilingual header mapping (PT, ES, FR, IT, EN)
            home_team = (
                row.get('Equipa Casa') or row.get('equipa_casa') or 
                row.get('Equipo Local') or row.get('equipo_local') or
                row.get('Équipe Domicile') or row.get('equipe_domicile') or
                row.get('Squadra Casa') or row.get('squadra_casa') or
                row.get('Home Team') or row.get('home_team') or ""
            )
            
            opponent_team = (
                row.get('Adversário') or row.get('adversario') or row.get('Equipa Visitante') or
                row.get('Rival') or row.get('Equipo Visitante') or
                row.get('Adversaire') or row.get('Équipe Extérieur') or
                row.get('Avversario') or row.get('Squadra Ospite') or
                row.get('Opponent') or row.get('Away Team') or row.get('opponent') or ""
            )
            
            match_date_str = (
                row.get('Data') or row.get('data') or
                row.get('Fecha') or row.get('fecha') or
                row.get('Date') or row.get('date') or ""
            )
            
            match_time = (
                row.get('Hora') or row.get('hora') or
                row.get('Heure') or row.get('heure') or
                row.get('Ora') or row.get('ora') or
                row.get('Time') or row.get('time') or ""
            )
            
            venue = (
                row.get('Local') or row.get('local') or row.get('Pavilhão') or
                row.get('Lugar') or row.get('Pabellón') or
                row.get('Lieu') or row.get('Pavillon') or
                row.get('Luogo') or row.get('Palazzetto') or
                row.get('Venue') or row.get('venue') or row.get('Pavilion') or ""
            )
            
            location_str = (
                row.get('Localização') or row.get('localizacao') or
                row.get('Ubicación') or row.get('ubicacion') or
                row.get('Localisation') or row.get('localisation') or
                row.get('Localizzazione') or row.get('localizzazione') or
                row.get('Location') or row.get('location') or "casa"
            )
            
            matchday_str = (
                row.get('Jornada') or row.get('jornada') or
                row.get('Journée') or row.get('journee') or
                row.get('Giornata') or row.get('giornata') or
                row.get('Round') or row.get('round') or row.get('Matchday') or ""
            )
            
            if not opponent_team:
                results["errors"].append(f"Linha sem adversário: {row}")
                continue
            
            # Parse date
            try:
                if isinstance(match_date_str, datetime):
                    match_date = match_date_str
                elif match_date_str:
                    from dateutil import parser
                    match_date = parser.parse(str(match_date_str))
                else:
                    match_date = datetime.now(timezone.utc)
            except:
                match_date = datetime.now(timezone.utc)
            
            # Normalize location
            location_map = {
                'casa': 'casa', 'home': 'casa', 'domicile': 'casa', 'local': 'casa',
                'fora': 'fora', 'away': 'fora', 'extérieur': 'fora', 'visitante': 'fora', 'trasferta': 'fora',
                'neutral': 'neutro', 'neutro': 'neutro', 'neutre': 'neutro'
            }
            location = location_map.get(str(location_str).lower().strip(), 'casa')
            
            # Parse matchday
            matchday = None
            if matchday_str:
                try:
                    matchday = int(matchday_str)
                except:
                    pass
            
            # Determine if club match
            team_name = championship.get('team_name', '')
            is_club_match = home_team.lower() == team_name.lower() if team_name else True
            
            # Create match
            match = ChampionshipMatch(
                championship_id=championship_id,
                team_id=championship['team_id'],
                home_team=home_team if home_team else None,
                opponent_team=opponent_team,
                match_date=match_date,
                match_time=str(match_time).strip() if match_time else None,
                location=location,
                venue=str(venue).strip() if venue else None,
                is_club_match=is_club_match,
                matchday=matchday
            )
            
            match_dict = match.model_dump()
            match_dict['match_date'] = match_dict['match_date'].isoformat()
            match_dict['created_at'] = match_dict['created_at'].isoformat()
            
            await db.championship_matches.insert_one(match_dict)
            match_dict.pop('_id', None)
            
            results["success"] += 1
            results["imported"].append({
                "home_team": home_team,
                "opponent_team": opponent_team,
                "match_date": str(match_date)[:10]
            })
            
        except Exception as e:
            results["errors"].append(f"Erro na linha: {str(e)}")
    
    return results

class GameSheetImport(BaseModel):
    url: str
    match_id: str
    confirm_conflict: bool = False

@api_router.post("/championships/matches/import-gamesheet")
async def import_gamesheet(data: GameSheetImport, current_user: dict = Depends(get_current_user)):
    """Import match data from APL game sheet URL"""

    match = await db.championship_matches.find_one({"id": data.match_id}, {"_id": 0})
    if not match:
        raise HTTPException(status_code=404, detail="Jogo não encontrado")

    championship = await db.championships.find_one(
        {"id": match["championship_id"]},
        {"_id": 0}
    )
    if not championship:
        raise HTTPException(status_code=404, detail="Competição não encontrada")

    if not await can_import_competition_gamesheet(current_user, championship):
        raise HTTPException(
            status_code=403,
            detail="Sem permissão para importar boletins desta competição"
        )
    
    # Fetch and parse the game sheet
    
    try:
        async with httpx.AsyncClient(timeout=30.0) as client:
            response = await client.get(data.url)
            response.raise_for_status()
            html = response.text
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Erro ao aceder à ficha de jogo: {str(e)}")
    
    soup = BeautifulSoup(html, 'html.parser')
    
    # Extract match result
    result_data = {}
    
    try:
        # Find the final score from the styled span (e.g., "7 - 2")
        score_span = soup.find('span', style=lambda x: x and 'background-color:#000000' in x if x else False)
        if score_span:
            score_text = score_span.get_text(strip=True)
            score_match = re.search(r'(\d+)\s*-\s*(\d+)', score_text)
            if score_match:
                result_data['home_score'] = int(score_match.group(1))
                result_data['away_score'] = int(score_match.group(2))
        
        # Fallback: search for score pattern in HTML
        if 'home_score' not in result_data:
            score_pattern = re.search(r'<b>(\d+)\s*-\s*(\d+)</b>', html)
            if score_pattern:
                result_data['home_score'] = int(score_pattern.group(1))
                result_data['away_score'] = int(score_pattern.group(2))
        
        # Find stats tables - look for tables with class="estadisticas" or containing player stats
        player_stats = []
        current_team = None
        
        # Find all statistics divs
        stats_divs = soup.find_all('div', class_='estadisticas')
        
        for stats_div in stats_divs:
            tables = stats_div.find_all('table')
            for table in tables:
                rows = table.find_all('tr')
                
                for row in rows:
                    cells = row.find_all('td')
                    if not cells:
                        continue
                    
                    # Check for team header (contains logo and team name)
                    first_cell = cells[0]
                    if first_cell.get('class') and 'fondo1' in first_cell.get('class', []):
                        # This is a team header row
                        team_name_span = first_cell.find('span')
                        if team_name_span:
                            current_team = team_name_span.get_text(strip=True)
                        continue
                    
                    # Check for header row (contains column names like G, AG, etc.)
                    if first_cell.get('class') and 'fondo3' in first_cell.get('class', []):
                        continue
                    
                    # Skip "Técnicos" section
                    row_text = row.get_text()
                    if 'Técnicos' in row_text or 'Total da equipa' in row_text:
                        continue
                    
                    # Parse player row - structure:
                    # [0]=nº, [1]=5I, [2]=flag, [3]=Nome, [4]=G, [5]=AG, [6]=D, [7]=Pe, [8]=LD, [9]=Amarelo, [10]=Azul, [11]=Vermelho
                    if len(cells) >= 12:
                        try:
                            jersey_text = cells[0].get_text(strip=True)
                            
                            # Skip if first column is a staff role (D, T, T2, MAS, MEC) or empty
                            if jersey_text in ['D', 'T', 'T2', 'MAS', 'MEC', '']:
                                continue
                            
                            # Get jersey number - can be "1", "01", etc.
                            jersey = jersey_text.lstrip('0') or '0'
                            
                            # Get player name from column 3
                            name_cell = cells[3]
                            name = name_cell.get_text(strip=True)
                            # Clean up name (remove captain marker ©)
                            name = re.sub(r'\s*©\s*', '', name).strip()
                            name = re.sub(r'\s+', ' ', name)  # Normalize spaces
                            
                            if not name:
                                continue
                            
                            # Extract stats - handle both numbers and "--" for staff
                            def parse_stat(cell_index):
                                """Parse a stat cell, returning 0 for non-numeric values"""
                                if cell_index >= len(cells):
                                    return 0
                                text = cells[cell_index].get_text(strip=True)
                                if text.isdigit():
                                    return int(text)
                                return 0
                            
                            def parse_fraction_stat(cell_index):
                                """Parse a fraction stat like '1/2' returning (scored, failed)
                                Format: X/Y where X = scored, Y = total attempts
                                Returns: (scored, failed) where failed = Y - X
                                """
                                if cell_index >= len(cells):
                                    return (0, 0)
                                text = cells[cell_index].get_text(strip=True)
                                if '/' in text:
                                    parts = text.split('/')
                                    try:
                                        scored = int(parts[0]) if parts[0].isdigit() else 0
                                        total = int(parts[1]) if len(parts) > 1 and parts[1].isdigit() else scored
                                        failed = max(0, total - scored)
                                        return (scored, failed)
                                    except (ValueError, IndexError):
                                        return (0, 0)
                                return (0, 0)
                            
                            goals = parse_stat(4)       # G (Golos)
                            assists = parse_stat(5)     # AG (Assistências)
                            defenses = parse_stat(6)    # D (Defesas)
                            penalties_tuple = parse_fraction_stat(7)   # Pe (Penáltis - X/Y format)
                            free_kicks_tuple = parse_fraction_stat(8)  # LD (Livres Diretos - X/Y format)
                            yellow_cards = parse_stat(9)   # Amarelo
                            blue_cards = parse_stat(10)    # Azul
                            red_cards = parse_stat(11)     # Vermelho
                            
                            if current_team:
                                player_stats.append({
                                    'team': current_team,
                                    'jersey_number': jersey,
                                    'name': name,
                                    'G': goals,
                                    'AG': assists,
                                    'D': defenses,
                                    'PM': penalties_tuple[0],      # Penáltis Marcados
                                    'PF': penalties_tuple[1],      # Penáltis Falhados
                                    'LDM': free_kicks_tuple[0],    # Livres Diretos Marcados
                                    'LDF': free_kicks_tuple[1],    # Livres Diretos Falhados
                                    'yellow': yellow_cards,
                                    'blue': blue_cards,
                                    'red': red_cards,
                                    # Legacy fields for backwards compatibility
                                    'goals': goals,
                                    'assists': assists,
                                    'defenses': defenses,
                                    'penalties_scored': penalties_tuple[0],
                                    'free_kicks_scored': free_kicks_tuple[0],
                                    'yellow_cards': yellow_cards,
                                    'blue_cards': blue_cards,
                                    'red_cards': red_cards
                                })
                        except Exception as e:
                            logging.warning(f"Error parsing player row: {e}")
                            continue
        
        result_data['player_stats'] = player_stats
        
        # Extract competition and date info
        text_content = soup.get_text()
        
        # Find date
        date_pattern = re.search(r'(\d{1,2})/(\d{1,2})/(\d{4})', text_content)
        if date_pattern:
            result_data['match_date'] = f"{date_pattern.group(3)}-{date_pattern.group(2).zfill(2)}-{date_pattern.group(1).zfill(2)}"
        
        # Find venue - stop at line break or special characters
        venue_pattern = re.search(r'Recinto:\s*([A-Z0-9\s\.\-]+)', text_content)
        if venue_pattern:
            result_data['venue'] = venue_pattern.group(1).strip()
        
        # Find referee
        referee_pattern = re.search(r'Árbitros?:\s*([^\n,]+)', text_content)
        if referee_pattern:
            result_data['referee'] = referee_pattern.group(1).strip()
            
    except Exception as e:
        logging.error(f"Error parsing gamesheet: {e}")
        raise HTTPException(status_code=400, detail=f"Erro ao processar ficha de jogo: {str(e)}")
    
    # Update match with imported data
    match = await db.championship_matches.find_one({"id": data.match_id}, {"_id": 0})
    if not match:
        raise HTTPException(status_code=404, detail="Jogo não encontrado")

    # Sprint 2.3C — deteção de conflito entre
    # o resultado atual e o resultado oficial.
    current_home_score = match.get(
        "home_score"
    )
    
    current_away_score = match.get(
        "away_score"
    )
    
    official_home_score = result_data.get(
        "home_score"
    )
    
    official_away_score = result_data.get(
        "away_score"
    )
    
    has_current_result = (
        current_home_score is not None
        and current_away_score is not None
        and bool(match.get("is_completed"))
    )
    
    has_official_result = (
        official_home_score is not None
        and official_away_score is not None
    )
    
    result_conflict = (
        has_current_result
        and has_official_result
        and (
            int(current_home_score)
            != int(official_home_score)
            or int(current_away_score)
            != int(official_away_score)
        )
    )
    
    if (
        result_conflict
        and not data.confirm_conflict
    ):
        conflict_payload = {
            "type": "gamesheet_result_conflict",
            "message": (
                "O resultado oficial é diferente "
                "do resultado atual"
            ),
            "match_id": data.match_id,
            "source": infer_match_source(
                data.url
            ),
            "url": data.url,
            "current": {
                "home_score": current_home_score,
                "away_score": current_away_score,
                "result": (
                    f"{current_home_score} - "
                    f"{current_away_score}"
                ),
                "result_source": (
                    match.get("result_source")
                    or "manual"
                ),
                "updated_at": match.get(
                    "updated_at"
                ),
                "updated_by": match.get(
                    "updated_by"
                ),
            },
            "official": {
                "home_score": official_home_score,
                "away_score": official_away_score,
                "result": (
                    f"{official_home_score} - "
                    f"{official_away_score}"
                ),
                "venue": result_data.get(
                    "venue"
                ),
                "referee": result_data.get(
                    "referee"
                ),
            },
        }
    
        now_iso = datetime.now(
            timezone.utc
        ).isoformat()
    
        await db.championship_matches.update_one(
            {"id": data.match_id},
            {
                "$set": {
                    "sync_status": "conflict",
                    "last_sync_error": (
                        "Resultado oficial diferente "
                        "do resultado atual"
                    ),
                    "conflict_detected_at": now_iso,
                    "conflict_data": conflict_payload,
                }
            }
        )
    
        raise HTTPException(
            status_code=409,
            detail=conflict_payload
        )
    
    update_fields = {
        "home_score": result_data.get(
            "home_score",
            0
        ),
        "away_score": result_data.get(
            "away_score",
            0
        ),
        "is_completed": True,
    
        "result_source": "gamesheet",
        "result_updated_at": datetime.now(
            timezone.utc
        ).isoformat(),
        "result_updated_by": current_user["id"],
    
        "gamesheet_url": data.url,
        "gamesheet_imported_at": datetime.now(
            timezone.utc
        ).isoformat(),
    }
    
    if result_data.get('venue'):
        update_fields['venue'] = result_data['venue']
    if result_data.get('referee'):
        update_fields['referee'] = result_data['referee']
    
    await db.championship_matches.update_one({"id": data.match_id}, {"$set": update_fields})
    
    # Update player statistics
    championship = await db.championships.find_one({"id": match['championship_id']}, {"_id": 0})
    team_id = championship['team_id'] if championship else None
    
    stats_updated = 0
    unmatched_players = []
    
    if team_id and result_data.get('player_stats'):
        # Get all team members once for efficient matching
        team_members = await db.users.find({"team_ids": team_id}, {"_id": 0}).to_list(200)
        
        # Helper function to normalize text (remove accents, uppercase)
        def normalize_name(name):
            """Remove accents and normalize name for comparison"""
            import unicodedata
            if not name:
                return ""
            # Normalize unicode and remove diacritics
            normalized = unicodedata.normalize('NFD', name)
            ascii_text = ''.join(c for c in normalized if unicodedata.category(c) != 'Mn')
            return ascii_text.upper().strip()
        
        for ps in result_data['player_stats']:
            player = None
            ps_name_normalized = normalize_name(ps['name'])
            ps_name_parts = ps_name_normalized.split()
            
            # Method 1: Try to find by exact jersey number
            for member in team_members:
                profile = member.get('profile') or {}
                sports_info = profile.get('sports_info') or {}
                jersey = sports_info.get('jersey_number')
                if jersey and str(jersey) == str(ps['jersey_number']):
                    player = member
                    break
            
            # Method 2: Try by full name match (normalized, no accents)
            if not player:
                for member in team_members:
                    member_name_normalized = normalize_name(member.get('name', ''))
                    if member_name_normalized == ps_name_normalized:
                        player = member
                        break
            
            # Method 3: Try by first + last name match (ignoring middle names)
            # "ANTONIO PEREIRA" should match "António Matias Pereira"
            if not player and len(ps_name_parts) >= 2:
                ps_first = ps_name_parts[0]
                ps_last = ps_name_parts[-1]
                for member in team_members:
                    member_name_normalized = normalize_name(member.get('name', ''))
                    member_parts = member_name_normalized.split()
                    if len(member_parts) >= 2:
                        mem_first = member_parts[0]
                        mem_last = member_parts[-1]
                        # Match if first AND last name match
                        if ps_first == mem_first and ps_last == mem_last:
                            player = member
                            break
            
            # Method 4: Try by partial name match (any matching part with min 3 chars)
            if not player:
                for member in team_members:
                    member_name_normalized = normalize_name(member.get('name', ''))
                    member_parts = member_name_normalized.split()
                    matches = 0
                    for ps_part in ps_name_parts:
                        if len(ps_part) >= 3:
                            for mem_part in member_parts:
                                if ps_part == mem_part:
                                    matches += 1
                                    break
                    # Need at least 2 matching parts for a positive match
                    if matches >= 2:
                        player = member
                        break
            
            # Method 5: Try fuzzy match - name contains or is contained
            if not player:
                for member in team_members:
                    member_name_normalized = normalize_name(member.get('name', ''))
                    if len(ps_name_normalized) >= 4 and len(member_name_normalized) >= 4:
                        if ps_name_normalized in member_name_normalized or member_name_normalized in ps_name_normalized:
                            player = member
                            break
            
            if player:
                # Create or update player match stats
                stat_id = f"{data.match_id}_{player['id']}"
                stat_doc = {
                    "id": stat_id,
                    "match_id": data.match_id,
                    "championship_id": match['championship_id'],
                    "player_id": player['id'],
                    "team_id": team_id,
                    "goals": ps['goals'],
                    "assists": ps['assists'],
                    "defenses": ps.get('defenses', 0),
                    "penalties_scored": ps.get('penalties_scored', 0),
                    "free_kicks_scored": ps.get('free_kicks_scored', 0),
                    "yellow_cards": ps['yellow_cards'],
                    "blue_cards": ps['blue_cards'],
                    "red_cards": ps['red_cards'],
                    "imported_from_gamesheet": True,
                    "created_at": datetime.now(timezone.utc).isoformat()
                }
                
                # Use the same collection as the rest of the app: player_match_stats
                await db.player_match_stats.update_one(
                    {"id": stat_id},
                    {"$set": stat_doc},
                    upsert=True
                )
                stats_updated += 1
            else:
                # Track unmatched players (with stats) for feedback
                if ps['goals'] > 0 or ps['assists'] > 0 or ps['yellow_cards'] > 0 or ps['blue_cards'] > 0 or ps['red_cards'] > 0:
                    unmatched_players.append(f"#{ps['jersey_number']} {ps['name']}")

    now_iso = datetime.now(
        timezone.utc
    ).isoformat()
    
    await db.championship_matches.update_one(
        {"id": data.match_id},
        {
            "$set": {
                "official_match_url": data.url,
                "gamesheet_url": data.url,
                "source_url": data.url,
    
                "source": infer_match_source(
                    data.url
                ),
    
                "sync_status": "synced",
    
                "last_synced_at": now_iso,
                "last_sync_error": None,

                "conflict_detected_at": None,
                "conflict_data": None,
    
                "is_verified": True,
                "verified_at": now_iso,
                "verified_by": current_user["id"],
            }
        }
    )
    
    # Guardar sempre o snapshot da ficha oficial,
    # mesmo quando não existem estatísticas de jogadores.
    gamesheet_update = {
        "gamesheet_raw_data": {
            "home_score": result_data.get(
                "home_score",
                0
            ),
            "away_score": result_data.get(
                "away_score",
                0
            ),
            "result": (
                f"{result_data.get('home_score', 0)}-"
                f"{result_data.get('away_score', 0)}"
            ),
            "venue": result_data.get(
                "venue"
            ),
            "referee": result_data.get(
                "referee"
            ),
            "total_players": len(
                result_data.get(
                    "player_stats",
                    []
                )
            ),
            "source": infer_match_source(
                data.url
            ),
            "imported_at": now_iso,
        }
    }
    
    # Guardar estatísticas individuais apenas
    # quando existem na ficha.
    if result_data.get("player_stats"):
        gamesheet_update[
            "gamesheet_player_stats"
        ] = result_data["player_stats"]
    
    await db.championship_matches.update_one(
        {"id": data.match_id},
        {
            "$set": gamesheet_update
        }
    )

    # Recuperar o jogo já com o resultado oficial aplicado.
    updated_match = await db.championship_matches.find_one(
        {"id": data.match_id},
        {"_id": 0}
    )
    
    if updated_match:
        await sync_calendar_event_from_match(
            data.match_id,
            championship=championship,
            match_data=updated_match,
            current_user_id=current_user["id"]
        )    
        
    response = {
        "message": "Ficha de jogo importada com sucesso",
        "result": (
            f"{result_data.get('home_score', 0)} - "
            f"{result_data.get('away_score', 0)}"
        ),
        "players_found": len(
            result_data.get("player_stats", [])
        ),
        "stats_updated": stats_updated,
    }

    if unmatched_players:
        response["unmatched_players"] = unmatched_players
        response["message"] += f" ({len(unmatched_players)} jogadores com estatísticas não encontrados na equipa)"
    
    return response


@api_router.get("/championships/matches/{match_id}/gamesheet-stats")
async def get_match_gamesheet_stats(match_id: str, current_user: dict = Depends(get_current_user)):
    """Get the raw gamesheet stats for a match"""
    match = await db.championship_matches.find_one({"id": match_id}, {"_id": 0})
    if not match:
        raise HTTPException(status_code=404, detail="Jogo não encontrado")
    
    return {
        "match_id": match_id,
        "home_score": match.get('home_score'),
        "away_score": match.get('away_score'),
        "gamesheet_url": match.get('gamesheet_url'),
        "gamesheet_imported_at": match.get('gamesheet_imported_at'),
        "player_stats": match.get('gamesheet_player_stats', []),
        "raw_data": match.get('gamesheet_raw_data', {})
    }


class GameSheetExtract(BaseModel):
    url: str

@api_router.post("/championships/extract-gamesheet-stats")
async def extract_gamesheet_stats(data: GameSheetExtract, current_user: dict = Depends(get_current_user)):
    """
    Extract and transform player statistics from a game sheet URL.
    Returns a structured list of players with all stats fields.
    
    Output schema per player:
    - player_name: str
    - team: str
    - jersey_number: str
    - G: int (goals)
    - AG: int (assists)
    - D: int (saves/defenses)
    - PM: int (penalties scored)
    - PF: int (penalties failed)
    - LDM: int (direct free hits scored)
    - LDF: int (direct free hits failed)
    - yellow: int (yellow cards)
    - blue: int (blue cards)
    - red: int (red cards)
    """
    checker = get_permission_checker(current_user)
    
    if not checker.can_manage_stats:
        raise HTTPException(status_code=403, detail="Sem permissão para extrair estatísticas")
    
    # Fetch the game sheet
    try:
        async with httpx.AsyncClient(timeout=30.0, follow_redirects=True) as client:
            response = await client.get(data.url)
            response.raise_for_status()
            # Try to detect and fix encoding issues
            # APL uses Windows-1252 encoding for Portuguese characters
            try:
                html = response.content.decode('windows-1252')
            except:
                try:
                    html = response.content.decode('iso-8859-1')
                except:
                    html = response.text
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Erro ao aceder à ficha de jogo: {str(e)}")
    
    soup = BeautifulSoup(html, 'html.parser')
    
    # Detect language from headers (multilingual support)
    def detect_language(soup):
        """Detect language from common header patterns"""
        text = soup.get_text().lower()
        if 'golos' in text or 'assistências' in text:
            return 'pt'
        elif 'goles' in text or 'asistencias' in text:
            return 'es'
        elif 'buts' in text or 'passes décisives' in text:
            return 'fr'
        elif 'gol' in text or 'assist' in text:
            return 'it'
        return 'en'
    
    language = detect_language(soup)
    
    # Normalize player name (remove accents, lowercase, trim)
    def normalize_player_name(name):
        import unicodedata
        # Remove captain marker
        name = re.sub(r'\s*©\s*', '', name).strip()
        # Normalize unicode characters
        name = unicodedata.normalize('NFD', name)
        name = ''.join(c for c in name if unicodedata.category(c) != 'Mn')
        # Clean up spaces
        name = re.sub(r'\s+', ' ', name).strip()
        return name
    
    result = {
        'home_score': None,
        'away_score': None,
        'teams': [],
        'players': [],
        'language_detected': language
    }
    
    try:
        # Find the final score
        score_span = soup.find('span', style=lambda x: x and 'background-color:#000000' in x if x else False)
        if score_span:
            score_text = score_span.get_text(strip=True)
            score_match = re.search(r'(\d+)\s*-\s*(\d+)', score_text)
            if score_match:
                result['home_score'] = int(score_match.group(1))
                result['away_score'] = int(score_match.group(2))
        
        # Fallback score search
        if result['home_score'] is None:
            score_pattern = re.search(r'<b>(\d+)\s*-\s*(\d+)</b>', html)
            if score_pattern:
                result['home_score'] = int(score_pattern.group(1))
                result['away_score'] = int(score_pattern.group(2))
        
        # Find all statistics divs
        stats_divs = soup.find_all('div', class_='estadisticas')
        current_team = None
        
        for stats_div in stats_divs:
            tables = stats_div.find_all('table')
            for table in tables:
                rows = table.find_all('tr')
                
                for row in rows:
                    cells = row.find_all('td')
                    if not cells:
                        continue
                    
                    # Check for team header
                    first_cell = cells[0]
                    if first_cell.get('class') and 'fondo1' in first_cell.get('class', []):
                        team_name_span = first_cell.find('span')
                        if team_name_span:
                            current_team = team_name_span.get_text(strip=True)
                            if current_team and current_team not in result['teams']:
                                result['teams'].append(current_team)
                        continue
                    
                    # Skip header rows and technical staff
                    if first_cell.get('class') and 'fondo3' in first_cell.get('class', []):
                        continue
                    
                    row_text = row.get_text()
                    if 'Técnicos' in row_text or 'Total da equipa' in row_text:
                        continue
                    
                    # Parse player row
                    if len(cells) >= 12:
                        try:
                            jersey_text = cells[0].get_text(strip=True)
                            
                            # Skip staff roles
                            if jersey_text in ['D', 'T', 'T2', 'MAS', 'MEC', '']:
                                continue
                            
                            jersey = jersey_text.lstrip('0') or '0'
                            
                            # Get player name
                            name_cell = cells[3]
                            name = name_cell.get_text(strip=True)
                            name = re.sub(r'\s*©\s*', '', name).strip()
                            name = re.sub(r'\s+', ' ', name)
                            
                            if not name:
                                continue
                            
                            # Parse stat functions
                            def parse_stat(cell_index):
                                if cell_index >= len(cells):
                                    return 0
                                text = cells[cell_index].get_text(strip=True)
                                if text == '--' or text == '-':
                                    return 0
                                if text.isdigit():
                                    return int(text)
                                return 0
                            
                            def parse_fraction(cell_index):
                                """Parse X/Y format, return (scored, failed)"""
                                if cell_index >= len(cells):
                                    return (0, 0)
                                text = cells[cell_index].get_text(strip=True)
                                if text == '--' or text == '-' or text == '':
                                    return (0, 0)
                                if '/' in text:
                                    parts = text.split('/')
                                    try:
                                        scored = int(parts[0]) if parts[0].strip().isdigit() else 0
                                        total = int(parts[1]) if len(parts) > 1 and parts[1].strip().isdigit() else scored
                                        failed = max(0, total - scored)
                                        return (scored, failed)
                                    except (ValueError, IndexError):
                                        return (0, 0)
                                return (0, 0)
                            
                            # Extract all stats
                            G = parse_stat(4)
                            AG = parse_stat(5)
                            D = parse_stat(6)
                            Pe = parse_fraction(7)
                            LD = parse_fraction(8)
                            yellow = parse_stat(9)
                            blue = parse_stat(10)
                            red = parse_stat(11)
                            
                            player_data = {
                                'player_name': name,
                                'player_name_normalized': normalize_player_name(name).lower(),
                                'team': current_team or 'Unknown',
                                'jersey_number': jersey,
                                'G': G,
                                'AG': AG,
                                'D': D,
                                'PM': Pe[0],
                                'PF': Pe[1],
                                'LDM': LD[0],
                                'LDF': LD[1],
                                'yellow': yellow,
                                'blue': blue,
                                'red': red
                            }
                            
                            # Validation
                            assert player_data['PM'] >= 0, "PM must be >= 0"
                            assert player_data['PF'] >= 0, "PF must be >= 0"
                            assert player_data['LDM'] >= 0, "LDM must be >= 0"
                            assert player_data['LDF'] >= 0, "LDF must be >= 0"
                            
                            result['players'].append(player_data)
                            
                        except Exception as e:
                            logging.warning(f"Error parsing player row: {e}")
                            continue
        
    except Exception as e:
        logging.error(f"Error extracting gamesheet stats: {e}")
        raise HTTPException(status_code=400, detail=f"Erro ao processar ficha de jogo: {str(e)}")
    
    return result


class APLCalendarImport(BaseModel):
    """Import championship calendar from APL website"""
    url: str  # URL of the APL division calendar page
    championship_id: str

@api_router.post("/championships/import-apl-calendar")
async def import_apl_calendar(
    data: APLCalendarImport,
    current_user: dict = Depends(get_current_user)
):
    """Import matches calendar from APL division page"""

    championship = await db.championships.find_one(
        {"id": data.championship_id},
        {"_id": 0}
    )

    if not championship:
        raise HTTPException(
            status_code=404,
            detail="Campeonato não encontrado"
        )

    if not await can_create_competition_game(current_user, championship):
        raise HTTPException(
            status_code=403,
            detail="Sem permissão para importar calendário desta competição"
        )

    team_id = championship.get("team_id")
    
    # Fetch the APL calendar page
    try:
        async with httpx.AsyncClient(timeout=30.0) as client:
            response = await client.get(data.url)
            response.raise_for_status()
            html = response.text
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Erro ao aceder ao calendário APL: {str(e)}")
    
    soup = BeautifulSoup(html, 'html.parser')
    
    matches_imported = 0
    matches_data = []
    
    try:
        # Find all match rows in the calendar tables
        # APL calendars typically have tables with match info
        tables = soup.find_all('table')
        
        for table in tables:
            rows = table.find_all('tr')
            current_matchday = None
            
            for row in rows:
                cells = row.find_all(['td', 'th'])
                row_text = row.get_text(strip=True)
                
                # Check for matchday header (e.g., "Jornada 1", "J1", "1ª Jornada")
                matchday_match = re.search(r'[Jj]ornada\s*(\d+)|[Jj](\d+)|(\d+)[ªº]?\s*[Jj]ornada', row_text)
                if matchday_match:
                    current_matchday = int(matchday_match.group(1) or matchday_match.group(2) or matchday_match.group(3))
                    continue
                
                # Skip header rows
                if len(cells) < 3:
                    continue
                
                # Try to extract match info from cells
                # Common patterns: [Date] [Time] [Home Team] [Score/vs] [Away Team] [Venue]
                cell_texts = [cell.get_text(strip=True) for cell in cells]
                
                # Look for date pattern (DD/MM/YYYY or DD-MM-YYYY)
                date_str = None
                time_str = None
                home_team = None
                away_team = None
                venue = None
                score_home = None
                score_away = None
                
                for i, text in enumerate(cell_texts):
                    # Date pattern
                    date_match = re.search(r'(\d{1,2})[/-](\d{1,2})[/-](\d{2,4})', text)
                    if date_match:
                        day, month, year = date_match.groups()
                        if len(year) == 2:
                            year = "20" + year
                        date_str = f"{year}-{month.zfill(2)}-{day.zfill(2)}"
                        continue
                    
                    # Time pattern
                    time_match = re.search(r'(\d{1,2})[h:.](\d{2})', text)
                    if time_match and not date_str:
                        time_str = f"{time_match.group(1).zfill(2)}:{time_match.group(2)}"
                        continue
                    
                    # Score pattern (e.g., "3-2", "3 - 2")
                    score_match = re.search(r'^(\d+)\s*[-x]\s*(\d+)$', text)
                    if score_match:
                        score_home = int(score_match.group(1))
                        score_away = int(score_match.group(2))
                        continue
                    
                    # Team names (non-empty text that's not date/time/score)
                    if text and len(text) >= 2 and not date_match and not time_match and not score_match:
                        if text.lower() not in ['vs', 'x', '-', 'local', 'hora', 'data', 'resultado']:
                            if not home_team:
                                home_team = text
                            elif not away_team:
                                away_team = text
                            elif not venue and len(text) > 3:
                                venue = text
                
                # Create match if we have minimum required data
                if home_team and away_team and date_str:
                    match_data = {
                        "home_team": home_team,
                        "away_team": away_team,
                        "date": date_str,
                        "time": time_str or "15:00",
                        "matchday": current_matchday,
                        "venue": venue,
                        "score_home": score_home,
                        "score_away": score_away
                    }
                    matches_data.append(match_data)
        
        # Import matches to database
        club = await db.clubs.find_one({}, {"_id": 0, "id": 1, "name": 1})
        club_name = club.get('name', '') if club else ''
        team = await db.teams.find_one({"id": team_id}, {"_id": 0, "name": 1})
        team_name = team.get('name', '') if team else ''
        
        for match_data in matches_data:
            # Determine if this is a club match
            is_club_match = (
                club_name.lower() in match_data['home_team'].lower() or
                club_name.lower() in match_data['away_team'].lower() or
                team_name.lower() in match_data['home_team'].lower() or
                team_name.lower() in match_data['away_team'].lower()
            )
            
            # Determine location (home/away/neutral)
            location = 'casa'
            if is_club_match:
                if club_name.lower() in match_data['home_team'].lower() or team_name.lower() in match_data['home_team'].lower():
                    location = 'casa'
                else:
                    location = 'fora'
            
            # Create match record
            match_id = str(uuid.uuid4())
            match_datetime = f"{match_data['date']}T{match_data['time']}:00"
            
            match_doc = {
                "id": match_id,
                "championship_id": data.championship_id,
                "team_id": team_id,
                "home_team": match_data['home_team'] if not is_club_match else None,
                "opponent_team": match_data['away_team'] if is_club_match and location == 'casa' else match_data['home_team'],
                "match_date": match_datetime,
                "location": location,
                "venue": match_data['venue'],
                "is_club_match": is_club_match,
                "matchday": match_data['matchday'],
                "home_score": match_data['score_home'],
                "away_score": match_data['score_away'],
                "is_completed": match_data['score_home'] is not None,
                "imported_from_apl": True,
                "apl_import_url": data.url,
                "created_at": datetime.now(timezone.utc).isoformat()
            }
            
            # Check if match already exists (same teams and date)
            existing = await db.championship_matches.find_one({
                "championship_id": data.championship_id,
                "match_date": match_datetime
            })
            
            if not existing:
                await db.championship_matches.insert_one(match_doc)
                matches_imported += 1
        
    except Exception as e:
        logging.error(f"Error parsing APL calendar: {e}")
        raise HTTPException(status_code=400, detail=f"Erro ao processar calendário APL: {str(e)}")
    
    return {
        "message": "Calendário APL importado com sucesso",
        "matches_found": len(matches_data),
        "matches_imported": matches_imported,
        "matches_skipped": len(matches_data) - matches_imported
    }


# =====================
# Match Lineup Endpoints
# =====================

@api_router.get("/championships/matches/{match_id}/lineup")
async def get_match_lineup(
    match_id: str,
    current_user: dict = Depends(get_current_user)
):
    match = await db.championship_matches.find_one({"id": match_id}, {"_id": 0})
    if not match:
        raise HTTPException(status_code=404, detail="Jogo não encontrado")

    championship = await db.championships.find_one(
        {"id": match["championship_id"]},
        {"_id": 0}
    )
    if not championship:
        raise HTTPException(status_code=404, detail="Competição não encontrada")

    if not await can_view_competition(current_user, championship):
        raise HTTPException(status_code=403, detail="Sem acesso a este jogo")

    lineup = await db.match_lineups.find_one({"match_id": match_id}, {"_id": 0})

    if not lineup:
        return {
            "match_id": match_id,
            "championship_id": match["championship_id"],
            "team_id": match["team_id"],
            "starting_five": [],
            "bench": [],
            "captain_id": None,
            "goalkeeper_id": None,
            "exists": False
        }

    lineup["exists"] = True
    return lineup


@api_router.post("/championships/matches/{match_id}/lineup")
async def save_match_lineup(
    match_id: str,
    lineup_data: dict,
    current_user: dict = Depends(get_current_user)
):
    match = await db.championship_matches.find_one({"id": match_id}, {"_id": 0})
    if not match:
        raise HTTPException(status_code=404, detail="Jogo não encontrado")

    championship = await db.championships.find_one(
        {"id": match["championship_id"]},
        {"_id": 0}
    )
    if not championship:
        raise HTTPException(status_code=404, detail="Competição não encontrada")

    if not await can_edit_competition_game(current_user, championship):
        raise HTTPException(
            status_code=403,
            detail="Sem permissão para gerir line-up deste jogo"
        )

    starting_five = lineup_data.get("starting_five", [])
    bench = lineup_data.get("bench", [])
    
    captain_id = lineup_data.get("captain_id")
    vice_captain_id = lineup_data.get("vice_captain_id")
    
    goalkeeper_starting_id = lineup_data.get("goalkeeper_starting_id")
    goalkeeper_bench_id = lineup_data.get("goalkeeper_bench_id")
    
    penalty_order = lineup_data.get("penalty_order", [])
    free_kick_order = lineup_data.get("free_kick_order", [])

    ball_center_id = lineup_data.get("ball_center_id")
    last_free_kick_id = lineup_data.get("last_free_kick_id")
    timeout_leader_id = lineup_data.get("timeout_leader_id")
    penalty_main_id = lineup_data.get("penalty_main_id")
    free_kick_main_id = lineup_data.get("free_kick_main_id")

    rotation_plan = lineup_data.get("rotation_plan", [])
    
    tactical_plan = lineup_data.get("tactical_plan")
    coach_notes = lineup_data.get("coach_notes")
    assistant_notes = lineup_data.get("assistant_notes")
    
    status = lineup_data.get("status", "draft")

    if len(starting_five) > 5:
        raise HTTPException(
            status_code=400,
            detail="O 5 inicial não pode ter mais de 5 jogadores"
        )

    existing = await db.match_lineups.find_one({"match_id": match_id}, {"_id": 0})

    now = datetime.now(timezone.utc)

    if existing:
        await db.match_lineups.update_one(
            {"match_id": match_id},
            {
                "$set": {
                    "starting_five": starting_five,
                    "bench": bench,
                    "captain_id": captain_id,
                    "vice_captain_id": vice_captain_id,
                    "goalkeeper_starting_id": goalkeeper_starting_id,
                    "goalkeeper_bench_id": goalkeeper_bench_id,
                    "penalty_order": penalty_order,
                    "free_kick_order": free_kick_order,
                    "rotation_plan": rotation_plan,
                    "tactical_plan": tactical_plan,
                    "coach_notes": coach_notes,
                    "assistant_notes": assistant_notes,
                    "status": status,
                    "version": int(existing.get("version", 1)) + 1,
                    "updated_at": now.isoformat(),
                    "updated_by": current_user["id"],
                    "ball_center_id": ball_center_id,
                    "last_free_kick_id": last_free_kick_id,
                    "timeout_leader_id": timeout_leader_id,
                    "penalty_main_id": penalty_main_id,
                    "free_kick_main_id": free_kick_main_id,
                }
            }
        )
    else:
        lineup = MatchLineup(
            match_id=match_id,
            championship_id=match["championship_id"],
            team_id=match["team_id"],
            starting_five=starting_five,
            bench=bench,
            captain_id=captain_id,
            vice_captain_id=vice_captain_id,
            goalkeeper_starting_id=goalkeeper_starting_id,
            goalkeeper_bench_id=goalkeeper_bench_id,
            penalty_order=penalty_order,
            free_kick_order=free_kick_order,
            rotation_plan=rotation_plan,
            tactical_plan=tactical_plan,
            coach_notes=coach_notes,
            assistant_notes=assistant_notes,
            status=status,
            version=1,
            created_by=current_user["id"],
            updated_by=current_user["id"],
            ball_center_id=ball_center_id,
            last_free_kick_id=last_free_kick_id,
            timeout_leader_id=timeout_leader_id,
            penalty_main_id=penalty_main_id,
            free_kick_main_id=free_kick_main_id,
        )

        lineup_dict = lineup.model_dump()
        lineup_dict["created_at"] = lineup_dict["created_at"].isoformat()
        lineup_dict["updated_at"] = lineup_dict["updated_at"].isoformat()

        await db.match_lineups.insert_one(lineup_dict)

    updated = await db.match_lineups.find_one({"match_id": match_id}, {"_id": 0})
    updated["exists"] = True

    return updated


@api_router.delete("/championships/matches/{match_id}/lineup")
async def delete_match_lineup(
    match_id: str,
    current_user: dict = Depends(get_current_user)
):
    match = await db.championship_matches.find_one({"id": match_id}, {"_id": 0})
    if not match:
        raise HTTPException(status_code=404, detail="Jogo não encontrado")

    championship = await db.championships.find_one(
        {"id": match["championship_id"]},
        {"_id": 0}
    )
    if not championship:
        raise HTTPException(status_code=404, detail="Competição não encontrada")

    if not await can_edit_competition_game(current_user, championship):
        raise HTTPException(
            status_code=403,
            detail="Sem permissão para eliminar line-up deste jogo"
        )

    await db.match_lineups.delete_one({"match_id": match_id})

    return {"message": "Line-up eliminado"}


# =====================
# Competition Teams Endpoints (equipas participantes)
# =====================

@api_router.get("/championships/{championship_id}/teams")
async def get_competition_teams(championship_id: str, current_user: dict = Depends(get_current_user)):
    """Get all teams participating in a championship"""
    championship = await db.championships.find_one({"id": championship_id}, {"_id": 0})
    if not championship:
        raise HTTPException(status_code=404, detail="Campeonato não encontrado")
    
    teams = await db.competition_teams.find({"championship_id": championship_id}, {"_id": 0}).to_list(100)
    return teams

@api_router.post("/championships/{championship_id}/teams")
async def create_competition_team(championship_id: str, data: CompetitionTeamCreate, current_user: dict = Depends(get_current_user)):
    """Create a new team in the championship"""
    checker = get_permission_checker(current_user)
    
    if not checker.can_manage_events:
        raise HTTPException(status_code=403, detail="Sem permissão para gerir equipas")
    
    championship = await db.championships.find_one({"id": championship_id}, {"_id": 0})
    if not championship:
        raise HTTPException(status_code=404, detail="Campeonato não encontrado")
    
    # Check team access
    if not checker.is_admin and not checker.can_access_team(championship.get('team_id')):
        raise HTTPException(status_code=403, detail="Sem acesso a este campeonato")
    
    # Check for duplicate team name
    existing = await db.competition_teams.find_one({
        "championship_id": championship_id,
        "name": {"$regex": f"^{data.name}$", "$options": "i"}
    })
    if existing:
        raise HTTPException(status_code=400, detail="Já existe uma equipa com este nome")
    
    team = CompetitionTeam(
        championship_id=championship_id,
        name=data.name,
        pavilion_name=data.pavilion_name,
        pavilion_address=data.pavilion_address,
        field_player_kit=data.field_player_kit.model_dump() if data.field_player_kit else None,
        goalkeeper_kit=data.goalkeeper_kit.model_dump() if data.goalkeeper_kit else None,
        created_by=current_user['id']
    )
    
    await db.competition_teams.insert_one(team.model_dump())
    
    # Add to championship's participating_teams list
    await db.championships.update_one(
        {"id": championship_id},
        {"$addToSet": {"participating_teams": data.name}}
    )
    
    return {**team.model_dump(), "_id": None}

@api_router.put("/championships/teams/{team_id}")
async def update_competition_team(team_id: str, data: dict, current_user: dict = Depends(get_current_user)):
    """Update a competition team"""
    checker = get_permission_checker(current_user)
    
    if not checker.can_manage_events:
        raise HTTPException(status_code=403, detail="Sem permissão para gerir equipas")
    
    team = await db.competition_teams.find_one({"id": team_id}, {"_id": 0})
    if not team:
        raise HTTPException(status_code=404, detail="Equipa não encontrada")
    
    allowed_fields = ['name', 'pavilion_name', 'pavilion_address', 'field_player_kit', 'goalkeeper_kit']
    update_data = {k: v for k, v in data.items() if k in allowed_fields}
    
    if not update_data:
        raise HTTPException(status_code=400, detail="Sem dados para atualizar")
    
    await db.competition_teams.update_one(
        {"id": team_id},
        {"$set": update_data}
    )
    
    updated = await db.competition_teams.find_one({"id": team_id}, {"_id": 0})
    return updated

@api_router.delete("/championships/teams/{team_id}")
async def delete_competition_team(team_id: str, current_user: dict = Depends(get_current_user)):
    """Delete a competition team"""
    checker = get_permission_checker(current_user)
    
    if not checker.can_manage_events:
        raise HTTPException(status_code=403, detail="Sem permissão para gerir equipas")
    
    team = await db.competition_teams.find_one({"id": team_id}, {"_id": 0})
    if not team:
        raise HTTPException(status_code=404, detail="Equipa não encontrada")
    
    # Remove from championship's participating_teams
    await db.championships.update_one(
        {"id": team.get('championship_id')},
        {"$pull": {"participating_teams": team.get('name')}}
    )
    
    await db.competition_teams.delete_one({"id": team_id})
    return {"message": "Equipa eliminada"}

@api_router.post("/championships/{championship_id}/teams/import")
async def import_competition_teams(championship_id: str, file: UploadFile = File(...), current_user: dict = Depends(get_current_user)):
    """Import competition teams from Excel file"""
    checker = get_permission_checker(current_user)
    
    if not checker.can_import_data:
        raise HTTPException(status_code=403, detail="Sem permissão para importar dados")
    
    championship = await db.championships.find_one({"id": championship_id}, {"_id": 0})
    if not championship:
        raise HTTPException(status_code=404, detail="Campeonato não encontrado")
    
    if not file.filename.endswith(('.xlsx', '.xls', '.csv')):
        raise HTTPException(status_code=400, detail="Formato de ficheiro não suportado. Use Excel (.xlsx, .xls) ou CSV")
    
    try:
        content = await file.read()
        
        if file.filename.endswith('.csv'):
            import io
            df = pd.read_csv(io.BytesIO(content))
        else:
            df = pd.read_excel(io.BytesIO(content))
        
        # Normalize column names
        df.columns = [str(col).lower().strip().replace(' ', '_') for col in df.columns]
        
        teams_imported = 0
        errors = []
        
        for idx, row in df.iterrows():
            try:
                # Get team name
                team_name = None
                for col in ['nome', 'name', 'equipa', 'team', 'nome_equipa', 'team_name']:
                    if col in df.columns and pd.notna(row.get(col)):
                        team_name = str(row[col]).strip()
                        break
                
                if not team_name:
                    errors.append(f"Linha {idx + 2}: Nome da equipa não encontrado")
                    continue
                
                # Check for duplicate
                existing = await db.competition_teams.find_one({
                    "championship_id": championship_id,
                    "name": {"$regex": f"^{team_name}$", "$options": "i"}
                })
                if existing:
                    errors.append(f"Linha {idx + 2}: Equipa '{team_name}' já existe")
                    continue
                
                # Extract other fields
                pavilion_name = None
                pavilion_address = None
                
                for col in ['pavilhao', 'pavilion', 'recinto', 'venue']:
                    if col in df.columns and pd.notna(row.get(col)):
                        pavilion_name = str(row[col]).strip()
                        break
                
                for col in ['morada', 'address', 'endereco', 'pavilion_address']:
                    if col in df.columns and pd.notna(row.get(col)):
                        pavilion_address = str(row[col]).strip()
                        break
                
                # Create team
                team = CompetitionTeam(
                    championship_id=championship_id,
                    name=team_name,
                    pavilion_name=pavilion_name,
                    pavilion_address=pavilion_address,
                    created_by=current_user['id']
                )
                
                await db.competition_teams.insert_one(team.model_dump())
                
                # Add to participating_teams list
                await db.championships.update_one(
                    {"id": championship_id},
                    {"$addToSet": {"participating_teams": team_name}}
                )
                
                teams_imported += 1
                
            except Exception as e:
                errors.append(f"Linha {idx + 2}: {str(e)}")
        
        return {
            "message": f"Importação concluída: {teams_imported} equipas importadas",
            "teams_imported": teams_imported,
            "errors": errors[:10]  # Return first 10 errors
        }
        
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Erro ao processar ficheiro: {str(e)}")


@api_router.get("/championships/{championship_id}/standings")
async def get_championship_standings(championship_id: str, current_user: dict = Depends(get_current_user)):
    championship = await db.championships.find_one({"id": championship_id}, {"_id": 0})
    if not championship:
        raise HTTPException(status_code=404, detail="Campeonato não encontrado")
    
    matches = await db.championship_matches.find({"championship_id": championship_id, "is_completed": True}, {"_id": 0}).to_list(100)
    
    # Build standings
    standings = {}
    team_data = await db.teams.find_one({"id": championship['team_id']}, {"_id": 0, "name": 1})
    our_team_name = team_data['name'] if team_data else "Nossa Equipa"
    
    # Initialize our team
    standings[our_team_name] = {"team": our_team_name, "played": 0, "won": 0, "drawn": 0, "lost": 0, "goals_for": 0, "goals_against": 0, "bonus": 0, "penalty": 0, "points": 0}
    
    # Initialize all teams from matches (including external teams)
    for match in matches:
        is_club_match = match.get('is_club_match', True)
        
        if is_club_match:
            # For club matches, add opponent
            opp = match['opponent_team']
            if opp not in standings:
                standings[opp] = {"team": opp, "played": 0, "won": 0, "drawn": 0, "lost": 0, "goals_for": 0, "goals_against": 0, "bonus": 0, "penalty": 0, "points": 0}
        else:
            # For external matches, add both teams
            home_team = match.get('home_team', 'Equipa A')
            away_team = match['opponent_team']
            if home_team not in standings:
                standings[home_team] = {"team": home_team, "played": 0, "won": 0, "drawn": 0, "lost": 0, "goals_for": 0, "goals_against": 0, "bonus": 0, "penalty": 0, "points": 0}
            if away_team not in standings:
                standings[away_team] = {"team": away_team, "played": 0, "won": 0, "drawn": 0, "lost": 0, "goals_for": 0, "goals_against": 0, "bonus": 0, "penalty": 0, "points": 0}
    
    # Calculate stats
    for match in matches:
        home_score = match.get('home_score', 0)
        away_score = match.get('away_score', 0)
        bonus = match.get('bonus_points', 0)
        penalty = match.get('penalty_points', 0)
        is_club_match = match.get('is_club_match', True)
        
        if is_club_match:
            # Club match: our team vs opponent
            opp = match['opponent_team']
            loc = match.get('location', 'casa')
            
            if loc == 'casa':
                our_goals = home_score
                their_goals = away_score
            else:
                our_goals = away_score
                their_goals = home_score
            
            standings[our_team_name]['played'] += 1
            standings[our_team_name]['goals_for'] += our_goals
            standings[our_team_name]['goals_against'] += their_goals
            standings[our_team_name]['bonus'] += bonus
            standings[our_team_name]['penalty'] += penalty
            
            standings[opp]['played'] += 1
            standings[opp]['goals_for'] += their_goals
            standings[opp]['goals_against'] += our_goals
            
            if our_goals > their_goals:
                standings[our_team_name]['won'] += 1
                standings[our_team_name]['points'] += 3
                standings[opp]['lost'] += 1
            elif our_goals < their_goals:
                standings[our_team_name]['lost'] += 1
                standings[opp]['won'] += 1
                standings[opp]['points'] += 3
            else:
                standings[our_team_name]['drawn'] += 1
                standings[our_team_name]['points'] += 1
                standings[opp]['drawn'] += 1
                standings[opp]['points'] += 1
        else:
            # External match: two external teams
            home_team = match.get('home_team', 'Equipa A')
            away_team = match['opponent_team']
            
            standings[home_team]['played'] += 1
            standings[home_team]['goals_for'] += home_score
            standings[home_team]['goals_against'] += away_score
            standings[home_team]['bonus'] += bonus
            standings[home_team]['penalty'] += penalty
            
            standings[away_team]['played'] += 1
            standings[away_team]['goals_for'] += away_score
            standings[away_team]['goals_against'] += home_score
            
            if home_score > away_score:
                standings[home_team]['won'] += 1
                standings[home_team]['points'] += 3
                standings[away_team]['lost'] += 1
            elif home_score < away_score:
                standings[home_team]['lost'] += 1
                standings[away_team]['won'] += 1
                standings[away_team]['points'] += 3
            else:
                standings[home_team]['drawn'] += 1
                standings[home_team]['points'] += 1
                standings[away_team]['drawn'] += 1
                standings[away_team]['points'] += 1
    
    # Apply bonus/penalty
    for team in standings.values():
        team['points'] += team['bonus'] - team['penalty']
        team['goal_diff'] = team['goals_for'] - team['goals_against']
    
    # Sort by points, then goal difference, then goals for
    sorted_standings = sorted(standings.values(), key=lambda x: (-x['points'], -x['goal_diff'], -x['goals_for']))
    
    return sorted_standings

class MatchTimelineEventCreate(BaseModel):
    event_type: str
    period: int = 1
    minute: int = 0
    second: int = 0
    team_side: Optional[str] = None
    player_id: Optional[str] = None
    player_name: Optional[str] = None
    secondary_player_id: Optional[str] = None
    secondary_player_name: Optional[str] = None
    notes: Optional[str] = None
    score_home: Optional[int] = None
    score_away: Optional[int] = None
    source: str = "manual"


class MatchTimelineEventUpdate(BaseModel):
    event_type: Optional[str] = None
    period: Optional[int] = None
    minute: Optional[int] = None
    second: Optional[int] = None
    team_side: Optional[str] = None
    player_id: Optional[str] = None
    player_name: Optional[str] = None
    secondary_player_id: Optional[str] = None
    secondary_player_name: Optional[str] = None
    notes: Optional[str] = None
    score_home: Optional[int] = None
    score_away: Optional[int] = None


MATCH_TIMELINE_EVENT_TYPES = {
    "match_start",
    "goal",
    "own_goal",
    "yellow_card",
    "blue_card",
    "red_card",
    "substitution",
    "timeout",
    "penalty_scored",
    "penalty_missed",
    "direct_free_kick_scored",
    "direct_free_kick_missed",
    "save",
    "period_start",
    "period_end",
    "halftime",
    "match_end",
    "technical_note",
}


def validate_match_timeline_payload(payload: dict):
    event_type = payload.get("event_type")
    period = payload.get("period", 1)
    minute = payload.get("minute", 0)
    second = payload.get("second", 0)
    team_side = payload.get("team_side")

    if event_type not in MATCH_TIMELINE_EVENT_TYPES:
        raise HTTPException(status_code=400, detail="Tipo de acontecimento inválido")

    if period not in [1, 2, 3, 4]:
        raise HTTPException(status_code=400, detail="Período inválido")

    if minute < 0 or minute > 120:
        raise HTTPException(status_code=400, detail="Minuto inválido")

    if second < 0 or second > 59:
        raise HTTPException(status_code=400, detail="Segundo inválido")

    if team_side not in [None, "home", "away", "neutral"]:
        raise HTTPException(status_code=400, detail="Lado da equipa inválido")


async def get_timeline_match_context(match_id: str):
    match = await db.championship_matches.find_one(
        {"id": match_id, "archived": {"$ne": True}},
        {"_id": 0}
    )

    if not match:
        raise HTTPException(status_code=404, detail="Jogo não encontrado")

    championship = await db.championships.find_one(
        {"id": match.get("championship_id")},
        {"_id": 0}
    )

    if not championship:
        raise HTTPException(status_code=404, detail="Competição não encontrada")

    return match, championship


@api_router.get("/championships/matches/{match_id}/timeline")
async def get_match_timeline(
    match_id: str,
    current_user: dict = Depends(get_current_user)
):
    match, championship = await get_timeline_match_context(match_id)

    if not await can_view_competition(current_user, championship):
        raise HTTPException(status_code=403, detail="Sem acesso a este jogo")

    events = await db.match_timeline_events.find(
        {"match_id": match_id},
        {"_id": 0}
    ).sort([
        ("period", 1),
        ("minute", 1),
        ("second", 1),
        ("created_at", 1),
    ]).to_list(1000)

    return {
        "match_id": match_id,
        "events": events,
        "total": len(events),
    }


@api_router.post("/championships/matches/{match_id}/timeline")
async def create_match_timeline_event(
    match_id: str,
    event_data: MatchTimelineEventCreate,
    current_user: dict = Depends(get_current_user)
):
    match, championship = await get_timeline_match_context(match_id)

    if not await can_edit_competition_statistics(current_user, championship):
        raise HTTPException(
            status_code=403,
            detail="Sem permissão para gerir acontecimentos deste jogo"
        )

    payload = event_data.model_dump()
    validate_match_timeline_payload(payload)

    if payload.get("player_id"):
        player = await db.users.find_one(
            {"id": payload["player_id"]},
            {"_id": 0, "id": 1, "name": 1}
        )
        if not player:
            raise HTTPException(status_code=404, detail="Atleta não encontrado")
        payload["player_name"] = payload.get("player_name") or player.get("name")

    if payload.get("secondary_player_id"):
        secondary_player = await db.users.find_one(
            {"id": payload["secondary_player_id"]},
            {"_id": 0, "id": 1, "name": 1}
        )
        if not secondary_player:
            raise HTTPException(status_code=404, detail="Segundo atleta não encontrado")
        payload["secondary_player_name"] = (
            payload.get("secondary_player_name") or secondary_player.get("name")
        )

    now = datetime.now(timezone.utc)

    timeline_event = {
        "id": str(uuid.uuid4()),
        "match_id": match_id,
        "championship_id": match.get("championship_id"),
        "team_id": match.get("team_id"),
        **payload,
        "created_by": current_user["id"],
        "created_at": now.isoformat(),
        "updated_by": current_user["id"],
        "updated_at": now.isoformat(),
    }

    await db.match_timeline_events.insert_one(timeline_event)
    timeline_event.pop("_id", None)

    await db.championship_matches.update_one(
        {"id": match_id},
        {"$set": {
            "timeline_updated_at": now.isoformat(),
            "timeline_updated_by": current_user["id"],
        }}
    )

    event_type = data.event_type

    if event_type in {"match_start", "period_start"}:
        await update_match_status(
            match_id=match_id,
            status="live",
            updated_by=current_user["id"],
        )

    elif event_type == "halftime":
        await update_match_status(
            match_id=match_id,
            status="half_time",
            updated_by=current_user["id"],
            sync_workflow=False,
        )

    elif event_type == "match_end":
        await update_match_status(
            match_id=match_id,
            status="finished",
            updated_by=current_user["id"],
        )
    
    return timeline_event


@api_router.put("/championships/matches/{match_id}/timeline/{event_id}")
async def update_match_timeline_event(
    match_id: str,
    event_id: str,
    updates: MatchTimelineEventUpdate,
    current_user: dict = Depends(get_current_user)
):
    match, championship = await get_timeline_match_context(match_id)

    if not await can_edit_competition_statistics(current_user, championship):
        raise HTTPException(
            status_code=403,
            detail="Sem permissão para editar acontecimentos deste jogo"
        )

    existing = await db.match_timeline_events.find_one(
        {"id": event_id, "match_id": match_id},
        {"_id": 0}
    )

    if not existing:
        raise HTTPException(status_code=404, detail="Acontecimento não encontrado")

    update_data = updates.model_dump(exclude_unset=True)
    validate_match_timeline_payload({**existing, **update_data})

    if update_data.get("player_id"):
        player = await db.users.find_one(
            {"id": update_data["player_id"]},
            {"_id": 0, "id": 1, "name": 1}
        )
        if not player:
            raise HTTPException(status_code=404, detail="Atleta não encontrado")
        update_data["player_name"] = player.get("name")

    if update_data.get("secondary_player_id"):
        secondary_player = await db.users.find_one(
            {"id": update_data["secondary_player_id"]},
            {"_id": 0, "id": 1, "name": 1}
        )
        if not secondary_player:
            raise HTTPException(status_code=404, detail="Segundo atleta não encontrado")
        update_data["secondary_player_name"] = secondary_player.get("name")

    now = datetime.now(timezone.utc)
    update_data["updated_by"] = current_user["id"]
    update_data["updated_at"] = now.isoformat()

    await db.match_timeline_events.update_one(
        {"id": event_id, "match_id": match_id},
        {"$set": update_data}
    )

    updated = await db.match_timeline_events.find_one(
        {"id": event_id, "match_id": match_id},
        {"_id": 0}
    )

    await db.championship_matches.update_one(
        {"id": match_id},
        {"$set": {
            "timeline_updated_at": now.isoformat(),
            "timeline_updated_by": current_user["id"],
        }}
    )

    return updated


@api_router.delete("/championships/matches/{match_id}/timeline/{event_id}")
async def delete_match_timeline_event(
    match_id: str,
    event_id: str,
    current_user: dict = Depends(get_current_user)
):
    match, championship = await get_timeline_match_context(match_id)

    if not await can_edit_competition_statistics(current_user, championship):
        raise HTTPException(
            status_code=403,
            detail="Sem permissão para eliminar acontecimentos deste jogo"
        )

    existing = await db.match_timeline_events.find_one(
        {"id": event_id, "match_id": match_id},
        {"_id": 0}
    )

    if not existing:
        raise HTTPException(status_code=404, detail="Acontecimento não encontrado")

    await db.match_timeline_events.delete_one(
        {"id": event_id, "match_id": match_id}
    )

    now = datetime.now(timezone.utc)

    await db.championship_matches.update_one(
        {"id": match_id},
        {"$set": {
            "timeline_updated_at": now.isoformat(),
            "timeline_updated_by": current_user["id"],
        }}
    )

    return {"message": "Acontecimento eliminado"}

# ==================== MATCH STATUS ENGINE ====================

MATCH_STATUS_VALUES = {
    "scheduled",
    "pre_match",
    "ready",
    "live",
    "half_time",
    "finished",
    "archived",
}

MATCH_STATUS_LABELS = {
    "scheduled": "Agendado",
    "pre_match": "Pré-jogo",
    "ready": "Pronto",
    "live": "Em jogo",
    "half_time": "Intervalo",
    "finished": "Finalizado",
    "archived": "Arquivado",
}

MATCH_STATUS_WORKFLOW_STAGE = {
    "scheduled": "draft",
    "pre_match": "convocation",
    "ready": "ready",
    "live": "live",
    "half_time": "live",
    "finished": "finished",
    "archived": None,
}


def infer_match_status(match: dict) -> str:
    if not match:
        return "scheduled"

    if match.get("archived") is True or match.get("is_archived") is True:
        return "archived"

    stored_status = match.get("match_status")
    if stored_status in MATCH_STATUS_VALUES:
        return stored_status

    if match.get("is_completed") is True:
        return "finished"

    workflow_stage = (match.get("workflow") or {}).get("stage")

    if workflow_stage in {
        "finished",
        "stats",
        "assistant",
        "evaluation",
        "feedback",
        "closed",
    }:
        return "finished"

    if workflow_stage == "live":
        return "live"

    if workflow_stage == "ready":
        return "ready"

    if workflow_stage in {"convocation", "lineup"}:
        return "pre_match"

    match_date = parse_match_datetime(match.get("match_date"))

    if match_date:
        now = datetime.now(timezone.utc)

        if match_date.tzinfo is None:
            match_date = match_date.replace(tzinfo=timezone.utc)

        hours_until_match = (match_date - now).total_seconds() / 3600

        if -6 <= hours_until_match <= 48:
            return "pre_match"

    return "scheduled"


def build_match_status_payload(
    status: str,
    updated_by: Optional[str] = None,
) -> dict:
    if status not in MATCH_STATUS_VALUES:
        raise HTTPException(
            status_code=400,
            detail="Estado de jogo inválido",
        )

    now_iso = datetime.now(timezone.utc).isoformat()

    return {
        "match_status": status,
        "match_status_label": MATCH_STATUS_LABELS[status],
        "match_status_updated_at": now_iso,
        "match_status_updated_by": updated_by,
    }


async def ensure_match_status(
    match: dict,
    updated_by: Optional[str] = None,
) -> dict:
    if not match:
        return match

    current_status = match.get("match_status")

    if current_status in MATCH_STATUS_VALUES:
        if not match.get("match_status_label"):
            payload = build_match_status_payload(
                current_status,
                updated_by=updated_by,
            )

            await db.championship_matches.update_one(
                {"id": match.get("id")},
                {"$set": payload},
            )

            match.update(payload)

        return match

    inferred_status = infer_match_status(match)

    payload = build_match_status_payload(
        inferred_status,
        updated_by=updated_by,
    )

    await db.championship_matches.update_one(
        {"id": match.get("id")},
        {"$set": payload},
    )

    match.update(payload)

    return match


async def update_match_status(
    match_id: str,
    status: str,
    updated_by: Optional[str] = None,
    sync_workflow: bool = True,
) -> dict:
    if status not in MATCH_STATUS_VALUES:
        raise HTTPException(
            status_code=400,
            detail="Estado de jogo inválido",
        )

    match = await db.championship_matches.find_one(
        {"id": match_id},
        {"_id": 0},
    )

    if not match:
        raise HTTPException(
            status_code=404,
            detail="Jogo não encontrado",
        )

    payload = build_match_status_payload(
        status,
        updated_by=updated_by,
    )

    if status == "archived":
        payload.update({
            "archived": True,
            "archived_at": datetime.now(timezone.utc).isoformat(),
            "archived_by": updated_by,
        })

    await db.championship_matches.update_one(
        {"id": match_id},
        {
            "$set": {
                **payload,
                "updated_at": datetime.now(timezone.utc).isoformat(),
                "updated_by": updated_by,
            }
        },
    )

    workflow_stage = MATCH_STATUS_WORKFLOW_STAGE.get(status)

    if (
        sync_workflow
        and workflow_stage
        and "update_match_workflow" in globals()
    ):
        await update_match_workflow(
            match_id=match_id,
            stage=workflow_stage,
            updated_by=updated_by,
        )

    return payload


class MatchStatusUpdate(BaseModel):
    status: Literal[
        "scheduled",
        "pre_match",
        "ready",
        "live",
        "half_time",
        "finished",
        "archived",
    ]

# ==================== MATCH WORKFLOW ENGINE ====================

MATCH_WORKFLOW_STAGES = [
    "draft",
    "convocation",
    "lineup",
    "ready",
    "live",
    "finished",
    "stats",
    "assistant",
    "evaluation",
    "feedback",
    "closed",
]


MATCH_WORKFLOW_PROGRESS = {
    "draft": 0,
    "convocation": 10,
    "lineup": 20,
    "ready": 30,
    "live": 45,
    "finished": 60,
    "stats": 75,
    "assistant": 85,
    "evaluation": 92,
    "feedback": 97,
    "closed": 100,
}


MATCH_WORKFLOW_LABELS = {
    "draft": "Pré-convocatória",
    "convocation": "Convocatória",
    "lineup": "Line-up",
    "ready": "Pronto para iniciar",
    "live": "Ao vivo",
    "finished": "Jogo terminado",
    "stats": "Estatísticas validadas",
    "assistant": "Assistente atualizado",
    "evaluation": "Avaliações concluídas",
    "feedback": "Feedback concluído",
    "closed": "Jogo encerrado",
}


def build_match_workflow(
    stage: str = "draft",
    updated_by: Optional[str] = None,
) -> dict:
    """
    Cria uma estrutura de workflow válida para um jogo.
    """

    valid_stage = (
        stage
        if stage in MATCH_WORKFLOW_PROGRESS
        else "draft"
    )

    now_iso = datetime.now(timezone.utc).isoformat()

    return {
        "stage": valid_stage,
        "label": MATCH_WORKFLOW_LABELS[valid_stage],
        "progress": MATCH_WORKFLOW_PROGRESS[valid_stage],
        "updated_at": now_iso,
        "updated_by": updated_by,
    }


async def ensure_match_workflow(
    match: dict,
    updated_by: Optional[str] = None,
) -> dict:
    """
    Garante que jogos antigos também possuem workflow.

    Se o jogo ainda não tiver workflow, cria-o de forma automática
    sem obrigar a uma migração manual da base de dados.
    """

    if not match:
        return match

    workflow = match.get("workflow")

    if (
        isinstance(workflow, dict)
        and workflow.get("stage") in MATCH_WORKFLOW_PROGRESS
    ):
        return match

    initial_stage = (
        "finished"
        if match.get("is_completed")
        else "draft"
    )

    new_workflow = build_match_workflow(
        stage=initial_stage,
        updated_by=updated_by,
    )

    await db.championship_matches.update_one(
        {"id": match.get("id")},
        {
            "$set": {
                "workflow": new_workflow,
            }
        },
    )

    match["workflow"] = new_workflow

    return match


async def update_match_workflow(
    match_id: str,
    stage: str,
    updated_by: Optional[str] = None,
    allow_regression: bool = False,
) -> dict:
    """
    Atualiza o estado do workflow de um jogo.

    Por defeito, impede que o workflow recue acidentalmente.
    """

    if stage not in MATCH_WORKFLOW_PROGRESS:
        raise HTTPException(
            status_code=400,
            detail="Estado de workflow inválido",
        )

    match = await db.championship_matches.find_one(
        {
            "id": match_id,
            "archived": {"$ne": True},
        },
        {"_id": 0},
    )

    if not match:
        raise HTTPException(
            status_code=404,
            detail="Jogo não encontrado",
        )

    match = await ensure_match_workflow(
        match,
        updated_by=updated_by,
    )

    current_stage = match["workflow"].get(
        "stage",
        "draft",
    )

    current_index = MATCH_WORKFLOW_STAGES.index(
        current_stage
        if current_stage in MATCH_WORKFLOW_STAGES
        else "draft"
    )

    new_index = MATCH_WORKFLOW_STAGES.index(stage)

    if new_index < current_index and not allow_regression:
        return match["workflow"]

    workflow = build_match_workflow(
        stage=stage,
        updated_by=updated_by,
    )

    await db.championship_matches.update_one(
        {"id": match_id},
        {
            "$set": {
                "workflow": workflow,
                "updated_at": datetime.now(
                    timezone.utc
                ).isoformat(),
            }
        },
    )

    return workflow
    
# ==================== MATCH TIMELINE SYNC ====================

TIMELINE_SCORING_TYPES = {
    "goal",
    "own_goal",
    "penalty_scored",
    "direct_free_kick_scored",
}

TIMELINE_PLAYER_STAT_FIELDS = [
    "goals",
    "assists",
    "saves",
    "penalties_scored",
    "penalties_missed",
    "free_kicks_scored",
    "free_kicks_missed",
    "yellow_cards",
    "blue_cards",
    "red_cards",
]


def empty_timeline_player_stats():
    return {
        "goals": 0,
        "assists": 0,
        "saves": 0,
        "penalties_scored": 0,
        "penalties_missed": 0,
        "free_kicks_scored": 0,
        "free_kicks_missed": 0,
        "yellow_cards": 0,
        "blue_cards": 0,
        "red_cards": 0,
    }


def increment_timeline_player_stat(
    stats_by_player: dict,
    player_id: Optional[str],
    field: str,
    amount: int = 1
):
    if not player_id:
        return

    if player_id not in stats_by_player:
        stats_by_player[player_id] = empty_timeline_player_stats()

    stats_by_player[player_id][field] = (
        int(stats_by_player[player_id].get(field, 0))
        + amount
    )


async def calculate_timeline_sync_data(match_id: str):
    match = await db.championship_matches.find_one(
        {"id": match_id, "archived": {"$ne": True}},
        {"_id": 0}
    )

    if not match:
        raise HTTPException(
            status_code=404,
            detail="Jogo não encontrado"
        )

    championship = await db.championships.find_one(
        {"id": match.get("championship_id")},
        {"_id": 0}
    )

    if not championship:
        raise HTTPException(
            status_code=404,
            detail="Competição não encontrada"
        )

    timeline = await db.match_timeline_events.find(
        {"match_id": match_id},
        {"_id": 0}
    ).sort([
        ("period", 1),
        ("minute", 1),
        ("second", 1),
        ("created_at", 1),
    ]).to_list(1000)

    home_score = 0
    away_score = 0
    stats_by_player = {}
    warnings = []

    for event in timeline:
        event_type = event.get("event_type")
        team_side = event.get("team_side")
        player_id = event.get("player_id")
        secondary_player_id = event.get("secondary_player_id")

        if event_type in TIMELINE_SCORING_TYPES:
            scoring_side = team_side

            if event_type == "own_goal":
                if team_side == "home":
                    scoring_side = "away"
                elif team_side == "away":
                    scoring_side = "home"
                else:
                    warnings.append(
                        "Existe um autogolo sem equipa definida."
                    )
                    scoring_side = None

            if scoring_side == "home":
                home_score += 1
            elif scoring_side == "away":
                away_score += 1
            else:
                warnings.append(
                    f"Acontecimento de golo sem equipa definida: "
                    f"{event.get('id')}"
                )

        if event_type == "goal":
            increment_timeline_player_stat(
                stats_by_player,
                player_id,
                "goals"
            )
            increment_timeline_player_stat(
                stats_by_player,
                secondary_player_id,
                "assists"
            )

        elif event_type == "own_goal":
            # Mantém o autogolo apenas na timeline.
            # Não é somado a "goals" do atleta.
            pass

        elif event_type == "penalty_scored":
            increment_timeline_player_stat(
                stats_by_player,
                player_id,
                "goals"
            )
            increment_timeline_player_stat(
                stats_by_player,
                player_id,
                 "penalties_scored"
            )

        elif event_type == "penalty_missed":
            increment_timeline_player_stat(
                stats_by_player,
                player_id,
                "penalties_missed"
            )

        elif event_type == "direct_free_kick_scored":
            increment_timeline_player_stat(
                stats_by_player,
                player_id,
                "goals"
            )
            increment_timeline_player_stat(
                stats_by_player,
                player_id,
                "free_kicks_scored"
            )

        elif event_type == "direct_free_kick_missed":
            increment_timeline_player_stat(
                stats_by_player,
                player_id,
                "free_kicks_missed"
            )

        elif event_type == "save":
            increment_timeline_player_stat(
                stats_by_player,
                player_id,
                "saves"
            )

        elif event_type == "yellow_card":
            increment_timeline_player_stat(
                stats_by_player,
                player_id,
                "yellow_cards"
            )

        elif event_type == "blue_card":
            increment_timeline_player_stat(
                stats_by_player,
                player_id,
                "blue_cards"
            )

        elif event_type == "red_card":
            increment_timeline_player_stat(
                stats_by_player,
                player_id,
                "red_cards"
            )

    current_stats = await db.player_match_stats.find(
        {"match_id": match_id},
        {"_id": 0}
    ).to_list(500)

    current_stats_by_player = {
        row.get("player_id"): row
        for row in current_stats
        if row.get("player_id")
    }

    player_ids = list(stats_by_player.keys())
    players = []

    if player_ids:
        players = await db.users.find(
            {"id": {"$in": player_ids}},
            {"_id": 0, "id": 1, "name": 1}
        ).to_list(len(player_ids))

    player_name_map = {
        player.get("id"): player.get("name", "Atleta")
        for player in players
    }

    player_changes = []

    for player_id, calculated in stats_by_player.items():
        current = current_stats_by_player.get(player_id, {})

        changes = {}

        for field in TIMELINE_PLAYER_STAT_FIELDS:
            current_value = int(current.get(field, 0) or 0)
            calculated_value = int(calculated.get(field, 0) or 0)

            if current_value != calculated_value:
                changes[field] = {
                    "current": current_value,
                    "timeline": calculated_value,
                }

        player_changes.append({
            "player_id": player_id,
            "player_name": player_name_map.get(
                player_id,
                current.get("player", {}).get("name", "Atleta")
            ),
            "calculated": calculated,
            "changes": changes,
            "has_changes": bool(changes),
        })

    current_home_score = int(match.get("home_score", 0) or 0)
    current_away_score = int(match.get("away_score", 0) or 0)

    score_has_changes = (
        current_home_score != home_score
        or current_away_score != away_score
    )

    return {
        "match": match,
        "championship": championship,
        "timeline_events": timeline,
        "calculated": {
            "home_score": home_score,
            "away_score": away_score,
            "player_stats": stats_by_player,
        },
        "preview": {
            "timeline_events_count": len(timeline),
            "current_score": {
                "home": current_home_score,
                "away": current_away_score,
            },
            "timeline_score": {
                "home": home_score,
                "away": away_score,
            },
            "score_has_changes": score_has_changes,
            "player_changes": player_changes,
            "players_with_changes": len([
                item
                for item in player_changes
                if item["has_changes"]
            ]),
            "warnings": warnings,
        },
    }


@api_router.get("/championships/matches/{match_id}/timeline-sync-preview"
)
async def get_match_timeline_sync_preview(
    match_id: str,
    current_user: dict = Depends(get_current_user)
):
    sync_data = await calculate_timeline_sync_data(match_id)
    championship = sync_data["championship"]

    if not await can_view_competition(
        current_user,
        championship
    ):
        raise HTTPException(
            status_code=403,
            detail="Sem acesso a este jogo"
        )

    return sync_data["preview"]


@api_router.post(
    "/championships/matches/{match_id}/timeline-sync"
)
async def apply_match_timeline_sync(
    match_id: str,
    current_user: dict = Depends(get_current_user)
):
    sync_data = await calculate_timeline_sync_data(match_id)
    match = sync_data["match"]
    championship = sync_data["championship"]
    calculated = sync_data["calculated"]
    preview = sync_data["preview"]

    if not await can_edit_competition_statistics(
        current_user,
        championship
    ):
        raise HTTPException(
            status_code=403,
            detail="Sem permissão para sincronizar este jogo"
        )

    if preview["timeline_events_count"] == 0:
        raise HTTPException(
            status_code=400,
            detail="A timeline ainda não possui acontecimentos"
        )

    now = datetime.now(timezone.utc)
    now_iso = now.isoformat()

    await db.championship_matches.update_one(
        {"id": match_id},
        {
            "$set": {
                "home_score": calculated["home_score"],
                "away_score": calculated["away_score"],
                "is_completed": bool(
                    match.get("is_completed")
                    or any(
                        event.get("event_type") == "match_end"
                        for event in sync_data["timeline_events"]
                    )
                ),
                "timeline_sync_status": "synced",
                "timeline_synced_at": now_iso,
                "timeline_synced_by": current_user["id"],
                "timeline_sync_event_count": (
                    preview["timeline_events_count"]
                ),
                "updated_at": now_iso,
                "updated_by": current_user["id"],
            }
        }
    )

    for player_id, calculated_stats in (
        calculated["player_stats"].items()
    ):
        existing = await db.player_match_stats.find_one(
            {
                "match_id": match_id,
                "player_id": player_id,
            },
            {"_id": 0}
        )

        stat_document = {
            "id": (
                existing.get("id")
                if existing
                else f"{match_id}_{player_id}"
            ),
            "match_id": match_id,
            "championship_id": match.get("championship_id"),
            "team_id": match.get("team_id"),
            "player_id": player_id,
            **calculated_stats,
            # Campos não derivados da timeline são preservados.
            "started_match": (
                existing.get("started_match", False)
                if existing
                else False
            ),
            "timeline_synced": True,
            "timeline_synced_at": now_iso,
            "timeline_synced_by": current_user["id"],
            "updated_at": now_iso,
        }

        if existing:
            stat_document["created_at"] = existing.get(
                "created_at",
                now_iso
            )
        else:
            stat_document["created_at"] = now_iso

        await db.player_match_stats.update_one(
            {
                "match_id": match_id,
                "player_id": player_id,
            },
            {"$set": stat_document},
            upsert=True
        )

    await update_match_workflow(
        match_id=match_id,
        stage="stats",
        updated_by=current_user["id"],
    )
    
    updated_match = await db.championship_matches.find_one(
        {"id": match_id},
        {"_id": 0}
    )

    return {
        "message": (
            "Resultado e estatísticas sincronizados "
            "com a timeline"
        ),
        "match": updated_match,
        "preview": preview,
        "synced_at": now_iso,
    }


# ==================== MATCH CENTER / TECHNICAL ASSISTANT ROUTES ====================

async def get_match_or_404(match_id: str) -> dict:
    match = await db.championship_matches.find_one({"id": match_id}, {"_id": 0})
    if not match:
        raise HTTPException(status_code=404, detail="Jogo não encontrado")

    match = await ensure_match_workflow(match)
    match = await ensure_match_status(match)

    return match


async def get_championship_for_match_or_404(match: dict) -> dict:
    championship = await db.championships.find_one(
        {"id": match.get("championship_id")},
        {"_id": 0}
    )
    if not championship:
        raise HTTPException(status_code=404, detail="Competição não encontrada")
    return championship


async def build_technical_assistant(match_id: str, current_user: dict) -> dict:
    match = await get_match_or_404(match_id)
    championship = await get_championship_for_match_or_404(match)

    if not await can_view_competition(current_user, championship):
        raise HTTPException(status_code=403, detail="Sem acesso a este jogo")

    stats = await db.player_match_stats.find(
        {"match_id": match_id},
        {"_id": 0}
    ).to_list(500)

    team = await db.teams.find_one(
        {"id": match.get("team_id")},
        {"_id": 0, "name": 1}
    )

    total_goals = sum(s.get("goals", 0) for s in stats)
    total_assists = sum(s.get("assists", 0) for s in stats)
    total_saves = sum(s.get("saves", 0) for s in stats)
    total_blue_cards = sum(s.get("blue_cards", 0) for s in stats)
    total_yellow_cards = sum(s.get("yellow_cards", 0) for s in stats)
    total_red_cards = sum(s.get("red_cards", 0) for s in stats)

    top_scorer = None
    top_assistant = None
    top_goalkeeper = None

    if stats:
        top_scorer = max(stats, key=lambda s: s.get("goals", 0), default=None)
        top_assistant = max(stats, key=lambda s: s.get("assists", 0), default=None)
        top_goalkeeper = max(stats, key=lambda s: s.get("saves", 0), default=None)

    async def enrich_player(stat):
        if not stat:
            return None
        player = await db.users.find_one(
            {"id": stat.get("player_id")},
            {"_id": 0, "id": 1, "name": 1}
        )
        return {
            "player_id": stat.get("player_id"),
            "player_name": player.get("name") if player else "Jogador",
            "value": stat,
        }

    top_scorer = await enrich_player(top_scorer) if top_scorer and top_scorer.get("goals", 0) > 0 else None
    top_assistant = await enrich_player(top_assistant) if top_assistant and top_assistant.get("assists", 0) > 0 else None
    top_goalkeeper = await enrich_player(top_goalkeeper) if top_goalkeeper and top_goalkeeper.get("saves", 0) > 0 else None

    highlights = []

    if top_scorer:
        highlights.append({
            "type": "top_scorer",
            "title": "Melhor marcador",
            "text": f"{top_scorer['player_name']} marcou {top_scorer['value'].get('goals', 0)} golo(s).",
        })

    if top_assistant:
        highlights.append({
            "type": "top_assistant",
            "title": "Melhor assistente",
            "text": f"{top_assistant['player_name']} fez {top_assistant['value'].get('assists', 0)} assistência(s).",
        })

    if top_goalkeeper:
        highlights.append({
            "type": "goalkeeper",
            "title": "Guarda-redes",
            "text": f"{top_goalkeeper['player_name']} registou {top_goalkeeper['value'].get('saves', 0)} defesa(s).",
        })

    alerts = []

    if total_blue_cards > 0:
        alerts.append({
            "type": "discipline",
            "level": "warning",
            "text": f"A equipa registou {total_blue_cards} cartão/cartões azul(is).",
        })

    if total_red_cards > 0:
        alerts.append({
            "type": "discipline",
            "level": "danger",
            "text": f"A equipa registou {total_red_cards} cartão/cartões vermelho(s).",
        })

    if match.get("gamesheet_url") and not match.get("technical_assistant_validated"):
        alerts.append({
            "type": "validation",
            "level": "info",
            "text": "Boletim importado. Recomenda-se validação das estatísticas.",
        })

    summary = {
        "result": None,
        "outcome": None,
    }

    if match.get("is_completed"):
        home_score = match.get("home_score", 0)
        away_score = match.get("away_score", 0)
        summary["result"] = f"{home_score} - {away_score}"

        if home_score > away_score:
            summary["outcome"] = "Vitória da equipa da casa"
        elif home_score < away_score:
            summary["outcome"] = "Vitória da equipa visitante"
        else:
            summary["outcome"] = "Empate"

    assistant = {
        "match_id": match_id,
        "championship_id": match.get("championship_id"),
        "team_id": match.get("team_id"),
        "team_name": team.get("name") if team else None,
        "version": match.get("technical_assistant_version", 1),
        "status": match.get("technical_assistant_status", "draft"),
        "validated": match.get("technical_assistant_validated", False),
        "published": match.get("technical_assistant_published", False),
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "summary": summary,
        "totals": {
            "goals": total_goals,
            "assists": total_assists,
            "saves": total_saves,
            "yellow_cards": total_yellow_cards,
            "blue_cards": total_blue_cards,
            "red_cards": total_red_cards,
        },
        "highlights": highlights,
        "alerts": alerts,
        "development_notes": [
            {
                "type": "rotation",
                "text": "A análise de rotação e cumprimento RTP será ligada ao módulo de line-up."
            }
        ],
        "source": {
            "stats_count": len(stats),
            "gamesheet_url": match.get("gamesheet_url"),
            "gamesheet_imported_at": match.get("gamesheet_imported_at"),
        }
    }

    return assistant


@api_router.get("/championships/matches/{match_id}")
async def get_single_championship_match(
    match_id: str,
    current_user: dict = Depends(get_current_user)
):
    match = await get_match_or_404(match_id)
    championship = await get_championship_for_match_or_404(match)

    if not await can_view_competition(current_user, championship):
        raise HTTPException(status_code=403, detail="Sem acesso a este jogo")

    return match


@api_router.get("/matches/{match_id}/technical-assistant")
async def get_match_technical_assistant(
    match_id: str,
    current_user: dict = Depends(get_current_user)
):
    return await build_technical_assistant(match_id, current_user)


@api_router.post("/matches/{match_id}/technical-assistant/regenerate")
async def regenerate_match_technical_assistant(
    match_id: str,
    current_user: dict = Depends(get_current_user)
):
    match = await get_match_or_404(match_id)
    championship = await get_championship_for_match_or_404(match)

    if not await can_edit_competition_statistics(
        current_user,
        championship
    ):
        raise HTTPException(
            status_code=403,
            detail="Sem permissão para recalcular o Assistente Técnico"
        )

    version = int(
        match.get("technical_assistant_version", 1)
    ) + 1

    await db.championship_matches.update_one(
        {"id": match_id},
        {
            "$set": {
                "technical_assistant_version": version,
                "technical_assistant_status": "draft",
                "technical_assistant_validated": False,
                "technical_assistant_published": False,
                "technical_assistant_regenerated_at": datetime.now(
                    timezone.utc
                ).isoformat(),
                "technical_assistant_regenerated_by": current_user["id"],
            }
        }
    )

    assistant = await build_technical_assistant(
        match_id,
        current_user,
    )

    await update_match_workflow(
        match_id=match_id,
        stage="assistant",
        updated_by=current_user["id"],
    )

    return assistant

@api_router.put(
    "/championships/matches/{match_id}/status"
)
async def update_championship_match_status(
    match_id: str,
    data: MatchStatusUpdate,
    current_user: dict = Depends(get_current_user),
):
    match = await db.championship_matches.find_one(
        {
            "id": match_id,
            "archived": {"$ne": True},
        },
        {"_id": 0},
    )

    if not match:
        raise HTTPException(
            status_code=404,
            detail="Jogo não encontrado",
        )

    championship = await db.championships.find_one(
        {"id": match.get("championship_id")},
        {"_id": 0},
    )

    if not championship:
        raise HTTPException(
            status_code=404,
            detail="Competição não encontrada",
        )

    if not await can_edit_competition_game(
        current_user,
        championship,
    ):
        raise HTTPException(
            status_code=403,
            detail="Sem permissão para alterar o estado do jogo",
        )

    if data.status == "archived":
        if not await can_archive_competition(
            current_user,
            championship,
        ):
            raise HTTPException(
                status_code=403,
                detail="Sem permissão para arquivar este jogo",
            )

    status_payload = await update_match_status(
        match_id=match_id,
        status=data.status,
        updated_by=current_user["id"],
    )

    updated_match = await db.championship_matches.find_one(
        {"id": match_id},
        {"_id": 0},
    )

    return {
        "message": "Estado do jogo atualizado",
        "status": status_payload,
        "match": updated_match,
    }


@api_router.post("/matches/{match_id}/technical-assistant/publish")
async def publish_match_technical_assistant(
    match_id: str,
    current_user: dict = Depends(get_current_user)
):
    match = await get_match_or_404(match_id)
    championship = await get_championship_for_match_or_404(match)

    if not await can_edit_competition_statistics(current_user, championship):
        raise HTTPException(
            status_code=403,
            detail="Sem permissão para publicar o Assistente Técnico"
        )

    await db.championship_matches.update_one(
        {"id": match_id},
        {
            "$set": {
                "technical_assistant_status": "published",
                "technical_assistant_validated": True,
                "technical_assistant_published": True,
                "technical_assistant_published_at": datetime.now(timezone.utc).isoformat(),
                "technical_assistant_published_by": current_user["id"],
            }
        }
    )

    return await build_technical_assistant(match_id, current_user)

# ==================== PLAYER MATCH STATS ROUTES ====================

@api_router.post("/matches/{match_id}/player-stats")
async def create_player_match_stats(
    match_id: str,
    stats: PlayerMatchStatsCreate,
    current_user: dict = Depends(get_current_user)
):
    match = await db.championship_matches.find_one(
        {"id": match_id},
        {"_id": 0}
    )

    if not match:
        raise HTTPException(
            status_code=404,
            detail="Jogo não encontrado"
        )

    championship = await db.championships.find_one(
        {"id": match["championship_id"]},
        {"_id": 0}
    )

    if not championship:
        raise HTTPException(
            status_code=404,
            detail="Competição não encontrada"
        )

    if not await can_edit_competition_statistics(current_user, championship):
        raise HTTPException(
            status_code=403,
            detail="Sem permissão para editar estatísticas desta competição"
        )

    player_stats = PlayerMatchStats(
        **stats.model_dump(),
        team_id=match["team_id"],
        championship_id=match["championship_id"]
    )

    stats_dict = player_stats.model_dump()
    stats_dict["created_at"] = stats_dict["created_at"].isoformat()

    # Upsert - atualiza se existir, cria se não existir
    await db.player_match_stats.update_one(
        {
            "match_id": match_id,
            "player_id": stats.player_id
        },
        {
            "$set": stats_dict
        },
        upsert=True
    )

    return stats_dict
    
@api_router.get("/matches/{match_id}/player-stats")
async def get_match_player_stats(match_id: str, current_user: dict = Depends(get_current_user)):
    stats = await db.player_match_stats.find({"match_id": match_id}, {"_id": 0}).to_list(500)
    
    # Deduplicate - keep only the most recent record for each player
    # Group by player_id and keep the one with most recent created_at or non-empty data
    player_stats_map = {}
    for stat in stats:
        player_id = stat.get('player_id')
        if player_id not in player_stats_map:
            player_stats_map[player_id] = stat
        else:
            # Compare - prefer record with actual data
            existing = player_stats_map[player_id]
            existing_has_data = any([
                existing.get('goals', 0) > 0,
                existing.get('saves', 0) > 0,
                existing.get('penalties_scored', 0) > 0,
                existing.get('penalties_missed', 0) > 0,
                existing.get('yellow_cards', 0) > 0,
                existing.get('started_match', False)
            ])
            new_has_data = any([
                stat.get('goals', 0) > 0,
                stat.get('saves', 0) > 0,
                stat.get('penalties_scored', 0) > 0,
                stat.get('penalties_missed', 0) > 0,
                stat.get('yellow_cards', 0) > 0,
                stat.get('started_match', False)
            ])
            # Prefer record with data, or more recent if both have data or neither has
            if new_has_data and not existing_has_data:
                player_stats_map[player_id] = stat
            elif new_has_data == existing_has_data:
                # Compare created_at
                new_time = stat.get('created_at', '')
                existing_time = existing.get('created_at', '')
                if new_time > existing_time:
                    player_stats_map[player_id] = stat
    
    stats = list(player_stats_map.values())
    
    # Enrich with player info
    for stat in stats:
        player = await db.users.find_one({"id": stat['player_id']}, {"_id": 0, "password": 0})
        if player:
            stat['player'] = player
    
    return stats

@api_router.get("/players/{player_id}/match-stats")
async def get_player_all_match_stats(player_id: str, championship_id: Optional[str] = None, current_user: dict = Depends(get_current_user)):
    query = {"player_id": player_id}
    if championship_id:
        query["championship_id"] = championship_id
    
    stats = await db.player_match_stats.find(query, {"_id": 0}).to_list(500)
    return stats

# ==================== EVENT ROUTES ====================

@api_router.post("/events")
async def create_event(event_data: EventCreate, current_user: dict = Depends(get_current_user)):
    checker = get_permission_checker(current_user)
    
    if not checker.can_manage_events:
        raise HTTPException(status_code=403, detail="Sem permissão para criar eventos")
    
    # Calendar V2 compatibility: support both team_id and team_ids
    event_payload = event_data.model_dump()
    
    team_ids = event_payload.get("team_ids") or []
    
    if not team_ids and event_payload.get("team_id"):
        team_ids = [event_payload.get("team_id")]
    
    if team_ids:
        for team_id in team_ids:
            if team_id and not checker.is_admin and not checker.can_access_team(team_id):
                raise HTTPException(status_code=403, detail="Sem acesso a esta equipa")
    
    if team_ids and not event_payload.get("team_id"):
        event_payload["team_id"] = team_ids[0]

    event_payload_for_model = {
        key: value
        for key, value in event_payload.items()
        if key != "team_ids"
    }
    
    event = Event(**event_payload_for_model, created_by=current_user['id'])
    event_dict = event.model_dump(mode="json")
    event_dict["team_ids"] = team_ids
    
    result = await db.events.insert_one(event_dict)

    event_dict["_id"] = str(result.inserted_id)
    
    return event_dict 

    
    # Notify guardians (parents) of team members about the new event
    if team_ids:
        # Format event time for notification
        event_time = event_dict['start_time']
        if isinstance(event_time, str):
            try:
                dt = datetime.fromisoformat(event_time.replace('Z', '+00:00'))
                event_time = dt.strftime('%d/%m/%Y às %H:%M')
            except:
                pass
        
        # Run notification in background (don't block response)
        import asyncio
        for team_id in team_ids:
            asyncio.create_task(notify_guardians_of_team_event(
                team_id=team_id,
                event_title=event_data.title,
                event_type=event_data.event_type,
                event_time=event_time
            ))
    
    return event_dict



def get_convocation_visibility_value(value) -> str:
    if value is None:
        return ConvocationVisibility.all.value
    return value.value if hasattr(value, "value") else str(value)


def is_private_convocation_document(convocation: Optional[dict]) -> bool:
    if not convocation:
        return False

    visibility = get_convocation_visibility_value(convocation.get("visibility"))
    return bool(
        visibility == ConvocationVisibility.private.value
        or convocation.get("is_private") is True
        or convocation.get("privacy") == "private"
    )


def can_view_full_convocation_for_event(current_user: dict, event: Optional[dict]) -> bool:
    if not current_user or not event:
        return False

    checker = get_permission_checker(current_user)
    team_id = event.get("team_id")

    return bool(
        checker.is_admin
        or (
            (checker.is_staff or checker.can_create_convocations or checker.can_manage_attendance)
            and team_id
            and checker.can_access_team(team_id)
        )
    )


def sanitize_convocation_for_user(convocation: Optional[dict], event: Optional[dict], current_user: dict) -> Optional[dict]:
    if not convocation:
        return None

    normalized = normalize_convocation_document(convocation)

    if not is_convocation_visible_to_user(normalized, event, current_user):
        return None

    if not is_private_convocation_document(normalized):
        return normalized

    if can_view_full_convocation_for_event(current_user, event):
        return normalized

    sanitized = dict(normalized)
    # Convocatória privada: atletas/responsáveis sabem que existe, mas não veem a lista de convocados.
    sanitized["player_ids"] = []
    sanitized["is_private"] = True
    sanitized["privacy"] = "private"
    return sanitized


def is_convocation_visible_to_user(convocation: Optional[dict], event: Optional[dict], current_user: dict) -> bool:
    """Visibility gate used by Dashboard, Calendar and /convocations/my.

    Staff with event/team access can see every lifecycle state.
    Athletes/responsáveis only see published or closed convocations in which
    they have an attendance record. Draft/cancelled convocations remain internal.
    """
    if not convocation:
        return False

    normalized = normalize_convocation_document(convocation)
    status = normalized.get("status") or ConvocationStatus.published.value

    if can_view_full_convocation_for_event(current_user, event):
        return True

    return status in [
        ConvocationStatus.published.value,
        ConvocationStatus.closed.value,
    ]


def convocation_event_badge_status(convocation: Optional[dict]) -> Optional[str]:
    if not convocation:
        return None

    lifecycle_status = convocation.get("status") or ConvocationStatus.published.value

    if lifecycle_status == ConvocationStatus.cancelled.value:
        return "cancelled"

    if lifecycle_status == ConvocationStatus.closed.value:
        return "closed"

    if lifecycle_status == ConvocationStatus.draft.value:
        return "draft"

    if is_private_convocation_document(convocation):
        return "private"

    if lifecycle_status == ConvocationStatus.published.value:
        return "launched"

    return "launched"


def _parse_dt(value):
    """Parse ISO datetime strings defensively while keeping existing datetime objects."""
    if isinstance(value, datetime):
        return value
    if isinstance(value, str) and value:
        try:
            return datetime.fromisoformat(value.replace("Z", "+00:00"))
        except Exception:
            return value
    return value


def _event_team_ids(event: dict) -> List[str]:
    team_ids = event.get("team_ids") or []
    if isinstance(team_ids, str):
        team_ids = [team_ids]
    if not team_ids and event.get("team_id"):
        team_ids = [event.get("team_id")]
    return [tid for tid in team_ids if tid]


async def attach_teams_to_events(events: List[dict]) -> List[dict]:
    """Attach team / teams to events with one Mongo query instead of N queries."""
    if not events:
        return events

    all_team_ids = set()
    for event in events:
        for tid in _event_team_ids(event):
            all_team_ids.add(tid)

    team_map: Dict[str, dict] = {}
    if all_team_ids:
        teams = await db.teams.find(
            {"id": {"$in": list(all_team_ids)}},
            {"_id": 0}
        ).to_list(len(all_team_ids))
        team_map = {team.get("id"): team for team in teams if team.get("id")}

    for event in events:
        event["start_time"] = _parse_dt(event.get("start_time"))
        if event.get("end_time"):
            event["end_time"] = _parse_dt(event.get("end_time"))

        team_ids = _event_team_ids(event)
        event["team_ids"] = team_ids
        event["teams"] = [team_map[tid] for tid in team_ids if tid in team_map]

        if not event.get("team") and event["teams"]:
            event["team"] = event["teams"][0]

    return events


async def enrich_event_with_convocation_context(event: dict, current_user: dict) -> dict:
    """Backward-compatible single-event wrapper.

    Internally uses the batch implementation so future routes can call the same
    optimized logic without duplicating database queries.
    """
    if not event or not event.get("id"):
        return event

    enriched = await enrich_events_with_convocation_context([event], current_user)
    return enriched[0] if enriched else event


async def enrich_events_with_convocation_context(events: List[dict], current_user: dict) -> List[dict]:
    """Batch-enrich events with convocation and attendance context.

    Sprint 4.0 Performance:
    Previous implementation executed 2 Mongo queries per event
    (convocation + attendance), which made Calendar/Convocations very slow.
    This version performs bounded batch reads and O(1) in-memory lookups.
    """
    if not events:
        return []

    event_ids = [event.get("id") for event in events if event and event.get("id")]
    if not event_ids:
        return events

    player_ids = get_accessible_player_ids(current_user)

    convocations = await db.convocations.find(
        {"event_id": {"$in": event_ids}},
        {"_id": 0}
    ).sort("created_at", -1).to_list(max(len(event_ids) * 2, 100))

    convocation_by_event: Dict[str, dict] = {}
    for convocation in convocations:
        event_id = convocation.get("event_id")
        if event_id and event_id not in convocation_by_event:
            convocation_by_event[event_id] = normalize_convocation_document(convocation)

    attendance_by_event: Dict[str, dict] = {}
    if player_ids:
        attendances = await db.attendance.find(
            {
                "event_id": {"$in": event_ids},
                "player_id": {"$in": player_ids},
            },
            {"_id": 0}
        ).sort("updated_at", -1).to_list(max(len(event_ids) * len(player_ids), 100))

        for attendance in attendances:
            event_id = attendance.get("event_id")
            if event_id and event_id not in attendance_by_event:
                attendance_by_event[event_id] = attendance

    enriched_events = []
    for event in events:
        if not event or not event.get("id"):
            enriched_events.append(event)
            continue

        enriched = dict(event)
        event_id = event.get("id")

        normalized_convocation = convocation_by_event.get(event_id)
        sanitized_convocation = sanitize_convocation_for_user(
            normalized_convocation,
            event,
            current_user
        )

        visible_convocation = sanitized_convocation
        enriched["has_convocation"] = bool(visible_convocation)
        enriched["convocation"] = visible_convocation
        enriched["convocation_id"] = visible_convocation.get("id") if visible_convocation else None
        enriched["convocation_visibility"] = visible_convocation.get("visibility") if visible_convocation else None
        enriched["convocation_lifecycle_status"] = visible_convocation.get("status") if visible_convocation else None
        enriched["convocation_status"] = convocation_event_badge_status(visible_convocation)
        enriched["is_private_convocation"] = is_private_convocation_document(visible_convocation)

        my_attendance = attendance_by_event.get(event_id)
        if my_attendance:
            enriched["my_attendance"] = my_attendance
            enriched["my_attendance_id"] = my_attendance.get("id")
            enriched["attendance_id"] = my_attendance.get("id")
            enriched["my_attendance_status"] = my_attendance.get("status")
            enriched["attendance_status"] = my_attendance.get("status")

        enriched_events.append(enriched)

    return enriched_events


@api_router.get("/events")
async def get_events(
    team_id: Optional[str] = None,
    event_type: Optional[str] = None,
    championship_id: Optional[str] = None,
    profile_type: Optional[str] = None,
    profile_user_id: Optional[str] = None,
    profile_role: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    checker = get_permission_checker(current_user)
    query = {}

    effective_user = current_user
    effective_checker = checker

    if profile_type == "associated" and profile_user_id:
        allowed_accounts = current_user.get("associated_accounts", [])

        if profile_user_id not in allowed_accounts:
            raise HTTPException(status_code=403, detail="Perfil associado não autorizado")

        associated_user = await db.users.find_one(
            {"id": profile_user_id},
            {"_id": 0, "hashed_password": 0, "password": 0}
        )

        if not associated_user:
            raise HTTPException(status_code=404, detail="Perfil associado não encontrado")

        effective_user = associated_user
        effective_checker = get_permission_checker(effective_user)

    elif profile_type == "self" and profile_role:
        effective_user = {
            **current_user,
            "role": profile_role
        }
        effective_checker = get_permission_checker(effective_user)

    if team_id:
        if not effective_checker.is_admin and not effective_checker.can_access_team(team_id):
            raise HTTPException(status_code=403, detail="Sem acesso a esta equipa")

        query["$or"] = [
            {"team_id": team_id},
            {"team_ids": team_id},
            {"team_ids": {"$in": [team_id]}}
        ]

    elif not effective_checker.is_admin:
        user_teams = list(effective_checker.team_ids)

        if user_teams:
            query["$or"] = [
                {"team_id": {"$in": user_teams}},
                {"team_ids": {"$in": user_teams}}
            ]
        else:
            return []

    if event_type:
        query["event_type"] = event_type

    if championship_id:
        query["championship_id"] = championship_id

    events = await db.events.find(query, {"_id": 0}).sort("start_time", 1).to_list(500)

    events = await attach_teams_to_events(events)
    events = await enrich_events_with_convocation_context(events, effective_user)

    return events


# NOTE: This route MUST be defined BEFORE /events/{event_id} to avoid route conflicts
@api_router.get("/events/upcoming-without-convocation")
async def get_upcoming_events_without_convocation(current_user: dict = Depends(get_current_user)):
    """Get upcoming events (next 24h) that don't have convocations - for coach notifications"""
    checker = get_permission_checker(current_user)
    
    if not checker.can_create_convocations:
        return []
    
    now = datetime.now(timezone.utc)
    next_24h = now + timedelta(hours=24)
    
    # Build query based on user's teams
    query = {
        "start_time": {"$gte": now.isoformat(), "$lte": next_24h.isoformat()},
        "status": "scheduled"
    }
    
    if not checker.is_admin:
        user_teams = list(checker.team_ids)
        if not user_teams:
            return []
        query["team_id"] = {"$in": user_teams}
    
    events = await db.events.find(query, {"_id": 0}).to_list(50)
    
    # Filter events without convocations
    events_without_conv = []
    for event in events:
        convocation = await db.convocations.find_one({"event_id": event['id']}, {"_id": 0})
        if not convocation:
            events_without_conv.append(event)
    
    return events_without_conv
@api_router.get("/events/birthdays")
async def get_birthday_events(
    year: Optional[int] = None,
    team_id: Optional[str] = None,
    team_ids: Optional[str] = None,
    profile_type: Optional[str] = None,
    profile_user_id: Optional[str] = None,
    profile_role: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    checker = get_permission_checker(current_user)

    effective_user = current_user
    effective_checker = checker

    if profile_type == "associated" and profile_user_id:
        allowed_accounts = current_user.get("associated_accounts", [])

        if profile_user_id not in allowed_accounts:
            raise HTTPException(status_code=403, detail="Perfil associado não autorizado")

        associated_user = await db.users.find_one(
            {"id": profile_user_id},
            {"_id": 0, "hashed_password": 0, "password": 0}
        )

        if not associated_user:
            raise HTTPException(status_code=404, detail="Perfil associado não encontrado")

        effective_user = associated_user
        effective_checker = get_permission_checker(effective_user)

    elif profile_type == "self" and profile_role:
        effective_user = {
            **current_user,
            "role": profile_role
        }
        effective_checker = get_permission_checker(effective_user)

    target_year = year or datetime.utcnow().year

    requested_team_ids = [
        item.strip()
        for item in (team_ids or "").split(",")
        if item.strip()
    ]    
    accessible_team_ids = []

    if requested_team_ids:
        accessible_team_ids = requested_team_ids
    
    elif team_id:
        if not effective_checker.is_admin and not effective_checker.can_access_team(team_id):
            raise HTTPException(status_code=403, detail="Sem acesso a esta equipa")
    
        accessible_team_ids = [team_id]
    
    elif effective_checker.is_admin:
        all_teams = await db.teams.find({}, {"_id": 0, "id": 1}).to_list(1000)
        accessible_team_ids = [team.get("id") for team in all_teams if team.get("id")]
    
    else:
        accessible_team_ids = list(effective_checker.team_ids)

    if not accessible_team_ids:
        return []

    birth_date_filter = {
        "$or": [
            {"profile.birth_date": {"$exists": True, "$ne": ""}},
            {"profile.date_of_birth": {"$exists": True, "$ne": ""}},
            {"birth_date": {"$exists": True, "$ne": ""}},
            {"date_of_birth": {"$exists": True, "$ne": ""}},
        ]
    }
    
    if effective_checker.is_admin and not team_id:
        users_query = birth_date_filter
    else:
        users_query = {
            "$and": [
                birth_date_filter,
                {
                    "$or": [
                        {"id": effective_user.get("id")},
                        {"team_ids": {"$in": accessible_team_ids}},
                        {"team_id": {"$in": accessible_team_ids}},
                        {"teams": {"$in": accessible_team_ids}},
                        {"profile.team_ids": {"$in": accessible_team_ids}},
                        {"profile.team_id": {"$in": accessible_team_ids}},
                    ]
                }
            ]
        }

    users = await db.users.find(
        users_query,
        {
            "_id": 0,
            "id": 1,
            "name": 1,
            "role": 1,
            "team_id": 1,
            "team_ids": 1,
            "teams": 1,
            "profile": 1,
            "avatar_url": 1,
        }
    ).to_list(1000)

    birthday_events = []

    for person in users:
        profile = person.get("profile") or {}

        raw_birth_date = (
            profile.get("birth_date")
            or profile.get("date_of_birth")
            or person.get("birth_date")
            or person.get("date_of_birth")
        )

        if not raw_birth_date:
            continue

        try:
            if isinstance(raw_birth_date, datetime):
                birth_date = raw_birth_date.date()
            else:
                raw_value = str(raw_birth_date).strip()
        
                if "/" in raw_value:
                    birth_date = datetime.strptime(raw_value[:10], "%d/%m/%Y").date()
                else:
                    birth_date = datetime.fromisoformat(raw_value[:10]).date()
        except Exception:
            continue

        try:
            birthday_date = birth_date.replace(year=target_year)
        except ValueError:
            birthday_date = birth_date.replace(year=target_year, day=28)

        age = target_year - birth_date.year

        person_team_ids = person.get("team_ids") or []

        if not person_team_ids and person.get("team_id"):
            person_team_ids = [person.get("team_id")]

        if not person_team_ids and person.get("teams"):
            person_team_ids = (
                person.get("teams")
                if isinstance(person.get("teams"), list)
                else [person.get("teams")]
            )

        visible_team_ids = [
            tid for tid in person_team_ids if tid in accessible_team_ids
        ]

        if not visible_team_ids and person.get("id") == effective_user.get("id"):
            visible_team_ids = accessible_team_ids[:1] or person_team_ids[:1] or ["birthday-profile"]
        
        if not visible_team_ids:
            if effective_checker.is_admin:
                visible_team_ids = person_team_ids[:1] or ["birthday-admin"]
            else:
                continue

        main_team_id = visible_team_ids[0]

        team_docs = []

        for birthday_team_id in visible_team_ids:
            team = await db.teams.find_one({"id": birthday_team_id}, {"_id": 0})
            if team:
                team_docs.append(team)

        birthday_events.append({
            "id": f"birthday-{person.get('id')}-{target_year}",
            "event_type": "birthday",
            "title": f"🎂 {person.get('name', 'Aniversário')}",
            "description": f"{age} anos",
            "location": "",
            "start_time": datetime(
                birthday_date.year,
                birthday_date.month,
                birthday_date.day,
                9,
                0,
                0
            ),
            "end_time": datetime(
                birthday_date.year,
                birthday_date.month,
                birthday_date.day,
                9,
                30,
                0
            ),
            "status": "scheduled",
            "team_id": main_team_id,
            "team_ids": visible_team_ids,
            "teams": team_docs,
            "team": team_docs[0] if team_docs else None,
            "is_virtual": True,
            "virtual_type": "birthday",
            "editable": False,
            "person_id": person.get("id"),
            "person_name": person.get("name"),
            "person_role": person.get("role"),
            "age": age,
        })

    birthday_events.sort(key=lambda item: item["start_time"])

    return birthday_events
    
@api_router.get("/events/{event_id}")
async def get_event(event_id: str, current_user: dict = Depends(get_current_user)):
    checker = get_permission_checker(current_user)
    
    event = await db.events.find_one({"id": event_id}, {"_id": 0})
    if not event:
        raise HTTPException(status_code=404, detail="Evento não encontrado")
    
    # Check team access
    event_team_ids = event.get("team_ids") or []

    if not event_team_ids and event.get("team_id"):
        event_team_ids = [event.get("team_id")]
    
    if event_team_ids and not checker.is_admin:
        has_access = any(checker.can_access_team(team_id) for team_id in event_team_ids)
        if not has_access:
            raise HTTPException(status_code=403, detail="Sem acesso a este evento")
    
    event = await enrich_event_with_convocation_context(event, current_user)

    return event

@api_router.put("/events/{event_id}")
async def update_event(event_id: str, updates: dict, current_user: dict = Depends(get_current_user)):
    checker = get_permission_checker(current_user)
    
    if not checker.can_manage_events:
        raise HTTPException(status_code=403, detail="Sem permissão para editar eventos")
    
    event = await db.events.find_one({"id": event_id})
    if not event:
        raise HTTPException(status_code=404, detail="Evento não encontrado")
    
    allowed_fields = [
        'event_type',
        'title',
        'description',
        'location',
        'start_time',
        'end_time',
        'opponent',
        'status',
        'team_id',
        'team_ids',
        'postponed_to_start_time',
        'postponed_to_end_time',
        'postponement_reason',
        'original_event_id',
        'remove_postponed_copy',
    ]
    filtered_updates = {}
    
    for key, value in updates.items():
        if key in allowed_fields:
            if key in [
                'start_time',
                'end_time',
                'postponed_to_start_time',
                'postponed_to_end_time'
            ] and value:
                if isinstance(value, str):
                    filtered_updates[key] = value
                else:
                    filtered_updates[key] = value.isoformat() if hasattr(value, 'isoformat') else value
            else:
                filtered_updates[key] = value
    
    remove_postponed_copy = bool(
        filtered_updates.pop("remove_postponed_copy", False)
    )
    
    if filtered_updates:
        await db.events.update_one(
            {"id": event_id},
            {"$set": filtered_updates}
        )
    
        if remove_postponed_copy or (
            filtered_updates.get("status") == "scheduled"
            and event.get("status") == "postponed"
        ):
            await db.events.delete_many(
                {"original_event_id": event_id}
            )
    
        updated_event = await db.events.find_one({"id": event_id}, {"_id": 0})
    
        try:
            recipients = await recipient_service.get_event_recipients(updated_event)
            logger.info(f"[COMM] Recipients encontrados: {recipients}")
        
            if filtered_updates.get("status") == "cancelled":
                await communication_service.notify_event_cancelled(updated_event)
                logger.info("[COMM] A iniciar envio de emails de cancelamento...")
        
                for recipient in recipients:
                    logger.info(f"[COMM] Enviar cancelamento para {recipient.get('email')}")
                    if recipient.get("email"):
                        result = await communication_service.send_event_postponed_email(
                        to_email=recipient["email"],
                        event=updated_event,
                        club_name="StickPro",
                        recipient_user_id=recipient.get("user_id"),
                    )
                    
                    logger.info(f"[COMM] Resultado envio adiamento: {result}")
        
            elif filtered_updates.get("status") == "postponed":
                await communication_service.notify_event_postponed(updated_event)
                logger.info("[COMM] A iniciar envio de emails de adiamento...")
            
                for recipient in recipients:
                    logger.info(f"[COMM] Enviar adiamento para {recipient.get('email')}")
                    if recipient.get("email"):
                        result = await communication_service.send_event_postponed_email(
                            to_email=recipient["email"],
                            event=updated_event,
                            club_name="StickPro",
                            recipient_user_id=recipient.get("user_id"),
                        )
            
                        logger.info(f"[COMM] Resultado envio adiamento: {result}")
        
            elif filtered_updates.get("status") == "scheduled" and event.get("status") in ["cancelled", "postponed"]:
                await communication_service.notify_event_restored(updated_event)
        
        except Exception as notification_error:
            logger.error(f"Erro ao comunicar alteração do evento {event_id}: {notification_error}")
        
    return {"message": "Evento atualizado"}

@api_router.delete("/events/{event_id}")
async def delete_event(event_id: str, current_user: dict = Depends(get_current_user)):
    checker = get_permission_checker(current_user)
    
    if not checker.can_manage_team:
        raise HTTPException(status_code=403, detail="Sem permissão para eliminar eventos")
    
    event = await db.events.find_one({"id": event_id}, {"_id": 0})
    if not event:
        raise HTTPException(status_code=404, detail="Evento não encontrado")
    
    # Check team access
    event_team_ids = event.get("team_ids") or []

    if not event_team_ids and event.get("team_id"):
        event_team_ids = [event.get("team_id")]
    
    if event_team_ids and not checker.is_admin:
        has_access = any(checker.can_access_team(team_id) for team_id in event_team_ids)
        if not has_access:
            raise HTTPException(status_code=403, detail="Sem acesso a este evento")
    
    await db.events.delete_one({"id": event_id})
    await db.convocations.delete_many({"event_id": event_id})
    await db.attendance.delete_many({"event_id": event_id})
    
    return {"message": "Evento eliminado"}

# ==================== CONVOCATION ROUTES ====================

def serialize_datetime_fields(document: dict, fields: List[str]) -> dict:
    """Return a copy with datetime fields converted to ISO strings for Mongo/API safety."""
    if not document:
        return document

    serialized = dict(document)
    for field in fields:
        value = serialized.get(field)
        if isinstance(value, datetime):
            serialized[field] = value.isoformat()
    return serialized


def normalize_convocation_document(convocation: Optional[dict]) -> Optional[dict]:
    """Backfill lifecycle fields for legacy convocations without mutating the database."""
    if not convocation:
        return None

    normalized = dict(convocation)
    normalized.setdefault("status", ConvocationStatus.published.value)
    normalized.setdefault("published_at", normalized.get("created_at"))
    normalized.setdefault("updated_at", normalized.get("created_at"))

    visibility = get_convocation_visibility_value(normalized.get("visibility"))
    normalized["visibility"] = visibility
    normalized["is_private"] = is_private_convocation_document(normalized)
    normalized["privacy"] = "private" if normalized["is_private"] else "public"

    return serialize_datetime_fields(normalized, ["created_at", "updated_at", "published_at"])


async def get_convocation_with_event_or_404(convocation_id: str):
    convocation = await db.convocations.find_one({"id": convocation_id}, {"_id": 0})
    if not convocation:
        raise HTTPException(status_code=404, detail="Convocatória não encontrada")

    event = await db.events.find_one({"id": convocation.get("event_id")}, {"_id": 0})
    if not event:
        raise HTTPException(status_code=404, detail="Evento associado não encontrado")

    return convocation, event


def ensure_can_manage_convocation(current_user: dict, event: dict):
    checker = get_permission_checker(current_user)

    if not checker.can_create_convocations:
        raise HTTPException(status_code=403, detail="Sem permissão para gerir convocatórias")

    team_id = event.get("team_id")
    if not checker.is_admin and team_id and not checker.can_access_team(team_id):
        raise HTTPException(status_code=403, detail="Sem acesso a este evento")

    return checker


@api_router.post("/convocations")
async def create_convocation(conv_data: ConvocationCreate, current_user: dict = Depends(get_current_user)):
    """Create or update the convocation for an event.

    Sprint 3.3.3: this endpoint is now idempotent per event. Calling it again
    for the same event updates the existing convocation and synchronizes
    attendance rows instead of creating duplicated convocations.
    """
    checker = get_permission_checker(current_user)

    if not checker.can_create_convocations:
        raise HTTPException(status_code=403, detail="Sem permissão para criar convocatórias")

    event = await db.events.find_one({"id": conv_data.event_id}, {"_id": 0})
    if not event:
        raise HTTPException(status_code=404, detail="Evento não encontrado")

    if not checker.is_admin and event.get('team_id') and not checker.can_access_team(event.get('team_id')):
        raise HTTPException(status_code=403, detail="Sem acesso a este evento")

    event_date = event['start_time'] if isinstance(event['start_time'], datetime) else datetime.fromisoformat(str(event['start_time']).replace("Z", "+00:00"))
    if event_date.tzinfo is None:
        event_date = event_date.replace(tzinfo=timezone.utc)

    # Check for unavailable players and warn
    unavailable_player_ids = []
    for player_id in conv_data.player_ids:
        unavails = await db.unavailabilities.find({
            "user_id": player_id,
            "start_date": {"$lte": event_date.isoformat()},
            "end_date": {"$gte": event_date.isoformat()}
        }, {"_id": 0}).to_list(1)
        if unavails:
            unavailable_player_ids.append(player_id)

    available_player_ids = [pid for pid in conv_data.player_ids if pid not in unavailable_player_ids]
    visibility_value = get_convocation_visibility_value(conv_data.visibility)

    if conv_data.is_private or str(conv_data.privacy or '').lower() == 'private':
        visibility_value = ConvocationVisibility.private.value

    now = datetime.now(timezone.utc)
    now_iso = now.isoformat()
    event_started = event_date <= now

    existing_convocation = await db.convocations.find_one(
        {"event_id": conv_data.event_id},
        {"_id": 0},
        sort=[("created_at", -1)]
    )

    if existing_convocation:
        convocation_id = existing_convocation["id"]
        next_status = (
            ConvocationStatus.published.value
            if conv_data.publish_immediately
            else existing_convocation.get("status", ConvocationStatus.draft.value)
        )

        update_data = {
            "player_ids": available_player_ids,
            "message": conv_data.message,
            "visibility": visibility_value,
            "is_private": visibility_value == ConvocationVisibility.private.value,
            "privacy": "private" if visibility_value == ConvocationVisibility.private.value else "public",
            "status": next_status,
            "updated_at": now_iso,
        }

        if next_status == ConvocationStatus.published.value and not existing_convocation.get("published_at"):
            update_data["published_at"] = now_iso

        await db.convocations.update_one(
            {"id": convocation_id},
            {"$set": update_data}
        )
    else:
        initial_status = (
            ConvocationStatus.published.value
            if conv_data.publish_immediately
            else ConvocationStatus.draft.value
        )

        convocation = Convocation(
            event_id=conv_data.event_id,
            player_ids=available_player_ids,
            message=conv_data.message,
            visibility=visibility_value,
            status=initial_status,
            published_at=now if initial_status == ConvocationStatus.published.value else None,
            created_by=current_user['id'],
            updated_at=now
        )

        convocation_id = convocation.id
        conv_dict = convocation.model_dump()
        conv_dict['visibility'] = visibility_value
        conv_dict['status'] = conv_dict['status'].value if hasattr(conv_dict['status'], 'value') else conv_dict['status']
        conv_dict['is_private'] = visibility_value == ConvocationVisibility.private.value
        conv_dict['privacy'] = 'private' if conv_dict['is_private'] else 'public'
        conv_dict['created_at'] = conv_dict['created_at'].isoformat()
        conv_dict['updated_at'] = conv_dict['updated_at'].isoformat()
        conv_dict['published_at'] = conv_dict['published_at'].isoformat() if conv_dict.get('published_at') else None
        await db.convocations.insert_one(conv_dict)

    # Synchronize attendance records for this convocation.
    existing_attendances = await db.attendance.find(
        {
            "event_id": conv_data.event_id,
            "$or": [
                {"convocation_id": convocation_id},
                {"convocation_id": {"$exists": False}},
                {"convocation_id": None},
            ],
        },
        {"_id": 0}
    ).to_list(1000)

    existing_by_player = {
        attendance.get("player_id"): attendance
        for attendance in existing_attendances
        if attendance.get("player_id")
    }

    new_player_ids = set(available_player_ids)
    old_player_ids = set(existing_by_player.keys())

    added_player_ids = [pid for pid in available_player_ids if pid not in old_player_ids]
    removed_player_ids = list(old_player_ids - new_player_ids)

    # Keep old history after the event has started; before it starts, removed
    # players should disappear from this convocation everywhere.
    if removed_player_ids and not event_started:
        await db.attendance.delete_many({
            "event_id": conv_data.event_id,
            "convocation_id": convocation_id,
            "player_id": {"$in": removed_player_ids}
        })

    # Ensure existing rows point to the current convocation.
    await db.attendance.update_many(
        {"event_id": conv_data.event_id, "player_id": {"$in": list(new_player_ids)}},
        {"$set": {"convocation_id": convocation_id, "updated_at": now_iso}}
    )

    for player_id in added_player_ids:
        attendance = Attendance(
            event_id=conv_data.event_id,
            convocation_id=convocation_id,
            player_id=player_id,
            team_id=event['team_id'],
            event_type=event['event_type'],
            championship_id=event.get('championship_id'),
            event_date=event_date
        )
        att_dict = attendance.model_dump()
        att_dict['event_date'] = att_dict['event_date'].isoformat()
        att_dict['updated_at'] = att_dict['updated_at'].isoformat()
        await db.attendance.insert_one(att_dict)

        player = await db.users.find_one({"id": player_id}, {"_id": 0})
        if player and player.get("email"):
            await send_email_notification(
                player['email'],
                f"Convocatória: {event.get('title', 'Evento')}",
                f"<h1>Foste convocado!</h1><p>{conv_data.message or 'Por favor confirma a tua presença.'}</p>"
            )

    updated_convocation = await db.convocations.find_one({"id": convocation_id}, {"_id": 0})
    result = normalize_convocation_document(updated_convocation)
    result["skipped_unavailable_players"] = unavailable_player_ids
    result["added_player_ids"] = added_player_ids
    result["removed_player_ids"] = removed_player_ids if not event_started else []
    result["updated_existing"] = bool(existing_convocation)

    return result


@api_router.get("/convocations")
async def get_convocations(
    event_id: Optional[str] = None,
    team_id: Optional[str] = None,
    status: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """List convocations accessible to staff/admin. Keeps existing frontend getAll() compatible."""
    checker = get_permission_checker(current_user)

    if not checker.can_create_convocations and not checker.is_staff and not checker.is_admin:
        raise HTTPException(status_code=403, detail="Sem permissão para ver convocatórias")

    query: Dict[str, Any] = {}
    if event_id:
        query["event_id"] = event_id
    if status:
        query["status"] = status

    allowed_event_ids: Optional[List[str]] = None

    if team_id:
        if not checker.is_admin and not checker.can_access_team(team_id):
            raise HTTPException(status_code=403, detail="Sem acesso a esta equipa")
        events = await db.events.find({"team_id": team_id}, {"_id": 0, "id": 1}).to_list(1000)
        allowed_event_ids = [event["id"] for event in events]
    elif not checker.is_admin:
        user_team_ids = list(checker.team_ids)
        if not user_team_ids:
            return []
        events = await db.events.find({"team_id": {"$in": user_team_ids}}, {"_id": 0, "id": 1}).to_list(1000)
        allowed_event_ids = [event["id"] for event in events]

    if allowed_event_ids is not None:
        if not allowed_event_ids:
            return []
        if event_id and event_id not in allowed_event_ids:
            raise HTTPException(status_code=403, detail="Sem acesso a este evento")
        query["event_id"] = event_id or {"$in": allowed_event_ids}

    convocations = await db.convocations.find(query, {"_id": 0}).sort("created_at", -1).to_list(500)

    result = []
    for convocation in convocations:
        event = await db.events.find_one({"id": convocation.get("event_id")}, {"_id": 0})
        result.append({
            **normalize_convocation_document(convocation),
            "event": event,
        })

    return result


async def update_convocation_lifecycle(convocation_id: str, next_status: str, current_user: dict, reset_published_at: bool = False):
    convocation, event = await get_convocation_with_event_or_404(convocation_id)
    ensure_can_manage_convocation(current_user, event)

    now_iso = datetime.now(timezone.utc).isoformat()
    update_data = {
        "status": next_status,
        "updated_at": now_iso,
    }

    if next_status == ConvocationStatus.published.value:
        update_data["published_at"] = now_iso
    elif reset_published_at:
        update_data["published_at"] = None

    await db.convocations.update_one(
        {"id": convocation_id},
        {"$set": update_data}
    )

    updated = await db.convocations.find_one({"id": convocation_id}, {"_id": 0})
    return normalize_convocation_document(updated)


@api_router.post("/convocations/{convocation_id}/publish")
async def publish_convocation(convocation_id: str, current_user: dict = Depends(get_current_user)):
    convocation = await update_convocation_lifecycle(
        convocation_id,
        ConvocationStatus.published.value,
        current_user,
    )
    return {"message": "Convocatória publicada", "convocation": convocation}


@api_router.post("/convocations/{convocation_id}/close")
async def close_convocation(convocation_id: str, current_user: dict = Depends(get_current_user)):
    convocation = await update_convocation_lifecycle(
        convocation_id,
        ConvocationStatus.closed.value,
        current_user,
    )
    return {"message": "Convocatória fechada", "convocation": convocation}


@api_router.post("/convocations/{convocation_id}/cancel")
async def cancel_convocation(convocation_id: str, current_user: dict = Depends(get_current_user)):
    convocation = await update_convocation_lifecycle(
        convocation_id,
        ConvocationStatus.cancelled.value,
        current_user,
    )
    return {"message": "Convocatória cancelada", "convocation": convocation}


@api_router.post("/convocations/{convocation_id}/reopen")
async def reopen_convocation(convocation_id: str, current_user: dict = Depends(get_current_user)):
    convocation, _event = await get_convocation_with_event_or_404(convocation_id)
    current_status = convocation.get("status", ConvocationStatus.draft.value)

    if current_status not in [ConvocationStatus.closed.value, ConvocationStatus.cancelled.value]:
        raise HTTPException(
            status_code=400,
            detail="Só é possível reabrir convocatórias fechadas ou canceladas"
        )

    convocation = await update_convocation_lifecycle(
        convocation_id,
        ConvocationStatus.draft.value,
        current_user,
        reset_published_at=True,
    )
    return {"message": "Convocatória reaberta", "convocation": convocation}


def get_accessible_player_ids(current_user: dict):
    player_ids = [str(current_user["id"])]

    linked_player_id = current_user.get("linked_player_id")
    if linked_player_id:
        player_ids.append(str(linked_player_id))

    linked_player_ids = current_user.get("linked_player_ids") or []
    for player_id in linked_player_ids:
        if player_id:
            player_ids.append(str(player_id))

    # remover duplicados
    return list(set(player_ids))

@api_router.get("/training-feedback/my-pending")
async def get_my_pending_training_feedback(current_user: dict = Depends(get_current_user)):
    now = datetime.now(timezone.utc)
    player_ids = get_accessible_player_ids(current_user)

    attendances = await db.attendance.find(
        {
            "player_id": {"$in": player_ids},
            "status": "confirmado",
            "event_type": {"$in": ["treino", "training"]},
        },
        {"_id": 0}
    ).to_list(1000)

    pending = []

    for attendance in attendances:
        event = await db.events.find_one({"id": attendance["event_id"]}, {"_id": 0})
        if not event:
            continue

        event_time = event.get("end_time") or event.get("start_time")

        if isinstance(event_time, str):
            event_time = datetime.fromisoformat(event_time.replace("Z", "+00:00"))

        if event_time.tzinfo is None:
            event_time = event_time.replace(tzinfo=timezone.utc)

        feedback_available_at = event_time + timedelta(minutes=30)

        if feedback_available_at > now:
            continue

        existing_feedback = await db.training_feedback.find_one(
            {
                "event_id": attendance["event_id"],
                "player_id": attendance["player_id"],
            },
            {"_id": 0}
        )

        if existing_feedback:
            continue

        pending.append(
            {
                "attendance": attendance,
                "event": event,
            }
        )

    return pending


@api_router.post("/training-feedback")
async def create_training_feedback(
    feedback: TrainingFeedbackCreate,
    current_user: dict = Depends(get_current_user)
):
    player_ids = get_accessible_player_ids(current_user)

    attendance = await db.attendance.find_one(
        {
            "event_id": feedback.event_id,
            "player_id": {"$in": player_ids},
            "status": "confirmado",
        },
        {"_id": 0}
    )

    if not attendance:
        raise HTTPException(
            status_code=403,
            detail="Só é possível dar feedback de treinos em que estiveste presente."
        )

    if attendance.get("event_type") not in ["treino", "training"]:
        raise HTTPException(
            status_code=400,
            detail="O feedback só está disponível para treinos."
        )

    real_player_id = attendance["player_id"]

    existing_feedback = await db.training_feedback.find_one(
        {
            "event_id": feedback.event_id,
            "player_id": real_player_id,
        },
        {"_id": 0}
    )

    if existing_feedback:
        raise HTTPException(
            status_code=400,
            detail="Já submeteste feedback para este treino."
        )

    if feedback.rating not in ["positive", "neutral", "negative"]:
        raise HTTPException(
            status_code=400,
            detail="Classificação inválida."
        )

    feedback_dict = TrainingFeedback(
        event_id=feedback.event_id,
        player_id=real_player_id,
        team_id=attendance["team_id"],
        rating=feedback.rating,
        comment=feedback.comment.strip() if feedback.comment else None,
    ).model_dump()

    feedback_dict["created_at"] = feedback_dict["created_at"].isoformat()

    await db.training_feedback.insert_one(feedback_dict)

    feedback_dict.pop("_id", None)

    return {
        "message": "Feedback submetido com sucesso.",
        "feedback": feedback_dict,
    }
    
@api_router.get("/training-feedback/team/{team_id}")
async def get_team_training_feedback(
    team_id: str,
    current_user: dict = Depends(get_current_user)
):

    feedbacks = await db.training_feedback.find(
        {"team_id": team_id},
        {"_id": 0}
    ).sort("created_at", -1).to_list(500)

    enriched_feedbacks = []

    for feedback in feedbacks:
        event = await db.events.find_one(
            {"id": feedback.get("event_id")},
            {"_id": 0}
        )

        player = await db.users.find_one(
            {"id": feedback.get("player_id")},
            {
                "_id": 0,
                "id": 1,
                "name": 1,
                "surname": 1,
                "email": 1,
                "avatar_url": 1
            }
        )

        enriched_feedbacks.append({
            **feedback,
            "event": event,
            "player": player
        })

    return enriched_feedbacks

@api_router.get("/training-feedback/event/{event_id}")
async def get_event_training_feedback(
    event_id: str,
    current_user: dict = Depends(get_current_user)
):
    checker = get_permission_checker(current_user)

    if not checker.is_admin and not checker.is_coach and not checker.is_staff:
        raise HTTPException(
            status_code=403,
            detail="Sem permissões para consultar feedback."
        )

    event = await db.events.find_one(
        {"id": event_id},
        {"_id": 0}
    )

    if not event:
        raise HTTPException(
            status_code=404,
            detail="Treino não encontrado."
        )

    if event.get("event_type") not in ["treino", "training"]:
        raise HTTPException(
            status_code=400,
            detail="O feedback só está disponível para treinos."
        )

    if not checker.is_admin and event.get("team_id"):
        if not checker.can_access_team(event.get("team_id")):
            raise HTTPException(
                status_code=403,
                detail="Sem acesso a esta equipa."
            )

    feedbacks = await db.training_feedback.find(
        {"event_id": event_id},
        {"_id": 0}
    ).sort("created_at", -1).to_list(500)

    total = len(feedbacks)

    positive = len([
        f for f in feedbacks
        if f.get("rating") == "positive"
    ])

    neutral = len([
        f for f in feedbacks
        if f.get("rating") == "neutral"
    ])

    negative = len([
        f for f in feedbacks
        if f.get("rating") == "negative"
    ])

    satisfaction_score = 0
    if total > 0:
        satisfaction_score = round(
            ((positive * 100) + (neutral * 50)) / total
        )

    comments = [
        {
            "player_id": f.get("player_id"),
            "rating": f.get("rating"),
            "comment": f.get("comment"),
            "created_at": f.get("created_at"),
        }
        for f in feedbacks
        if f.get("comment")
    ]

    return {
        "event_id": event_id,
        "team_id": event.get("team_id"),
        "total": total,
        "positive": positive,
        "neutral": neutral,
        "negative": negative,
        "satisfaction_score": satisfaction_score,
        "comments": comments,
    }

async def build_my_convocations_for_user(
    current_user: dict,
    *,
    only_pending: bool = False,
    include_past: bool = True,
    limit: int = 500
) -> List[dict]:
    """Single source of truth for personal convocations.

    Sprint 4.0 Performance:
    Loads attendances, events and convocations in batches instead of executing
    multiple find_one calls for each attendance row.
    """
    player_ids = get_accessible_player_ids(current_user)
    if not player_ids:
        return []

    query: Dict[str, Any] = {"player_id": {"$in": player_ids}}
    if only_pending:
        query["status"] = "pendente"

    if not include_past:
        query["event_date"] = {"$gte": datetime.now(timezone.utc).isoformat()}

    attendances = await db.attendance.find(
        query,
        {"_id": 0}
    ).sort("event_date", -1).to_list(limit)

    if not attendances:
        return []

    event_ids = list({att.get("event_id") for att in attendances if att.get("event_id")})
    convocation_ids = list({
        att.get("convocation_id")
        for att in attendances
        if att.get("convocation_id")
    })

    events = await db.events.find(
        {"id": {"$in": event_ids}},
        {"_id": 0}
    ).to_list(len(event_ids))

    events = await attach_teams_to_events(events)
    enriched_events = await enrich_events_with_convocation_context(events, current_user)
    event_map = {event.get("id"): event for event in enriched_events if event.get("id")}

    convocations_by_id: Dict[str, dict] = {}
    if convocation_ids:
        convocations = await db.convocations.find(
            {"id": {"$in": convocation_ids}},
            {"_id": 0}
        ).to_list(len(convocation_ids))
        convocations_by_id = {
            conv.get("id"): normalize_convocation_document(conv)
            for conv in convocations
            if conv.get("id")
        }

    convocations_by_event: Dict[str, dict] = {}
    fallback_needed_event_ids = [
        event_id
        for event_id in event_ids
        if event_id and event_id not in {
            conv.get("event_id") for conv in convocations_by_id.values()
        }
    ]

    if fallback_needed_event_ids:
        fallback_convocations = await db.convocations.find(
            {"event_id": {"$in": fallback_needed_event_ids}},
            {"_id": 0}
        ).sort("created_at", -1).to_list(max(len(fallback_needed_event_ids) * 2, 100))

        for convocation in fallback_convocations:
            event_id = convocation.get("event_id")
            if event_id and event_id not in convocations_by_event:
                convocations_by_event[event_id] = normalize_convocation_document(convocation)

    result = []
    for att in attendances:
        event = event_map.get(att.get("event_id"))
        if not event:
            continue

        convocation = None
        if att.get("convocation_id"):
            convocation = convocations_by_id.get(att.get("convocation_id"))

        if not convocation:
            convocation = convocations_by_event.get(att.get("event_id"))

        normalized_convocation = normalize_convocation_document(convocation)
        sanitized_convocation = sanitize_convocation_for_user(normalized_convocation, event, current_user)

        if normalized_convocation and not sanitized_convocation:
            continue

        if only_pending and normalized_convocation:
            if normalized_convocation.get("status") in [
                ConvocationStatus.closed.value,
                ConvocationStatus.cancelled.value,
            ]:
                continue

        result.append({
            "attendance": att,
            "event": event,
            "convocation": sanitized_convocation,
        })

    return result



@api_router.get("/convocations/my")
async def get_my_convocations(current_user: dict = Depends(get_current_user)):
    return await build_my_convocations_for_user(current_user, only_pending=False, include_past=True, limit=500)

def calculate_commitment_medal(rate: float):
    if rate >= 100:
        return "gold"
    if rate >= 90:
        return "silver"
    if rate >= 75:
        return "bronze"
    return "none"


def calculate_commitment_rate(total: int, confirmed: int):
    if total <= 0:
        return 0
    return round((confirmed / total) * 100)


def calculate_missing_for_next_medal(total: int, confirmed: int):
    if total <= 0:
        return {
            "target": "bronze",
            "missing": 1,
        }

    current_rate = calculate_commitment_rate(total, confirmed)

    if current_rate >= 100:
        return {
            "target": "gold",
            "missing": 0,
        }

    if current_rate >= 90:
        target_rate = 100
        target = "gold"
    elif current_rate >= 75:
        target_rate = 90
        target = "silver"
    else:
        target_rate = 75
        target = "bronze"

    missing = 0
    future_confirmed = confirmed
    future_total = total

    while calculate_commitment_rate(future_total, future_confirmed) < target_rate:
        future_confirmed += 1
        future_total += 1
        missing += 1

        if missing > 50:
            break

    return {
        "target": target,
        "missing": missing,
    }


@api_router.get("/commitment/my")
async def get_my_commitment(current_user: dict = Depends(get_current_user)):
    attendances = await db.attendance.find(
        {"player_id": current_user["id"]},
        {"_id": 0}
    ).to_list(1000)

    training_items = [
        att for att in attendances
        if att.get("event_type") in ["treino", "training"]
    ]

    game_items = [
        att for att in attendances
        if att.get("event_type") in ["jogo", "game", "championship_game", "friendly_game"]
    ]

    training_total = len(training_items)
    training_confirmed = len([
        att for att in training_items
        if att.get("status") == "confirmado"
    ])

    game_total = len(game_items)
    game_confirmed = len([
        att for att in game_items
        if att.get("status") == "confirmado"
    ])

    training_rate = calculate_commitment_rate(training_total, training_confirmed)
    game_rate = calculate_commitment_rate(game_total, game_confirmed)

    return {
        "training": {
            "total": training_total,
            "confirmed": training_confirmed,
            "rate": training_rate,
            "medal": calculate_commitment_medal(training_rate),
            "next_goal": calculate_missing_for_next_medal(
                training_total,
                training_confirmed
            ),
        },
        "games": {
            "total": game_total,
            "confirmed": game_confirmed,
            "rate": game_rate,
            "medal": calculate_commitment_medal(game_rate),
            "next_goal": calculate_missing_for_next_medal(
                game_total,
                game_confirmed
            ),
        },
    }

@api_router.put("/attendance/{attendance_id}")
async def update_attendance(attendance_id: str, update: AttendanceUpdate, current_user: dict = Depends(get_current_user)):
    checker = get_permission_checker(current_user)
    
    attendance = await db.attendance.find_one({"id": attendance_id})
    if not attendance:
        raise HTTPException(status_code=404, detail="Registo de presença não encontrado")
    
    # Get event to check if it has started
    event = await db.events.find_one({"id": attendance['event_id']}, {"_id": 0})
    event_started = False
    if event:
        event_time = event.get('start_time')
        if isinstance(event_time, str):
            event_time = datetime.fromisoformat(event_time.replace('Z', '+00:00'))
        # Ensure event_time is timezone-aware
        if event_time.tzinfo is None:
            event_time = event_time.replace(tzinfo=timezone.utc)
        event_started = datetime.now(timezone.utc) >= event_time
    
    # Determine who can update
    can_update = False
    is_self_or_family = False
    
    if checker.is_admin:
        can_update = True
    elif checker.is_coach and attendance.get('team_id'):
        # Coaches can always update for their teams (even after event started)
        can_update = checker.can_access_team(attendance.get('team_id'))
    elif attendance['player_id'] == current_user['id']:
        # Players can update their own attendance
        is_self_or_family = True
        can_update = True
    elif checker.is_staff and attendance.get('team_id'):
        # Other staff can update attendance for their teams
        can_update = checker.can_access_team(attendance.get('team_id'))
    elif checker.is_family_member and checker.linked_player_id:
        # Family members can update linked player's attendance
        is_self_or_family = True
        can_update = attendance['player_id'] == checker.linked_player_id
    
    if not can_update:
        raise HTTPException(status_code=403, detail="Sem permissão para atualizar esta presença")
    
    # After event started, only admin/coach can edit
    if event_started and is_self_or_family:
        raise HTTPException(
            status_code=403, 
            detail="O evento já começou. Apenas treinadores podem atualizar a presença."
        )
    
    now_iso = datetime.now(timezone.utc).isoformat()
    await db.attendance.update_one(
        {"id": attendance_id},
        {"$set": {"status": update.status, "reason": update.reason, "updated_at": now_iso}}
    )

    if attendance.get("convocation_id"):
        await db.convocations.update_one(
            {"id": attendance.get("convocation_id")},
            {"$set": {"updated_at": now_iso}}
        )
    
    return {"message": "Presença atualizada"}

# ==================== ATTENDANCE ANALYTICS ROUTES ====================

@api_router.get("/teams/{team_id}/attendance")
async def get_team_attendance(team_id: str, season: Optional[str] = None, month: Optional[int] = None, 
                              event_type: Optional[str] = None, championship_id: Optional[str] = None,
                              current_user: dict = Depends(get_current_user)):
    checker = get_permission_checker(current_user)
    
    # Check team access
    if not checker.is_admin and not checker.can_access_team(team_id):
        raise HTTPException(status_code=403, detail="Sem acesso a esta equipa")
    
    query = {"team_id": team_id}
    
    if event_type:
        query["event_type"] = event_type
    if championship_id:
        query["championship_id"] = championship_id
    
    attendances = await db.attendance.find(query, {"_id": 0}).to_list(5000)
    
    # For players, filter to only show their own attendance
    if checker.is_player and not checker.is_staff and not checker.is_admin:
        attendances = [a for a in attendances if a['player_id'] == current_user['id']]
    # For family members, filter to only show linked player's attendance
    elif checker.is_family_member and checker.linked_player_id:
        attendances = [a for a in attendances if a['player_id'] == checker.linked_player_id]
    
    # Filter by month if specified
    if month:
        attendances = [a for a in attendances if datetime.fromisoformat(a['event_date']).month == month]
    
    # Group by player
    player_stats = {}
    for att in attendances:
        pid = att['player_id']
        if pid not in player_stats:
            player_stats[pid] = {"total": 0, "confirmado": 0, "ausente": 0, "pendente": 0, "faltou_sem_aviso": 0}
        player_stats[pid]["total"] += 1
        status = att.get('status', 'pendente')
        if status in player_stats[pid]:
            player_stats[pid][status] += 1
    
    # Enrich with player info
    result = []
    for pid, stats in player_stats.items():
        player = await db.users.find_one({"id": pid}, {"_id": 0, "password": 0})
        if player:
            stats['player'] = player
            stats['attendance_rate'] = round((stats['confirmado'] / stats['total']) * 100, 1) if stats['total'] > 0 else 0
            result.append(stats)
    
    result.sort(key=lambda x: -x['attendance_rate'])
    return result

@api_router.get("/teams/{team_id}/attendance/summary")
async def get_team_attendance_summary(team_id: str, current_user: dict = Depends(get_current_user)):
    checker = get_permission_checker(current_user)
    
    # Check team access
    if not checker.is_admin and not checker.can_access_team(team_id):
        raise HTTPException(status_code=403, detail="Sem acesso a esta equipa")
    
    # Get all attendance for this team
    attendances = await db.attendance.find({"team_id": team_id}, {"_id": 0}).to_list(5000)
    
    # For family members, filter to only show linked player's attendance
    if checker.is_family_member and checker.linked_player_id:
        attendances = [a for a in attendances if a['player_id'] == checker.linked_player_id]
    
    # Group by month
    monthly = {}
    by_event_type = {
        "treino": {"total": 0, "confirmado": 0}, 
        "jogo_campeonato": {"total": 0, "confirmado": 0}, 
        "jogo_amigavel": {"total": 0, "confirmado": 0},
        "torneio": {"total": 0, "confirmado": 0},
        "outro": {"total": 0, "confirmado": 0}
    }
    
    for att in attendances:
        event_date = datetime.fromisoformat(att['event_date']) if isinstance(att['event_date'], str) else att['event_date']
        month_key = event_date.strftime("%Y-%m")
        
        if month_key not in monthly:
            monthly[month_key] = {"total": 0, "confirmado": 0}
        
        monthly[month_key]["total"] += 1
        if att['status'] == 'confirmado':
            monthly[month_key]["confirmado"] += 1
        
        et = att.get('event_type', 'treino')
        if et in by_event_type:
            by_event_type[et]["total"] += 1
            if att['status'] == 'confirmado':
                by_event_type[et]["confirmado"] += 1
    
    return {"monthly": monthly, "by_event_type": by_event_type, "total_records": len(attendances)}

@api_router.get("/teams/{team_id}/attendance/search")
async def search_team_attendance(team_id: str, query: str, current_user: dict = Depends(get_current_user)):
    """Search attendance by player name"""
    checker = get_permission_checker(current_user)
    
    # Check team access
    if not checker.is_admin and not checker.can_access_team(team_id):
        raise HTTPException(status_code=403, detail="Sem acesso a esta equipa")
    
    # For players and family members, they can only search their own data
    if checker.is_player and not checker.is_staff:
        # Get own attendance
        attendances = await db.attendance.find({
            "team_id": team_id,
            "player_id": current_user['id']
        }, {"_id": 0}).to_list(500)
        
        # Filter by search term in own name
        user = await db.users.find_one({"id": current_user['id']}, {"_id": 0, "name": 1})
        if user and query.lower() not in user.get('name', '').lower():
            return []
    elif checker.is_family_member and checker.linked_player_id:
        attendances = await db.attendance.find({
            "team_id": team_id,
            "player_id": checker.linked_player_id
        }, {"_id": 0}).to_list(500)
        
        player = await db.users.find_one({"id": checker.linked_player_id}, {"_id": 0, "name": 1})
        if player and query.lower() not in player.get('name', '').lower():
            return []
    else:
        # Staff and admin can search all players
        # First find players matching the query
        players = await db.users.find({
            "team_ids": team_id,
            "name": {"$regex": query, "$options": "i"}
        }, {"_id": 0, "id": 1, "name": 1}).to_list(50)
        
        player_ids = [p['id'] for p in players]
        
        if not player_ids:
            return []
        
        attendances = await db.attendance.find({
            "team_id": team_id,
            "player_id": {"$in": player_ids}
        }, {"_id": 0}).to_list(2000)
    
    # Group by player
    player_stats = {}
    for att in attendances:
        pid = att['player_id']
        if pid not in player_stats:
            player_stats[pid] = {"total": 0, "confirmado": 0, "ausente": 0, "pendente": 0, "faltou_sem_aviso": 0}
        player_stats[pid]["total"] += 1
        status = att.get('status', 'pendente')
        if status in player_stats[pid]:
            player_stats[pid][status] += 1
    
    # Enrich with player info
    result = []
    for pid, stats in player_stats.items():
        player = await db.users.find_one({"id": pid}, {"_id": 0, "password": 0})
        if player:
            stats['player'] = player
            stats['attendance_rate'] = round((stats['confirmado'] / stats['total']) * 100, 1) if stats['total'] > 0 else 0
            result.append(stats)
    
    result.sort(key=lambda x: -x['attendance_rate'])
    return result

@api_router.get("/teams/{team_id}/attendance/unavailabilities")
async def get_team_attendance_unavailabilities(team_id: str, current_user: dict = Depends(get_current_user)):
    """Get unavailability periods relevant to attendance for this team"""
    checker = get_permission_checker(current_user)
    
    # Check team access
    if not checker.is_admin and not checker.can_access_team(team_id):
        raise HTTPException(status_code=403, detail="Sem acesso a esta equipa")
    
    # Get team members
    team_members = await db.users.find({"team_ids": team_id}, {"_id": 0, "id": 1, "name": 1, "role": 1}).to_list(100)
    member_ids = [m['id'] for m in team_members]
    
    # For players, only show own unavailabilities
    if checker.is_player and not checker.is_staff:
        unavailabilities = await db.unavailabilities.find({
            "user_id": current_user['id']
        }, {"_id": 0}).sort("start_date", -1).to_list(50)
    # For family members, show linked player's unavailabilities
    elif checker.is_family_member and checker.linked_player_id:
        unavailabilities = await db.unavailabilities.find({
            "user_id": checker.linked_player_id
        }, {"_id": 0}).sort("start_date", -1).to_list(50)
    else:
        # Staff and admin can see all team members' unavailabilities
        unavailabilities = await db.unavailabilities.find({
            "user_id": {"$in": member_ids}
        }, {"_id": 0}).sort("start_date", -1).to_list(200)
    
    # Enrich with user info
    for unav in unavailabilities:
        user = await db.users.find_one({"id": unav['user_id']}, {"_id": 0, "name": 1, "role": 1})
        if user:
            unav['user_name'] = user.get('name', 'Unknown')
            unav['user_role'] = user.get('role', 'jogador')
    
    return unavailabilities

@api_router.get("/attendance/my/detailed")
async def get_my_detailed_attendance(current_user: dict = Depends(get_current_user)):
    """Get current user's detailed attendance with event info and unavailabilities"""
    checker = get_permission_checker(current_user)
    
    # Determine which player's data to show
    player_id = current_user['id']
    if checker.is_family_member and checker.linked_player_id:
        player_id = checker.linked_player_id
    
    # Get attendance
    attendances = await db.attendance.find({"player_id": player_id}, {"_id": 0}).sort("event_date", -1).to_list(200)
    
    result = []
    now = datetime.now(timezone.utc)
    
    for att in attendances:
        event = await db.events.find_one({"id": att['event_id']}, {"_id": 0})
        if event:
            # Check if event has started
            event_time = event.get('start_time')
            if isinstance(event_time, str):
                event_time = datetime.fromisoformat(event_time.replace('Z', '+00:00'))
            
            # Ensure event_time is timezone-aware
            if event_time.tzinfo is None:
                event_time = event_time.replace(tzinfo=timezone.utc)
            
            event_started = now >= event_time
            
            result.append({
                "attendance": att,
                "event": event,
                "event_started": event_started,
                "can_edit": not event_started or checker.is_admin or checker.is_coach
            })
    
    # Get unavailabilities
    unavailabilities = await db.unavailabilities.find({
        "user_id": player_id
    }, {"_id": 0}).sort("start_date", -1).to_list(50)
    
    return {
        "attendance": result,
        "unavailabilities": unavailabilities
    }

@api_router.get("/events/{event_id}/attendance")
async def get_event_attendance(event_id: str, current_user: dict = Depends(get_current_user)):
    """Get attendance records for a specific event"""
    checker = get_permission_checker(current_user)
    
    event = await db.events.find_one({"id": event_id}, {"_id": 0})
    if not event:
        raise HTTPException(status_code=404, detail="Event not found")
    
    # Check team access
    if not checker.is_admin and event.get('team_id') and not checker.can_access_team(event.get('team_id')):
        raise HTTPException(status_code=403, detail="Sem acesso a este evento")
    
    convocation = await db.convocations.find_one({"event_id": event_id}, {"_id": 0})
    convocation = normalize_convocation_document(convocation)

    attendances = await db.attendance.find({"event_id": event_id}, {"_id": 0}).to_list(100)
    
    # Convocatória privada: atletas/responsáveis apenas veem o próprio registo.
    if is_private_convocation_document(convocation) and not can_view_full_convocation_for_event(current_user, event):
        accessible_player_ids = set(get_accessible_player_ids(current_user))
        attendances = [a for a in attendances if a.get('player_id') in accessible_player_ids]
    # For family members, filter to only show linked player's attendance
    elif checker.is_family_member and checker.linked_player_id:
        attendances = [a for a in attendances if a['player_id'] == checker.linked_player_id]
    
    # Enrich with player info
    result = []
    for att in attendances:
        player = await db.users.find_one({"id": att['player_id']}, {"_id": 0, "password": 0})
        if player:
            att['player'] = player
        result.append(att)
    
    # Calculate summary
    summary = {
        "total": len(result),
        "confirmado": len([a for a in result if a['status'] == 'confirmado']),
        "ausente": len([a for a in result if a['status'] == 'ausente']),
        "pendente": len([a for a in result if a['status'] == 'pendente'])
    }
    
    return {
        "attendance": result,
        "summary": summary,
        "convocation": sanitize_convocation_for_user(convocation, event, current_user),
        "convocation_status": convocation.get("status") if convocation else None,
        "convocation_lifecycle_status": convocation.get("status") if convocation else None,
        "convocation_visibility": convocation.get("visibility") if convocation else None,
        "is_private": is_private_convocation_document(convocation),
        "is_private_convocation": is_private_convocation_document(convocation),
    }

# ==================== CONVOCATION STATUS ROUTES ====================

class ConvocationStatusUpdate(BaseModel):
    player_id: str
    status: str  # 'confirmado', 'ausente', 'pendente'

@api_router.get("/events/{event_id}/convocation-status")
async def get_convocation_status(event_id: str, current_user: dict = Depends(get_current_user)):
    """Get convocation status for an event - grouped by present/absent/pending"""
    checker = get_permission_checker(current_user)
    
    event = await db.events.find_one({"id": event_id}, {"_id": 0})
    if not event:
        raise HTTPException(status_code=404, detail="Evento não encontrado")
    
    team_id = event.get('team_id')
    if not checker.is_admin and not checker.can_access_team(team_id):
        raise HTTPException(status_code=403, detail="Sem acesso a este evento")

    convocation = await db.convocations.find_one({"event_id": event_id}, {"_id": 0})
    convocation = normalize_convocation_document(convocation)
    
    # Get all attendance records for this event
    attendances = await db.attendance.find({"event_id": event_id}, {"_id": 0}).to_list(500)

    private_restricted_view = (
        is_private_convocation_document(convocation)
        and not can_view_full_convocation_for_event(current_user, event)
    )

    if private_restricted_view:
        accessible_player_ids = set(get_accessible_player_ids(current_user))
        attendances = [a for a in attendances if a.get('player_id') in accessible_player_ids]
    
    # Get player info
    player_ids = [a['player_id'] for a in attendances]
    players = await db.users.find({"id": {"$in": player_ids}}, {"_id": 0, "password": 0}).to_list(500)
    players_map = {p['id']: p for p in players}
    
    # Group by status
    present = []
    absent = []
    pending = []
    
    for att in attendances:
        player = players_map.get(att['player_id'])
        if not player:
            continue
        
        player_data = {
            "id": player['id'],
            "name": player.get('name'),
            "email": player.get('email'),
            "avatar_url": player.get('avatar_url'),
            "jersey_number": player.get('profile', {}).get('sports_info', {}).get('jersey_number'),
            "status": att.get('status', 'pendente'),
            "reason": att.get('reason'),
            "attendance_id": att.get('id')
        }
        
        status = att.get('status', 'pendente')
        if status == 'confirmado':
            present.append(player_data)
        elif status == 'ausente' or status == 'faltou_sem_aviso':
            absent.append(player_data)
        else:
            pending.append(player_data)
    
    # Check if event has passed
    event_date = event.get('date') or event.get('start_date')
    event_passed = False
    if event_date:
        try:
            event_dt = datetime.fromisoformat(event_date.replace('Z', '+00:00'))
            event_passed = datetime.now(timezone.utc) > event_dt
        except:
            pass
    
    return {
        "event_id": event_id,
        "event_title": event.get('title'),
        "event_date": event_date,
        "event_passed": event_passed,
        "convocation": sanitize_convocation_for_user(convocation, event, current_user),
        "convocation_status": convocation.get("status") if convocation else None,
        "convocation_lifecycle_status": convocation.get("status") if convocation else None,
        "convocation_visibility": convocation.get("visibility") if convocation else None,
        "is_private": is_private_convocation_document(convocation),
        "is_private_convocation": is_private_convocation_document(convocation),
        "private_restricted_view": private_restricted_view,
        "present": sorted(present, key=lambda x: x['name'] or ''),
        "absent": sorted(absent, key=lambda x: x['name'] or ''),
        "pending": sorted(pending, key=lambda x: x['name'] or ''),
        "total": len(present) + len(absent) + len(pending),
        "confirmed_count": len(present)
    }

@api_router.put("/events/{event_id}/convocation-status")
async def update_convocation_status(event_id: str, update: ConvocationStatusUpdate, current_user: dict = Depends(get_current_user)):
    """Update a player's convocation status - Coach/Admin only"""
    checker = get_permission_checker(current_user)
    
    event = await db.events.find_one({"id": event_id}, {"_id": 0})
    if not event:
        raise HTTPException(status_code=404, detail="Evento não encontrado")
    
    team_id = event.get('team_id')
    
    # Equipa técnica/admin pode atualizar sempre o estado de convocatória da sua equipa.
    if not checker.is_admin and not (checker.is_staff and team_id and checker.can_access_team(team_id)):
        raise HTTPException(status_code=403, detail="Apenas equipa técnica e administradores podem atualizar o estado")
    
    # Validate status
    valid_statuses = ['confirmado', 'ausente', 'pendente', 'faltou_sem_aviso']
    if update.status not in valid_statuses:
        raise HTTPException(status_code=400, detail=f"Estado inválido. Use: {', '.join(valid_statuses)}")
    
    # Find and update attendance
    attendance = await db.attendance.find_one({
        "event_id": event_id,
        "player_id": update.player_id
    })
    
    if not attendance:
        raise HTTPException(status_code=404, detail="Registo de presença não encontrado")
    
    now_iso = datetime.now(timezone.utc).isoformat()
    await db.attendance.update_one(
        {"id": attendance['id']},
        {"$set": {
            "status": update.status,
            "updated_at": now_iso,
            "updated_by": current_user['id']
        }}
    )

    if attendance.get("convocation_id"):
        await db.convocations.update_one(
            {"id": attendance.get("convocation_id")},
            {"$set": {"updated_at": now_iso}}
        )
    
    return {"message": "Estado atualizado", "status": update.status}

@api_router.post("/events/{event_id}/send-reminder")
async def send_convocation_reminder(event_id: str, current_user: dict = Depends(get_current_user)):
    """Send reminder to all pending players for an event"""
    checker = get_permission_checker(current_user)
    
    event = await db.events.find_one({"id": event_id}, {"_id": 0})
    if not event:
        raise HTTPException(status_code=404, detail="Evento não encontrado")
    
    team_id = event.get('team_id')
    
    # Only coach and admin can send reminders
    if not checker.is_admin and not (checker.is_coach and checker.can_access_team(team_id)):
        raise HTTPException(status_code=403, detail="Apenas treinadores e administradores podem enviar lembretes")
    
    # Get pending attendances
    pending_attendances = await db.attendance.find({
        "event_id": event_id,
        "status": "pendente"
    }, {"_id": 0}).to_list(200)
    
    if not pending_attendances:
        return {"message": "Sem jogadores pendentes", "sent_count": 0}
    
    player_ids = [a['player_id'] for a in pending_attendances]
    players = await db.users.find({"id": {"$in": player_ids}}, {"_id": 0, "id": 1, "email": 1, "name": 1}).to_list(200)
    
    sent_count = 0
    errors = []
    
    # Try to send email reminders
    resend_key = os.environ.get('RESEND_API_KEY')
    
    for player in players:
        try:
            # Create app notification
            notification = {
                "id": str(uuid.uuid4()),
                "user_id": player['id'],
                "type": "convocation_reminder",
                "title": "Lembrete de Convocatória",
                "message": f"Por favor confirma a tua presença no evento: {event.get('title')}",
                "event_id": event_id,
                "is_read": False,
                "created_at": datetime.now(timezone.utc).isoformat()
            }
            await db.notifications.insert_one(notification)
            
            # Try email if available
            if resend_key and player.get('email'):
                try:
                    import resend
                    resend.api_key = resend_key
                    resend.Emails.send({
                        "from": "StickPro <noreply@stickpro.app>",
                        "to": [player['email']],
                        "subject": f"Lembrete: {event.get('title')}",
                        "html": f"""
                        <h2>Lembrete de Convocatória</h2>
                        <p>Olá {player.get('name', 'Atleta')},</p>
                        <p>Por favor confirma a tua presença no evento <strong>{event.get('title')}</strong>.</p>
                        <p>Data: {event.get('date', event.get('start_date', 'N/A'))[:10]}</p>
                        <p>Acede à app para confirmar ou indicar ausência.</p>
                        """
                    })
                except Exception as e:
                    logging.warning(f"Email send error: {e}")
            
            sent_count += 1
        except Exception as e:
            errors.append(f"{player.get('name')}: {str(e)}")
    
    return {
        "message": "Lembretes enviados",
        "sent_count": sent_count,
        "errors": errors
    }

@api_router.post("/events/{event_id}/auto-mark-absent")
async def auto_mark_absent(event_id: str, current_user: dict = Depends(get_current_user)):
    """Auto-mark all pending players as absent for past events"""
    checker = get_permission_checker(current_user)
    
    event = await db.events.find_one({"id": event_id}, {"_id": 0})
    if not event:
        raise HTTPException(status_code=404, detail="Evento não encontrado")
    
    team_id = event.get('team_id')
    
    # Only coach and admin can auto-mark
    if not checker.is_admin and not (checker.is_coach and checker.can_access_team(team_id)):
        raise HTTPException(status_code=403, detail="Apenas treinadores e administradores podem marcar como ausente")
    
    # Check if event has passed
    event_date = event.get('date') or event.get('start_date')
    if event_date:
        try:
            event_dt = datetime.fromisoformat(event_date.replace('Z', '+00:00'))
            if datetime.now(timezone.utc) <= event_dt:
                raise HTTPException(status_code=400, detail="O evento ainda não passou")
        except HTTPException:
            raise
        except:
            pass
    
    # Update all pending to 'faltou_sem_aviso' (absent without notice)
    result = await db.attendance.update_many(
        {"event_id": event_id, "status": "pendente"},
        {"$set": {
            "status": "faltou_sem_aviso",
            "updated_at": datetime.now(timezone.utc).isoformat(),
            "updated_by": current_user['id'],
            "auto_marked": True
        }}
    )
    
    return {
        "message": "Pendentes marcados como ausentes",
        "updated_count": result.modified_count
    }

# ==================== STATISTICS ROUTES ====================

@api_router.get("/teams/{team_id}/stats")
async def get_team_stats(team_id: str, championship_id: Optional[str] = None, current_user: dict = Depends(get_current_user)):
    query = {"team_id": team_id}
    if championship_id:
        query["championship_id"] = championship_id
    
    all_stats = await db.player_match_stats.find(query, {"_id": 0}).to_list(5000)
    
    # Aggregate by player
    player_totals = {}
    for stat in all_stats:
        pid = stat['player_id']
        if pid not in player_totals:
            player_totals[pid] = {
                "player_id": pid, "games_played": 0, "minutes_played": 0, "goals": 0, "assists": 0,
                "penalties_scored": 0, "penalties_missed": 0, "penalties_saved": 0, "penalties_conceded": 0,
                "free_kicks_scored": 0, "free_kicks_missed": 0, "free_kicks_saved": 0, "free_kicks_conceded": 0,
                "saves": 0, "blue_cards": 0, "yellow_cards": 0, "white_cards": 0, "red_cards": 0
            }
        
        pt = player_totals[pid]
        pt['games_played'] += 1
        for key in ['minutes_played', 'goals', 'assists', 'penalties_scored', 'penalties_missed', 
                    'penalties_saved', 'penalties_conceded', 'free_kicks_scored', 'free_kicks_missed',
                    'free_kicks_saved', 'free_kicks_conceded', 'saves', 'blue_cards', 'yellow_cards', 
                    'white_cards', 'red_cards']:
            pt[key] += stat.get(key, 0)
    
    # Enrich with player info
    result = []
    for pid, stats in player_totals.items():
        player = await db.users.find_one({"id": pid}, {"_id": 0, "password": 0})
        if player:
            stats['player'] = player
            result.append(stats)
    
    return result

@api_router.get("/player-stats/{player_id}/consolidated")
async def get_player_consolidated_stats(player_id: str, current_user: dict = Depends(get_current_user)):
    player = await db.users.find_one({"id": player_id}, {"_id": 0, "password": 0})
    if not player:
        raise HTTPException(status_code=404, detail="Jogador não encontrado")
    
    # Get all match stats for this player
    all_stats = await db.player_match_stats.find({"player_id": player_id}, {"_id": 0}).to_list(1000)
    
    # Consolidated totals
    consolidated = {
        "games_played": 0, "minutes_played": 0, "goals": 0, "assists": 0,
        "penalties_scored": 0, "penalties_missed": 0, "penalties_saved": 0, "penalties_conceded": 0,
        "free_kicks_scored": 0, "free_kicks_missed": 0, "free_kicks_saved": 0, "free_kicks_conceded": 0,
        "saves": 0, "blue_cards": 0, "yellow_cards": 0, "white_cards": 0, "red_cards": 0
    }
    
    # Per team stats
    team_stats = {}
    for stat in all_stats:
        tid = stat.get('team_id')
        if tid not in team_stats:
            team_stats[tid] = {k: 0 for k in consolidated.keys()}
            team_stats[tid]['team_id'] = tid
        
        consolidated['games_played'] += 1
        team_stats[tid]['games_played'] += 1
        
        for key in list(consolidated.keys())[1:]:
            consolidated[key] += stat.get(key, 0)
            team_stats[tid][key] += stat.get(key, 0)
    
    # Enrich team stats with team info
    per_team_stats = []
    for tid, ts in team_stats.items():
        team = await db.teams.find_one({"id": tid}, {"_id": 0})
        if team:
            ts['team'] = team
            per_team_stats.append(ts)
    
    # Get teams the player belongs to
    teams = []
    for tid in player.get('team_ids', []):
        team = await db.teams.find_one({"id": tid}, {"_id": 0})
        if team:
            teams.append(team)
    
    return {
        "player": player,
        "consolidated": consolidated,
        "per_team_stats": per_team_stats,
        "teams": teams,
        "teams_count": len(teams)
    }

# ==================== MESSAGE ROUTES ====================

@api_router.post("/messages")
async def send_message(msg_data: MessageCreate, current_user: dict = Depends(get_current_user)):
    message = Message(
        team_id=msg_data.team_id,
        sender_id=current_user['id'],
        sender_name=current_user['name'],
        content=msg_data.content,
        recipient_ids=msg_data.recipient_ids,
        attachment_name=msg_data.attachment_name
    )
    
    # Handle attachment (store as base64 in DB for simplicity - in production use object storage)
    if msg_data.attachment_data:
        message.attachment_url = f"data:{msg_data.attachment_name};base64,{msg_data.attachment_data}"
    
    msg_dict = message.model_dump()
    msg_dict['created_at'] = msg_dict['created_at'].isoformat()
    
    await db.messages.insert_one(msg_dict)
    # Remove MongoDB _id before returning
    msg_dict.pop('_id', None)
    
    # Send email notifications
    if msg_data.recipient_ids:
        recipients = await db.users.find({"id": {"$in": msg_data.recipient_ids}}, {"_id": 0, "email": 1, "name": 1}).to_list(100)
    else:
        # Send to all team members
        team = await db.teams.find_one({"id": msg_data.team_id}, {"_id": 0})
        if team:
            all_ids = team.get('coach_ids', []) + team.get('delegate_ids', []) + team.get('player_ids', [])
            recipients = await db.users.find({"id": {"$in": all_ids}}, {"_id": 0, "email": 1, "name": 1}).to_list(100)
        else:
            recipients = []
    
    for recipient in recipients:
        if recipient['email'] != current_user['email']:
            await send_email_notification(
                recipient['email'],
                f"Nova mensagem de {current_user['name']}",
                f"<p>{msg_data.content}</p>"
            )
    
    return msg_dict

@api_router.get("/messages/{team_id}")
async def get_messages(team_id: str, limit: int = 50, current_user: dict = Depends(get_current_user)):
    messages = await db.messages.find(
        {"team_id": team_id},
        {"_id": 0}
    ).sort("created_at", -1).limit(limit).to_list(limit)
    
    return list(reversed(messages))

@api_router.get("/teams/{team_id}/members-for-message")
async def get_members_for_message(team_id: str, current_user: dict = Depends(get_current_user)):
    team = await db.teams.find_one({"id": team_id}, {"_id": 0})
    if not team:
        raise HTTPException(status_code=404, detail="Equipa não encontrada")
    
    all_ids = team.get('coach_ids', []) + team.get('delegate_ids', []) + team.get('player_ids', [])
    members = await db.users.find({"id": {"$in": all_ids}}, {"_id": 0, "password": 0, "id": 1, "name": 1, "email": 1, "role": 1}).to_list(100)
    
    return members

# ==================== DASHBOARD ROUTE ====================

@api_router.get("/dashboard")
async def get_dashboard(
    profile_type: Optional[str] = None,
    profile_user_id: Optional[str] = None,
    profile_role: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """
    Get dashboard data filtered by user role:
    - Admin/Gestor Desportivo: ALL club events
    - Coach/Delegate/Player: ONLY events from their teams
    - Parent/Guardian: Events of their linked children
    """
        
    effective_user = current_user

    if profile_type == "associated" and profile_user_id:
        allowed_accounts = current_user.get("associated_accounts", [])
    
        if profile_user_id not in allowed_accounts:
            raise HTTPException(status_code=403, detail="Perfil associado não autorizado")
    
        associated_user = await db.users.find_one(
            {"id": profile_user_id},
            {"_id": 0, "hashed_password": 0, "password": 0}
        )
    
        if not associated_user:
            raise HTTPException(status_code=404, detail="Perfil associado não encontrado")
    
        effective_user = {
            **associated_user,
            "role": "responsavel",
            "linked_player_id": associated_user["id"],
            "linked_player_ids": [associated_user["id"]],
            "team_ids": associated_user.get("team_ids", [])
        }
    
    elif profile_type == "self" and profile_role:
        effective_user = {
            **current_user,
            "role": profile_role
        }
        
    user_role = effective_user.get('role')
    user_teams = effective_user.get('team_ids', [])
    linked_player_ids = effective_user.get('linked_player_ids', [])
    linked_player_id = effective_user.get('linked_player_id')
    
    # Build event query based on role
    now = datetime.now(timezone.utc).isoformat()
    upcoming_query = {"start_time": {"$gte": now}}
    
    if is_admin_role(user_role):
        # Admin/Gestor Desportivo: see ALL club events (no team filter)
        pass
    elif user_role == 'responsavel':
        # Parent/Guardian: see events of their children
        # Get team_ids from linked players
        child_team_ids = set()
        all_linked = linked_player_ids if linked_player_ids else ([linked_player_id] if linked_player_id else [])
        
        for player_id in all_linked:
            player = await db.users.find_one({"id": player_id}, {"_id": 0, "team_ids": 1})
            if player and player.get('team_ids'):
                child_team_ids.update(player['team_ids'])
        
        if child_team_ids:
            upcoming_query["team_id"] = {"$in": list(child_team_ids)}
        else:
            # No linked children or children have no teams - return empty
            upcoming_query["team_id"] = {"$in": []}
    else:
        # Coach/Delegate/Player: see ONLY events from their teams
        if user_teams:
            upcoming_query["team_id"] = {"$in": user_teams}
        else:
            # User has no teams - return empty
            upcoming_query["team_id"] = {"$in": []}
    
    upcoming_events = await db.events.find(upcoming_query, {"_id": 0}).sort("start_time", 1).limit(5).to_list(5)

    upcoming_events = await attach_teams_to_events(upcoming_events)
    upcoming_events = await enrich_events_with_convocation_context(upcoming_events, effective_user)

    # Pending convocations - single source of truth shared with /convocations/my.
    pending_convocations = await build_my_convocations_for_user(
        effective_user,
        only_pending=True,
        include_past=False,
        limit=10,
    )

    # Teams count based on role
    if is_admin_role(user_role):
        teams_count = await db.teams.count_documents({})
    elif user_role == 'responsavel':
        # Count teams of linked children
        child_team_ids = set()
        all_linked = linked_player_ids if linked_player_ids else ([linked_player_id] if linked_player_id else [])
        if all_linked:
            linked_players = await db.users.find(
                {"id": {"$in": all_linked}},
                {"_id": 0, "team_ids": 1}
            ).to_list(len(all_linked))
            for player in linked_players:
                if player and player.get('team_ids'):
                    child_team_ids.update(player['team_ids'])
        teams_count = len(child_team_ids)
    else:
        teams_count = len(user_teams)
    
    # Recent messages - filter by accessible teams
    recent_messages = []
    if is_admin_role(user_role):
        # Admin sees all messages
        recent_messages = await db.messages.find({}, {"_id": 0}).sort("created_at", -1).limit(5).to_list(5)
    elif user_role == 'responsavel':
        # Parent sees messages from children's teams
        child_team_ids = set()
        all_linked = linked_player_ids if linked_player_ids else ([linked_player_id] if linked_player_id else [])
        if all_linked:
            linked_players = await db.users.find(
                {"id": {"$in": all_linked}},
                {"_id": 0, "team_ids": 1}
            ).to_list(len(all_linked))
            for player in linked_players:
                if player and player.get('team_ids'):
                    child_team_ids.update(player['team_ids'])
        if child_team_ids:
            recent_messages = await db.messages.find({"team_id": {"$in": list(child_team_ids)}}, {"_id": 0}).sort("created_at", -1).limit(5).to_list(5)
    elif user_teams:
        recent_messages = await db.messages.find({"team_id": {"$in": user_teams}}, {"_id": 0}).sort("created_at", -1).limit(5).to_list(5)
    
    return {
        "upcoming_events": upcoming_events,
        "pending_convocations": pending_convocations,
        "teams_count": teams_count,
        "recent_messages": recent_messages
    }

# ==================== ROOT ROUTE ====================

@api_router.get("/")
async def root():
    return {"message": "Roller Hockey Hub API", "version": "2.0.0"}

# Include router
# ==================== FILE UPLOAD ====================

@api_router.post("/upload/image")
async def upload_image(file: UploadFile = File(...), current_user: dict = Depends(get_current_user)):
    """Upload an image file and return the URL"""
    # Validate file type
    allowed_types = ["image/jpeg", "image/png", "image/gif", "image/webp"]
    if file.content_type not in allowed_types:
        raise HTTPException(status_code=400, detail="Tipo de ficheiro não permitido. Use JPEG, PNG, GIF ou WebP.")
    
    # Validate file size (max 5MB)
    content = await file.read()
    if len(content) > 5 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="Ficheiro muito grande. Máximo 5MB.")
    
    # Generate unique filename
    ext = file.filename.split('.')[-1] if '.' in file.filename else 'jpg'
    filename = f"{uuid.uuid4()}.{ext}"
    filepath = UPLOADS_DIR / filename
    
    # Save file
    with open(filepath, "wb") as f:
        f.write(content)
    
    # Return URL with /api prefix so it's accessible via proxy
    return {"url": f"/api/uploads/{filename}", "filename": filename}

@api_router.delete("/upload/{filename}")
async def delete_image(filename: str, current_user: dict = Depends(get_current_user)):
    """Delete an uploaded image"""
    filepath = UPLOADS_DIR / filename
    if filepath.exists():
        filepath.unlink()
        return {"message": "Ficheiro eliminado"}
    raise HTTPException(status_code=404, detail="Ficheiro não encontrado")

# =====================
# Library Endpoints
# =====================

@api_router.get("/library")
async def get_library_items(
    category: Optional[str] = None,
    item_type: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Get all library items with optional filters"""
    query = {}
    if category:
        query["category"] = category
    if item_type:
        query["item_type"] = item_type
    
    items = await db.library_items.find(query, {"_id": 0}).sort("created_at", -1).to_list(500)
    return items

@api_router.get("/library/categories")
async def get_library_categories(current_user: dict = Depends(get_current_user)):
    """Get all unique categories"""
    categories = await db.library_items.distinct("category")
    return [c for c in categories if c]

@api_router.post("/library")
async def create_library_item(item: LibraryItemCreate, current_user: dict = Depends(get_current_user)):
    """Create a new library item"""
    checker = get_permission_checker(current_user)
    
    if not checker.can_manage_team:
        raise HTTPException(status_code=403, detail="Sem permissão para criar recursos")
    
    # Generate thumbnail for videos
    thumbnail_url = None
    if item.item_type == "video":
        # Extract YouTube/Vimeo thumbnail
        if "youtube.com" in item.url or "youtu.be" in item.url:
            video_id = None
            if "youtu.be" in item.url:
                video_id = item.url.split("/")[-1].split("?")[0]
            elif "v=" in item.url:
                video_id = item.url.split("v=")[1].split("&")[0]
            if video_id:
                thumbnail_url = f"https://img.youtube.com/vi/{video_id}/hqdefault.jpg"
        elif "vimeo.com" in item.url:
            video_id = item.url.split("/")[-1]
            thumbnail_url = f"https://vumbnail.com/{video_id}.jpg"
    
    library_item = LibraryItem(
        title=item.title,
        description=item.description,
        item_type=item.item_type,
        url=item.url,
        category=item.category,
        tags=item.tags,
        thumbnail_url=thumbnail_url,
        created_by=current_user['id']
    )
    
    await db.library_items.insert_one(library_item.model_dump())
    return {**library_item.model_dump(), "_id": None}

@api_router.put("/library/{item_id}")
async def update_library_item(item_id: str, item: LibraryItemCreate, current_user: dict = Depends(get_current_user)):
    """Update a library item"""
    checker = get_permission_checker(current_user)
    
    if not checker.can_manage_team:
        raise HTTPException(status_code=403, detail="Sem permissão para editar recursos")
    
    existing = await db.library_items.find_one({"id": item_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Item não encontrado")
    
    update_data = item.model_dump()
    await db.library_items.update_one({"id": item_id}, {"$set": update_data})
    
    updated = await db.library_items.find_one({"id": item_id}, {"_id": 0})
    return updated

@api_router.delete("/library/{item_id}")
async def delete_library_item(item_id: str, current_user: dict = Depends(get_current_user)):
    """Delete a library item"""
    checker = get_permission_checker(current_user)
    
    if not checker.can_manage_team:
        raise HTTPException(status_code=403, detail="Sem permissão para eliminar recursos")
    
    result = await db.library_items.delete_one({"id": item_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Item não encontrado")
    
    return {"message": "Item eliminado com sucesso"}

# =====================
# AI Assistant Endpoints
# =====================
@api_router.post("/ai/chat")
async def ai_chat(request: AIChatRequest, current_user: dict = Depends(get_current_user)):
    """Chat with AI assistant about roller hockey and app help"""
    from openai import OpenAI

    api_key = os.environ.get("OPENAI_API_KEY")
    if not api_key:
        raise HTTPException(status_code=500, detail="OPENAI_API_KEY não configurada")

    client = OpenAI(api_key=api_key)

    session_id = request.session_id or f"user_{current_user['id']}_{datetime.now().strftime('%Y%m%d')}"

    language_names = {
        "pt": "Português",
        "en": "English",
        "es": "Español",
        "fr": "Français",
        "it": "Italiano",
    }

    selected_language = request.language if request.language in language_names else "pt"
    selected_language_name = language_names[selected_language]

    history = await db.ai_chat_history.find(
        {"session_id": session_id}
    ).sort("timestamp", 1).to_list(50)

    system_message = f"""Tu és o Assistente StickPro, um especialista em hóquei em patins e na aplicação StickPro.

IDIOMA DE RESPOSTA:
- Responde sempre em {selected_language_name}.
- O idioma foi definido pela aplicação através do campo language="{selected_language}".
- Nunca mudes de idioma, exceto se o utilizador pedir explicitamente.

SOBRE A APP STICKPRO:
- Gestão de equipas de hóquei em patins
- Calendário de eventos: treinos, jogos, torneios, reuniões e eventos do clube
- Convocatórias e presenças
- Estatísticas de jogadores: golos, assistências, cartões e desempenho
- Campeonatos 5x5 e 3x3
- Importação de fichas de jogo da APL
- Gestão de membros, atletas, treinadores, encarregados de educação e dirigentes
- Gestão de clubes, equipas e épocas desportivas
- Biblioteca de documentos
- Apoio à organização operacional do clube

SOBRE HÓQUEI EM PATINS:
- É um desporto coletivo jogado com patins de quatro rodas, stick e bola
- Cada equipa joga normalmente com 5 jogadores em pista: 4 jogadores de campo e 1 guarda-redes
- Nos escalões seniores, a duração habitual é de 2 partes de 25 minutos
- Existem penáltis, livres diretos, faltas de equipa e cartões
- O cartão azul implica exclusão temporária e pode originar livre direto
- O cartão vermelho implica expulsão definitiva
- As regras podem variar conforme escalão, competição e regulamento aplicável

ESPECIALIDADES:
- Explicar como usar a aplicação StickPro
- Ajudar treinadores, dirigentes e gestores desportivos
- Apoiar tarefas de gestão de equipas
- Explicar conceitos de hóquei em patins
- Apoiar interpretação geral de regulamentos FPP e World Skate
- Sugerir boas práticas de gestão de clubes

LIMITAÇÕES:
- Ainda não tens acesso direto aos dados reais do clube, pagamentos, presenças ou estatísticas internas.
- Ainda não tens acesso documental completo aos regulamentos oficiais carregados em base de conhecimento.
- Quando não tiveres certeza, deves dizer que não tens certeza.
- Quando a pergunta depender de regulamentos oficiais, recomenda confirmação no regulamento aplicável.

ESTILO:
- Responde de forma clara, útil e prática.
- Usa linguagem simples.
- Dá passos concretos quando o utilizador perguntar como fazer algo.
- Evita respostas demasiado longas, salvo se o utilizador pedir detalhe."""

    try:
        messages_for_openai = [
            {"role": "system", "content": system_message}
        ]

        for msg in history[-10:]:
            role = msg.get("role")
            content = msg.get("content")

            if role in ["user", "assistant"] and content:
                messages_for_openai.append({
                    "role": role,
                    "content": content
                })

        messages_for_openai.append({
            "role": "user",
            "content": request.message
        })

        completion = client.chat.completions.create(
            model="gpt-4o-mini",
            messages=messages_for_openai,
            temperature=0.4,
            max_tokens=700,
        )

        response = completion.choices[0].message.content

        await db.ai_chat_history.insert_one({
            "session_id": session_id,
            "user_id": current_user["id"],
            "role": "user",
            "content": request.message,
            "language": selected_language,
            "timestamp": datetime.now(timezone.utc)
        })

        await db.ai_chat_history.insert_one({
            "session_id": session_id,
            "user_id": current_user["id"],
            "role": "assistant",
            "content": response,
            "language": selected_language,
            "timestamp": datetime.now(timezone.utc)
        })

        return {
            "response": response,
            "session_id": session_id,
            "language": selected_language
        }

    except Exception as e:
        logging.error(f"AI Chat error: {e}")
        raise HTTPException(status_code=500, detail=f"Erro no assistente: {str(e)}")

# =====================
# Push Notifications Endpoints
# =====================

@api_router.get("/notifications/vapid-public-key")
async def get_vapid_public_key():
    """Get VAPID public key for push subscription"""
    public_key = os.environ.get('VAPID_PUBLIC_KEY')
    if not public_key:
        raise HTTPException(status_code=500, detail="Push notifications não configuradas")
    return {"publicKey": public_key}

@api_router.post("/notifications/subscribe")
async def subscribe_to_notifications(subscription: dict, current_user: dict = Depends(get_current_user)):
    """Subscribe user to push notifications"""
    # Store subscription in database
    subscription_data = {
        "user_id": current_user['id'],
        "endpoint": subscription.get('endpoint'),
        "keys": subscription.get('keys'),
        "created_at": datetime.now(timezone.utc)
    }
    
    # Update or insert subscription
    await db.push_subscriptions.update_one(
        {"user_id": current_user['id'], "endpoint": subscription.get('endpoint')},
        {"$set": subscription_data},
        upsert=True
    )
    
    return {"message": "Subscribed to notifications"}

@api_router.delete("/notifications/unsubscribe")
async def unsubscribe_from_notifications(current_user: dict = Depends(get_current_user)):
    """Unsubscribe user from push notifications"""
    await db.push_subscriptions.delete_many({"user_id": current_user['id']})
    return {"message": "Unsubscribed from notifications"}

@api_router.post("/notifications/send")
async def send_notification(notification_data: dict, current_user: dict = Depends(get_current_user)):
    """Send push notification - Admin/Coach only"""
    checker = get_permission_checker(current_user)
    
    if not checker.can_create_convocations:
        raise HTTPException(status_code=403, detail="Sem permissão para enviar notificações")
    
    from pywebpush import webpush, WebPushException
    import json
    
    vapid_private_key = os.environ.get('VAPID_PRIVATE_KEY')
    vapid_claims_email = os.environ.get('VAPID_CLAIMS_EMAIL', 'noreply@stickpro.com')
    
    if not vapid_private_key:
        raise HTTPException(status_code=500, detail="Push notifications não configuradas")
    
    # Get target user subscriptions
    user_ids = notification_data.get('user_ids', [])
    team_id = notification_data.get('team_id')
    
    query = {}
    if user_ids:
        query["user_id"] = {"$in": user_ids}
    elif team_id:
        # Get all team members
        members = await db.team_members.find({"team_id": team_id}, {"user_id": 1}).to_list(1000)
        member_ids = [m['user_id'] for m in members]
        query["user_id"] = {"$in": member_ids}
    else:
        # Send to all users
        pass
    
    subscriptions = await db.push_subscriptions.find(query, {"_id": 0}).to_list(1000)
    
    payload = json.dumps({
        "title": notification_data.get('title', 'Stick Pro'),
        "body": notification_data.get('body', 'Nova notificação'),
        "url": notification_data.get('url', '/'),
        "icon": "/icons/icon-192x192.png"
    })
    
    success_count = 0
    failed_count = 0
    
    for sub in subscriptions:
        try:
            webpush(
                subscription_info={
                    "endpoint": sub['endpoint'],
                    "keys": sub['keys']
                },
                data=payload,
                vapid_private_key=vapid_private_key,
                vapid_claims={"sub": f"mailto:{vapid_claims_email}"}
            )
            success_count += 1
        except WebPushException as e:
            logging.error(f"Push failed: {e}")
            # Remove invalid subscriptions
            if e.response and e.response.status_code in [404, 410]:
                await db.push_subscriptions.delete_one({"endpoint": sub['endpoint']})
            failed_count += 1
        except Exception as e:
            logging.error(f"Push error: {e}")
            failed_count += 1
    
    return {
        "message": "Notificações enviadas",
        "success": success_count,
        "failed": failed_count
    }


# ==================== EVALUATION ROUTES — Sprint 4.2.1 ====================

def calculate_evaluation_overall_score(scores: List[dict], criteria_map: Dict[str, dict]) -> Optional[float]:
    """Weighted average for evaluation scores. Returns None when no valid scores exist."""
    weighted_total = 0.0
    weight_total = 0.0

    for score_item in scores or []:
        criterion_id = score_item.get("criterion_id")
        raw_score = score_item.get("score")

        if criterion_id is None or raw_score is None:
            continue

        criterion = criteria_map.get(criterion_id, {})
        weight = float(criterion.get("weight", 1.0) or 1.0)

        try:
            score_value = float(raw_score)
        except (TypeError, ValueError):
            continue

        weighted_total += score_value * weight
        weight_total += weight

    if weight_total <= 0:
        return None

    return round(weighted_total / weight_total, 2)


async def get_evaluation_player_or_404(player_id: str) -> dict:
    player = await db.users.find_one(
        {"id": player_id},
        {"_id": 0, "hashed_password": 0, "password": 0}
    )

    if not player:
        raise HTTPException(status_code=404, detail="Atleta não encontrado")

    if player.get("role") not in ["jogador", "atleta", "player"]:
        raise HTTPException(status_code=400, detail="A avaliação só pode ser associada a atletas")

    return player


def is_evaluation_shared_with_development_circle(evaluation: Optional[dict]) -> bool:
    if not evaluation:
        return False

    return bool(
        evaluation.get("share_with_player") or
        evaluation.get("share_with_guardian") or
        evaluation.get("visibility") in ["player", "guardian", "all"]
    )


def can_view_player_evaluations(
    current_user: dict,
    player_id: str,
    team_id: Optional[str] = None,
    evaluation: Optional[dict] = None
) -> bool:
    checker = get_permission_checker(current_user)

    if checker.is_admin:
        return True

    if team_id and checker.is_staff and checker.can_access_team(team_id):
        return True

    shared = is_evaluation_shared_with_development_circle(evaluation)

    if current_user.get("id") == player_id:
        return shared

    linked_player_ids = current_user.get("linked_player_ids", []) or []
    linked_player_id = current_user.get("linked_player_id")

    if linked_player_id and linked_player_id not in linked_player_ids:
        linked_player_ids.append(linked_player_id)

    if player_id in linked_player_ids:
        return shared

    return False


def is_development_circle_viewer(current_user: dict, player_id: str) -> bool:
    if current_user.get("id") == player_id:
        return True

    linked_player_ids = current_user.get("linked_player_ids", []) or []
    linked_player_id = current_user.get("linked_player_id")

    if linked_player_id and linked_player_id not in linked_player_ids:
        linked_player_ids.append(linked_player_id)

    return player_id in linked_player_ids


def build_public_player_evaluation_view(evaluation: dict) -> dict:
    return {
        "id": evaluation.get("id"),
        "player_id": evaluation.get("player_id"),
        "team_id": evaluation.get("team_id"),
        "event_id": evaluation.get("event_id"),
        "period_label": evaluation.get("period_label"),
        "created_at": evaluation.get("created_at"),
        "updated_at": evaluation.get("updated_at"),
        "shared": True,
        "view_type": "development",
        "public_summary": evaluation.get("public_summary"),
        "strengths": evaluation.get("strengths") or [],
        "improvement_goals": evaluation.get("improvement_goals") or [],
        "motivational_message": evaluation.get("motivational_message"),
    }


@api_router.post("/evaluations/criteria")
async def create_evaluation_criterion(
    criterion_data: EvaluationCriterionCreate,
    current_user: dict = Depends(get_current_user)
):
    checker = get_permission_checker(current_user)

    if not checker.is_staff and not checker.is_admin:
        raise HTTPException(status_code=403, detail="Sem permissão para criar critérios de avaliação")

    if criterion_data.team_id and not checker.is_admin and not checker.can_access_team(criterion_data.team_id):
        raise HTTPException(status_code=403, detail="Sem acesso a esta equipa")

    if criterion_data.scale_min >= criterion_data.scale_max:
        raise HTTPException(status_code=400, detail="A escala mínima deve ser inferior à escala máxima")

    criterion = EvaluationCriterion(
        **criterion_data.model_dump(),
        club_id=current_user.get("club_id"),
        created_by=current_user["id"]
    )

    criterion_dict = criterion.model_dump()
    criterion_dict["created_at"] = criterion_dict["created_at"].isoformat()
    criterion_dict["updated_at"] = criterion_dict["updated_at"].isoformat()

    await db.evaluation_criteria.insert_one(dict(criterion_dict))

    criterion_dict.pop("_id", None)
    return criterion_dict


@api_router.get("/evaluations/criteria")
async def get_evaluation_criteria(
    team_id: Optional[str] = None,
    include_inactive: bool = False,
    current_user: dict = Depends(get_current_user)
):
    checker = get_permission_checker(current_user)

    query: Dict[str, Any] = {}

    if not include_inactive:
        query["is_active"] = {"$ne": False}

    if team_id:
        if not checker.is_admin and not checker.can_access_team(team_id):
            raise HTTPException(status_code=403, detail="Sem acesso a esta equipa")
        query["$or"] = [{"team_id": team_id}, {"team_id": None}, {"team_id": {"$exists": False}}]
    elif not checker.is_admin:
        accessible_team_ids = list(checker.team_ids)
        query["$or"] = [
            {"team_id": {"$in": accessible_team_ids}},
            {"team_id": None},
            {"team_id": {"$exists": False}},
        ]

    criteria = await db.evaluation_criteria.find(query, {"_id": 0}).sort("category", 1).sort("name", 1).to_list(500)
    return criteria


@api_router.put("/evaluations/criteria/{criterion_id}")
async def update_evaluation_criterion(
    criterion_id: str,
    updates: EvaluationCriterionUpdate,
    current_user: dict = Depends(get_current_user)
):
    checker = get_permission_checker(current_user)

    criterion = await db.evaluation_criteria.find_one({"id": criterion_id}, {"_id": 0})
    if not criterion:
        raise HTTPException(status_code=404, detail="Critério não encontrado")

    if not checker.is_admin:
        if criterion.get("created_by") != current_user.get("id") and not checker.can_access_team(criterion.get("team_id")):
            raise HTTPException(status_code=403, detail="Sem permissão para editar este critério")

    update_data = updates.model_dump(exclude_unset=True)

    if "scale_min" in update_data or "scale_max" in update_data:
        next_min = update_data.get("scale_min", criterion.get("scale_min", 1))
        next_max = update_data.get("scale_max", criterion.get("scale_max", 5))
        if next_min >= next_max:
            raise HTTPException(status_code=400, detail="A escala mínima deve ser inferior à escala máxima")

    if update_data.get("team_id") and not checker.is_admin and not checker.can_access_team(update_data["team_id"]):
        raise HTTPException(status_code=403, detail="Sem acesso a esta equipa")

    if update_data:
        update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
        await db.evaluation_criteria.update_one(
            {"id": criterion_id},
            {"$set": update_data}
        )

    updated = await db.evaluation_criteria.find_one({"id": criterion_id}, {"_id": 0})
    return updated


@api_router.delete("/evaluations/criteria/{criterion_id}")
async def archive_evaluation_criterion(
    criterion_id: str,
    current_user: dict = Depends(get_current_user)
):
    checker = get_permission_checker(current_user)

    criterion = await db.evaluation_criteria.find_one({"id": criterion_id}, {"_id": 0})
    if not criterion:
        raise HTTPException(status_code=404, detail="Critério não encontrado")

    if not checker.is_admin:
        if criterion.get("created_by") != current_user.get("id") and not checker.can_access_team(criterion.get("team_id")):
            raise HTTPException(status_code=403, detail="Sem permissão para arquivar este critério")

    await db.evaluation_criteria.update_one(
        {"id": criterion_id},
        {
            "$set": {
                "is_active": False,
                "updated_at": datetime.now(timezone.utc).isoformat()
            }
        }
    )

    return {"message": "Critério arquivado"}


@api_router.post("/evaluations")
async def create_player_evaluation(
    evaluation_data: PlayerEvaluationCreate,
    current_user: dict = Depends(get_current_user)
):
    checker = get_permission_checker(current_user)

    if not checker.is_staff and not checker.is_admin:
        raise HTTPException(status_code=403, detail="Sem permissão para avaliar atletas")

    if not checker.is_admin and not checker.can_access_team(evaluation_data.team_id):
        raise HTTPException(status_code=403, detail="Sem acesso a esta equipa")

    player = await get_evaluation_player_or_404(evaluation_data.player_id)

    if evaluation_data.team_id not in (player.get("team_ids") or []):
        raise HTTPException(status_code=400, detail="O atleta não pertence a esta equipa")

    criterion_ids = [item.criterion_id for item in evaluation_data.scores]
    criteria = []
    if criterion_ids:
        criteria = await db.evaluation_criteria.find(
            {"id": {"$in": criterion_ids}},
            {"_id": 0}
        ).to_list(500)

    criteria_map = {criterion["id"]: criterion for criterion in criteria}

    for item in evaluation_data.scores:
        criterion = criteria_map.get(item.criterion_id)
        if not criterion:
            raise HTTPException(status_code=400, detail=f"Critério inválido: {item.criterion_id}")

        scale_min = criterion.get("scale_min", 1)
        scale_max = criterion.get("scale_max", 5)

        if item.score < scale_min or item.score > scale_max:
            raise HTTPException(
                status_code=400,
                detail=f"Pontuação fora da escala para {criterion.get('name', item.criterion_id)}"
            )

    score_dicts = [item.model_dump() for item in evaluation_data.scores]
    overall_score = calculate_evaluation_overall_score(score_dicts, criteria_map)

    share_with_development_circle = bool(
        evaluation_data.share_with_player or
        evaluation_data.share_with_guardian or
        evaluation_data.visibility in ["player", "guardian", "all"]
    )    
    
    evaluation = PlayerEvaluation(
        player_id=evaluation_data.player_id,
        team_id=evaluation_data.team_id,
        event_id=evaluation_data.event_id,
        period_label=evaluation_data.period_label,
        visibility=evaluation_data.visibility,
        scores=score_dicts,
        general_comment=evaluation_data.general_comment,
        share_with_player=share_with_development_circle,
        share_with_guardian=share_with_development_circle,
        public_summary=evaluation_data.public_summary,
        strengths=evaluation_data.strengths,
        improvement_goals=evaluation_data.improvement_goals,
        motivational_message=evaluation_data.motivational_message,
        overall_score=overall_score,
        created_by=current_user["id"]
    )

    evaluation_dict = evaluation.model_dump()
    evaluation_dict["created_at"] = evaluation_dict["created_at"].isoformat()
    evaluation_dict["updated_at"] = evaluation_dict["updated_at"].isoformat()

    await db.player_evaluations.insert_one(dict(evaluation_dict))

    evaluation_dict.pop("_id", None)
    return evaluation_dict


@api_router.get("/evaluations/player/{player_id}")
async def get_player_evaluations(
    player_id: str,
    team_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    await get_evaluation_player_or_404(player_id)

    query: Dict[str, Any] = {"player_id": player_id}
    if team_id:
        query["team_id"] = team_id

    evaluations = await db.player_evaluations.find(query, {"_id": 0}).sort("created_at", -1).to_list(500)

    visible = []
    for evaluation in evaluations:
        if can_view_player_evaluations(
            current_user,
            player_id,
            evaluation.get("team_id"),
            evaluation
        ):
            if is_development_circle_viewer(current_user, player_id):
                visible.append(build_public_player_evaluation_view(evaluation))
            else:
                visible.append(evaluation)

    return visible

@api_router.put("/evaluations/{evaluation_id}")
async def update_player_evaluation(
    evaluation_id: str,
    updates: PlayerEvaluationUpdate,
    current_user: dict = Depends(get_current_user)
):
    checker = get_permission_checker(current_user)

    evaluation = await db.player_evaluations.find_one({"id": evaluation_id}, {"_id": 0})
    if not evaluation:
        raise HTTPException(status_code=404, detail="Avaliação não encontrada")

    if not checker.is_admin:
        if evaluation.get("created_by") != current_user.get("id") and not checker.can_access_team(evaluation.get("team_id")):
            raise HTTPException(status_code=403, detail="Sem permissão para editar esta avaliação")

    update_data = updates.model_dump(exclude_unset=True)

    if "share_with_player" in update_data or "share_with_guardian" in update_data:
        share_with_development_circle = bool(
            update_data.get("share_with_player") or
            update_data.get("share_with_guardian")
        )
        update_data["share_with_player"] = share_with_development_circle
        update_data["share_with_guardian"] = share_with_development_circle
    
    if update_data.get("visibility") in ["player", "guardian", "all"]:
        update_data["share_with_player"] = True
        update_data["share_with_guardian"] = True
    
    if "scores" in update_data and update_data["scores"] is not None:
        score_dicts = [
            item.model_dump() if hasattr(item, "model_dump") else item
            for item in update_data["scores"]
        ]

        criterion_ids = [item.get("criterion_id") for item in score_dicts]
        criteria = await db.evaluation_criteria.find(
            {"id": {"$in": criterion_ids}},
            {"_id": 0}
        ).to_list(500)
        criteria_map = {criterion["id"]: criterion for criterion in criteria}

        for item in score_dicts:
            criterion = criteria_map.get(item.get("criterion_id"))
            if not criterion:
                raise HTTPException(status_code=400, detail=f"Critério inválido: {item.get('criterion_id')}")

            if item.get("score") < criterion.get("scale_min", 1) or item.get("score") > criterion.get("scale_max", 5):
                raise HTTPException(status_code=400, detail="Pontuação fora da escala")

        update_data["scores"] = score_dicts
        update_data["overall_score"] = calculate_evaluation_overall_score(score_dicts, criteria_map)

    if update_data:
        update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
        await db.player_evaluations.update_one(
            {"id": evaluation_id},
            {"$set": update_data}
        )

    updated = await db.player_evaluations.find_one({"id": evaluation_id}, {"_id": 0})
    return updated


@api_router.delete("/evaluations/{evaluation_id}")
async def delete_player_evaluation(
    evaluation_id: str,
    current_user: dict = Depends(get_current_user)
):
    checker = get_permission_checker(current_user)

    evaluation = await db.player_evaluations.find_one({"id": evaluation_id}, {"_id": 0})
    if not evaluation:
        raise HTTPException(status_code=404, detail="Avaliação não encontrada")

    if not checker.is_admin and evaluation.get("created_by") != current_user.get("id"):
        raise HTTPException(status_code=403, detail="Sem permissão para eliminar esta avaliação")

    await db.player_evaluations.delete_one({"id": evaluation_id})

    return {"message": "Avaliação eliminada"}




# ==================== EVALUATION PLAN ROUTES — Sprint 4.2.3.1 ====================

async def validate_evaluation_plan_access(current_user: dict, team_id: Optional[str] = None, *, write: bool = False):
    checker = get_permission_checker(current_user)

    if write and not checker.is_staff and not checker.is_admin:
        raise HTTPException(status_code=403, detail="Sem permissão para gerir planos de avaliação")

    if team_id and not checker.is_admin and not checker.can_access_team(team_id):
        raise HTTPException(status_code=403, detail="Sem acesso a esta equipa")

    return checker


async def validate_evaluation_plan_criteria(criteria: List[EvaluationPlanCriterion], team_id: Optional[str], checker) -> List[dict]:
    if not criteria:
        raise HTTPException(status_code=400, detail="O plano deve incluir pelo menos um critério")

    criterion_ids = [item.criterion_id for item in criteria]
    if len(criterion_ids) != len(set(criterion_ids)):
        raise HTTPException(status_code=400, detail="O plano contém critérios repetidos")

    criteria_docs = await db.evaluation_criteria.find(
        {"id": {"$in": criterion_ids}, "is_active": {"$ne": False}},
        {"_id": 0}
    ).to_list(500)

    found_ids = {criterion["id"] for criterion in criteria_docs}
    missing_ids = [criterion_id for criterion_id in criterion_ids if criterion_id not in found_ids]
    if missing_ids:
        raise HTTPException(status_code=400, detail=f"Critérios inválidos ou arquivados: {', '.join(missing_ids)}")

    for criterion in criteria_docs:
        criterion_team_id = criterion.get("team_id")
        if criterion_team_id and team_id and criterion_team_id != team_id:
            raise HTTPException(status_code=400, detail=f"O critério '{criterion.get('name')}' pertence a outra equipa")
        if criterion_team_id and not team_id:
            raise HTTPException(status_code=400, detail=f"O critério '{criterion.get('name')}' é específico de uma equipa")
        if criterion_team_id and not checker.is_admin and not checker.can_access_team(criterion_team_id):
            raise HTTPException(status_code=403, detail=f"Sem acesso ao critério '{criterion.get('name')}'")

    criteria_dicts = []
    for index, item in enumerate(criteria):
        item_dict = item.model_dump()
        item_dict["order"] = item_dict.get("order", index)
        item_dict["weight"] = float(item_dict.get("weight", 1.0) or 1.0)
        item_dict["required"] = bool(item_dict.get("required", True))
        criteria_dicts.append(item_dict)

    return criteria_dicts


async def enrich_evaluation_plan(plan: dict) -> dict:
    if not plan:
        return plan

    criterion_ids = [item.get("criterion_id") for item in plan.get("criteria", []) if item.get("criterion_id")]
    criteria_docs = []
    if criterion_ids:
        criteria_docs = await db.evaluation_criteria.find({"id": {"$in": criterion_ids}}, {"_id": 0}).to_list(500)

    criteria_map = {criterion["id"]: criterion for criterion in criteria_docs}
    enriched_criteria = []
    total_weight = 0.0

    for item in sorted(plan.get("criteria", []), key=lambda value: value.get("order", 0)):
        weight = float(item.get("weight", 1.0) or 1.0)
        total_weight += weight
        enriched_criteria.append({**item, "criterion": criteria_map.get(item.get("criterion_id"))})

    plan["criteria"] = enriched_criteria
    plan["criteria_count"] = len(enriched_criteria)
    plan["total_weight"] = round(total_weight, 2)
    return plan


@api_router.post("/evaluations/plans")
async def create_evaluation_plan(plan_data: EvaluationPlanCreate, current_user: dict = Depends(get_current_user)):
    checker = await validate_evaluation_plan_access(current_user, plan_data.team_id, write=True)

    if not plan_data.name.strip():
        raise HTTPException(status_code=400, detail="Indica o nome do plano")

    criteria_dicts = await validate_evaluation_plan_criteria(plan_data.criteria, plan_data.team_id, checker)

    plan = EvaluationPlan(
        name=plan_data.name.strip(),
        description=plan_data.description.strip() if plan_data.description else None,
        category=plan_data.category,
        team_id=plan_data.team_id,
        club_id=current_user.get("club_id"),
        criteria=criteria_dicts,
        estimated_minutes=plan_data.estimated_minutes,
        is_active=plan_data.is_active,
        created_by=current_user["id"]
    )

    plan_dict = plan.model_dump()
    plan_dict["created_at"] = plan_dict["created_at"].isoformat()
    plan_dict["updated_at"] = plan_dict["updated_at"].isoformat()

    await db.evaluation_plans.insert_one(dict(plan_dict))

    plan_dict.pop("_id", None)
    return await enrich_evaluation_plan(plan_dict)


@api_router.get("/evaluations/plans")
async def get_evaluation_plans(
    team_id: Optional[str] = None,
    category: Optional[str] = None,
    include_inactive: bool = False,
    current_user: dict = Depends(get_current_user)
):
    checker = await validate_evaluation_plan_access(current_user, team_id)

    query: Dict[str, Any] = {}

    if not include_inactive:
        query["is_active"] = {"$ne": False}

    if category and category != "all":
        query["category"] = category

    if team_id:
        query["$or"] = [{"team_id": team_id}, {"team_id": None}, {"team_id": {"$exists": False}}]
    elif not checker.is_admin:
        accessible_team_ids = list(checker.team_ids)
        query["$or"] = [
            {"team_id": {"$in": accessible_team_ids}},
            {"team_id": None},
            {"team_id": {"$exists": False}},
        ]

    plans = await db.evaluation_plans.find(query, {"_id": 0}).sort("category", 1).sort("name", 1).to_list(500)
    return [await enrich_evaluation_plan(plan) for plan in plans]


@api_router.get("/evaluations/plans/{plan_id}")
async def get_evaluation_plan(plan_id: str, current_user: dict = Depends(get_current_user)):
    plan = await db.evaluation_plans.find_one({"id": plan_id}, {"_id": 0})
    if not plan:
        raise HTTPException(status_code=404, detail="Plano não encontrado")

    await validate_evaluation_plan_access(current_user, plan.get("team_id"))
    return await enrich_evaluation_plan(plan)


@api_router.put("/evaluations/plans/{plan_id}")
async def update_evaluation_plan(
    plan_id: str,
    updates: EvaluationPlanUpdate,
    current_user: dict = Depends(get_current_user)
):
    checker = get_permission_checker(current_user)

    plan = await db.evaluation_plans.find_one({"id": plan_id}, {"_id": 0})
    if not plan:
        raise HTTPException(status_code=404, detail="Plano não encontrado")

    next_team_id = updates.team_id if updates.team_id is not None else plan.get("team_id")
    await validate_evaluation_plan_access(current_user, next_team_id, write=True)

    if not checker.is_admin:
        if plan.get("created_by") != current_user.get("id") and not checker.can_access_team(plan.get("team_id")):
            raise HTTPException(status_code=403, detail="Sem permissão para editar este plano")

    update_data = updates.model_dump(exclude_unset=True)

    if "name" in update_data:
        if not update_data["name"].strip():
            raise HTTPException(status_code=400, detail="Indica o nome do plano")
        update_data["name"] = update_data["name"].strip()

    if "description" in update_data and update_data["description"]:
        update_data["description"] = update_data["description"].strip()

    if "criteria" in update_data and update_data["criteria"] is not None:
        update_data["criteria"] = await validate_evaluation_plan_criteria(update_data["criteria"], next_team_id, checker)

    if update_data:
        update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
        await db.evaluation_plans.update_one({"id": plan_id}, {"$set": update_data})

    updated = await db.evaluation_plans.find_one({"id": plan_id}, {"_id": 0})
    return await enrich_evaluation_plan(updated)


@api_router.post("/evaluations/plans/{plan_id}/duplicate")
async def duplicate_evaluation_plan(plan_id: str, current_user: dict = Depends(get_current_user)):
    plan = await db.evaluation_plans.find_one({"id": plan_id}, {"_id": 0})
    if not plan:
        raise HTTPException(status_code=404, detail="Plano não encontrado")

    await validate_evaluation_plan_access(current_user, plan.get("team_id"), write=True)

    duplicated = dict(plan)
    duplicated["id"] = str(uuid.uuid4())
    duplicated["name"] = f"{plan.get('name', 'Plano')} (cópia)"
    duplicated["created_by"] = current_user["id"]
    duplicated["created_at"] = datetime.now(timezone.utc).isoformat()
    duplicated["updated_at"] = datetime.now(timezone.utc).isoformat()
    duplicated["is_active"] = True

    await db.evaluation_plans.insert_one(dict(duplicated))

    duplicated.pop("_id", None)
    return await enrich_evaluation_plan(duplicated)


@api_router.delete("/evaluations/plans/{plan_id}")
async def archive_evaluation_plan(plan_id: str, current_user: dict = Depends(get_current_user)):
    checker = get_permission_checker(current_user)

    plan = await db.evaluation_plans.find_one({"id": plan_id}, {"_id": 0})
    if not plan:
        raise HTTPException(status_code=404, detail="Plano não encontrado")

    if not checker.is_admin:
        if plan.get("created_by") != current_user.get("id") and not checker.can_access_team(plan.get("team_id")):
            raise HTTPException(status_code=403, detail="Sem permissão para arquivar este plano")

    await db.evaluation_plans.update_one(
        {"id": plan_id},
        {"$set": {"is_active": False, "updated_at": datetime.now(timezone.utc).isoformat()}}
    )

    return {"message": "Plano arquivado"}




# ==================== EVALUATION EXECUTION ROUTES — Sprint 4.2.4.1 ====================

async def get_team_players_for_evaluation(team_id: str, current_user: dict) -> List[dict]:
    checker = get_permission_checker(current_user)

    if not checker.is_admin and not checker.can_access_team(team_id):
        raise HTTPException(status_code=403, detail="Sem acesso a esta equipa")

    players = await db.users.find(
        {
            "team_ids": team_id,
            "role": {"$in": ["jogador", "atleta", "player"]}
        },
        {
            "_id": 0,
            "hashed_password": 0,
            "password": 0,
            "reset_token": 0,
            "reset_token_expires": 0,
        }
    ).sort("name", 1).to_list(1000)

    return players


def normalize_plan_scores_for_player(
    scores: List[EvaluationFromPlanScore],
    plan: dict,
    criteria_map: Dict[str, dict]
) -> List[dict]:
    plan_criterion_ids = {
        item.get("criterion_id")
        for item in plan.get("criteria", [])
        if item.get("criterion_id")
    }

    score_dicts = []
    received_ids = set()

    for item in scores or []:
        if item.criterion_id not in plan_criterion_ids:
            raise HTTPException(
                status_code=400,
                detail=f"O critério {item.criterion_id} não pertence ao plano selecionado"
            )

        criterion = criteria_map.get(item.criterion_id)
        if not criterion:
            raise HTTPException(
                status_code=400,
                detail=f"Critério inválido: {item.criterion_id}"
            )

        scale_min = criterion.get("scale_min", 1)
        scale_max = criterion.get("scale_max", 5)

        if item.score < scale_min or item.score > scale_max:
            raise HTTPException(
                status_code=400,
                detail=f"Pontuação fora da escala para {criterion.get('name', item.criterion_id)}"
            )

        received_ids.add(item.criterion_id)
        score_dicts.append(item.model_dump())

    required_missing = []
    for plan_item in plan.get("criteria", []):
        if plan_item.get("required", True) and plan_item.get("criterion_id") not in received_ids:
            required_missing.append(plan_item.get("criterion_id"))

    if required_missing:
        raise HTTPException(
            status_code=400,
            detail=f"Existem critérios obrigatórios sem pontuação: {', '.join(required_missing)}"
        )

    return score_dicts


@api_router.get("/evaluations/teams/{team_id}/players")
async def get_evaluation_team_players(
    team_id: str,
    current_user: dict = Depends(get_current_user)
):
    return await get_team_players_for_evaluation(team_id, current_user)


@api_router.post("/evaluations/from-plan")
async def create_bulk_evaluations_from_plan(
    payload: BulkEvaluationFromPlanCreate,
    current_user: dict = Depends(get_current_user)
):
    checker = get_permission_checker(current_user)

    if not checker.is_staff and not checker.is_admin:
        raise HTTPException(status_code=403, detail="Sem permissão para avaliar atletas")

    if not checker.is_admin and not checker.can_access_team(payload.team_id):
        raise HTTPException(status_code=403, detail="Sem acesso a esta equipa")

    if not payload.evaluations:
        raise HTTPException(status_code=400, detail="Seleciona pelo menos um atleta para avaliar")

    plan = await db.evaluation_plans.find_one(
        {"id": payload.plan_id, "is_active": {"$ne": False}},
        {"_id": 0}
    )
    if not plan:
        raise HTTPException(status_code=404, detail="Plano de avaliação não encontrado")

    if plan.get("team_id") and plan.get("team_id") != payload.team_id:
        raise HTTPException(status_code=400, detail="Este plano pertence a outra equipa")

    if payload.event_id:
        event = await db.events.find_one({"id": payload.event_id}, {"_id": 0})
        if not event:
            raise HTTPException(status_code=404, detail="Evento não encontrado")
        if event.get("team_id") and event.get("team_id") != payload.team_id:
            raise HTTPException(status_code=400, detail="O evento pertence a outra equipa")

    team_players = await get_team_players_for_evaluation(payload.team_id, current_user)
    valid_player_ids = {player["id"] for player in team_players}

    requested_player_ids = [item.player_id for item in payload.evaluations]
    invalid_players = [player_id for player_id in requested_player_ids if player_id not in valid_player_ids]

    if invalid_players:
        raise HTTPException(
            status_code=400,
            detail=f"Atletas inválidos para esta equipa: {', '.join(invalid_players)}"
        )

    criterion_ids = [
        item.get("criterion_id")
        for item in plan.get("criteria", [])
        if item.get("criterion_id")
    ]

    criteria_docs = await db.evaluation_criteria.find(
        {"id": {"$in": criterion_ids}},
        {"_id": 0}
    ).to_list(500)

    criteria_map = {criterion["id"]: criterion for criterion in criteria_docs}

    created = []
    now_iso = datetime.now(timezone.utc).isoformat()

    for item in payload.evaluations:
        score_dicts = normalize_plan_scores_for_player(item.scores, plan, criteria_map)
        overall_score = calculate_evaluation_overall_score(score_dicts, criteria_map)

        share_with_development_circle = bool(
        item.share_with_player or
        item.share_with_guardian or
        payload.visibility in ["player", "guardian", "all"]
    )
    
        evaluation = PlayerEvaluation(
            player_id=item.player_id,
            team_id=payload.team_id,
            event_id=payload.event_id,
            period_label=payload.period_label,
            visibility=payload.visibility,
            scores=score_dicts,
            general_comment=item.general_comment,
            share_with_player=share_with_development_circle,
            share_with_guardian=share_with_development_circle,
            public_summary=item.public_summary,
            strengths=item.strengths,
            improvement_goals=item.improvement_goals,
            motivational_message=item.motivational_message,
            overall_score=overall_score,
            created_by=current_user["id"]
        )

        evaluation_dict = evaluation.model_dump()
        evaluation_dict["plan_id"] = payload.plan_id
        evaluation_dict["created_at"] = evaluation_dict["created_at"].isoformat()
        evaluation_dict["updated_at"] = evaluation_dict["updated_at"].isoformat()
        evaluation_dict["source"] = "plan"
        evaluation_dict["created_batch_at"] = now_iso

        await db.player_evaluations.insert_one(dict(evaluation_dict))

        evaluation_dict.pop("_id", None)
        created.append(evaluation_dict)

    return {
        "message": "Avaliações criadas",
        "created_count": len(created),
        "evaluations": created,
    }


@api_router.get("/evaluations/team/{team_id}/summary")
async def get_team_evaluation_summary(
    team_id: str,
    current_user: dict = Depends(get_current_user)
):
    checker = get_permission_checker(current_user)

    if not checker.is_admin and not checker.can_access_team(team_id):
        raise HTTPException(status_code=403, detail="Sem acesso a esta equipa")

    evaluations = await db.player_evaluations.find(
        {"team_id": team_id},
        {"_id": 0}
    ).sort("created_at", -1).to_list(1000)

    players = await get_team_players_for_evaluation(team_id, current_user)
    players_map = {player["id"]: player for player in players}

    by_player: Dict[str, dict] = {}

    for evaluation in evaluations:
        player_id = evaluation.get("player_id")
        if not player_id:
            continue

        if player_id not in by_player:
            by_player[player_id] = {
                "player": players_map.get(player_id),
                "evaluations_count": 0,
                "latest_evaluation": None,
                "average_score": None,
                "scores": [],
            }

        by_player[player_id]["evaluations_count"] += 1
        if by_player[player_id]["latest_evaluation"] is None:
            by_player[player_id]["latest_evaluation"] = evaluation

        if evaluation.get("overall_score") is not None:
            by_player[player_id]["scores"].append(float(evaluation["overall_score"]))

    result = []

    for player_id, item in by_player.items():
        scores = item.pop("scores", [])
        if scores:
            item["average_score"] = round(sum(scores) / len(scores), 2)

        result.append(item)

    return {
        "team_id": team_id,
        "players_evaluated": len(result),
        "evaluations_count": len(evaluations),
        "items": result,
    }

@api_router.get("/evaluations/{evaluation_id}")
async def get_player_evaluation(
    evaluation_id: str,
    current_user: dict = Depends(get_current_user)
):
    evaluation = await db.player_evaluations.find_one({"id": evaluation_id}, {"_id": 0})
    if not evaluation:
        raise HTTPException(status_code=404, detail="Avaliação não encontrada")

    if not can_view_player_evaluations(
        current_user,
        evaluation.get("player_id"),
        evaluation.get("team_id"),
        evaluation
    ):
        raise HTTPException(status_code=403, detail="Sem permissão para ver esta avaliação")

    if is_development_circle_viewer(current_user, evaluation.get("player_id")):
        return build_public_player_evaluation_view(evaluation)

    criterion_ids = [
        item.get("criterion_id")
        for item in evaluation.get("scores", [])
        if item.get("criterion_id")
    ]

    criteria = []
    if criterion_ids:
        criteria = await db.evaluation_criteria.find(
            {"id": {"$in": criterion_ids}},
            {"_id": 0}
        ).to_list(500)

    criteria_map = {criterion["id"]: criterion for criterion in criteria}

    enriched_scores = []
    for item in evaluation.get("scores", []):
        criterion = criteria_map.get(item.get("criterion_id"))
        enriched_scores.append({
            **item,
            "criterion": criterion
        })

    evaluation["scores"] = enriched_scores
    return evaluation


# ==================== UNAVAILABILITY ROUTES ====================

@api_router.get("/unavailabilities")
async def get_unavailabilities(team_id: Optional[str] = None, user_id: Optional[str] = None, current_user: dict = Depends(get_current_user)):
    """Get unavailability periods - filtered by team/user and role permissions"""
    checker = get_permission_checker(current_user)
    query = {}
    
    if user_id:
        # Specific user's unavailabilities
        if user_id == current_user['id']:
            query["user_id"] = user_id
        elif checker.is_admin or checker.is_staff:
            query["user_id"] = user_id
        else:
            raise HTTPException(status_code=403, detail="Sem permissão para ver indisponibilidades de outros utilizadores")
    elif team_id:
        # Team's unavailabilities
        if not checker.is_admin and not checker.can_access_team(team_id):
            raise HTTPException(status_code=403, detail="Sem acesso a esta equipa")
        query["team_ids"] = team_id
    elif not checker.is_admin:
        # Filter by user's accessible teams
        user_teams = list(checker.team_ids)
        if user_teams:
            query["team_ids"] = {"$in": user_teams}
        else:
            # Show only own unavailabilities
            query["user_id"] = current_user['id']
    
    unavailabilities = await db.unavailabilities.find(query, {"_id": 0}).sort("start_date", 1).to_list(200)
    
    # Enrich with user info
    for unav in unavailabilities:
        user = await db.users.find_one({"id": unav['user_id']}, {"_id": 0, "name": 1, "role": 1})
        if user:
            unav['user_name'] = user.get('name', 'Unknown')
            unav['user_role'] = user.get('role', 'jogador')
    
    return unavailabilities

@api_router.get("/unavailabilities/my")
async def get_my_unavailabilities(current_user: dict = Depends(get_current_user)):
    """Get current user's unavailabilities"""
    unavailabilities = await db.unavailabilities.find({"user_id": current_user['id']}, {"_id": 0}).sort("start_date", 1).to_list(100)
    return unavailabilities

@api_router.post("/unavailabilities")
async def create_unavailability(data: UnavailabilityCreate, current_user: dict = Depends(get_current_user)):
    """Create unavailability period - players, coaches and delegates can create their own"""
    if data.start_date >= data.end_date:
        raise HTTPException(status_code=400, detail="Data inicial deve ser anterior à data final")
    
    # Get user's team IDs
    user = await db.users.find_one({"id": current_user['id']}, {"_id": 0})
    user_teams = user.get('team_ids', []) if user else []
    
    unavailability = Unavailability(
        user_id=current_user['id'],
        team_ids=user_teams,
        start_date=data.start_date,
        end_date=data.end_date,
        reason=data.reason,
        notes=data.notes
    )
    
    unav_dict = unavailability.model_dump()
    unav_dict['start_date'] = unav_dict['start_date'].isoformat()
    unav_dict['end_date'] = unav_dict['end_date'].isoformat()
    unav_dict['created_at'] = unav_dict['created_at'].isoformat()
    
    await db.unavailabilities.insert_one(unav_dict)
    unav_dict.pop('_id', None)
    
    # Notify coaches of the affected teams
    for team_id in user_teams:
        # Get coaches for this team
        team = await db.teams.find_one({"id": team_id}, {"_id": 0})
        if team:
            coach_ids = team.get('coach_ids', [])
            # Also get users with coach role assigned to this team
            coaches = await db.users.find({
                "$or": [
                    {"id": {"$in": coach_ids}},
                    {"team_ids": team_id, "role": {"$in": ["treinador", "treinador_adjunto"]}}
                ]
            }, {"_id": 0, "id": 1}).to_list(20)
            
            coach_user_ids = [c['id'] for c in coaches if c['id'] != current_user['id']]
            
            if coach_user_ids:
                reason_labels = {
                    'ferias': 'Férias',
                    'doenca': 'Doença/Consulta Médica',
                    'escola': 'Atividades Escolares',
                    'outro': 'Outro Motivo'
                }
                reason_label = reason_labels.get(data.reason, data.reason)
                
                # Send push notification
                try:
                    await send_push_to_users(
                        user_ids=coach_user_ids,
                        title="Jogador Indisponível",
                        body=f"{current_user.get('name', 'Jogador')} está indisponível ({reason_label}) de {data.start_date.strftime('%d/%m')} a {data.end_date.strftime('%d/%m')}",
                        url="/attendance"
                    )
                except Exception as e:
                    logging.error(f"Failed to notify coaches of unavailability: {e}")
                
                # Send email to coaches
                try:
                    for coach in coaches:
                        if coach['id'] != current_user['id']:
                            coach_data = await db.users.find_one({"id": coach['id']}, {"_id": 0, "email": 1, "name": 1})
                            if coach_data and coach_data.get('email'):
                                email_content = f"""
                                    <p>Olá <strong>{coach_data.get('name', 'Treinador')}</strong>,</p>
                                    <p>O atleta <strong>{current_user.get('name', 'Jogador')}</strong> registou um período de indisponibilidade:</p>
                                    <table style="margin: 20px 0; border-collapse: collapse;">
                                        <tr style="background-color: #fef3c7;">
                                            <td style="padding: 12px 16px; border: 1px solid #fcd34d; font-weight: 600;">Período</td>
                                            <td style="padding: 12px 16px; border: 1px solid #fcd34d;">{data.start_date.strftime('%d/%m/%Y')} a {data.end_date.strftime('%d/%m/%Y')}</td>
                                        </tr>
                                        <tr>
                                            <td style="padding: 12px 16px; border: 1px solid #e5e7eb; font-weight: 600;">Motivo</td>
                                            <td style="padding: 12px 16px; border: 1px solid #e5e7eb;">{reason_label}</td>
                                        </tr>
                                    </table>
                                    {f'<p><strong>Notas:</strong> {data.notes}</p>' if data.notes else ''}
                                    <p style="color: #6b7280; font-size: 14px;">O atleta não será incluído nas convocatórias durante este período.</p>
                                """
                                
                                await send_email_notification(
                                    coach_data.get('email'),
                                    f"Indisponibilidade: {current_user.get('name', 'Jogador')}",
                                    build_email_template("Atleta Indisponível", email_content)
                                )
                except Exception as e:
                    logging.error(f"Failed to send unavailability email to coaches: {e}")
    
    return unav_dict

@api_router.put("/unavailabilities/{unavailability_id}")
async def update_unavailability(unavailability_id: str, data: UnavailabilityCreate, current_user: dict = Depends(get_current_user)):
    """Update unavailability - only owner can update"""
    unavailability = await db.unavailabilities.find_one({"id": unavailability_id}, {"_id": 0})
    if not unavailability:
        raise HTTPException(status_code=404, detail="Indisponibilidade não encontrada")
    
    if unavailability['user_id'] != current_user['id']:
        checker = get_permission_checker(current_user)
        if not checker.is_admin:
            raise HTTPException(status_code=403, detail="Sem permissão para editar esta indisponibilidade")
    
    if data.start_date >= data.end_date:
        raise HTTPException(status_code=400, detail="Data inicial deve ser anterior à data final")
    
    update_data = {
        "start_date": data.start_date.isoformat(),
        "end_date": data.end_date.isoformat(),
        "reason": data.reason,
        "notes": data.notes
    }
    
    await db.unavailabilities.update_one({"id": unavailability_id}, {"$set": update_data})
    return {"message": "Indisponibilidade atualizada"}

@api_router.delete("/unavailabilities/{unavailability_id}")
async def delete_unavailability(unavailability_id: str, current_user: dict = Depends(get_current_user)):
    """Delete unavailability - only owner or admin can delete"""
    unavailability = await db.unavailabilities.find_one({"id": unavailability_id}, {"_id": 0})
    if not unavailability:
        raise HTTPException(status_code=404, detail="Indisponibilidade não encontrada")
    
    if unavailability['user_id'] != current_user['id']:
        checker = get_permission_checker(current_user)
        if not checker.is_admin:
            raise HTTPException(status_code=403, detail="Sem permissão para eliminar esta indisponibilidade")
    
    await db.unavailabilities.delete_one({"id": unavailability_id})
    return {"message": "Indisponibilidade eliminada"}

@api_router.get("/unavailabilities/check")
async def check_unavailability(player_ids: str, event_date: str, current_user: dict = Depends(get_current_user)):
    """Check if players are unavailable for a specific date - used during convocation creation"""
    checker = get_permission_checker(current_user)
    
    if not checker.can_create_convocations:
        raise HTTPException(status_code=403, detail="Sem permissão")
    
    player_id_list = player_ids.split(',')
    event_dt = datetime.fromisoformat(event_date.replace('Z', '+00:00'))
    
    # Find unavailabilities that overlap with event date
    unavailable_players = []
    
    for player_id in player_id_list:
        unavails = await db.unavailabilities.find({
            "user_id": player_id,
            "start_date": {"$lte": event_dt.isoformat()},
            "end_date": {"$gte": event_dt.isoformat()}
        }, {"_id": 0}).to_list(10)
        
        if unavails:
            player = await db.users.find_one({"id": player_id}, {"_id": 0, "name": 1})
            unavailable_players.append({
                "player_id": player_id,
                "player_name": player.get('name', 'Unknown') if player else 'Unknown',
                "unavailabilities": unavails
            })
    
    return {"unavailable_players": unavailable_players}

# ==================== EVENT REMINDER SYSTEM ====================

async def process_event_reminders():
    """
    Process events approaching without convocation and send reminders to coaches.
    This function is idempotent - it can be called multiple times safely.
    """
    now = datetime.now(timezone.utc)
    
    # Window: events starting between 3.5h and 4.5h from now
    # This gives a 1-hour window to catch events even if the job runs slightly off schedule
    window_start = now + timedelta(hours=3, minutes=30)
    window_end = now + timedelta(hours=4, minutes=30)
    
    logging.info(f"Processing event reminders for events between {window_start} and {window_end}")
    
    # Find upcoming events without convocation in the window
    events = await db.events.find({
        "start_time": {
            "$gte": window_start.isoformat(),
            "$lte": window_end.isoformat()
        },
        "status": "scheduled"
    }, {"_id": 0}).to_list(100)
    
    reminders_sent = 0
    reminders_skipped = 0
    
    for event in events:
        event_id = event['id']
        team_id = event.get('team_id')
        
        if not team_id:
            continue
        
        # Check if this event already has a convocation
        convocation = await db.convocations.find_one({"event_id": event_id}, {"_id": 0})
        if convocation:
            reminders_skipped += 1
            continue
        
        # Check if reminder was already sent for this event
        existing_reminder = await db.event_reminders.find_one({
            "event_id": event_id,
            "reminder_type": "no_convocation_4h"
        }, {"_id": 0})
        
        if existing_reminder:
            reminders_skipped += 1
            continue
        
        # Get coaches for this team
        team = await db.teams.find_one({"id": team_id}, {"_id": 0})
        if not team:
            continue
        
        coach_ids = team.get('coach_ids', [])
        
        # Also get users with coach role assigned to this team
        coaches = await db.users.find({
            "$or": [
                {"id": {"$in": coach_ids}},
                {"team_ids": team_id, "role": {"$in": ["treinador", "treinador_adjunto"]}}
            ]
        }, {"_id": 0, "id": 1, "email": 1, "name": 1}).to_list(20)
        
        if not coaches:
            logging.warning(f"No coaches found for team {team_id}, skipping reminder for event {event_id}")
            continue
        
        coach_user_ids = list(set([c['id'] for c in coaches]))
        
        # Parse event time for display
        event_time = event['start_time']
        if isinstance(event_time, str):
            event_time = datetime.fromisoformat(event_time.replace('Z', '+00:00'))
        
        # Send push notification to coaches
        try:
            await send_push_to_users(
                user_ids=coach_user_ids,
                title="⚠️ Evento sem Convocatória!",
                body=f"{event.get('title', 'Evento')} começa às {event_time.strftime('%H:%M')} e ainda não tem convocatória",
                url="/calendar"
            )
        except Exception as e:
            logging.error(f"Failed to send push reminder for event {event_id}: {e}")
        
        # Send email notification to coaches
        for coach in coaches:
            try:
                email_content = f"""
                    <p>Olá <strong>{coach.get('name', 'Treinador')}</strong>,</p>
                    <p>O evento <strong>{event.get('title', 'Evento')}</strong> começa dentro de aproximadamente 4 horas e ainda não tem convocatória criada.</p>
                    <div style="margin: 20px 0; padding: 16px; background-color: #fef3c7; border-radius: 8px; border-left: 4px solid #f59e0b;">
                        <p style="margin: 0 0 8px 0; font-weight: 600; color: #92400e;">⚠️ Ação Necessária</p>
                        <p style="margin: 0; color: #78350f;">Por favor, crie a convocatória para que os jogadores possam confirmar presença.</p>
                    </div>
                    <table style="margin: 20px 0; border-collapse: collapse; width: 100%;">
                        <tr style="background-color: #f8fafc;">
                            <td style="padding: 12px; border: 1px solid #e5e7eb; font-weight: 600;">Evento</td>
                            <td style="padding: 12px; border: 1px solid #e5e7eb;">{event.get('title', 'Evento')}</td>
                        </tr>
                        <tr>
                            <td style="padding: 12px; border: 1px solid #e5e7eb; font-weight: 600;">Data/Hora</td>
                            <td style="padding: 12px; border: 1px solid #e5e7eb;">{event_time.strftime('%d/%m/%Y às %H:%M')}</td>
                        </tr>
                        <tr style="background-color: #f8fafc;">
                            <td style="padding: 12px; border: 1px solid #e5e7eb; font-weight: 600;">Tipo</td>
                            <td style="padding: 12px; border: 1px solid #e5e7eb;">{event.get('event_type', 'Outro')}</td>
                        </tr>
                        <tr>
                            <td style="padding: 12px; border: 1px solid #e5e7eb; font-weight: 600;">Local</td>
                            <td style="padding: 12px; border: 1px solid #e5e7eb;">{event.get('location', 'N/A')}</td>
                        </tr>
                        <tr style="background-color: #f8fafc;">
                            <td style="padding: 12px; border: 1px solid #e5e7eb; font-weight: 600;">Equipa</td>
                            <td style="padding: 12px; border: 1px solid #e5e7eb;">{team.get('name', 'N/A')}</td>
                        </tr>
                    </table>
                """
                
                await send_email_notification(
                    coach.get('email'),
                    f"⚠️ Lembrete: {event.get('title', 'Evento')} sem convocatória",
                    build_email_template("Evento sem Convocatória", email_content)
                )
            except Exception as e:
                logging.error(f"Failed to send email reminder to {coach.get('email')}: {e}")
        
        # Record that reminder was sent
        reminder = EventReminder(
            event_id=event_id,
            team_id=team_id,
            reminder_type="no_convocation_4h",
            notified_user_ids=coach_user_ids
        )
        reminder_dict = reminder.model_dump()
        reminder_dict['sent_at'] = reminder_dict['sent_at'].isoformat()
        await db.event_reminders.insert_one(reminder_dict)
        
        reminders_sent += 1
        logging.info(f"Sent reminder for event {event_id} to {len(coach_user_ids)} coach(es)")
    
    return {
        "processed": len(events),
        "reminders_sent": reminders_sent,
        "reminders_skipped": reminders_skipped
    }

@api_router.post("/reminders/process")
async def trigger_reminder_processing(current_user: dict = Depends(get_current_user)):
    """
    Manually trigger reminder processing.
    Admin only - this endpoint can be called by a cron job or scheduler.
    """
    checker = get_permission_checker(current_user)
    
    if not checker.is_admin:
        raise HTTPException(status_code=403, detail="Apenas administradores podem processar lembretes")
    
    result = await process_event_reminders()
    return result

@api_router.get("/reminders/status")
async def get_reminder_status(event_id: Optional[str] = None, current_user: dict = Depends(get_current_user)):
    """Get reminder status for events"""
    checker = get_permission_checker(current_user)
    
    if not checker.is_staff and not checker.is_admin:
        raise HTTPException(status_code=403, detail="Sem permissão")
    
    query = {}
    if event_id:
        query["event_id"] = event_id
    
    reminders = await db.event_reminders.find(query, {"_id": 0}).sort("sent_at", -1).to_list(50)
    return reminders

@api_router.get("/reminders/pending")
async def get_pending_reminders(current_user: dict = Depends(get_current_user)):
    """
    Get events that will need reminders soon (next 6 hours) but don't have convocations yet.
    Useful for coaches to see what events need attention.
    """
    checker = get_permission_checker(current_user)
    
    if not checker.can_create_convocations:
        return []
    
    now = datetime.now(timezone.utc)
    next_6h = now + timedelta(hours=6)
    
    # Build query based on user's teams
    query = {
        "start_time": {"$gte": now.isoformat(), "$lte": next_6h.isoformat()},
        "status": "scheduled"
    }
    
    if not checker.is_admin:
        user_teams = list(checker.team_ids)
        if not user_teams:
            return []
        query["team_id"] = {"$in": user_teams}
    
    events = await db.events.find(query, {"_id": 0}).to_list(50)
    
    # Filter events without convocations and check reminder status
    result = []
    for event in events:
        convocation = await db.convocations.find_one({"event_id": event['id']}, {"_id": 0})
        if not convocation:
            # Check if reminder was already sent
            reminder = await db.event_reminders.find_one({
                "event_id": event['id'],
                "reminder_type": "no_convocation_4h"
            }, {"_id": 0})
            
            event['reminder_sent'] = reminder is not None
            event['reminder_sent_at'] = reminder.get('sent_at') if reminder else None
            result.append(event)
    
    return result

# ==================== PAYMENTS AND MONTHLY FEES ROUTES ====================

def get_payment_status(due_date: datetime, paid_at: Optional[datetime]) -> str:
    """Calculate payment status based on due date and payment date"""
    if paid_at:
        return "paid"
    now = datetime.now(timezone.utc)
    if isinstance(due_date, str):
        due_date = datetime.fromisoformat(due_date.replace('Z', '+00:00'))
    if due_date.tzinfo is None:
        due_date = due_date.replace(tzinfo=timezone.utc)
    if now > due_date:
        return "overdue"
    return "pending"

@api_router.get("/payments/my")
async def get_my_payments(season: Optional[str] = None, current_user: dict = Depends(get_current_user)):
    """Get current user's payments - accessible by all roles for their own data"""
    query = {"user_id": current_user['id']}
    
    # Get monthly fees
    fees = await db.monthly_fees.find(query, {"_id": 0}).sort("due_date", -1).to_list(100)
    
    # Get custom payments
    custom = await db.custom_payments.find(query, {"_id": 0}).sort("due_date", -1).to_list(100)
    
    # Update status for each payment
    for fee in fees:
        fee['status'] = get_payment_status(fee.get('due_date'), fee.get('paid_at'))
        fee['type'] = 'monthly_fee'
    
    for payment in custom:
        payment['status'] = get_payment_status(payment.get('due_date'), payment.get('paid_at'))
        payment['type'] = 'custom'
    
    # Combine and sort by due_date
    all_payments = fees + custom
    all_payments.sort(key=lambda x: x.get('due_date', ''), reverse=True)
    
    return all_payments

@api_router.get("/payments/status")
async def get_my_payment_status(current_user: dict = Depends(get_current_user)):
    """Get overall payment status for dashboard indicator"""
    now = datetime.now(timezone.utc)
    
    # Check if user has payments disabled
    user = await db.users.find_one({"id": current_user['id']}, {"_id": 0})
    if user and user.get('payments_disabled'):
        return {"status": "disabled", "message": "Pagamentos desativados"}
    
    # Get unpaid fees
    unpaid_fees = await db.monthly_fees.find({
        "user_id": current_user['id'],
        "paid_at": None
    }, {"_id": 0}).to_list(50)
    
    # Get unpaid custom payments
    unpaid_custom = await db.custom_payments.find({
        "user_id": current_user['id'],
        "paid_at": None
    }, {"_id": 0}).to_list(50)
    
    overdue_count = 0
    pending_count = 0
    total_overdue = 0
    total_pending = 0
    
    for fee in unpaid_fees:
        due_date = fee.get('due_date')
        if isinstance(due_date, str):
            due_date = datetime.fromisoformat(due_date.replace('Z', '+00:00'))
        if due_date and due_date.tzinfo is None:
            due_date = due_date.replace(tzinfo=timezone.utc)
        if due_date and now > due_date:
            overdue_count += 1
            total_overdue += fee.get('amount', 0)
        else:
            pending_count += 1
            total_pending += fee.get('amount', 0)
    
    for payment in unpaid_custom:
        due_date = payment.get('due_date')
        if isinstance(due_date, str):
            due_date = datetime.fromisoformat(due_date.replace('Z', '+00:00'))
        if due_date and due_date.tzinfo is None:
            due_date = due_date.replace(tzinfo=timezone.utc)
        if due_date and now > due_date:
            overdue_count += 1
            total_overdue += payment.get('amount', 0)
        else:
            pending_count += 1
            total_pending += payment.get('amount', 0)
    
    if overdue_count > 0:
        return {
            "status": "overdue",
            "overdue_count": overdue_count,
            "pending_count": pending_count,
            "total_overdue": total_overdue,
            "total_pending": total_pending
        }
    elif pending_count > 0:
        return {
            "status": "pending",
            "pending_count": pending_count,
            "total_pending": total_pending
        }
    else:
        return {"status": "paid", "message": "Todos os pagamentos em dia"}

@api_router.get("/payments/admin")
async def get_all_payments(user_id: Optional[str] = None, status: Optional[str] = None, current_user: dict = Depends(get_current_user)):
    """Get all payments - admin only"""
    checker = get_permission_checker(current_user)
    
    if not checker.is_admin:
        raise HTTPException(status_code=403, detail="Apenas administradores podem aceder a todos os pagamentos")
    
    fee_query = {}
    custom_query = {}
    
    if user_id:
        fee_query["user_id"] = user_id
        custom_query["user_id"] = user_id
    
    # Get monthly fees
    fees = await db.monthly_fees.find(fee_query, {"_id": 0}).sort("due_date", -1).to_list(500)
    
    # Get custom payments
    custom = await db.custom_payments.find(custom_query, {"_id": 0}).sort("due_date", -1).to_list(500)
    
    # Enrich with user info and update status
    user_cache = {}
    for fee in fees:
        fee['status'] = get_payment_status(fee.get('due_date'), fee.get('paid_at'))
        fee['type'] = 'monthly_fee'
        uid = fee.get('user_id')
        if uid not in user_cache:
            user = await db.users.find_one({"id": uid}, {"_id": 0, "name": 1, "email": 1})
            user_cache[uid] = user
        if user_cache.get(uid):
            fee['user_name'] = user_cache[uid].get('name', 'Unknown')
            fee['user_email'] = user_cache[uid].get('email', '')
    
    for payment in custom:
        payment['status'] = get_payment_status(payment.get('due_date'), payment.get('paid_at'))
        payment['type'] = 'custom'
        uid = payment.get('user_id')
        if uid not in user_cache:
            user = await db.users.find_one({"id": uid}, {"_id": 0, "name": 1, "email": 1})
            user_cache[uid] = user
        if user_cache.get(uid):
            payment['user_name'] = user_cache[uid].get('name', 'Unknown')
            payment['user_email'] = user_cache[uid].get('email', '')
    
    # Filter by status if specified
    all_payments = fees + custom
    if status:
        all_payments = [p for p in all_payments if p.get('status') == status]
    
    all_payments.sort(key=lambda x: x.get('due_date', ''), reverse=True)
    
    return all_payments

@api_router.get("/payments/export")
async def export_payments_excel(
    status: Optional[str] = Query(None, description="Filter by status: paid, pending, overdue"),
    payment_type: Optional[str] = Query(None, description="Filter by type: monthly_fee, custom"),
    user_id: Optional[str] = Query(None, description="Filter by user ID"),
    team_id: Optional[str] = Query(None, description="Filter by team ID"),
    season: Optional[str] = Query(None, description="Filter by season (e.g., 2025/2026)"),
    due_date_from: Optional[str] = Query(None, description="Due date from (ISO format)"),
    due_date_to: Optional[str] = Query(None, description="Due date to (ISO format)"),
    paid_date_from: Optional[str] = Query(None, description="Payment date from (ISO format)"),
    paid_date_to: Optional[str] = Query(None, description="Payment date to (ISO format)"),
    search: Optional[str] = Query(None, description="Search by player name or email"),
    current_user: dict = Depends(get_current_user)
):
    """Export payments to Excel file - admin only"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    
    checker = get_permission_checker(current_user)
    
    if not checker.is_admin:
        raise HTTPException(status_code=403, detail="Apenas administradores podem exportar pagamentos")
    
    # Build queries for both collections
    fee_query = {}
    custom_query = {}
    
    if user_id:
        fee_query["user_id"] = user_id
        custom_query["user_id"] = user_id
    
    # Get monthly fees
    fees = await db.monthly_fees.find(fee_query, {"_id": 0}).sort("due_date", -1).to_list(1000)
    
    # Get custom payments
    custom = await db.custom_payments.find(custom_query, {"_id": 0}).sort("due_date", -1).to_list(1000)
    
    # Get all users for enrichment
    users = await db.users.find({}, {"_id": 0, "id": 1, "name": 1, "email": 1, "profile": 1, "team_ids": 1}).to_list(1000)
    user_map = {u["id"]: u for u in users}
    
    # Get all teams for team names
    teams = await db.teams.find({}, {"_id": 0, "id": 1, "name": 1, "season": 1}).to_list(100)
    team_map = {t["id"]: t for t in teams}
    
    # Process and enrich payments
    all_payments = []
    
    for fee in fees:
        fee['type'] = 'monthly_fee'
        fee['type_display'] = 'Mensalidade'
        fee['status'] = get_payment_status(fee.get('due_date'), fee.get('paid_at'))
        fee['description'] = f"Mensalidade {fee.get('month', '')}/{fee.get('year', '')}"
        
        user = user_map.get(fee.get('user_id'))
        if user:
            fee['user_name'] = user.get('name', 'Unknown')
            fee['user_email'] = user.get('email', '')
            fee['date_of_birth'] = user.get('profile', {}).get('identity', {}).get('birth_date', '')
            # Get team info
            team_ids = user.get('team_ids', [])
            team_names = [team_map.get(tid, {}).get('name', '') for tid in team_ids if team_map.get(tid)]
            fee['team'] = ', '.join(team_names) if team_names else ''
            team_seasons = [team_map.get(tid, {}).get('season', '') for tid in team_ids if team_map.get(tid)]
            fee['season'] = team_seasons[0] if team_seasons else ''
        
        all_payments.append(fee)
    
    for payment in custom:
        payment['type'] = 'custom'
        payment['type_display'] = 'Pagamento Personalizado'
        payment['status'] = get_payment_status(payment.get('due_date'), payment.get('paid_at'))
        
        user = user_map.get(payment.get('user_id'))
        if user:
            payment['user_name'] = user.get('name', 'Unknown')
            payment['user_email'] = user.get('email', '')
            payment['date_of_birth'] = user.get('profile', {}).get('identity', {}).get('birth_date', '')
            team_ids = user.get('team_ids', [])
            team_names = [team_map.get(tid, {}).get('name', '') for tid in team_ids if team_map.get(tid)]
            payment['team'] = ', '.join(team_names) if team_names else ''
            team_seasons = [team_map.get(tid, {}).get('season', '') for tid in team_ids if team_map.get(tid)]
            payment['season'] = team_seasons[0] if team_seasons else ''
        
        all_payments.append(payment)
    
    # Apply filters
    if status:
        all_payments = [p for p in all_payments if p.get('status') == status]
    
    if payment_type:
        all_payments = [p for p in all_payments if p.get('type') == payment_type]
    
    if team_id:
        team_name = team_map.get(team_id, {}).get('name', '')
        if team_name:
            all_payments = [p for p in all_payments if team_name in p.get('team', '')]
    
    if season:
        all_payments = [p for p in all_payments if p.get('season') == season or season in p.get('description', '')]
    
    if due_date_from:
        all_payments = [p for p in all_payments if p.get('due_date', '') >= due_date_from]
    
    if due_date_to:
        all_payments = [p for p in all_payments if p.get('due_date', '') <= due_date_to]
    
    if paid_date_from:
        all_payments = [p for p in all_payments if p.get('paid_at') and p.get('paid_at', '') >= paid_date_from]
    
    if paid_date_to:
        all_payments = [p for p in all_payments if p.get('paid_at') and p.get('paid_at', '') <= paid_date_to]
    
    if search:
        search_lower = search.lower()
        all_payments = [p for p in all_payments if 
                       search_lower in p.get('user_name', '').lower() or 
                       search_lower in p.get('user_email', '').lower()]
    
    # Sort by due date descending
    all_payments.sort(key=lambda x: x.get('due_date', ''), reverse=True)
    
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Pagamentos"
    
    # Define headers
    headers = [
        "Nome do Jogador",
        "Data de Nascimento",
        "Equipa",
        "Época",
        "Tipo de Pagamento",
        "Descrição",
        "Valor (€)",
        "Data de Criação",
        "Data de Vencimento",
        "Estado",
        "Data de Pagamento",
        "Comprovativo",
        "Notas"
    ]
    
    # Header styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="0D9488", end_color="0D9488", fill_type="solid")  # Teal color
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Status translations
    status_map = {
        'paid': 'Pago',
        'pending': 'Pendente',
        'overdue': 'Atrasado'
    }
    
    # Write data
    for row_idx, payment in enumerate(all_payments, 2):
        # Format dates
        due_date = payment.get('due_date', '')
        if due_date:
            try:
                dt = datetime.fromisoformat(due_date.replace('Z', '+00:00'))
                due_date = dt.strftime('%d/%m/%Y')
            except:
                pass
        
        created_at = payment.get('created_at', '')
        if created_at:
            try:
                dt = datetime.fromisoformat(created_at.replace('Z', '+00:00'))
                created_at = dt.strftime('%d/%m/%Y')
            except:
                pass
        
        paid_at = payment.get('paid_at', '')
        if paid_at:
            try:
                dt = datetime.fromisoformat(paid_at.replace('Z', '+00:00'))
                paid_at = dt.strftime('%d/%m/%Y')
            except:
                pass
        
        birth_date = payment.get('date_of_birth', '')
        if birth_date:
            try:
                dt = datetime.fromisoformat(birth_date.replace('Z', '+00:00'))
                birth_date = dt.strftime('%d/%m/%Y')
            except:
                pass
        
        row_data = [
            payment.get('user_name', ''),
            birth_date,
            payment.get('team', ''),
            payment.get('season', ''),
            payment.get('type_display', ''),
            payment.get('description', payment.get('title', '')),
            payment.get('amount', 0),
            created_at,
            due_date,
            status_map.get(payment.get('status'), payment.get('status', '')),
            paid_at,
            'Sim' if payment.get('proof_url') else 'Não',
            payment.get('notes', '')
        ]
        
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")
            
            # Format amount column
            if col == 7 and isinstance(value, (int, float)):
                cell.number_format = '€#,##0.00'
    
    # Adjust column widths
    column_widths = [25, 15, 20, 12, 20, 35, 12, 14, 14, 12, 14, 12, 30]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    
    # Freeze header row
    ws.freeze_panes = 'A2'
    
    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Generate filename with timestamp
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f"pagamentos_export_{timestamp}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@api_router.get("/payments/summary")
async def get_payments_summary(current_user: dict = Depends(get_current_user)):
    """Get payments summary - admin only"""
    checker = get_permission_checker(current_user)
    
    if not checker.is_admin:
        raise HTTPException(status_code=403, detail="Apenas administradores podem ver resumo de pagamentos")
    
    now = datetime.now(timezone.utc)
    
    # Get all unpaid
    unpaid_fees = await db.monthly_fees.find({"paid_at": None}, {"_id": 0}).to_list(500)
    unpaid_custom = await db.custom_payments.find({"paid_at": None}, {"_id": 0}).to_list(500)
    
    # Get paid this month
    month_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    paid_fees_month = await db.monthly_fees.find({
        "paid_at": {"$gte": month_start.isoformat()}
    }, {"_id": 0}).to_list(500)
    paid_custom_month = await db.custom_payments.find({
        "paid_at": {"$gte": month_start.isoformat()}
    }, {"_id": 0}).to_list(500)
    
    total_overdue = 0
    total_pending = 0
    overdue_count = 0
    pending_count = 0
    
    for fee in unpaid_fees:
        due_date = fee.get('due_date')
        if isinstance(due_date, str):
            due_date = datetime.fromisoformat(due_date.replace('Z', '+00:00'))
        if due_date and due_date.tzinfo is None:
            due_date = due_date.replace(tzinfo=timezone.utc)
        if due_date and now > due_date:
            overdue_count += 1
            total_overdue += fee.get('amount', 0)
        else:
            pending_count += 1
            total_pending += fee.get('amount', 0)
    
    for payment in unpaid_custom:
        due_date = payment.get('due_date')
        if isinstance(due_date, str):
            due_date = datetime.fromisoformat(due_date.replace('Z', '+00:00'))
        if due_date and due_date.tzinfo is None:
            due_date = due_date.replace(tzinfo=timezone.utc)
        if due_date and now > due_date:
            overdue_count += 1
            total_overdue += payment.get('amount', 0)
        else:
            pending_count += 1
            total_pending += payment.get('amount', 0)
    
    total_collected_month = sum(f.get('amount', 0) for f in paid_fees_month) + sum(p.get('amount', 0) for p in paid_custom_month)
    
    return {
        "overdue_count": overdue_count,
        "pending_count": pending_count,
        "total_overdue": total_overdue,
        "total_pending": total_pending,
        "collected_this_month": total_collected_month,
        "paid_count_this_month": len(paid_fees_month) + len(paid_custom_month)
    }

@api_router.post("/payments/monthly-fees")
async def create_monthly_fee(data: MonthlyFeeCreate, current_user: dict = Depends(get_current_user)):
    """Create a monthly fee - admin only"""
    checker = get_permission_checker(current_user)
    
    if not checker.is_admin:
        raise HTTPException(status_code=403, detail="Apenas administradores podem criar mensalidades")
    
    # Check if user exists
    user = await db.users.find_one({"id": data.user_id}, {"_id": 0})
    if not user:
        raise HTTPException(status_code=404, detail="Utilizador não encontrado")
    
    # Check if fee already exists for this month/year
    existing = await db.monthly_fees.find_one({
        "user_id": data.user_id,
        "month": data.month,
        "year": data.year
    })
    if existing:
        raise HTTPException(status_code=400, detail=f"Já existe mensalidade para {data.month}/{data.year}")
    
    fee = MonthlyFee(
        user_id=data.user_id,
        amount=data.amount,
        month=data.month,
        year=data.year,
        due_date=data.due_date,
        notes=data.notes,
        created_by=current_user['id']
    )
    
    fee_dict = fee.model_dump()
    fee_dict['due_date'] = fee_dict['due_date'].isoformat()
    fee_dict['created_at'] = fee_dict['created_at'].isoformat()
    
    await db.monthly_fees.insert_one(fee_dict)
    fee_dict.pop('_id', None)
    
    return fee_dict

@api_router.post("/payments/monthly-fees/bulk")
async def create_monthly_fees_bulk(month: int, year: int, amount: float, due_date: datetime, user_ids: List[str] = None, current_user: dict = Depends(get_current_user)):
    """Create monthly fees for multiple users - admin only"""
    checker = get_permission_checker(current_user)
    
    if not checker.is_admin:
        raise HTTPException(status_code=403, detail="Apenas administradores podem criar mensalidades")
    
    if not user_ids:
        # Get all active players
        users = await db.users.find({
            "role": "jogador",
            "is_archived": {"$ne": True},
            "payments_disabled": {"$ne": True}
        }, {"_id": 0, "id": 1}).to_list(500)
        user_ids = [u['id'] for u in users]
    
    created = 0
    skipped = 0
    
    for uid in user_ids:
        # Check if already exists
        existing = await db.monthly_fees.find_one({
            "user_id": uid,
            "month": month,
            "year": year
        })
        if existing:
            skipped += 1
            continue
        
        # Check if user has payments disabled
        user = await db.users.find_one({"id": uid}, {"_id": 0})
        if user and user.get('payments_disabled'):
            skipped += 1
            continue
        
        fee = MonthlyFee(
            user_id=uid,
            amount=amount,
            month=month,
            year=year,
            due_date=due_date,
            created_by=current_user['id']
        )
        
        fee_dict = fee.model_dump()
        fee_dict['due_date'] = fee_dict['due_date'].isoformat()
        fee_dict['created_at'] = fee_dict['created_at'].isoformat()
        
        await db.monthly_fees.insert_one(fee_dict)
        created += 1
    
    return {"message": f"Mensalidades criadas: {created}, ignoradas: {skipped}"}

@api_router.post("/payments/monthly-fees/import")
async def import_monthly_fees(file: UploadFile = File(...), current_user: dict = Depends(get_current_user)):
    """Import monthly fees from Excel - admin only"""
    checker = get_permission_checker(current_user)
    
    if not checker.is_admin:
        raise HTTPException(status_code=403, detail="Apenas administradores podem importar mensalidades")
    
    if not file.filename.endswith(('.xlsx', '.xls', '.csv')):
        raise HTTPException(status_code=400, detail="Formato não suportado. Use Excel ou CSV")
    
    try:
        content = await file.read()
        
        if file.filename.endswith('.csv'):
            df = pd.read_csv(io.BytesIO(content))
        else:
            df = pd.read_excel(io.BytesIO(content))
        
        df.columns = [str(col).lower().strip().replace(' ', '_') for col in df.columns]
        
        created = 0
        errors = []
        
        for idx, row in df.iterrows():
            try:
                # Find user by email or name
                email = None
                user = None
                
                for col in ['email', 'e-mail', 'correio']:
                    if col in df.columns and pd.notna(row.get(col)):
                        email = str(row[col]).strip()
                        break
                
                if email:
                    user = await db.users.find_one({"email": email}, {"_id": 0})
                
                if not user:
                    errors.append(f"Linha {idx + 2}: Utilizador não encontrado")
                    continue
                
                # Get amount
                amount = None
                for col in ['valor', 'amount', 'montante', 'quantia']:
                    if col in df.columns and pd.notna(row.get(col)):
                        amount = float(row[col])
                        break
                
                if not amount:
                    errors.append(f"Linha {idx + 2}: Valor não encontrado")
                    continue
                
                # Get month/year
                month = None
                year = None
                for col in ['mes', 'month', 'mês']:
                    if col in df.columns and pd.notna(row.get(col)):
                        month = int(row[col])
                        break
                
                for col in ['ano', 'year']:
                    if col in df.columns and pd.notna(row.get(col)):
                        year = int(row[col])
                        break
                
                if not month or not year:
                    errors.append(f"Linha {idx + 2}: Mês ou ano não encontrado")
                    continue
                
                # Get due date
                due_date = None
                for col in ['vencimento', 'due_date', 'data_limite']:
                    if col in df.columns and pd.notna(row.get(col)):
                        due_date = pd.to_datetime(row[col])
                        break
                
                if not due_date:
                    # Default to last day of month
                    if month == 12:
                        due_date = datetime(year + 1, 1, 1) - timedelta(days=1)
                    else:
                        due_date = datetime(year, month + 1, 1) - timedelta(days=1)
                
                # Check if exists
                existing = await db.monthly_fees.find_one({
                    "user_id": user['id'],
                    "month": month,
                    "year": year
                })
                if existing:
                    errors.append(f"Linha {idx + 2}: Já existe mensalidade para {month}/{year}")
                    continue
                
                fee = MonthlyFee(
                    user_id=user['id'],
                    amount=amount,
                    month=month,
                    year=year,
                    due_date=due_date,
                    created_by=current_user['id']
                )
                
                fee_dict = fee.model_dump()
                fee_dict['due_date'] = fee_dict['due_date'].isoformat()
                fee_dict['created_at'] = fee_dict['created_at'].isoformat()
                
                await db.monthly_fees.insert_one(fee_dict)
                created += 1
                
            except Exception as e:
                errors.append(f"Linha {idx + 2}: {str(e)}")
        
        return {
            "message": f"Importação concluída: {created} mensalidades criadas",
            "created": created,
            "errors": errors[:10]
        }
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Erro ao processar ficheiro: {str(e)}")

@api_router.post("/payments/custom")
async def create_custom_payment(data: CustomPaymentCreate, current_user: dict = Depends(get_current_user)):
    """Create a custom payment/charge - admin only"""
    checker = get_permission_checker(current_user)
    
    if not checker.is_admin:
        raise HTTPException(status_code=403, detail="Apenas administradores podem criar pagamentos")
    
    # Check if user exists
    user = await db.users.find_one({"id": data.user_id}, {"_id": 0})
    if not user:
        raise HTTPException(status_code=404, detail="Utilizador não encontrado")
    
    payment = CustomPayment(
        user_id=data.user_id,
        title=data.title,
        description=data.description,
        amount=data.amount,
        due_date=data.due_date,
        created_by=current_user['id']
    )
    
    payment_dict = payment.model_dump()
    payment_dict['due_date'] = payment_dict['due_date'].isoformat()
    payment_dict['created_at'] = payment_dict['created_at'].isoformat()
    
    await db.custom_payments.insert_one(payment_dict)
    payment_dict.pop('_id', None)
    
    # Notify the user via push and email
    try:
        await send_push_to_users(
            user_ids=[data.user_id],
            title="Novo Pagamento",
            body=f"Foi criado um novo pagamento: {data.title} - €{data.amount:.2f}",
            url="/payments"
        )
    except Exception as e:
        logging.error(f"Failed to notify user of new payment: {e}")
    
    # Send email notification
    try:
        due_date_str = data.due_date.strftime('%d/%m/%Y') if hasattr(data.due_date, 'strftime') else str(data.due_date)[:10]
        email_content = f"""
            <p>Olá <strong>{user.get('name', 'Atleta')}</strong>,</p>
            <p>Foi criado um novo pagamento para ti:</p>
            <table style="margin: 20px 0; border-collapse: collapse; width: 100%;">
                <tr style="background-color: #f8fafc;">
                    <td style="padding: 12px; border: 1px solid #e5e7eb; font-weight: 600;">Descrição</td>
                    <td style="padding: 12px; border: 1px solid #e5e7eb;">{data.title}</td>
                </tr>
                <tr>
                    <td style="padding: 12px; border: 1px solid #e5e7eb; font-weight: 600;">Valor</td>
                    <td style="padding: 12px; border: 1px solid #e5e7eb; font-size: 18px; color: #0f172a;"><strong>€{data.amount:.2f}</strong></td>
                </tr>
                <tr style="background-color: #f8fafc;">
                    <td style="padding: 12px; border: 1px solid #e5e7eb; font-weight: 600;">Vencimento</td>
                    <td style="padding: 12px; border: 1px solid #e5e7eb;">{due_date_str}</td>
                </tr>
            </table>
            {f'<p><strong>Detalhes:</strong> {data.description}</p>' if data.description else ''}
            <p style="margin-top: 20px;">Podes aceder à app para ver todos os teus pagamentos e carregar o comprovativo.</p>
        """
        
        await send_email_notification(
            user.get('email'),
            f"Novo Pagamento: {data.title}",
            build_email_template("Novo Pagamento Criado", email_content)
        )
    except Exception as e:
        logging.error(f"Failed to send payment email: {e}")
    
    return payment_dict

@api_router.put("/payments/{payment_type}/{payment_id}/mark-paid")
async def mark_payment_as_paid(payment_type: str, payment_id: str, current_user: dict = Depends(get_current_user)):
    """Mark a payment as paid - admin only"""
    checker = get_permission_checker(current_user)
    
    if not checker.is_admin:
        raise HTTPException(status_code=403, detail="Apenas administradores podem marcar pagamentos como pagos")
    
    collection = db.monthly_fees if payment_type == "monthly_fee" else db.custom_payments
    
    payment = await collection.find_one({"id": payment_id}, {"_id": 0})
    if not payment:
        raise HTTPException(status_code=404, detail="Pagamento não encontrado")
    
    now = datetime.now(timezone.utc)
    await collection.update_one(
        {"id": payment_id},
        {"$set": {"paid_at": now.isoformat(), "status": "paid"}}
    )
    
    # Send confirmation email to the user
    try:
        user = await db.users.find_one({"id": payment.get('user_id')}, {"_id": 0})
        if user and user.get('email'):
            # Get payment description
            if payment_type == "monthly_fee":
                months = ['Jan', 'Fev', 'Mar', 'Abr', 'Mai', 'Jun', 'Jul', 'Ago', 'Set', 'Out', 'Nov', 'Dez']
                month_name = months[payment.get('month', 1) - 1]
                payment_desc = f"Mensalidade {month_name}/{payment.get('year', '')}"
            else:
                payment_desc = payment.get('title', 'Pagamento')
            
            email_content = f"""
                <p>Olá <strong>{user.get('name', 'Atleta')}</strong>,</p>
                <p>Confirmamos que o teu pagamento foi registado com sucesso:</p>
                <div style="margin: 20px 0; padding: 20px; background-color: #ecfdf5; border-radius: 8px; border-left: 4px solid #10b981;">
                    <p style="margin: 0 0 8px 0; font-size: 14px; color: #065f46;">Pagamento Confirmado</p>
                    <p style="margin: 0; font-size: 18px; font-weight: 600; color: #047857;">{payment_desc}</p>
                    <p style="margin: 8px 0 0 0; font-size: 24px; font-weight: 700; color: #059669;">€{payment.get('amount', 0):.2f}</p>
                </div>
                <p style="color: #6b7280; font-size: 14px;">Data de confirmação: {now.strftime('%d/%m/%Y às %H:%M')}</p>
                <p style="margin-top: 20px;">Obrigado!</p>
            """
            
            await send_email_notification(
                user.get('email'),
                f"Pagamento Confirmado: {payment_desc}",
                build_email_template("Pagamento Confirmado ✓", email_content)
            )
    except Exception as e:
        logging.error(f"Failed to send payment confirmation email: {e}")
    
    return {"message": "Pagamento marcado como pago"}

@api_router.put("/payments/{payment_type}/{payment_id}/upload-proof")
async def upload_payment_proof(payment_type: str, payment_id: str, file: UploadFile = File(...), current_user: dict = Depends(get_current_user)):
    """Upload payment proof - owner or admin"""
    collection = db.monthly_fees if payment_type == "monthly_fee" else db.custom_payments
    
    payment = await collection.find_one({"id": payment_id}, {"_id": 0})
    if not payment:
        raise HTTPException(status_code=404, detail="Pagamento não encontrado")
    
    checker = get_permission_checker(current_user)
    if payment['user_id'] != current_user['id'] and not checker.is_admin:
        raise HTTPException(status_code=403, detail="Sem permissão para este pagamento")
    
    # Validate file type
    allowed_extensions = ['.pdf', '.jpg', '.jpeg', '.png', '.webp']
    file_ext = os.path.splitext(file.filename)[1].lower()
    if file_ext not in allowed_extensions:
        raise HTTPException(status_code=400, detail="Tipo de ficheiro não suportado. Use PDF, JPG ou PNG")
    
    # Read and save file
    content = await file.read()
    
    # Create uploads directory if needed
    uploads_dir = Path("/app/uploads/proofs")
    uploads_dir.mkdir(parents=True, exist_ok=True)
    
    # Generate unique filename
    filename = f"{payment_id}_{uuid.uuid4().hex[:8]}{file_ext}"
    file_path = uploads_dir / filename
    
    with open(file_path, 'wb') as f:
        f.write(content)
    
    # Store relative path
    proof_url = f"/uploads/proofs/{filename}"
    
    await collection.update_one(
        {"id": payment_id},
        {"$set": {"proof_url": proof_url, "proof_filename": file.filename}}
    )
    
    return {"message": "Comprovativo carregado", "proof_url": proof_url}

@api_router.delete("/payments/{payment_type}/{payment_id}")
async def delete_payment(payment_type: str, payment_id: str, current_user: dict = Depends(get_current_user)):
    """Delete a payment - admin only"""
    checker = get_permission_checker(current_user)
    
    if not checker.is_admin:
        raise HTTPException(status_code=403, detail="Apenas administradores podem eliminar pagamentos")
    
    collection = db.monthly_fees if payment_type == "monthly_fee" else db.custom_payments
    
    result = await collection.delete_one({"id": payment_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Pagamento não encontrado")
    
    return {"message": "Pagamento eliminado"}

@api_router.put("/users/{user_id}/payment-settings")
async def update_user_payment_settings(user_id: str, data: PaymentSettingsUpdate, current_user: dict = Depends(get_current_user)):
    """Update user payment settings - admin only"""
    checker = get_permission_checker(current_user)
    
    if not checker.is_admin:
        raise HTTPException(status_code=403, detail="Apenas administradores podem alterar definições de pagamento")
    
    user = await db.users.find_one({"id": user_id}, {"_id": 0})
    if not user:
        raise HTTPException(status_code=404, detail="Utilizador não encontrado")
    
    update_data = {}
    if data.payments_disabled is not None:
        update_data['payments_disabled'] = data.payments_disabled
    if data.default_monthly_fee is not None:
        update_data['default_monthly_fee'] = data.default_monthly_fee
    
    if update_data:
        await db.users.update_one({"id": user_id}, {"$set": update_data})
    
    return {"message": "Definições atualizadas"}

@api_router.get("/users/{user_id}/payments")
async def get_user_payments(user_id: str, current_user: dict = Depends(get_current_user)):
    """Get payments for a specific user - admin or owner only"""
    checker = get_permission_checker(current_user)
    
    if user_id != current_user['id'] and not checker.is_admin:
        raise HTTPException(status_code=403, detail="Sem permissão para ver pagamentos de outros utilizadores")
    
    # Get monthly fees
    fees = await db.monthly_fees.find({"user_id": user_id}, {"_id": 0}).sort("due_date", -1).to_list(100)
    
    # Get custom payments
    custom = await db.custom_payments.find({"user_id": user_id}, {"_id": 0}).sort("due_date", -1).to_list(100)
    
    # Update status
    for fee in fees:
        fee['status'] = get_payment_status(fee.get('due_date'), fee.get('paid_at'))
        fee['type'] = 'monthly_fee'
    
    for payment in custom:
        payment['status'] = get_payment_status(payment.get('due_date'), payment.get('paid_at'))
        payment['type'] = 'custom'
    
    all_payments = fees + custom
    all_payments.sort(key=lambda x: x.get('due_date', ''), reverse=True)
    
    return all_payments


# Background task runner for periodic reminder processing
async def start_reminder_scheduler():
    """
    Start a background task that processes reminders every 30 minutes.
    This is a simple scheduler - in production, consider using Celery or APScheduler.
    """
    while True:
        try:
            logging.info("Running scheduled reminder processing...")
            result = await process_event_reminders()
            logging.info(f"Reminder processing complete: {result}")
        except Exception as e:
            logging.error(f"Error in reminder scheduler: {e}")
        
        # Wait 30 minutes before next run
        await asyncio.sleep(30 * 60)

# Start the scheduler when the app starts
async def ensure_performance_indexes():
    """Create MongoDB indexes needed by Calendar, Convocations and Dashboard.

    Safe to run on every startup. MongoDB only creates missing indexes.
    """
    try:
        await db.events.create_index([("team_id", 1), ("start_time", 1)])
        await db.events.create_index([("team_ids", 1), ("start_time", 1)])
        await db.events.create_index([("id", 1)], unique=True)
        await db.events.create_index([("championship_id", 1)])

        await db.attendance.create_index([("event_id", 1), ("player_id", 1)])
        await db.attendance.create_index([("player_id", 1), ("event_date", -1)])
        await db.attendance.create_index([("convocation_id", 1)])

        await db.convocations.create_index([("event_id", 1), ("created_at", -1)])
        await db.convocations.create_index([("id", 1)], unique=True)

        await db.teams.create_index([("id", 1)], unique=True)

        await db.users.create_index([("id", 1)], unique=True)
        await db.users.create_index([("team_ids", 1)])
        await db.users.create_index([("associated_accounts", 1)])
        await db.users.create_index([("linked_player_ids", 1)])

        await db.messages.create_index([("team_id", 1), ("created_at", -1)])
        logger.info("Performance indexes ensured")
    except Exception as e:
        logger.error(f"Failed to ensure performance indexes: {e}")


@app.on_event("startup")
async def startup_event():
    await ensure_performance_indexes()
    # Start the reminder scheduler as a background task
    asyncio.create_task(start_reminder_scheduler())
    logging.info("Event reminder scheduler started")

# ==================== MAIN ====================

app.include_router(api_router)

# Static files for payment proofs
BASE_DIR = Path(__file__).resolve().parent
uploads_path = BASE_DIR / "uploads"
uploads_path.mkdir(parents=True, exist_ok=True)
if uploads_path.exists():
    app.mount("/uploads", StaticFiles(directory=str(uploads_path)), name="uploads")

cors_origins = [
    origin.strip()
    for origin in os.environ.get(
        "CORS_ORIGINS",
        "http://localhost:3000,https://stick-pro-projeto.vercel.app"
    ).split(",")
    if origin.strip()
]

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=cors_origins,
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
